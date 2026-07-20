"""Isolamento de sessão, IDs, JS e constantes do AgenteCompara vs Cleide."""
from __future__ import annotations

import importlib
import io
import os
import pathlib
import re
from types import SimpleNamespace

import openpyxl
import pytest

from app.agente_compara_doc_service import (
    AGENTE_COMPARA_CHAT_FLOW_TYPE,
    AGENTE_COMPARA_DOC_IDS_SESSION_KEY,
    AGENTE_COMPARA_DOMAIN,
    AGENTE_COMPARA_TEMP_TABLE_EXTRACTION_FLOW_TYPE,
    AGENTE_COMPARA_TEMP_TABLE_ID_SESSION_KEY,
    AGENTE_COMPARA_TEMPLATE_FILENAME,
    AUDIT_BATCH_SHEET_NAME,
    SOURCE_AGENT_AGENTE_COMPARA,
    TEMP_TABLE_VERSION_MARKER,
    _resolve_agente_compara_execution_id,
    agente_compara_batch_run_idempotency_key,
    agente_compara_batch_upload_idempotency_key,
    agente_compara_chat_idempotency_key,
    agente_compara_coverage_upload_idempotency_key,
    agente_compara_temp_table_extraction_idempotency_key,
    agente_compara_upload_idempotency_key,
    clear_documents_for_session,
)
from app.cleide_audit_doc_service import (
    CLEIDE_AUDIT_DOC_IDS_SESSION_KEY,
    CLEIDE_AUDIT_TEMP_TABLE_ID_SESSION_KEY,
)
from app.extensions import db
from app.models import CleitonBillingApropriacao
from app.services.agente_compara_config_service import (
    AgenteComparaConfig,
    DEFAULT_FALLBACK_MESSAGE,
)
from app.services.cleide_audit_config_service import CleideAuditConfig
from app.services.cleiton_upload_billing_service import apropriar_billing_agente_compara_operational_flow
from tests.cleiton_doc_fixtures import make_csv, make_txt, patch_cleiton_doc_cfg, patch_cleiton_doc_store
from tests.conftest import seed_cleiton_cost_config, seed_conta_franquia_cliente, seed_sistema_interno, seed_usuario


def _default_ac_cfg(**overrides):
    defaults = {
        "chat_enabled": True,
        "upload_enabled": True,
        "chat_max_history": 10,
        "document_context_max_chars": 24000,
        "max_documents_considered": 3,
        "question_max_chars": 4000,
        "fallback_message": DEFAULT_FALLBACK_MESSAGE,
        "no_documents_behavior": "allow_guided",
        "show_documents_used": True,
        "no_hallucination_instruction_enabled": True,
        "audited_file_max_bytes": None,
        "audited_file_max_rows": 2000,
    }
    defaults.update(overrides)
    return AgenteComparaConfig(**defaults)


def _default_cleide_cfg(**overrides):
    defaults = {
        "chat_enabled": True,
        "upload_enabled": True,
        "chat_max_history": 10,
        "document_context_max_chars": 24000,
        "max_documents_considered": 3,
        "question_max_chars": 4000,
        "fallback_message": DEFAULT_FALLBACK_MESSAGE,
        "no_documents_behavior": "allow_guided",
        "show_documents_used": True,
        "no_hallucination_instruction_enabled": True,
        "audited_file_max_bytes": None,
        "audited_file_max_rows": 2000,
    }
    defaults.update(overrides)
    return CleideAuditConfig(**defaults)


def _patch_ac_cfg(monkeypatch, **overrides):
    cfg = _default_ac_cfg(**overrides)
    for target in (
        "app.agente_compara_api_routes.get_agente_compara_config",
        "app.agente_compara_doc_service.get_agente_compara_config",
    ):
        monkeypatch.setattr(target, lambda _cfg=cfg: _cfg)
    return cfg


def _patch_cleide_cfg(monkeypatch, **overrides):
    cfg = _default_cleide_cfg(**overrides)
    for target in (
        "app.cleide_audit_routes.get_cleide_audit_config",
        "app.cleide_audit_doc_service.get_cleide_audit_config",
    ):
        monkeypatch.setattr(target, lambda _cfg=cfg: _cfg)
    return cfg


def _load_web_module():
    os.environ.setdefault("APP_ENV", "dev")
    os.environ.setdefault("DATABASE_URL", "postgresql://user:pass@localhost:5432/testdb")
    os.environ.setdefault("SECRET_KEY", "test-secret")
    return importlib.import_module("app.web")


def _setup_doc_env(monkeypatch, tmp_path, **cfg_overrides):
    patch_cleiton_doc_store(tmp_path, monkeypatch)
    cfg = patch_cleiton_doc_cfg(monkeypatch, **cfg_overrides)
    monkeypatch.setattr("app.agente_compara_doc_service.get_cleiton_doc_config", lambda: cfg)
    monkeypatch.setattr("app.agente_compara_api_routes.get_cleiton_doc_config", lambda: cfg)
    monkeypatch.setattr("app.cleide_audit_doc_service.get_cleiton_doc_config", lambda: cfg)
    monkeypatch.setattr("app.cleide_audit_routes.get_cleiton_doc_config", lambda: cfg)
    _patch_ac_cfg(monkeypatch)
    _patch_cleide_cfg(monkeypatch)
    return cfg


def _authorized(monkeypatch, web, *, authz=None):
    fake_user = SimpleNamespace(is_authenticated=True, conta_id=1, franquia_id=1, id=1)
    monkeypatch.setattr(web, "current_user", fake_user)
    monkeypatch.setattr("app.agente_compara_api_routes.current_user", fake_user)
    monkeypatch.setattr("app.cleide_audit_routes.current_user", fake_user)
    authz_payload = authz or {"permitido": True, "modo_operacao": "normal"}
    monkeypatch.setattr(
        web,
        "avaliar_autorizacao_operacao_por_franquia",
        lambda _u: authz_payload,
    )
    monkeypatch.setattr(
        "app.agente_compara_api_routes.avaliar_autorizacao_operacao_por_franquia",
        lambda _u: authz_payload,
    )
    monkeypatch.setattr(
        "app.cleide_audit_routes.avaliar_autorizacao_operacao_por_franquia",
        lambda _u: authz_payload,
    )


def _upload_ac(client, filename: str, content: bytes, mime: str = "text/csv"):
    return client.post(
        "/api/agente-compara/documents/upload",
        data={"file": (io.BytesIO(content), filename, mime)},
        content_type="multipart/form-data",
    )


def _upload_cleide(client, filename: str, content: bytes, mime: str = "text/plain"):
    return client.post(
        "/api/cleide-auditoria/documents/upload",
        data={"file": (io.BytesIO(content), filename, mime)},
        content_type="multipart/form-data",
    )


@pytest.fixture
def web_client(app, tmp_path, monkeypatch, ctx):
    with app.app_context():
        _setup_doc_env(monkeypatch, tmp_path)
    web = _load_web_module()
    _authorized(monkeypatch, web)
    monkeypatch.setattr(
        "app.run_agente_compara_temp_table.trigger_temp_table_extraction_for_session",
        lambda **_k: None,
    )
    monkeypatch.setattr(
        "app.run_cleide_audit_temp_table.trigger_temp_table_extraction_for_session",
        lambda **_k: None,
    )
    web.app.config["TESTING"] = True
    return web.app.test_client()


def test_session_keys_are_agente_compara_prefixed_and_distinct_from_cleide():
    assert AGENTE_COMPARA_DOC_IDS_SESSION_KEY == "agente_compara_doc_ids"
    assert AGENTE_COMPARA_TEMP_TABLE_ID_SESSION_KEY == "agente_compara_temp_table_id"
    assert AGENTE_COMPARA_DOC_IDS_SESSION_KEY.startswith("agente_compara_")
    assert AGENTE_COMPARA_DOC_IDS_SESSION_KEY != CLEIDE_AUDIT_DOC_IDS_SESSION_KEY
    assert AGENTE_COMPARA_TEMP_TABLE_ID_SESSION_KEY != CLEIDE_AUDIT_TEMP_TABLE_ID_SESSION_KEY
    assert not AGENTE_COMPARA_DOC_IDS_SESSION_KEY.startswith("cleide_audit_")


def test_clear_agente_compara_docs_does_not_clear_cleide_session_keys(app, monkeypatch):
    app.config["SECRET_KEY"] = "test-secret-agente-compara-isolation"
    monkeypatch.setattr(
        "app.agente_compara_doc_service._agente_compara_document_owned_by_session",
        lambda doc_id: doc_id == "ac-doc-1",
    )
    monkeypatch.setattr(
        "app.agente_compara_doc_service.remove_document_record",
        lambda _doc_id: {"removed": True},
    )
    monkeypatch.setattr(
        "app.agente_compara_doc_service.invalidate_temp_table_for_session",
        lambda **_k: None,
    )
    with app.app_context():
        with app.test_request_context("/"):
            from flask import session

            session[AGENTE_COMPARA_DOC_IDS_SESSION_KEY] = ["ac-doc-1"]
            session[AGENTE_COMPARA_TEMP_TABLE_ID_SESSION_KEY] = "ac-tt-1"
            session[CLEIDE_AUDIT_DOC_IDS_SESSION_KEY] = ["cleide-doc-1"]
            session[CLEIDE_AUDIT_TEMP_TABLE_ID_SESSION_KEY] = "cleide-tt-1"

            clear_documents_for_session()

            assert AGENTE_COMPARA_DOC_IDS_SESSION_KEY not in session
            assert session.get(CLEIDE_AUDIT_DOC_IDS_SESSION_KEY) == ["cleide-doc-1"]
            assert session.get(CLEIDE_AUDIT_TEMP_TABLE_ID_SESSION_KEY) == "cleide-tt-1"


def test_idempotency_prefixes_start_with_agente_compara():
    keys = [
        agente_compara_upload_idempotency_key("req-1"),
        agente_compara_chat_idempotency_key("req-2"),
        agente_compara_coverage_upload_idempotency_key("sess", "v1"),
        agente_compara_batch_upload_idempotency_key("sess", "batch-1"),
        agente_compara_batch_run_idempotency_key("sess", "batch-1", "run-1"),
        agente_compara_temp_table_extraction_idempotency_key(["doc-a"]),
    ]
    for key in keys:
        assert key.startswith("agente-compara-")
        assert not key.startswith("cleide-audit-")


def test_js_has_zero_cleide_auditoria_api():
    js = pathlib.Path("app/static/js/agente_compara.js").read_text(encoding="utf-8")
    assert "/api/cleide-auditoria" not in js
    assert js.count("/api/cleide-auditoria") == 0


def test_html_ids_do_not_collide_with_cleide_auditoria():
    html = pathlib.Path("app/templates/agente_compara.html").read_text(encoding="utf-8")
    assert "cleideAuditoria" not in html
    assert re.search(r'id="agenteCompara\w*"', html)
    assert 'id="agenteComparaShell"' in html
    assert 'id="agenteComparaActionsMenu"' in html


def test_agent_domain_flow_types_source_and_temp_table_marker():
    assert AGENTE_COMPARA_DOMAIN == "agente_compara"
    assert SOURCE_AGENT_AGENTE_COMPARA == "agente_compara"
    assert TEMP_TABLE_VERSION_MARKER == "agente_compara_temp_table_v1"
    assert AGENTE_COMPARA_CHAT_FLOW_TYPE.startswith("agente_compara_")
    assert AGENTE_COMPARA_TEMP_TABLE_EXTRACTION_FLOW_TYPE.startswith("agente_compara_")
    assert AGENTE_COMPARA_TEMP_TABLE_EXTRACTION_FLOW_TYPE == "agente_compara_temp_table_extraction"


def test_agente_compara_cannot_delete_cleide_document(web_client, tmp_path):
    cleide_up = _upload_cleide(web_client, "cleide.txt", make_txt("cleide")).get_json()
    cleide_doc_id = cleide_up["document"]["doc_id"]
    assert (tmp_path / f"{cleide_doc_id}.json").is_file()

    resp = web_client.delete(f"/api/agente-compara/documents/{cleide_doc_id}")
    assert resp.status_code == 200
    body = resp.get_json()
    assert body["ok"] is True
    assert body["removed_from_session"] is False
    assert body["removed_from_store"] is False
    assert (tmp_path / f"{cleide_doc_id}.json").is_file()

    with web_client.session_transaction() as sess:
        assert cleide_doc_id in (sess.get(CLEIDE_AUDIT_DOC_IDS_SESSION_KEY) or [])
        assert cleide_doc_id not in (sess.get(AGENTE_COMPARA_DOC_IDS_SESSION_KEY) or [])


def test_cleide_cannot_delete_agente_compara_document(web_client, tmp_path):
    ac_up = _upload_ac(web_client, "ac.csv", make_csv([["a", "b"], ["1", "2"]]), "text/csv").get_json()
    ac_doc_id = ac_up["document"]["doc_id"]
    assert (tmp_path / f"{ac_doc_id}.json").is_file()

    resp = web_client.delete(f"/api/cleide-auditoria/documents/{ac_doc_id}")
    assert resp.status_code == 200
    body = resp.get_json()
    assert body["removed_from_session"] is False
    assert body["removed_from_store"] is False
    assert (tmp_path / f"{ac_doc_id}.json").is_file()

    with web_client.session_transaction() as sess:
        assert ac_doc_id in (sess.get(AGENTE_COMPARA_DOC_IDS_SESSION_KEY) or [])
        assert ac_doc_id not in (sess.get(CLEIDE_AUDIT_DOC_IDS_SESSION_KEY) or [])


def test_agente_compara_cannot_delete_document_from_other_session_scope(app, tmp_path, monkeypatch, ctx):
    with app.app_context():
        _setup_doc_env(monkeypatch, tmp_path)
    app.config["SECRET_KEY"] = "test-secret-agente-compara-scope"
    with app.test_request_context("/"):
        from app.agente_compara_doc_service import prepare_and_register_document

        foreign = prepare_and_register_document(
            display_name="foreign.csv",
            file_bytes=make_csv([["x"], ["1"]]),
            mime_type="text/csv",
            extension=".csv",
        )
        foreign_doc_id = foreign["doc_id"]

    with app.test_request_context("/"):
        from app.agente_compara_doc_service import prepare_and_register_document, remove_document_from_session

        prepare_and_register_document(
            display_name="local.csv",
            file_bytes=make_csv([["y"], ["2"]]),
            mime_type="text/csv",
            extension=".csv",
        )
        result = remove_document_from_session(foreign_doc_id)
        assert result["removed_from_session"] is False
        assert result["removed_from_store"] is False
        assert (tmp_path / f"{foreign_doc_id}.json").is_file()


def test_agente_compara_legitimate_delete_removes_store_and_session(web_client, tmp_path):
    up = _upload_ac(web_client, "ok.csv", make_csv([["c"], ["1"]]), "text/csv").get_json()
    doc_id = up["document"]["doc_id"]

    resp = web_client.delete(f"/api/agente-compara/documents/{doc_id}")
    assert resp.status_code == 200
    body = resp.get_json()
    assert body["removed_from_session"] is True
    assert body["removed_from_store"] is True
    assert not (tmp_path / f"{doc_id}.json").exists()

    with web_client.session_transaction() as sess:
        assert doc_id not in (sess.get(AGENTE_COMPARA_DOC_IDS_SESSION_KEY) or [])


def test_agente_compara_delete_nonexistent_doc_is_safe(web_client, tmp_path):
    resp = web_client.delete("/api/agente-compara/documents/00000000000000000000000000000000")
    assert resp.status_code == 200
    body = resp.get_json()
    assert body["ok"] is True
    assert body["removed_from_session"] is False
    assert body["removed_from_store"] is False

    with web_client.session_transaction() as sess:
        assert AGENTE_COMPARA_DOC_IDS_SESSION_KEY not in sess or not sess.get(AGENTE_COMPARA_DOC_IDS_SESSION_KEY)


def test_clear_agente_compara_preserves_cleide_documents(web_client, tmp_path):
    cleide_up = _upload_cleide(web_client, "cleide.txt", make_txt("cleide")).get_json()
    cleide_doc_id = cleide_up["document"]["doc_id"]
    ac_up = _upload_ac(web_client, "ac.csv", make_csv([["a"], ["1"]]), "text/csv").get_json()
    ac_doc_id = ac_up["document"]["doc_id"]

    resp = web_client.post("/api/agente-compara/documents/clear")
    assert resp.status_code == 200
    body = resp.get_json()
    assert body["removed_from_session"] == 1
    assert not (tmp_path / f"{ac_doc_id}.json").exists()
    assert (tmp_path / f"{cleide_doc_id}.json").is_file()

    with web_client.session_transaction() as sess:
        assert cleide_doc_id in (sess.get(CLEIDE_AUDIT_DOC_IDS_SESSION_KEY) or [])
        assert ac_doc_id not in (sess.get(AGENTE_COMPARA_DOC_IDS_SESSION_KEY) or [])


def test_clear_cleide_preserves_agente_compara_documents(web_client, tmp_path):
    ac_up = _upload_ac(web_client, "ac.csv", make_csv([["a"], ["1"]]), "text/csv").get_json()
    ac_doc_id = ac_up["document"]["doc_id"]
    cleide_up = _upload_cleide(web_client, "cleide.txt", make_txt("cleide")).get_json()
    cleide_doc_id = cleide_up["document"]["doc_id"]

    resp = web_client.post("/api/cleide-auditoria/documents/clear")
    assert resp.status_code == 200
    assert resp.get_json()["removed_from_session"] == 1
    assert not (tmp_path / f"{cleide_doc_id}.json").exists()
    assert (tmp_path / f"{ac_doc_id}.json").is_file()

    with web_client.session_transaction() as sess:
        assert ac_doc_id in (sess.get(AGENTE_COMPARA_DOC_IDS_SESSION_KEY) or [])
        assert cleide_doc_id not in (sess.get(CLEIDE_AUDIT_DOC_IDS_SESSION_KEY) or [])


def test_agente_compara_template_xlsx_sheet_and_columns():
    template_path = pathlib.Path("app/protected_files/templates") / AGENTE_COMPARA_TEMPLATE_FILENAME
    assert template_path.is_file()
    wb = openpyxl.load_workbook(template_path, read_only=True, data_only=True)
    try:
        assert AUDIT_BATCH_SHEET_NAME in wb.sheetnames
        assert "Modelo Cleide" not in wb.sheetnames
        ws = wb[AUDIT_BATCH_SHEET_NAME]
        header = [str(cell).strip() if cell is not None else "" for cell in next(ws.iter_rows(values_only=True))]
    finally:
        wb.close()

    cleide_path = pathlib.Path("app/protected_files/templates/template_cleide_auditoria_frete.xlsx")
    wb_c = openpyxl.load_workbook(cleide_path, read_only=True, data_only=True)
    try:
        ws_c = wb_c["Modelo Cleide"]
        cleide_header = [
            str(cell).strip() if cell is not None else "" for cell in next(ws_c.iter_rows(values_only=True))
        ]
    finally:
        wb_c.close()
    assert header == cleide_header


def test_html_has_no_bi_cleide_label():
    html = pathlib.Path("app/templates/agente_compara.html").read_text(encoding="utf-8")
    assert "BI Cleide" not in html


def test_execution_id_requires_client_reuse_for_idempotency(app):
    with app.test_request_context("/api/agente-compara/audit/run", method="POST"):
        first = _resolve_agente_compara_execution_id()
        second = _resolve_agente_compara_execution_id()
        assert first != second

    with app.test_request_context(
        "/api/agente-compara/audit/run",
        method="POST",
        headers={"X-Execution-ID": "client-stable-key"},
    ):
        third = _resolve_agente_compara_execution_id()
        fourth = _resolve_agente_compara_execution_id()
        assert third == fourth == "client-stable-key"


def test_billing_same_execution_idempotency_key_deduplicates(app):
    with app.app_context():
        seed_sistema_interno()
        seed_cleiton_cost_config()
        conta, franquia = seed_conta_franquia_cliente(slug="conta-ac-exec-id")
        user = seed_usuario(franquia.id, conta.id, email="ac-exec-id@test.com", categoria="pro")

        with app.test_request_context("/api/agente-compara/audit/upload", method="POST"):
            from flask import g

            g.identidade = {
                "conta_id": conta.id,
                "franquia_id": franquia.id,
                "usuario_id": user.id,
                "tipo_origem": "http_usuario",
                "origem_sistema": False,
            }
            key = agente_compara_coverage_upload_idempotency_key("sess-1", "exec-shared")
            first = apropriar_billing_agente_compara_operational_flow(
                flow_type="agente_compara_coverage_upload",
                idempotency_key=key,
                rows_processed=3,
                processing_time_ms=50,
                status="success",
                execution_id="exec-shared",
            )
            second = apropriar_billing_agente_compara_operational_flow(
                flow_type="agente_compara_coverage_upload",
                idempotency_key=key,
                rows_processed=3,
                processing_time_ms=50,
                status="success",
                execution_id="exec-shared",
            )

        assert first["duplicado"] is False
        assert second["duplicado"] is True
        assert CleitonBillingApropriacao.query.filter_by(idempotency_key=key).count() == 1


def test_billing_different_execution_idempotency_keys_are_distinct(app):
    with app.app_context():
        seed_sistema_interno()
        seed_cleiton_cost_config()
        conta, franquia = seed_conta_franquia_cliente(slug="conta-ac-exec-diff")
        user = seed_usuario(franquia.id, conta.id, email="ac-exec-diff@test.com", categoria="pro")

        with app.test_request_context("/api/agente-compara/audit/upload", method="POST"):
            from flask import g

            g.identidade = {
                "conta_id": conta.id,
                "franquia_id": franquia.id,
                "usuario_id": user.id,
                "tipo_origem": "http_usuario",
                "origem_sistema": False,
            }
            first = apropriar_billing_agente_compara_operational_flow(
                flow_type="agente_compara_coverage_upload",
                idempotency_key=agente_compara_coverage_upload_idempotency_key("sess-1", "exec-a"),
                rows_processed=3,
                processing_time_ms=50,
                status="success",
                execution_id="exec-a",
            )
            second = apropriar_billing_agente_compara_operational_flow(
                flow_type="agente_compara_coverage_upload",
                idempotency_key=agente_compara_coverage_upload_idempotency_key("sess-1", "exec-b"),
                rows_processed=3,
                processing_time_ms=50,
                status="success",
                execution_id="exec-b",
            )

        assert first["duplicado"] is False
        assert second["duplicado"] is False
        assert CleitonBillingApropriacao.query.count() == 2
