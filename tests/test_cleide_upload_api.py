import io
import importlib
import os
from pathlib import Path
from types import SimpleNamespace

import flask_login.utils
import pytest
import openpyxl


def _load_web_module():
    os.environ.setdefault("APP_ENV", "dev")
    os.environ.setdefault("DATABASE_URL", "postgresql://user:pass@localhost:5432/testdb")
    os.environ.setdefault("SECRET_KEY", "test-secret")
    return importlib.import_module("app.web")


def _authorized(monkeypatch):
    fake_user = SimpleNamespace(
        is_authenticated=True,
        is_active=True,
        is_anonymous=False,
        get_id=lambda: "1",
        conta_id=1,
        franquia_id=1,
        categoria="pro",
        full_name="Teste Cleide",
        email="cleide@example.com",
        franquia=None,
    )
    monkeypatch.setattr(flask_login.utils, "_get_user", lambda: fake_user)
    monkeypatch.setattr("app.cleide_routes.current_user", fake_user)
    monkeypatch.setattr(
        "app.cleide_routes.avaliar_autorizacao_operacao_por_franquia",
        lambda _u: {"permitido": True, "modo_operacao": "normal"},
    )
    monkeypatch.setattr("app.cleide_routes.get_cleide_config", lambda: SimpleNamespace(layout_version=3))


def _cfg(monkeypatch, max_bytes=2 * 1024 * 1024):
    monkeypatch.setattr(
        "app.cleide_upload_pipeline.get_cleide_config",
        lambda: SimpleNamespace(
            upload_total_max=10000,
            upload_max_file_size_bytes=max_bytes,
            upload_ttl_minutes=30,
            csv_delimiter_default=",",
            structural_max_rows=10000,
            structural_max_columns=120,
            analytics_max_rows=10000,
            analytics_group_limit=25,
        ),
    )


def _tmp_store(monkeypatch, tmp_path: Path):
    monkeypatch.setattr("app.cleide_upload_store.get_cleide_upload_tmp_dir", lambda: str(tmp_path))
    monkeypatch.setattr("app.cleide_routes.get_cleide_upload_tmp_dir", lambda: str(tmp_path))


def _xlsx_bytes():
    return _xlsx_bytes_with_headers(
        ("transportadora", "uf_origem", "uf_destino", "valor_frete", "peso", "data_emissao"),
        ("XP", "SP", "RJ", 100, 10, "2026-01-01"),
    )


def _xlsx_bytes_with_headers(*headers_and_rows):
    import openpyxl

    wb = openpyxl.Workbook()
    ws = wb.active
    for ridx, row in enumerate(headers_and_rows, start=1):
        for cidx, cell in enumerate(row, start=1):
            ws.cell(row=ridx, column=cidx, value=cell)
    stream = io.BytesIO()
    wb.save(stream)
    return stream.getvalue()


def test_upload_exige_login(monkeypatch, tmp_path):
    web = _load_web_module()
    monkeypatch.setattr("app.cleide_routes.current_user", SimpleNamespace(is_authenticated=False))
    _cfg(monkeypatch)
    _tmp_store(monkeypatch, tmp_path)
    client = web.app.test_client()
    resp = client.post(
        "/api/cleide/upload",
        data={"file": (io.BytesIO(_xlsx_bytes()), "base.xlsx")},
        content_type="multipart/form-data",
    )
    assert resp.status_code == 401
    assert resp.get_json()["error"] == "Autenticacao necessaria."


def test_upload_status_exige_login(monkeypatch, tmp_path):
    web = _load_web_module()
    monkeypatch.setattr("app.cleide_routes.current_user", SimpleNamespace(is_authenticated=False))
    _cfg(monkeypatch)
    _tmp_store(monkeypatch, tmp_path)
    client = web.app.test_client()
    resp = client.get("/api/cleide/upload/status")
    assert resp.status_code == 401
    assert resp.get_json()["error"] == "Autenticacao necessaria."


def test_template_download_nao_exige_login_nem_consumo_franquia(monkeypatch, tmp_path):
    web = _load_web_module()
    calls = {"authz": 0}

    def _authz(_u):
        calls["authz"] += 1
        return {"permitido": True, "modo_operacao": "normal"}

    monkeypatch.setattr("app.cleide_routes.current_user", SimpleNamespace(is_authenticated=False))
    monkeypatch.setattr("app.cleide_routes.avaliar_autorizacao_operacao_por_franquia", _authz)
    _cfg(monkeypatch)
    _tmp_store(monkeypatch, tmp_path)

    client = web.app.test_client()
    resp = client.get("/api/cleide/template")

    assert resp.status_code == 200
    disposition = resp.headers.get("Content-Disposition", "")
    assert "template_cleide_auditoria_frete.xlsx" in disposition
    assert calls["authz"] == 0
    with client.session_transaction() as sess:
        assert sess.get("cleide_upload_ref") is None


def test_template_download_contem_colunas_planejadas_e_obrigatorias(monkeypatch):
    web = _load_web_module()
    monkeypatch.setattr("app.cleide_routes.current_user", SimpleNamespace(is_authenticated=False))
    client = web.app.test_client()

    resp = client.get("/api/cleide/template")
    assert resp.status_code == 200

    wb = openpyxl.load_workbook(io.BytesIO(resp.data), read_only=True, data_only=True)
    try:
        assert "Modelo Cleide" in wb.sheetnames
        ws = wb["Modelo Cleide"]
        header = [str(cell).strip() if cell is not None else "" for cell in next(ws.iter_rows(values_only=True))]
    finally:
        wb.close()

    expected_columns = [
        "transportadora",
        "numero_documento",
        "cidade_origem",
        "uf_origem",
        "cidade_destino",
        "uf_destino",
        "valor_nf",
        "valor_frete",
        "peso",
        "modal",
        "data_emissao",
        "data_entrega",
    ]
    assert header == expected_columns


def test_upload_clear_exige_login(monkeypatch, tmp_path):
    web = _load_web_module()
    monkeypatch.setattr("app.cleide_routes.current_user", SimpleNamespace(is_authenticated=False))
    _cfg(monkeypatch)
    _tmp_store(monkeypatch, tmp_path)
    client = web.app.test_client()
    resp = client.post("/api/cleide/upload/clear")
    assert resp.status_code == 401
    assert resp.get_json()["error"] == "Autenticacao necessaria."


def test_upload_xlsx_valido(monkeypatch, tmp_path):
    web = _load_web_module()
    _authorized(monkeypatch)
    _cfg(monkeypatch)
    _tmp_store(monkeypatch, tmp_path)
    client = web.app.test_client()
    resp = client.post(
        "/api/cleide/upload",
        data={"file": (io.BytesIO(_xlsx_bytes()), "base.xlsx")},
        content_type="multipart/form-data",
    )
    assert resp.status_code == 200
    body = resp.get_json()
    assert body["success"] is True
    assert body["extension"] == ".xlsx"
    assert body["replaced_previous_upload"] is False
    assert body["dataset_validado"] is True
    assert body["sheet_detectada"] == "Sheet"
    assert body["linhas_detectadas"] >= 1
    assert isinstance(body["upload_ref"], str)
    assert len(list(tmp_path.iterdir())) >= 1


def test_upload_csv_utf8_valido(monkeypatch, tmp_path):
    web = _load_web_module()
    _authorized(monkeypatch)
    _cfg(monkeypatch)
    _tmp_store(monkeypatch, tmp_path)
    client = web.app.test_client()
    payload = "transportadora,uf_origem,uf_destino,valor_frete,peso,data_emissao\nXP,SP,RJ,123.45,10,2026-01-01\n".encode(
        "utf-8"
    )
    resp = client.post(
        "/api/cleide/upload",
        data={"file": (io.BytesIO(payload), "base.csv")},
        content_type="multipart/form-data",
    )
    assert resp.status_code == 200
    body = resp.get_json()
    assert body["success"] is True
    assert body["extension"] == ".csv"
    assert body["encoding"] in {"utf-8", "utf-8-sig"}
    assert body["dataset_validado"] is True
    assert body["linhas_detectadas"] == 1
    assert "transportadora" in body["aliases_resolvidos"]
    assert body["analytics_ready"] is True
    assert "kpis" in body
    assert "aggregate_counts" in body
    assert "dataset_summary" in body
    assert "pareto_fretes_zerados_uf_destino" in body
    assert "pareto_fretes_zerados_transportadora" in body
    assert "cleide_contexto_operacional" in body
    contexto = body["cleide_contexto_operacional"]
    assert contexto["schema_version"] == "cleide_contexto_operacional.v1"
    assert contexto["phase"] == "8_context_prep_no_ai"
    assert contexto["agent"] == "cleide"
    assert contexto["namespace"] == "cleide"
    assert contexto["security_guards"]["contains_raw_dataset"] is False
    assert contexto["security_guards"]["contains_ai_output"] is False
    assert body["kpis"]["total_documentos"] == 1
    assert body["kpis"]["valor_total_frete"] == 123.45


def test_upload_csv_latin1_valido(monkeypatch, tmp_path):
    web = _load_web_module()
    _authorized(monkeypatch)
    _cfg(monkeypatch)
    _tmp_store(monkeypatch, tmp_path)
    client = web.app.test_client()
    payload = (
        "transportadora,uf_origem,uf_destino,valor_frete,peso,data_emissao\n"
        "transportadora acao,sp,pb,150.50,23,2026-01-01\n"
    ).encode("latin1")
    resp = client.post(
        "/api/cleide/upload",
        data={"file": (io.BytesIO(payload), "base.csv")},
        content_type="multipart/form-data",
    )
    assert resp.status_code == 200
    body = resp.get_json()
    assert body["success"] is True
    assert body["encoding"] in {"latin1", "utf-8", "utf-8-sig"}
    assert body["dataset_validado"] is True


def test_upload_csv_utf8_sig_valido(monkeypatch, tmp_path):
    web = _load_web_module()
    _authorized(monkeypatch)
    _cfg(monkeypatch)
    _tmp_store(monkeypatch, tmp_path)
    client = web.app.test_client()
    payload = (
        "transportadora,uf_origem,uf_destino,valor_frete,peso,data_emissao\n"
        "xp,sp,rj,100,1,2026-01-01\n"
    ).encode("utf-8-sig")
    resp = client.post(
        "/api/cleide/upload",
        data={"file": (io.BytesIO(payload), "base.csv")},
        content_type="multipart/form-data",
    )
    assert resp.status_code == 200
    body = resp.get_json()
    assert body["success"] is True
    assert body["encoding"] == "utf-8-sig"
    assert body["dataset_validado"] is True


def test_upload_csv_aliases_validos(monkeypatch, tmp_path):
    web = _load_web_module()
    _authorized(monkeypatch)
    _cfg(monkeypatch)
    _tmp_store(monkeypatch, tmp_path)
    client = web.app.test_client()
    payload = (
        "nome transportadora,origem_uf,destino_uf,vl_frete,peso_kg,dt_emissao\n"
        "xp,sp,rj,100,2,2026-01-01\n"
    ).encode("utf-8")
    resp = client.post(
        "/api/cleide/upload",
        data={"file": (io.BytesIO(payload), "alias.csv")},
        content_type="multipart/form-data",
    )
    assert resp.status_code == 200
    body = resp.get_json()
    assert body["dataset_validado"] is True
    assert body["colunas_faltantes"] == []
    assert body["aliases_resolvidos"]["transportadora"] == "nome transportadora"


def test_upload_csv_colunas_obrigatorias_faltantes(monkeypatch, tmp_path):
    web = _load_web_module()
    _authorized(monkeypatch)
    _cfg(monkeypatch)
    _tmp_store(monkeypatch, tmp_path)
    client = web.app.test_client()
    payload = "transportadora,uf_origem\nxp,sp\n".encode("utf-8")
    resp = client.post(
        "/api/cleide/upload",
        data={"file": (io.BytesIO(payload), "missing.csv")},
        content_type="multipart/form-data",
    )
    assert resp.status_code == 200
    body = resp.get_json()
    assert body["dataset_validado"] is False
    assert "uf_destino" in body["colunas_faltantes"]
    assert "valor_frete" in body["colunas_faltantes"]
    assert body["analytics_ready"] is False


def test_upload_substitui_upload_anterior(monkeypatch, tmp_path):
    web = _load_web_module()
    _authorized(monkeypatch)
    _cfg(monkeypatch)
    _tmp_store(monkeypatch, tmp_path)
    client = web.app.test_client()
    first = client.post(
        "/api/cleide/upload",
        data={
            "file": (
                io.BytesIO(
                    "transportadora,uf_origem,uf_destino,valor_frete,peso,data_emissao\nx,sp,rj,10,1,2026-01-01\n".encode(
                        "utf-8"
                    )
                ),
                "f1.csv",
            )
        },
        content_type="multipart/form-data",
    )
    first_ref = first.get_json()["upload_ref"]
    second = client.post(
        "/api/cleide/upload",
        data={
            "file": (
                io.BytesIO(
                    "transportadora,uf_origem,uf_destino,valor_frete,peso,data_emissao\ny,sp,rj,20,2,2026-01-01\n".encode(
                        "utf-8"
                    )
                ),
                "f2.csv",
            )
        },
        content_type="multipart/form-data",
    )
    assert second.status_code == 200
    second_body = second.get_json()
    assert second_body["replaced_previous_upload"] is True
    assert second_body["upload_ref"] != first_ref
    assert not (tmp_path / f"{first_ref}.csv").exists()


def test_upload_extensao_invalida(monkeypatch, tmp_path):
    web = _load_web_module()
    _authorized(monkeypatch)
    _cfg(monkeypatch)
    _tmp_store(monkeypatch, tmp_path)
    client = web.app.test_client()
    resp = client.post(
        "/api/cleide/upload",
        data={"file": (io.BytesIO(b"{}"), "payload.json")},
        content_type="multipart/form-data",
    )
    assert resp.status_code == 400
    assert resp.get_json()["error_code"] == "invalid_extension"


@pytest.mark.parametrize("filename", ["base.xls", "base.pdf", "base.zip", "base.png"])
def test_upload_extensoes_bloqueadas(monkeypatch, tmp_path, filename):
    web = _load_web_module()
    _authorized(monkeypatch)
    _cfg(monkeypatch)
    _tmp_store(monkeypatch, tmp_path)
    client = web.app.test_client()
    resp = client.post(
        "/api/cleide/upload",
        data={"file": (io.BytesIO(b"stub"), filename)},
        content_type="multipart/form-data",
    )
    assert resp.status_code == 400
    assert resp.get_json()["error_code"] == "invalid_extension"


def test_upload_arquivo_vazio(monkeypatch, tmp_path):
    web = _load_web_module()
    _authorized(monkeypatch)
    _cfg(monkeypatch)
    _tmp_store(monkeypatch, tmp_path)
    client = web.app.test_client()
    resp = client.post(
        "/api/cleide/upload",
        data={"file": (io.BytesIO(b""), "base.csv")},
        content_type="multipart/form-data",
    )
    assert resp.status_code == 400
    assert resp.get_json()["error_code"] == "empty_file"


def test_upload_xlsx_corrompido(monkeypatch, tmp_path):
    web = _load_web_module()
    _authorized(monkeypatch)
    _cfg(monkeypatch)
    _tmp_store(monkeypatch, tmp_path)
    client = web.app.test_client()
    resp = client.post(
        "/api/cleide/upload",
        data={"file": (io.BytesIO(b"not-an-xlsx"), "base.xlsx")},
        content_type="multipart/form-data",
    )
    assert resp.status_code == 400
    assert resp.get_json()["error_code"] == "invalid_xlsx"


def test_upload_csv_invalido(monkeypatch, tmp_path):
    web = _load_web_module()
    _authorized(monkeypatch)
    _cfg(monkeypatch)
    _tmp_store(monkeypatch, tmp_path)
    client = web.app.test_client()
    binary = b"transportadora"
    resp = client.post(
        "/api/cleide/upload",
        data={"file": (io.BytesIO(binary), "base.csv")},
        content_type="multipart/form-data",
    )
    assert resp.status_code == 400
    assert resp.get_json()["error_code"] == "invalid_csv"


def test_upload_csv_delimitador_ruim(monkeypatch, tmp_path):
    web = _load_web_module()
    _authorized(monkeypatch)
    _cfg(monkeypatch)
    _tmp_store(monkeypatch, tmp_path)
    client = web.app.test_client()
    payload = "transportadora||uf_origem||uf_destino\nxp||sp||rj\n".encode("utf-8")
    resp = client.post(
        "/api/cleide/upload",
        data={"file": (io.BytesIO(payload), "delimiter.csv")},
        content_type="multipart/form-data",
    )
    assert resp.status_code == 400
    assert resp.get_json()["error_code"] == "invalid_csv"


def test_upload_sem_arquivo(monkeypatch, tmp_path):
    web = _load_web_module()
    _authorized(monkeypatch)
    _cfg(monkeypatch)
    _tmp_store(monkeypatch, tmp_path)
    client = web.app.test_client()
    resp = client.post("/api/cleide/upload", data={}, content_type="multipart/form-data")
    assert resp.status_code == 400
    assert resp.get_json()["error_code"] == "missing_file"


def test_upload_multipart_invalido(monkeypatch, tmp_path):
    web = _load_web_module()
    _authorized(monkeypatch)
    _cfg(monkeypatch)
    _tmp_store(monkeypatch, tmp_path)
    client = web.app.test_client()
    resp = client.post("/api/cleide/upload", data=b"raw-json", content_type="application/json")
    assert resp.status_code == 400
    assert resp.get_json()["error_code"] == "invalid_multipart"


def test_upload_path_traversal_sanitizado(monkeypatch, tmp_path):
    web = _load_web_module()
    _authorized(monkeypatch)
    _cfg(monkeypatch)
    _tmp_store(monkeypatch, tmp_path)
    client = web.app.test_client()
    resp = client.post(
        "/api/cleide/upload",
        data={
            "file": (
                io.BytesIO(
                    "transportadora,uf_origem,uf_destino,valor_frete,peso,data_emissao\nx,sp,rj,10,1,2026-01-01\n".encode(
                        "utf-8"
                    )
                ),
                "../../evil.csv",
            )
        },
        content_type="multipart/form-data",
    )
    assert resp.status_code == 200
    body = resp.get_json()
    assert "/" not in body["filename"]
    assert "\\" not in body["filename"]


def test_upload_limite_tamanho(monkeypatch, tmp_path):
    web = _load_web_module()
    _authorized(monkeypatch)
    _cfg(monkeypatch, max_bytes=4)
    _tmp_store(monkeypatch, tmp_path)
    client = web.app.test_client()
    resp = client.post(
        "/api/cleide/upload",
        data={
            "file": (
                io.BytesIO(
                    "transportadora,uf_origem,uf_destino,valor_frete,peso,data_emissao\nx,sp,rj,10,1,2026-01-01\n".encode(
                        "utf-8"
                    )
                ),
                "base.csv",
            )
        },
        content_type="multipart/form-data",
    )
    assert resp.status_code == 413
    assert resp.get_json()["error_code"] == "file_too_large"


def test_upload_csv_acima_upload_total_max_rejeita_sem_side_effects(monkeypatch, tmp_path):
    web = _load_web_module()
    _authorized(monkeypatch)
    _cfg(monkeypatch)
    _tmp_store(monkeypatch, tmp_path)
    monkeypatch.setattr(
        "app.cleide_upload_pipeline.get_cleide_config",
        lambda: SimpleNamespace(
            upload_total_max=2,
            upload_max_file_size_bytes=2 * 1024 * 1024,
            upload_ttl_minutes=30,
            csv_delimiter_default=",",
            structural_max_rows=10000,
            structural_max_columns=120,
            analytics_max_rows=10000,
            analytics_group_limit=25,
        ),
    )
    client = web.app.test_client()
    payload = (
        "transportadora,uf_origem,uf_destino,valor_frete,peso,data_emissao\n"
        "a,sp,rj,10,1,2026-01-01\n"
        "\n"
        "b,sp,rj,20,2,2026-01-02\n"
        "c,sp,rj,30,3,2026-01-03\n"
    ).encode("utf-8")
    resp = client.post(
        "/api/cleide/upload",
        data={"file": (io.BytesIO(payload), "above-limit.csv")},
        content_type="multipart/form-data",
    )
    assert resp.status_code == 413
    body = resp.get_json()
    assert body["success"] is False
    assert body["error_code"] == "upload_total_max_exceeded"
    assert body["upload_total_max"] == 2
    assert body["linhas_detectadas"] == 3
    assert list(tmp_path.iterdir()) == []
    with client.session_transaction() as sess:
        assert sess.get("cleide_upload_ref") is None
        assert sess.get("cleide_dataset_context") is None


def test_upload_xlsx_acima_upload_total_max_rejeita_sem_side_effects(monkeypatch, tmp_path):
    web = _load_web_module()
    _authorized(monkeypatch)
    _cfg(monkeypatch)
    _tmp_store(monkeypatch, tmp_path)
    monkeypatch.setattr(
        "app.cleide_upload_pipeline.get_cleide_config",
        lambda: SimpleNamespace(
            upload_total_max=1,
            upload_max_file_size_bytes=2 * 1024 * 1024,
            upload_ttl_minutes=30,
            csv_delimiter_default=",",
            structural_max_rows=10000,
            structural_max_columns=120,
            analytics_max_rows=10000,
            analytics_group_limit=25,
        ),
    )
    client = web.app.test_client()
    payload = _xlsx_bytes_with_headers(
        ("transportadora", "uf_origem", "uf_destino", "valor_frete", "peso", "data_emissao"),
        ("a", "SP", "RJ", 100, 10, "2026-01-01"),
        ("b", "SP", "RJ", 120, 12, "2026-01-02"),
    )
    resp = client.post(
        "/api/cleide/upload",
        data={"file": (io.BytesIO(payload), "above-limit.xlsx")},
        content_type="multipart/form-data",
    )
    assert resp.status_code == 413
    body = resp.get_json()
    assert body["success"] is False
    assert body["error_code"] == "upload_total_max_exceeded"
    assert body["upload_total_max"] == 1
    assert body["linhas_detectadas"] == 2
    assert list(tmp_path.iterdir()) == []
    with client.session_transaction() as sess:
        assert sess.get("cleide_upload_ref") is None
        assert sess.get("cleide_dataset_context") is None


def test_upload_csv_exatamente_no_upload_total_max_aceita(monkeypatch, tmp_path):
    web = _load_web_module()
    _authorized(monkeypatch)
    _cfg(monkeypatch)
    _tmp_store(monkeypatch, tmp_path)
    monkeypatch.setattr(
        "app.cleide_upload_pipeline.get_cleide_config",
        lambda: SimpleNamespace(
            upload_total_max=2,
            upload_max_file_size_bytes=2 * 1024 * 1024,
            upload_ttl_minutes=30,
            csv_delimiter_default=",",
            structural_max_rows=10000,
            structural_max_columns=120,
            analytics_max_rows=10000,
            analytics_group_limit=25,
        ),
    )
    client = web.app.test_client()
    payload = (
        "transportadora,uf_origem,uf_destino,valor_frete,peso,data_emissao\n"
        "a,sp,rj,10,1,2026-01-01\n"
        "b,sp,rj,20,2,2026-01-02\n"
    ).encode("utf-8")
    resp = client.post(
        "/api/cleide/upload",
        data={"file": (io.BytesIO(payload), "limit.csv")},
        content_type="multipart/form-data",
    )
    assert resp.status_code == 200
    assert resp.get_json()["success"] is True


def test_upload_csv_abaixo_upload_total_max_aceita(monkeypatch, tmp_path):
    web = _load_web_module()
    _authorized(monkeypatch)
    _cfg(monkeypatch)
    _tmp_store(monkeypatch, tmp_path)
    monkeypatch.setattr(
        "app.cleide_upload_pipeline.get_cleide_config",
        lambda: SimpleNamespace(
            upload_total_max=3,
            upload_max_file_size_bytes=2 * 1024 * 1024,
            upload_ttl_minutes=30,
            csv_delimiter_default=",",
            structural_max_rows=10000,
            structural_max_columns=120,
            analytics_max_rows=10000,
            analytics_group_limit=25,
        ),
    )
    client = web.app.test_client()
    payload = (
        "transportadora,uf_origem,uf_destino,valor_frete,peso,data_emissao\n"
        "a,sp,rj,10,1,2026-01-01\n"
        "b,sp,rj,20,2,2026-01-02\n"
    ).encode("utf-8")
    resp = client.post(
        "/api/cleide/upload",
        data={"file": (io.BytesIO(payload), "below-limit.csv")},
        content_type="multipart/form-data",
    )
    assert resp.status_code == 200
    assert resp.get_json()["success"] is True


def test_upload_status_e_clear(monkeypatch, tmp_path):
    web = _load_web_module()
    _authorized(monkeypatch)
    _cfg(monkeypatch)
    _tmp_store(monkeypatch, tmp_path)
    client = web.app.test_client()
    up = client.post(
        "/api/cleide/upload",
        data={
            "file": (
                io.BytesIO(
                    "transportadora,uf_origem,uf_destino,valor_frete,peso,data_emissao\nx,sp,rj,10,1,2026-01-01\n".encode(
                        "utf-8"
                    )
                ),
                "base.csv",
            )
        },
        content_type="multipart/form-data",
    )
    assert up.status_code == 200
    status1 = client.get("/api/cleide/upload/status")
    assert status1.status_code == 200
    status_body = status1.get_json()
    assert status_body["upload_ativo"] is True
    assert "dataset_validado" in status_body
    assert "colunas_detectadas" in status_body
    assert "colunas_faltantes" in status_body
    assert "linhas_detectadas" in status_body
    assert "sheet_detectada" in status_body
    assert "aliases_resolvidos" in status_body
    assert "analytics_ready" in status_body
    assert "kpis" in status_body
    assert "aggregate_counts" in status_body
    assert "dataset_summary" in status_body
    assert "transportadora_stats" in status_body
    assert "uf_origem_stats" in status_body
    assert "uf_destino_stats" in status_body
    assert "temporal_stats" in status_body
    assert "pareto_fretes_zerados_uf_destino" in status_body
    assert "pareto_fretes_zerados_transportadora" in status_body
    assert "cleide_contexto_operacional" in status_body
    contexto = status_body["cleide_contexto_operacional"]
    assert contexto["schema_version"] == "cleide_contexto_operacional.v1"
    assert contexto["filter_context"]["filter_mode"] == "aggregate_approximation"
    assert contexto["filter_context"]["kpi_scope"] == "global_session"
    assert contexto["semantic_limits"]["no_row_level_intersection"] is True
    assert contexto["semantic_limits"]["kpis_are_global_session_scope"] is True
    assert contexto["security_guards"]["contains_roberto_payload"] is False
    assert contexto["security_guards"]["contains_ai_output"] is False
    clear = client.post("/api/cleide/upload/clear")
    assert clear.status_code == 200
    status2 = client.get("/api/cleide/upload/status")
    assert status2.get_json()["upload_ativo"] is False


def test_upload_status_get_nao_limpa_ref_quando_arquivo_ausente(monkeypatch, tmp_path):
    web = _load_web_module()
    _authorized(monkeypatch)
    _cfg(monkeypatch)
    _tmp_store(monkeypatch, tmp_path)
    client = web.app.test_client()

    up = client.post(
        "/api/cleide/upload",
        data={
            "file": (
                io.BytesIO(
                    "transportadora,uf_origem,uf_destino,valor_frete,peso,data_emissao\nx,sp,rj,10,1,2026-01-01\n".encode(
                        "utf-8"
                    )
                ),
                "base.csv",
            )
        },
        content_type="multipart/form-data",
    )
    assert up.status_code == 200
    upload_ref = up.get_json()["upload_ref"]
    stored = tmp_path / f"{upload_ref}.csv"
    assert stored.exists()
    stored.unlink()

    with client.session_transaction() as sess:
        ref_before = sess.get("cleide_upload_ref")
    assert ref_before == upload_ref

    status = client.get("/api/cleide/upload/status")
    assert status.status_code == 200
    body = status.get_json()
    assert body["success"] is True
    assert body["upload_ativo"] is False
    assert body["stale_upload"] is True
    assert body["upload_ref"] == upload_ref

    with client.session_transaction() as sess:
        assert sess.get("cleide_upload_ref") == upload_ref


def test_upload_stale_permite_limpeza_explicita(monkeypatch, tmp_path):
    web = _load_web_module()
    _authorized(monkeypatch)
    _cfg(monkeypatch)
    _tmp_store(monkeypatch, tmp_path)
    client = web.app.test_client()

    up = client.post(
        "/api/cleide/upload",
        data={
            "file": (
                io.BytesIO(
                    "transportadora,uf_origem,uf_destino,valor_frete,peso,data_emissao\nx,sp,rj,10,1,2026-01-01\n".encode(
                        "utf-8"
                    )
                ),
                "base.csv",
            )
        },
        content_type="multipart/form-data",
    )
    assert up.status_code == 200
    upload_ref = up.get_json()["upload_ref"]
    (tmp_path / f"{upload_ref}.csv").unlink(missing_ok=True)

    stale_status = client.get("/api/cleide/upload/status")
    assert stale_status.status_code == 200
    assert stale_status.get_json()["stale_upload"] is True

    clear = client.post("/api/cleide/upload/clear")
    assert clear.status_code == 200

    with client.session_transaction() as sess:
        assert sess.get("cleide_upload_ref") is None

    status_after_clear = client.get("/api/cleide/upload/status")
    assert status_after_clear.status_code == 200
    after_body = status_after_clear.get_json()
    assert after_body["upload_ativo"] is False
    assert after_body.get("stale_upload") is None


def test_upload_status_get_nao_altera_filesystem_quando_stale(monkeypatch, tmp_path):
    web = _load_web_module()
    _authorized(monkeypatch)
    _cfg(monkeypatch)
    _tmp_store(monkeypatch, tmp_path)
    client = web.app.test_client()

    up = client.post(
        "/api/cleide/upload",
        data={
            "file": (
                io.BytesIO(
                    "transportadora,uf_origem,uf_destino,valor_frete,peso,data_emissao\nx,sp,rj,10,1,2026-01-01\n".encode(
                        "utf-8"
                    )
                ),
                "base.csv",
            )
        },
        content_type="multipart/form-data",
    )
    assert up.status_code == 200
    upload_ref = up.get_json()["upload_ref"]

    stale_file = tmp_path / f"{upload_ref}.csv"
    stale_file.unlink(missing_ok=True)
    sentinel = tmp_path / "sentinel.keep"
    sentinel.write_text("keep", encoding="utf-8")
    before_entries = sorted(p.name for p in tmp_path.iterdir())

    status = client.get("/api/cleide/upload/status")
    assert status.status_code == 200
    assert status.get_json()["stale_upload"] is True

    after_entries = sorted(p.name for p in tmp_path.iterdir())
    assert after_entries == before_entries
    assert sentinel.exists()


def test_upload_status_get_sem_side_effect_quando_sem_upload(monkeypatch, tmp_path):
    web = _load_web_module()
    _authorized(monkeypatch)
    _cfg(monkeypatch)
    _tmp_store(monkeypatch, tmp_path)
    client = web.app.test_client()

    sentinel = tmp_path / "sentinel.keep"
    sentinel.write_text("keep", encoding="utf-8")
    before_entries = sorted(p.name for p in tmp_path.iterdir())

    with client.session_transaction() as sess:
        sess["custom_marker"] = "preserve"
        ref_before = sess.get("cleide_upload_ref")

    status = client.get("/api/cleide/upload/status")
    assert status.status_code == 200
    body = status.get_json()
    assert body["success"] is True
    assert body["upload_ativo"] is False
    assert body.get("stale_upload") is None
    assert body["dataset_validado"] is False
    assert body["colunas_detectadas"] == []
    assert body["analytics_ready"] is False
    assert body["kpis"] == {}
    assert body["aggregate_counts"] == {}
    assert body["dataset_summary"] == {}
    assert body["transportadora_stats"] == []
    assert body["uf_origem_stats"] == []
    assert body["uf_destino_stats"] == []
    assert body["temporal_stats"] == []


def test_status_sem_dataset_bruto(monkeypatch, tmp_path):
    web = _load_web_module()
    _authorized(monkeypatch)
    _cfg(monkeypatch)
    _tmp_store(monkeypatch, tmp_path)
    client = web.app.test_client()

    up = client.post(
        "/api/cleide/upload",
        data={
            "file": (
                io.BytesIO(
                    "transportadora,uf_origem,uf_destino,valor_frete,peso,data_emissao\nx,sp,rj,10,1,2026-01-01\n".encode(
                        "utf-8"
                    )
                ),
                "base.csv",
            )
        },
        content_type="multipart/form-data",
    )
    assert up.status_code == 200
    body = client.get("/api/cleide/upload/status").get_json()
    assert "rows" not in body
    assert "dataset_raw" not in body
    assert "dataframe" not in body
    assert "raw_bytes" not in body
    assert "analytics_context" not in body
    contexto = body["cleide_contexto_operacional"]
    contexto_str = str(contexto).lower()
    assert "dataset_raw" not in contexto_str
    assert "dataframe" not in contexto_str
    assert "raw_bytes" not in contexto_str


def test_status_analytics_quality_campos_presentes(monkeypatch, tmp_path):
    web = _load_web_module()
    _authorized(monkeypatch)
    _cfg(monkeypatch)
    _tmp_store(monkeypatch, tmp_path)
    client = web.app.test_client()

    payload = (
        "transportadora,uf_origem,uf_destino,valor_frete,peso,data_emissao\n"
        "xp,sp,rj,100,2,2026-01-01\n"
        "xp,sp,rj,abc,2,2026-01-32\n"
        "xp,sp,rj,-5,2,2026-01-01\n"
    ).encode("utf-8")
    up = client.post(
        "/api/cleide/upload",
        data={"file": (io.BytesIO(payload), "quality.csv")},
        content_type="multipart/form-data",
    )
    assert up.status_code == 200
    body = client.get("/api/cleide/upload/status").get_json()
    summary = body.get("dataset_summary") or {}
    assert "linhas_processadas" in summary
    assert "invalid_numeric_rows" in summary
    assert "invalid_date_rows" in summary
    assert "negative_value_rows" in summary
    assert "numeric_issue_details" in summary
    details = summary["numeric_issue_details"]
    assert "invalid_rows_total" in details
    assert "by_column" in details
    assert "by_reason" in details
    assert "samples" in details
    assert details["by_column"]["valor_frete"] >= 0
    assert details["by_column"]["peso"] >= 0
    assert details["by_column"]["both"] >= 0
    assert details["by_reason"]["empty"] >= 0
    assert details["by_reason"]["invalid_format"] >= 0
    assert details["by_reason"]["negative"] >= 0
    assert isinstance(details["samples"], list)
    assert len(details["samples"]) <= 10
    assert summary["linhas_processadas"] == 3
    assert summary["invalid_numeric_rows"] >= 1
    assert summary["invalid_date_rows"] >= 1
    assert summary["negative_value_rows"] >= 1


def test_status_fallback_sem_agregados_com_dataset_validado(monkeypatch, tmp_path):
    web = _load_web_module()
    _authorized(monkeypatch)
    _cfg(monkeypatch)
    _tmp_store(monkeypatch, tmp_path)
    client = web.app.test_client()

    payload = (
        "transportadora,uf_origem,uf_destino,valor_frete,peso,data_emissao\n"
        ",,,abc,def,99/99/9999\n"
    ).encode("utf-8")
    up = client.post(
        "/api/cleide/upload",
        data={"file": (io.BytesIO(payload), "fallback.csv")},
        content_type="multipart/form-data",
    )
    assert up.status_code == 200
    body = client.get("/api/cleide/upload/status").get_json()
    assert body["dataset_validado"] is True
    assert body["analytics_ready"] is True
    counts = body["aggregate_counts"]
    assert counts["transportadora_stats"] == 0
    assert counts["uf_origem_stats"] == 0
    assert counts["uf_destino_stats"] == 0
    assert counts["temporal_stats"] == 0
    assert body["transportadora_stats"] == []
    assert body["uf_origem_stats"] == []
    assert body["uf_destino_stats"] == []
    assert body["temporal_stats"] == []


def test_upload_status_exige_autorizacao(monkeypatch, tmp_path):
    web = _load_web_module()
    monkeypatch.setattr("app.cleide_routes.current_user", SimpleNamespace(is_authenticated=True))
    monkeypatch.setattr(
        "app.cleide_routes.avaliar_autorizacao_operacao_por_franquia",
        lambda _u: {"permitido": False, "modo_operacao": "blocked"},
    )
    _cfg(monkeypatch)
    _tmp_store(monkeypatch, tmp_path)
    client = web.app.test_client()
    resp = client.get("/api/cleide/upload/status")
    assert resp.status_code == 403


def test_upload_clear_exige_autorizacao(monkeypatch, tmp_path):
    web = _load_web_module()
    monkeypatch.setattr("app.cleide_routes.current_user", SimpleNamespace(is_authenticated=True))
    monkeypatch.setattr(
        "app.cleide_routes.avaliar_autorizacao_operacao_por_franquia",
        lambda _u: {"permitido": False, "modo_operacao": "blocked"},
    )
    _cfg(monkeypatch)
    _tmp_store(monkeypatch, tmp_path)
    client = web.app.test_client()
    resp = client.post("/api/cleide/upload/clear")
    assert resp.status_code == 403


def test_upload_lock_concorrencia(monkeypatch, tmp_path):
    web = _load_web_module()
    _authorized(monkeypatch)
    _cfg(monkeypatch)
    _tmp_store(monkeypatch, tmp_path)
    client = web.app.test_client()
    status = client.get("/api/cleide/upload/status").get_json()
    assert status["upload_lock"] is None
    bootstrap = client.post(
        "/api/cleide/upload",
        data={
            "file": (
                io.BytesIO(
                    "transportadora,uf_origem,uf_destino,valor_frete,peso,data_emissao\nseed,sp,rj,5,1,2026-01-01\n".encode(
                        "utf-8"
                    )
                ),
                "base.csv",
            )
        },
        content_type="multipart/form-data",
    )
    assert bootstrap.status_code == 200
    lock = bootstrap.get_json()["upload_lock"]
    resp = client.post(
        "/api/cleide/upload",
        headers={"X-Cleide-Upload-Lock": "wrong-lock"},
        data={
            "file": (
                io.BytesIO(
                    "transportadora,uf_origem,uf_destino,valor_frete,peso,data_emissao\nx,sp,rj,10,1,2026-01-01\n".encode(
                        "utf-8"
                    )
                ),
                "base.csv",
            )
        },
        content_type="multipart/form-data",
    )
    assert resp.status_code == 409
    assert resp.get_json()["error_code"] == "invalid_upload_lock"

    ok = client.post(
        "/api/cleide/upload",
        headers={"X-Cleide-Upload-Lock": lock},
        data={
            "file": (
                io.BytesIO(
                    "transportadora,uf_origem,uf_destino,valor_frete,peso,data_emissao\ny,sp,rj,20,2,2026-01-01\n".encode(
                        "utf-8"
                    )
                ),
                "base.csv",
            )
        },
        content_type="multipart/form-data",
    )
    assert ok.status_code == 200

    clear_fail = client.post(
        "/api/cleide/upload/clear",
        headers={"X-Cleide-Upload-Lock": "wrong-lock"},
    )
    assert clear_fail.status_code == 409
    assert clear_fail.get_json()["error_code"] == "invalid_upload_lock"

    clear_ok = client.post(
        "/api/cleide/upload/clear",
        headers={"X-Cleide-Upload-Lock": lock},
    )
    assert clear_ok.status_code == 200


def test_upload_status_read_only_nao_cria_lock_em_sessao(monkeypatch, tmp_path):
    from app.cleide_contracts import SESSION_KEY_CLEIDE_UPLOAD_LOCK

    web = _load_web_module()
    _authorized(monkeypatch)
    _cfg(monkeypatch)
    _tmp_store(monkeypatch, tmp_path)
    client = web.app.test_client()

    with client.session_transaction() as flask_session:
        assert SESSION_KEY_CLEIDE_UPLOAD_LOCK not in flask_session

    resp = client.get("/api/cleide/upload/status")
    assert resp.status_code == 200
    body = resp.get_json()
    assert body["upload_ativo"] is False
    assert body["upload_lock"] is None

    with client.session_transaction() as flask_session:
        assert SESSION_KEY_CLEIDE_UPLOAD_LOCK not in flask_session


def test_upload_xlsx_sheet_vazia(monkeypatch, tmp_path):
    web = _load_web_module()
    _authorized(monkeypatch)
    _cfg(monkeypatch)
    _tmp_store(monkeypatch, tmp_path)
    client = web.app.test_client()
    payload = _xlsx_bytes_with_headers(
        ("", "", ""),
    )
    resp = client.post(
        "/api/cleide/upload",
        data={"file": (io.BytesIO(payload), "empty.xlsx")},
        content_type="multipart/form-data",
    )
    assert resp.status_code == 400
    assert resp.get_json()["error_code"] == "invalid_xlsx"


def test_upload_xlsx_aliases_validos(monkeypatch, tmp_path):
    web = _load_web_module()
    _authorized(monkeypatch)
    _cfg(monkeypatch)
    _tmp_store(monkeypatch, tmp_path)
    client = web.app.test_client()
    payload = _xlsx_bytes_with_headers(
        ("nome transportadora", "origem_uf", "destino_uf", "vl_frete", "peso_kg", "dt_emissao"),
        ("xp", "sp", "rj", 10, 1, "2026-01-01"),
    )
    resp = client.post(
        "/api/cleide/upload",
        data={"file": (io.BytesIO(payload), "alias.xlsx")},
        content_type="multipart/form-data",
    )
    assert resp.status_code == 200
    body = resp.get_json()
    assert body["dataset_validado"] is True
    assert body["sheet_detectada"] == "Sheet"


def test_cleanup_remove_csv_e_xlsx_suffix(monkeypatch, tmp_path):
    from app.cleide_upload_store import clear_cleide_upload_file

    _tmp_store(monkeypatch, tmp_path)
    ref = "abc123"
    (tmp_path / f"{ref}.csv").write_bytes(b"a,b\n1,2\n")
    (tmp_path / f"{ref}.xlsx").write_bytes(b"xlsx")
    clear_cleide_upload_file(ref)
    assert not (tmp_path / f"{ref}.csv").exists()
    assert not (tmp_path / f"{ref}.xlsx").exists()


def test_ttl_sweep_remove_expirados(monkeypatch, tmp_path):
    import time
    from app.cleide_upload_store import cleanup_expired_cleide_uploads

    _tmp_store(monkeypatch, tmp_path)
    old_file = tmp_path / "oldref.csv"
    old_file.write_bytes(b"a,b\n1,2\n")
    stale = time.time() - 7200
    os.utime(old_file, (stale, stale))
    removed = cleanup_expired_cleide_uploads(30)
    assert removed >= 1
    assert not old_file.exists()


def test_upload_nao_usa_roberto_contracts():
    import app.cleide_upload_pipeline as p

    source = Path(p.__file__).read_text(encoding="utf-8")
    assert "roberto_upload_ref" not in source
    assert "roberto_upload_tmp" not in source
    assert "/api/roberto/" not in source
    assert "run_roberto" not in source
    assert "iaconsumoevento" not in source.lower()


def test_health_status_fase_5():
    import app.cleide_routes as routes

    web = _load_web_module()
    client = web.app.test_client()
    with web.app.app_context():
        resp = client.get("/api/cleide/health")
    assert resp.status_code == 200
    body = resp.get_json()
    assert body["status"] == "ready_local_phase_8_2_controlled_context"
    assert body["status"] != "ready_local_phase_7_filters"


def test_dashboard_filter_endpoint_real_intersection(monkeypatch, tmp_path):
    web = _load_web_module()
    _authorized(monkeypatch)
    _cfg(monkeypatch)
    _tmp_store(monkeypatch, tmp_path)
    client = web.app.test_client()

    payload = (
        "transportadora,uf_origem,uf_destino,valor_frete,peso,data_emissao\n"
        "A,SP,RJ,0,10,2026-01-01\n"
        "A,SP,MG,100,20,2026-01-02\n"
        "B,PR,RJ,0,5,2026-01-03\n"
    ).encode("utf-8")
    up = client.post(
        "/api/cleide/upload",
        data={"file": (io.BytesIO(payload), "base.csv")},
        content_type="multipart/form-data",
    )
    assert up.status_code == 200

    resp = client.post(
        "/api/cleide/dashboard/filter",
        json={"filters": {"transportadora": "A", "uf_destino": "RJ"}},
    )
    assert resp.status_code == 200
    body = resp.get_json()
    assert body["success"] is True
    assert body["kpis"]["total_documentos"] == 1
    assert body["kpis"]["valor_total_frete"] == 0.0
    assert body["active_filters"]["transportadora"] == "A"
    assert body["active_filters"]["uf_destino"] == "RJ"
    assert body["active_filters"]["uf_origem"] is None
    assert body["active_filters"]["data_inicio"] is None
    assert body["active_filters"]["data_fim"] is None
    assert body["transportadora_stats"][0]["chave"] == "A"
    assert body["uf_destino_stats"][0]["chave"] == "RJ"
    assert body["uf_origem_stats"][0]["chave"] == "SP"
    assert body["temporal_stats"][0]["data"] == "2026-01-01"
    assert body["pareto_fretes_zerados_uf_destino"][0]["chave"] == "RJ"
    assert body["pareto_fretes_zerados_transportadora"][0]["chave"] == "A"
    assert "rows" not in body
    assert "raw_bytes" not in body
    assert "analytics_context" not in body
    contexto = body["cleide_contexto_operacional"]
    assert contexto["filter_context"]["filter_mode"] == "row_level_intersection_backend"
    assert contexto["filter_context"]["kpi_scope"] == "filtered_session_intersection"
    assert contexto["semantic_limits"]["no_row_level_intersection"] is False
    assert contexto["semantic_limits"]["multi_dimension_filters_are_approximate"] is False
    assert contexto["semantic_limits"]["kpis_are_global_session_scope"] is False


def test_dashboard_filter_falha_sem_upload(monkeypatch, tmp_path):
    web = _load_web_module()
    _authorized(monkeypatch)
    _cfg(monkeypatch)
    _tmp_store(monkeypatch, tmp_path)
    client = web.app.test_client()
    resp = client.post("/api/cleide/dashboard/filter", json={"filters": {"transportadora": "A"}})
    assert resp.status_code == 400
    body = resp.get_json()
    assert body["success"] is False
    assert body["error_code"] == "missing_upload_ref"


def test_dashboard_filter_exige_login(monkeypatch, tmp_path):
    web = _load_web_module()
    monkeypatch.setattr("app.cleide_routes.current_user", SimpleNamespace(is_authenticated=False))
    _cfg(monkeypatch)
    _tmp_store(monkeypatch, tmp_path)
    client = web.app.test_client()
    resp = client.post("/api/cleide/dashboard/filter", json={"filters": {"transportadora": "A"}})
    assert resp.status_code == 401
    assert resp.get_json()["error"] == "Autenticacao necessaria."


def test_dashboard_filter_falha_com_stale_upload_ref(monkeypatch, tmp_path):
    web = _load_web_module()
    _authorized(monkeypatch)
    _cfg(monkeypatch)
    _tmp_store(monkeypatch, tmp_path)
    client = web.app.test_client()

    with client.session_transaction() as sess:
        sess["cleide_upload_ref"] = "stale-ref"
        sess["cleide_dataset_context"] = {
            "dataset_validado": True,
            "aliases_resolvidos": {
                "transportadora": "transportadora",
                "uf_origem": "uf_origem",
                "uf_destino": "uf_destino",
                "valor_frete": "valor_frete",
                "peso": "peso",
                "data_emissao": "data_emissao",
            },
            "raw_headers": [
                "transportadora",
                "uf_origem",
                "uf_destino",
                "valor_frete",
                "peso",
                "data_emissao",
            ],
        }

    resp = client.post("/api/cleide/dashboard/filter", json={"filters": {"transportadora": "A"}})
    assert resp.status_code == 409
    body = resp.get_json()
    assert body["success"] is False
    assert body["error_code"] == "stale_upload_ref"
    assert "kpis" not in body
    assert "transportadora_stats" not in body
    assert "analytics_context" not in body
