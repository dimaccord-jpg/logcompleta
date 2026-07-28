"""Contrato do arquivo operacional do AgenteCompara (Arquivo para Comparação)."""
from __future__ import annotations

import importlib
import io
import os
from types import SimpleNamespace
from unittest.mock import MagicMock

import openpyxl
import pytest

from app.agente_compara_comparison_state import (
    AGENTE_COMPARA_COMPARISON_STATE_SESSION_KEY,
    STEP_CALCULATION_FILE,
    STEP_CONFIGURATION_READY,
    get_comparison_state,
    get_table_by_slot,
)
from app.agente_compara_doc_service import (
    AGENTE_COMPARA_TEMP_TABLE_ID_SESSION_KEY,
    AGENTE_COMPARA_TEMPLATE_FILENAME,
    AUDIT_BATCH_SHEET_NAME,
    AgenteComparaBatchError,
    ERROR_AUDIT_MISSING_COLUMNS,
    ERROR_AUDIT_PARSE_FAILED,
    _AUDIT_REQUIRED_FIELDS,
    _parse_audit_csv_bytes,
    _parse_audit_xlsx_bytes,
)
from app.services.agente_compara_config_service import (
    AgenteComparaConfig,
    DEFAULT_FALLBACK_MESSAGE,
)
from tests.cleiton_doc_fixtures import make_audit_xlsx, make_csv, patch_cleiton_doc_cfg, patch_cleiton_doc_store
from tests.test_agente_compara_comparison_journey import _set_comparison_at_step


EXPECTED_HEADERS = [
    "numero_documento",
    "cidade_origem",
    "uf_origem",
    "cidade_destino",
    "uf_destino",
    "valor_nf",
    "peso",
    "modal",
    "data_emissao",
]

NEW_CONTRACT_HEADERS = list(EXPECTED_HEADERS)
# Arquivo antigo (11 colunas) com transportadora + data_entrega ainda tolerado.
LEGACY_HEADERS_OLD_CONTRACT = [
    "transportadora",
    "numero_documento",
    "cidade_origem",
    "uf_origem",
    "cidade_destino",
    "uf_destino",
    "valor_nf",
    "peso",
    "modal",
    "data_emissao",
    "data_entrega",
]
# Legado com valor_frete adicional (já aprovado) + transportadora/data_entrega.
LEGACY_HEADERS = [
    "transportadora",
    "numero_documento",
    "cidade_origem",
    "uf_origem",
    "cidade_destino",
    "uf_destino",
    "valor_nf",
    "valor_frete",
    "peso",
    "modal",
    "data_emissao",
    "data_entrega",
]


def _load_web_module():
    os.environ.setdefault("APP_ENV", "dev")
    os.environ.setdefault("DATABASE_URL", "postgresql://user:pass@localhost:5432/testdb")
    os.environ.setdefault("SECRET_KEY", "test-secret")
    return importlib.import_module("app.web")


def _patch_ac_cfg(monkeypatch):
    cfg = AgenteComparaConfig(
        chat_enabled=True,
        upload_enabled=True,
        chat_max_history=10,
        document_context_max_chars=24000,
        max_documents_considered=3,
        question_max_chars=4000,
        fallback_message=DEFAULT_FALLBACK_MESSAGE,
        no_documents_behavior="allow_guided",
        show_documents_used=True,
        no_hallucination_instruction_enabled=True,
        audited_file_max_bytes=None,
        audited_file_max_rows=2000,
    )
    for target in (
        "app.agente_compara_api_routes.get_agente_compara_config",
        "app.agente_compara_doc_service.get_agente_compara_config",
        "app.agente_compara_doc_context.get_agente_compara_config",
    ):
        monkeypatch.setattr(target, lambda _cfg=cfg: _cfg)
    return cfg


def _setup_env(monkeypatch, tmp_path):
    patch_cleiton_doc_store(tmp_path, monkeypatch)
    cfg = patch_cleiton_doc_cfg(monkeypatch)
    monkeypatch.setattr("app.agente_compara_doc_service.get_cleiton_doc_config", lambda: cfg)
    monkeypatch.setattr("app.agente_compara_api_routes.get_cleiton_doc_config", lambda: cfg)
    monkeypatch.setattr("app.agente_compara_doc_context.get_cleiton_doc_config", lambda: cfg)
    _patch_ac_cfg(monkeypatch)
    return cfg


def _authorized(monkeypatch, web):
    fake_user = SimpleNamespace(is_authenticated=True, conta_id=1, franquia_id=1, id=1)
    monkeypatch.setattr(web, "current_user", fake_user)
    monkeypatch.setattr("app.agente_compara_api_routes.current_user", fake_user)
    monkeypatch.setattr(
        web,
        "avaliar_autorizacao_operacao_por_franquia",
        lambda _u: {"permitido": True, "modo_operacao": "normal"},
    )
    monkeypatch.setattr(
        "app.agente_compara_api_routes.avaliar_autorizacao_operacao_por_franquia",
        lambda _u: {"permitido": True, "modo_operacao": "normal"},
    )


@pytest.fixture
def web_client(app, tmp_path, monkeypatch, ctx):
    with app.app_context():
        _setup_env(monkeypatch, tmp_path)
    web = _load_web_module()
    _authorized(monkeypatch, web)
    web.app.config["TESTING"] = True
    web.app.config["SECRET_KEY"] = "test-secret"
    return web.app.test_client()


def _sample_row(**overrides) -> list[str]:
    row = {
        "numero_documento": "123",
        "cidade_origem": "São Paulo",
        "uf_origem": "SP",
        "cidade_destino": "Campinas",
        "uf_destino": "SP",
        "valor_nf": "1000",
        "peso": "48",
        "modal": "Rodo",
        "data_emissao": "2024-01-01",
    }
    row.update(overrides)
    return [row[h] for h in NEW_CONTRACT_HEADERS]


def _sample_legacy_old_contract_row(**overrides) -> list[str]:
    base = {
        "transportadora": "Transportadora X",
        "numero_documento": "123",
        "cidade_origem": "São Paulo",
        "uf_origem": "SP",
        "cidade_destino": "Campinas",
        "uf_destino": "SP",
        "valor_nf": "1000",
        "peso": "48",
        "modal": "Rodo",
        "data_emissao": "2024-01-01",
        "data_entrega": "2024-01-05",
    }
    base.update(overrides)
    return [base[h] for h in LEGACY_HEADERS_OLD_CONTRACT]


def _sample_legacy_row(**overrides) -> list[str]:
    base = {
        "transportadora": "Transportadora X",
        "numero_documento": "123",
        "cidade_origem": "São Paulo",
        "uf_origem": "SP",
        "cidade_destino": "Campinas",
        "uf_destino": "SP",
        "valor_nf": "1000",
        "valor_frete": "100.5",
        "peso": "48",
        "modal": "Rodo",
        "data_emissao": "2024-01-01",
        "data_entrega": "2024-01-05",
    }
    base.update(overrides)
    return [base[h] for h in LEGACY_HEADERS]


def _csv_new(*rows) -> bytes:
    data = rows or [_sample_row()]
    return make_csv([NEW_CONTRACT_HEADERS, *data])


def _csv_legacy(*rows) -> bytes:
    data = rows or [_sample_legacy_row()]
    return make_csv([LEGACY_HEADERS, *data])


def _csv_legacy_old_contract(*rows) -> bytes:
    data = rows or [_sample_legacy_old_contract_row()]
    return make_csv([LEGACY_HEADERS_OLD_CONTRACT, *data])


def _xlsx_new(*rows) -> bytes:
    data = rows or [_sample_row()]
    return make_audit_xlsx(
        [NEW_CONTRACT_HEADERS, *data],
        sheet_name=AUDIT_BATCH_SHEET_NAME,
    )


def _xlsx_legacy(*rows) -> bytes:
    data = rows or [_sample_legacy_row()]
    return make_audit_xlsx(
        [LEGACY_HEADERS, *data],
        sheet_name=AUDIT_BATCH_SHEET_NAME,
    )


def _operational_fields(row: dict) -> dict:
    keys = (
        "document_number",
        "destination_city",
        "destination_uf",
        "audited_weight",
        "origin_city",
        "origin_uf",
        "invoice_value",
        "modal",
        "issue_date",
    )
    return {k: row.get(k) for k in keys}


def test_template_xlsx_real_workbook_contract():
    path = f"app/protected_files/templates/{AGENTE_COMPARA_TEMPLATE_FILENAME}"
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        assert wb.sheetnames == [AUDIT_BATCH_SHEET_NAME]
        header = [
            str(cell).strip() if cell is not None else ""
            for cell in next(wb[AUDIT_BATCH_SHEET_NAME].iter_rows(values_only=True))
        ]
    finally:
        wb.close()
    assert header == EXPECTED_HEADERS
    assert len(header) == 9
    assert "" not in header
    assert "transportadora" not in header
    assert "data_entrega" not in header
    assert "valor_frete" not in header

    cleide = openpyxl.load_workbook(
        "app/protected_files/templates/template_cleide_auditoria_frete.xlsx",
        read_only=True,
        data_only=True,
    )
    try:
        assert "Modelo Cleide" in cleide.sheetnames
        cleide_header = [
            str(cell).strip() if cell is not None else ""
            for cell in next(cleide["Modelo Cleide"].iter_rows(values_only=True))
        ]
    finally:
        cleide.close()
    assert "valor_frete" in cleide_header
    assert "transportadora" in cleide_header


def test_template_download_route_returns_agente_compara_workbook(web_client):
    # Política atual: download do modelo não exige autenticação na rota.
    resp = web_client.get("/api/agente-compara/audit-template")
    assert resp.status_code == 200
    assert (
        resp.headers.get("Content-Disposition", "").find(AGENTE_COMPARA_TEMPLATE_FILENAME) >= 0
    )
    wb = openpyxl.load_workbook(io.BytesIO(resp.data), read_only=True, data_only=True)
    try:
        assert AUDIT_BATCH_SHEET_NAME in wb.sheetnames
        header = [
            str(cell).strip() if cell is not None else ""
            for cell in next(wb[AUDIT_BATCH_SHEET_NAME].iter_rows(values_only=True))
        ]
    finally:
        wb.close()
    assert header == EXPECTED_HEADERS
    assert "transportadora" not in header
    assert "data_entrega" not in header
    assert "valor_frete" not in header


def test_required_fields_exclude_valor_frete():
    assert "charged_freight" not in _AUDIT_REQUIRED_FIELDS
    assert "carrier" not in _AUDIT_REQUIRED_FIELDS
    assert "delivery_date" not in _AUDIT_REQUIRED_FIELDS
    assert "document_number" in _AUDIT_REQUIRED_FIELDS
    assert "destination_city" in _AUDIT_REQUIRED_FIELDS
    assert "destination_uf" in _AUDIT_REQUIRED_FIELDS
    assert "audited_weight" in _AUDIT_REQUIRED_FIELDS


def test_parse_xlsx_valid_without_valor_frete():
    rows, header_map, sheet = _parse_audit_xlsx_bytes(
        _xlsx_new(),
        source_file_name="ops.xlsx",
        max_bytes=5_000_000,
        max_rows=2000,
    )
    assert sheet == AUDIT_BATCH_SHEET_NAME
    assert len(rows) == 1
    assert rows[0]["document_number"] == "123"
    assert rows[0]["audited_weight"] == 48.0
    assert "charged_freight" not in rows[0]
    assert "carrier" not in rows[0]
    assert "delivery_date" not in rows[0]
    assert "valor_frete" not in header_map.values()
    assert "transportadora" not in header_map
    assert "data_entrega" not in header_map


def test_parse_csv_valid_without_valor_frete():
    rows, header_map, _ = _parse_audit_csv_bytes(
        _csv_new(),
        source_file_name="ops.csv",
        max_rows=2000,
    )
    assert len(rows) == 1
    assert rows[0]["destination_city"] == "Campinas"
    assert "charged_freight" not in rows[0]
    assert "carrier" not in rows[0]
    assert "delivery_date" not in rows[0]


def test_parse_legacy_file_with_valor_frete_accepted_and_ignored():
    rows, _, _ = _parse_audit_csv_bytes(
        _csv_legacy(),
        source_file_name="legacy.csv",
        max_rows=2000,
    )
    assert len(rows) == 1
    assert "charged_freight" not in rows[0]
    assert "carrier" not in rows[0]
    assert "delivery_date" not in rows[0]
    assert rows[0]["document_number"] == "123"


def test_parse_legacy_old_contract_with_transportadora_and_data_entrega_ignored():
    rows, header_map, _ = _parse_audit_csv_bytes(
        _csv_legacy_old_contract(),
        source_file_name="legacy_old.csv",
        max_rows=2000,
    )
    assert len(rows) == 1
    assert "carrier" not in rows[0]
    assert "delivery_date" not in rows[0]
    assert "charged_freight" not in rows[0]
    assert rows[0]["document_number"] == "123"
    assert rows[0]["audited_weight"] == 48.0
    assert "transportadora" not in header_map
    assert "data_entrega" not in header_map


def test_parse_extra_columns_tolerated():
    headers = NEW_CONTRACT_HEADERS + ["coluna_extra", "observacao"]
    row = _sample_row() + ["x", "y"]
    rows, _, _ = _parse_audit_csv_bytes(
        make_csv([headers, row]),
        source_file_name="extra.csv",
        max_rows=2000,
    )
    assert len(rows) == 1
    assert rows[0]["document_number"] == "123"


def test_parse_missing_numero_documento_column():
    headers = [h for h in NEW_CONTRACT_HEADERS if h != "numero_documento"]
    row = [v for h, v in zip(NEW_CONTRACT_HEADERS, _sample_row()) if h != "numero_documento"]
    with pytest.raises(AgenteComparaBatchError) as exc:
        _parse_audit_csv_bytes(make_csv([headers, row]), source_file_name="x.csv", max_rows=2000)
    assert exc.value.error_code == ERROR_AUDIT_MISSING_COLUMNS
    assert "numero_documento" in exc.value.message


def test_parse_missing_cidade_destino_column():
    headers = [h for h in NEW_CONTRACT_HEADERS if h != "cidade_destino"]
    row = [v for h, v in zip(NEW_CONTRACT_HEADERS, _sample_row()) if h != "cidade_destino"]
    with pytest.raises(AgenteComparaBatchError) as exc:
        _parse_audit_csv_bytes(make_csv([headers, row]), source_file_name="x.csv", max_rows=2000)
    assert exc.value.error_code == ERROR_AUDIT_MISSING_COLUMNS
    assert "cidade_destino" in exc.value.message


def test_parse_missing_uf_destino_column():
    headers = [h for h in NEW_CONTRACT_HEADERS if h != "uf_destino"]
    row = [v for h, v in zip(NEW_CONTRACT_HEADERS, _sample_row()) if h != "uf_destino"]
    with pytest.raises(AgenteComparaBatchError) as exc:
        _parse_audit_csv_bytes(make_csv([headers, row]), source_file_name="x.csv", max_rows=2000)
    assert exc.value.error_code == ERROR_AUDIT_MISSING_COLUMNS
    assert "uf_destino" in exc.value.message


def test_parse_missing_peso_column():
    headers = [h for h in NEW_CONTRACT_HEADERS if h != "peso"]
    row = [v for h, v in zip(NEW_CONTRACT_HEADERS, _sample_row()) if h != "peso"]
    with pytest.raises(AgenteComparaBatchError) as exc:
        _parse_audit_csv_bytes(make_csv([headers, row]), source_file_name="x.csv", max_rows=2000)
    assert exc.value.error_code == ERROR_AUDIT_MISSING_COLUMNS
    assert "peso" in exc.value.message


@pytest.mark.parametrize(
    "peso",
    ["", "abc", "-1"],
)
def test_parse_invalid_peso_row(peso):
    with pytest.raises(AgenteComparaBatchError) as exc:
        _parse_audit_csv_bytes(
            _csv_new(_sample_row(peso=peso)),
            source_file_name="x.csv",
            max_rows=2000,
        )
    assert exc.value.error_code == ERROR_AUDIT_PARSE_FAILED


def test_parse_empty_numero_documento_on_data_row():
    with pytest.raises(AgenteComparaBatchError) as exc:
        _parse_audit_csv_bytes(
            _csv_new(_sample_row(numero_documento="")),
            source_file_name="x.csv",
            max_rows=2000,
        )
    assert exc.value.error_code == ERROR_AUDIT_PARSE_FAILED


def test_parse_empty_valor_nf_accepted():
    rows, _, _ = _parse_audit_csv_bytes(
        _csv_new(_sample_row(valor_nf="")),
        source_file_name="x.csv",
        max_rows=2000,
    )
    assert len(rows) == 1
    assert "invoice_value" not in rows[0]


def test_parse_completely_empty_row_ignored():
    empty = [""] * len(NEW_CONTRACT_HEADERS)
    rows, _, _ = _parse_audit_csv_bytes(
        make_csv([NEW_CONTRACT_HEADERS, empty, _sample_row()]),
        source_file_name="x.csv",
        max_rows=2000,
    )
    assert len(rows) == 1
    assert rows[0]["row_index"] == 1


def test_parse_duplicate_document_numbers_preserved():
    rows, _, _ = _parse_audit_csv_bytes(
        _csv_new(_sample_row(numero_documento="100"), _sample_row(numero_documento="100")),
        source_file_name="x.csv",
        max_rows=2000,
    )
    assert len(rows) == 2
    assert rows[0]["document_number"] == "100"
    assert rows[1]["document_number"] == "100"


def test_parse_aliases_preserved():
    headers = [
        "nome transportadora",
        "numero do documento",
        "cidade de origem",
        "estado origem",
        "cidade de destino",
        "estado destino",
        "valor da nf",
        "peso kg",
        "modalidade",
        "data de emissao",
        "data de entrega",
    ]
    row = [
        "Carrier Y",
        "999",
        "Santos",
        "SP",
        "Curitiba",
        "PR",
        "1500",
        "20",
        "Aereo",
        "2024-02-01",
        "2024-02-03",
    ]
    rows, _, _ = _parse_audit_csv_bytes(
        make_csv([headers, row]),
        source_file_name="alias.csv",
        max_rows=2000,
    )
    assert "carrier" not in rows[0]
    assert "delivery_date" not in rows[0]
    assert rows[0]["document_number"] == "999"
    assert rows[0]["destination_city"] == "Curitiba"
    assert rows[0]["destination_uf"] == "PR"
    assert rows[0]["audited_weight"] == 20.0
    assert rows[0]["invoice_value"] == 1500.0
    assert rows[0]["origin_city"] == "Santos"
    assert rows[0]["issue_date"] == "2024-02-01"


def test_legacy_and_new_contract_produce_same_operational_fields():
    new_rows, _, _ = _parse_audit_csv_bytes(
        _csv_new(),
        source_file_name="a.csv",
        max_rows=2000,
    )
    legacy_rows, _, _ = _parse_audit_csv_bytes(
        _csv_legacy(),
        source_file_name="b.csv",
        max_rows=2000,
    )
    legacy_old_rows, _, _ = _parse_audit_csv_bytes(
        _csv_legacy_old_contract(),
        source_file_name="c.csv",
        max_rows=2000,
    )
    assert _operational_fields(new_rows[0]) == _operational_fields(legacy_rows[0])
    assert _operational_fields(new_rows[0]) == _operational_fields(legacy_old_rows[0])
    assert "charged_freight" not in new_rows[0]
    assert "charged_freight" not in legacy_rows[0]
    assert "carrier" not in new_rows[0]
    assert "carrier" not in legacy_rows[0]
    assert "carrier" not in legacy_old_rows[0]
    assert "delivery_date" not in new_rows[0]
    assert "delivery_date" not in legacy_rows[0]
    assert "delivery_date" not in legacy_old_rows[0]


def test_parse_optional_columns_absent_accepted():
    headers = ["numero_documento", "cidade_destino", "uf_destino", "peso"]
    row = ["123", "Campinas", "SP", "48"]
    rows, _, _ = _parse_audit_csv_bytes(
        make_csv([headers, row]),
        source_file_name="min.csv",
        max_rows=2000,
    )
    assert len(rows) == 1
    assert rows[0]["document_number"] == "123"
    assert "origin_city" not in rows[0]
    assert "invoice_value" not in rows[0]
    assert "modal" not in rows[0]
    assert "issue_date" not in rows[0]
    assert "carrier" not in rows[0]
    assert "delivery_date" not in rows[0]


def test_transportadora_and_data_entrega_not_required():
    assert "carrier" not in _AUDIT_REQUIRED_FIELDS
    assert "delivery_date" not in _AUDIT_REQUIRED_FIELDS
    rows, _, _ = _parse_audit_csv_bytes(
        _csv_new(),
        source_file_name="ok.csv",
        max_rows=2000,
    )
    assert "carrier" not in rows[0]
    assert "delivery_date" not in rows[0]


def test_upload_new_contract_xlsx_without_calculation_or_gemini(web_client, monkeypatch):
    run_mock = MagicMock(side_effect=AssertionError("cálculo não deve rodar"))
    gemini_mock = MagicMock(side_effect=AssertionError("Gemini não deve ser chamado"))
    monkeypatch.setattr("app.agente_compara_doc_service.run_audit_batch_for_session", run_mock)
    monkeypatch.setattr(
        "app.run_agente_compara_temp_table.cleiton_governed_generate_content",
        gemini_mock,
    )
    monkeypatch.setattr(
        "app.agente_compara_doc_service.compute_audit_outputs",
        MagicMock(side_effect=AssertionError("compute_audit_outputs não deve rodar")),
    )

    primary = _set_comparison_at_step(web_client, STEP_CALCULATION_FILE)
    resp = web_client.post(
        "/api/agente-compara/audit/upload",
        data={"file": (io.BytesIO(_xlsx_new()), "operacional.xlsx")},
        content_type="multipart/form-data",
    )
    assert resp.status_code == 200, resp.get_json()
    body = resp.get_json()
    batch = body["temp_table"]["audit_batch"]
    assert batch["status"] == "uploaded"
    assert batch["source_file_name"] == "operacional.xlsx"
    assert batch["row_count"] == 1
    assert batch["sheet_name"] == AUDIT_BATCH_SHEET_NAME
    assert not batch.get("results")
    assert body["temp_table"]["comparison"]["current_step"] == STEP_CONFIGURATION_READY
    assert primary["temp_table_id"]
    run_mock.assert_not_called()
    assert gemini_mock.call_count == 0


def test_upload_legacy_csv_with_valor_frete_accepted(web_client, monkeypatch):
    monkeypatch.setattr(
        "app.agente_compara_doc_service.run_audit_batch_for_session",
        MagicMock(side_effect=AssertionError("cálculo não deve rodar")),
    )
    _set_comparison_at_step(web_client, STEP_CALCULATION_FILE)
    resp = web_client.post(
        "/api/agente-compara/audit/upload",
        data={"file": (io.BytesIO(_csv_legacy()), "legado.csv", "text/csv")},
        content_type="multipart/form-data",
    )
    assert resp.status_code == 200, resp.get_json()
    batch = resp.get_json()["temp_table"]["audit_batch"]
    assert batch["status"] == "uploaded"
    assert batch["row_count"] == 1
    assert not batch.get("results")


def test_upload_rejects_missing_peso_without_requiring_valor_frete(web_client):
    _set_comparison_at_step(web_client, STEP_CALCULATION_FILE)
    headers = [h for h in NEW_CONTRACT_HEADERS if h != "peso"]
    row = [v for h, v in zip(NEW_CONTRACT_HEADERS, _sample_row()) if h != "peso"]
    resp = web_client.post(
        "/api/agente-compara/audit/upload",
        data={"file": (io.BytesIO(make_csv([headers, row])), "sem_peso.csv", "text/csv")},
        content_type="multipart/form-data",
    )
    assert resp.status_code == 400
    body = resp.get_json()
    assert body["error_code"] == ERROR_AUDIT_MISSING_COLUMNS
    assert "peso" in body["message"]
    assert "valor_frete" not in body["message"]


def test_upload_legacy_old_contract_ignores_transportadora_and_data_entrega(web_client, monkeypatch):
    monkeypatch.setattr(
        "app.agente_compara_doc_service.run_audit_batch_for_session",
        MagicMock(side_effect=AssertionError("cálculo não deve rodar")),
    )
    primary = _set_comparison_at_step(web_client, STEP_CALCULATION_FILE)
    resp = web_client.post(
        "/api/agente-compara/audit/upload",
        data={"file": (io.BytesIO(_csv_legacy_old_contract()), "legado_old.csv", "text/csv")},
        content_type="multipart/form-data",
    )
    assert resp.status_code == 200, resp.get_json()
    batch = resp.get_json()["temp_table"]["audit_batch"]
    assert batch["status"] == "uploaded"
    assert batch["row_count"] == 1
    header_map = batch.get("header_map") or {}
    assert "transportadora" not in header_map
    assert "data_entrega" not in header_map
    assert "valor_frete" not in header_map
    assert "numero_documento" in header_map or "document_number" in header_map.values()

    from app.agente_compara_doc_service import load_temp_table_record

    record = load_temp_table_record(primary["temp_table_id"], ttl_hours=24)
    assert record is not None
    stored_rows = (record.get("audit_batch") or {}).get("normalized_rows") or []
    assert len(stored_rows) == 1
    assert "carrier" not in stored_rows[0]
    assert "delivery_date" not in stored_rows[0]
    assert "charged_freight" not in stored_rows[0]


def test_fingerprint_ignores_legacy_transportadora_and_data_entrega():
    from app.agente_compara_calculation_execution_service import (
        build_calculation_fingerprint_payload,
        compute_calculation_fingerprint,
    )

    new_rows, _, _ = _parse_audit_csv_bytes(
        _csv_new(),
        source_file_name="ops.csv",
        max_rows=2000,
    )
    legacy_rows, _, _ = _parse_audit_csv_bytes(
        _csv_legacy_old_contract(),
        source_file_name="ops.csv",
        max_rows=2000,
    )
    assert "carrier" not in new_rows[0]
    assert "carrier" not in legacy_rows[0]
    assert "delivery_date" not in new_rows[0]
    assert "delivery_date" not in legacy_rows[0]

    state = {
        "comparison_id": "cmp-fp-contract",
        "tables": {
            "t1": {
                "table_id": "t1",
                "temp_table_id": "tt1",
                "slot_number": 1,
                "status": "confirmed",
                "confirmed": True,
                "carrier_name": "Alpha",
            },
            "t2": {
                "table_id": "t2",
                "temp_table_id": "tt2",
                "slot_number": 2,
                "status": "confirmed",
                "confirmed": True,
                "carrier_name": "Beta",
            },
        },
    }
    source_file_identity = {
        "audit_batch_id": "ab1",
        "source_file_name": "ops.csv",
        "sheet_name": None,
        "row_count": 1,
        "input_schema_version": 1,
        "temp_table_id": "tt1",
    }
    table_records = {
        "tt1": {"edit_version": 1, "updated_at": "2024-01-01T00:00:00Z", "human_review_status": "confirmed"},
        "tt2": {"edit_version": 1, "updated_at": "2024-01-01T00:00:00Z", "human_review_status": "confirmed"},
    }

    fp_new = compute_calculation_fingerprint(
        build_calculation_fingerprint_payload(
            comparison_id="cmp-fp-contract",
            state=state,
            normalized_rows=new_rows,
            table_records=table_records,
            tax_config=None,
            coverage_table={},
            source_file_identity=source_file_identity,
        )
    )
    fp_legacy = compute_calculation_fingerprint(
        build_calculation_fingerprint_payload(
            comparison_id="cmp-fp-contract",
            state=state,
            normalized_rows=legacy_rows,
            table_records=table_records,
            tax_config=None,
            coverage_table={},
            source_file_identity=source_file_identity,
        )
    )
    assert fp_new == fp_legacy

    mutated = [dict(new_rows[0])]
    mutated[0]["audited_weight"] = 99.0
    fp_mutated = compute_calculation_fingerprint(
        build_calculation_fingerprint_payload(
            comparison_id="cmp-fp-contract",
            state=state,
            normalized_rows=mutated,
            table_records=table_records,
            tax_config=None,
            coverage_table={},
            source_file_identity=source_file_identity,
        )
    )
    assert fp_mutated != fp_new

    mutated_dest = [dict(new_rows[0])]
    mutated_dest[0]["destination_city"] = "Outra Cidade"
    fp_dest = compute_calculation_fingerprint(
        build_calculation_fingerprint_payload(
            comparison_id="cmp-fp-contract",
            state=state,
            normalized_rows=mutated_dest,
            table_records=table_records,
            tax_config=None,
            coverage_table={},
            source_file_identity=source_file_identity,
        )
    )
    assert fp_dest != fp_new