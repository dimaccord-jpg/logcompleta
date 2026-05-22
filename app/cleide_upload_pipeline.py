from __future__ import annotations

import io
import logging
import time
import csv
from typing import Any
from uuid import uuid4

from flask import current_app, jsonify, request, session
from werkzeug.datastructures import FileStorage
from werkzeug.utils import secure_filename

from app.cleide_contracts import (
    clear_cleide_dataset_context,
    clear_cleide_upload_in_progress,
    get_cleide_upload_lock,
    clear_cleide_upload_ref,
    get_cleide_dataset_context,
    get_or_create_cleide_upload_lock,
    get_cleide_upload_ref,
    is_cleide_upload_in_progress,
    mark_cleide_upload_in_progress,
    set_cleide_dataset_context,
    set_cleide_upload_ref,
)
from app.cleide_analytics import AnalyticsProcessingError, build_analytics_context, build_filtered_analytics_context
from app.cleide_operational_context import build_cleide_operational_context
from app.cleide_structural_validation import (
    StructuralValidationError,
    analyze_structural_layout,
)
from app.cleide_upload_store import (
    clear_cleide_upload_file,
    get_upload_ref_extension,
    maybe_cleanup_expired_cleide_uploads,
    resolve_cleide_upload_file,
    save_cleide_upload_file,
)
from app.services.cleide_config_service import get_cleide_config
from app.services.cleiton_upload_billing_service import apropriar_billing_upload_cleide

logger = logging.getLogger(__name__)

ALLOWED_EXTENSIONS = {".xlsx", ".csv"}
ALLOWED_MIME = {
    ".csv": {
        "text/csv",
        "application/csv",
        "text/plain",
        "application/vnd.ms-excel",
    },
    ".xlsx": {
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        "application/zip",
    },
}

_CLEIDE_FILTER_KEYS = (
    "transportadora",
    "uf_origem",
    "uf_destino",
    "data_inicio",
    "data_fim",
)


def _count_non_empty_data_rows(*, raw_bytes: bytes, extension: str, delimiter_default: str) -> int:
    ext = (extension or "").strip().lower()
    if ext == ".csv":
        return _count_csv_data_rows(raw_bytes=raw_bytes, delimiter_default=delimiter_default)
    if ext == ".xlsx":
        return _count_xlsx_data_rows(raw_bytes=raw_bytes)
    return 0


def _count_csv_data_rows(*, raw_bytes: bytes, delimiter_default: str) -> int:
    from app.cleide_structural_validation import _decode_csv, _detect_delimiter

    try:
        decoded_text, _encoding = _decode_csv(raw_bytes)
    except Exception:
        return 0
    if not decoded_text.strip():
        return 0
    delimiter = _detect_delimiter(decoded_text, delimiter_default)
    try:
        rows = csv.reader(io.StringIO(decoded_text), delimiter=delimiter)
        return _count_non_empty_rows_excluding_header(rows)
    except csv.Error:
        return 0


def _count_xlsx_data_rows(*, raw_bytes: bytes) -> int:
    import zipfile
    from app.cleide_structural_validation import StructuralValidationError, _extract_tabular_structure
    try:
        import openpyxl
    except ImportError:
        return 0

    try:
        workbook = openpyxl.load_workbook(io.BytesIO(raw_bytes), read_only=True, data_only=True)
    except (OSError, ValueError, KeyError, RuntimeError, zipfile.BadZipFile):
        return 0

    try:
        for name in workbook.sheetnames:
            sheet = workbook[name]
            try:
                base = _extract_tabular_structure(
                    rows=sheet.iter_rows(values_only=True),
                    max_rows=200000,
                    max_columns=500,
                )
            except StructuralValidationError:
                continue
            if len(base.get("canonical_headers") or []) < 2:
                continue
            rows_count = int(base.get("linhas_detectadas") or 0)
            if rows_count > 0:
                return rows_count
    finally:
        workbook.close()
    return 0


def _count_non_empty_rows_excluding_header(rows) -> int:
    header_found = False
    data_rows = 0
    for row in rows:
        normalized_row = [str(cell).strip() if cell is not None else "" for cell in row]
        while normalized_row and not normalized_row[-1]:
            normalized_row.pop()
        if not any(cell.strip() for cell in normalized_row):
            continue
        if not header_found:
            header_found = True
            continue
        data_rows += 1
    return data_rows


def _json_error(message: str, *, code: str, status: int = 400):
    return jsonify({"success": False, "error": message, "error_code": code}), status


def _extract_file() -> FileStorage | None:
    return request.files.get("file") or request.files.get("arquivo")


def _stream_size(file: FileStorage) -> int:
    stream = file.stream
    cur = stream.tell()
    stream.seek(0, io.SEEK_END)
    end = stream.tell()
    stream.seek(cur, io.SEEK_SET)
    return max(0, int(end))


def _validate_extension(filename: str) -> str | None:
    safe_name = secure_filename(filename or "")
    if not safe_name:
        return None
    dot = safe_name.rfind(".")
    if dot < 0:
        return None
    ext = safe_name[dot:].lower()
    if ext not in ALLOWED_EXTENSIONS:
        return None
    return ext


def _validate_mime(ext: str, mimetype: str | None) -> bool:
    if not mimetype:
        return True
    return (mimetype or "").strip().lower() in ALLOWED_MIME.get(ext, set())


def _read_bytes(file: FileStorage) -> bytes:
    stream = file.stream
    stream.seek(0, io.SEEK_SET)
    raw = stream.read() or b""
    stream.seek(0, io.SEEK_SET)
    return raw


def _normalize_dashboard_filters(filters: dict[str, Any] | None) -> dict[str, str | None]:
    data = filters if isinstance(filters, dict) else {}
    out: dict[str, str | None] = {}
    for key in _CLEIDE_FILTER_KEYS:
        raw = data.get(key)
        if raw is None:
            out[key] = None
            continue
        text = str(raw).strip()
        out[key] = text if text else None
    return out


def _resolve_cleide_execution_id() -> str:
    execution_id = (request.headers.get("X-Execution-ID") or "").strip()
    if not execution_id:
        execution_id = (request.form.get("execution_id") or "").strip()
    if not execution_id:
        execution_id = str(uuid4())
    return execution_id[:120]


def process_cleide_upload() -> tuple[Any, int]:
    lock_token = get_or_create_cleide_upload_lock(session)
    request_token = (
        (request.headers.get("X-Cleide-Upload-Lock") or "").strip()
        or (request.form.get("upload_lock") or "").strip()
    )
    if request_token and request_token != lock_token:
        return _json_error(
            "Conflito de upload na sessao atual. Atualize a tela e tente novamente.",
            code="invalid_upload_lock",
            status=409,
        )

    if is_cleide_upload_in_progress(session):
        return _json_error(
            "Upload da Cleide ja esta em andamento. Aguarde a finalizacao.",
            code="upload_in_progress",
            status=409,
        )
    mark_cleide_upload_in_progress(session)
    started_at = time.perf_counter()
    emitted_processing_event = False
    upload_ref_for_key = _resolve_cleide_execution_id()

    def _resolve_rows_processed_from_analytics(analytics_payload: dict[str, Any] | None) -> int:
        if not isinstance(analytics_payload, dict):
            return 0
        summary = analytics_payload.get("dataset_summary")
        if isinstance(summary, dict):
            try:
                return max(0, int(summary.get("linhas_processadas") or 0))
            except (TypeError, ValueError):
                return 0
        return 0

    def _emit_processing_event(status: str, rows_processed: int, err: str | None = None) -> None:
        nonlocal emitted_processing_event
        if emitted_processing_event:
            return
        emitted_processing_event = True
        elapsed_ms = int((time.perf_counter() - started_at) * 1000)
        try:
            apropriar_billing_upload_cleide(
                idempotency_key=f"cleide-upload:{upload_ref_for_key}",
                rows_processed=max(0, int(rows_processed)),
                processing_time_ms=max(0, elapsed_ms),
                status=status,
                error_summary=err,
                execution_id=upload_ref_for_key,
            )
        except Exception:
            logger.exception("Falha ao apropriar billing do upload Cleide.")
            try:
                from app.run_cleiton_processing_governance import cleiton_register_processing_event

                cleiton_register_processing_event(
                    agent="cleide",
                    flow_type="upload_fretes",
                    processing_type="non_llm",
                    rows_processed=max(0, int(rows_processed)),
                    processing_time_ms=max(0, elapsed_ms),
                    status=status,
                    error_summary=err,
                    apply_operational_motor=False,
                )
            except Exception:
                logger.exception("Falha no fallback de ProcessingEvent do upload Cleide.")
    try:
        if not request.content_type or "multipart/form-data" not in request.content_type.lower():
            return _json_error("Requisicao deve ser multipart/form-data.", code="invalid_multipart")
        file = _extract_file()
        if file is None or not (file.filename or "").strip():
            return _json_error("Nenhum arquivo enviado.", code="missing_file")
        safe_name = secure_filename(file.filename or "")
        ext = _validate_extension(file.filename or "")
        if ext is None:
            return _json_error("Formato invalido. Envie apenas XLSX ou CSV.", code="invalid_extension")
        if not _validate_mime(ext, file.mimetype):
            return _json_error("Tipo de arquivo invalido para o formato informado.", code="invalid_mime")

        size = _stream_size(file)
        if size <= 0:
            return _json_error("Arquivo vazio.", code="empty_file")
        cfg = get_cleide_config()
        if size > int(cfg.upload_max_file_size_bytes):
            return _json_error(
                "Arquivo acima do limite configurado para upload.",
                code="file_too_large",
                status=413,
            )

        raw = _read_bytes(file)
        linhas_detectadas = _count_non_empty_data_rows(
            raw_bytes=raw,
            extension=ext,
            delimiter_default=cfg.csv_delimiter_default,
        )
        upload_total_max = int(cfg.upload_total_max)
        if linhas_detectadas > upload_total_max:
            error_message = "Arquivo excede o limite máximo de linhas permitido para a Cleide."
            logger.warning(
                "Upload Cleide bloqueado por upload_total_max. limite=%s linhas_detectadas=%s",
                upload_total_max,
                linhas_detectadas,
            )
            _emit_processing_event("failure", 0, error_message)
            return (
                jsonify(
                    {
                        "success": False,
                        "error": error_message,
                        "error_code": "upload_total_max_exceeded",
                        "upload_total_max": upload_total_max,
                        "linhas_detectadas": int(linhas_detectadas),
                    }
                ),
                413,
            )
        try:
            structural = analyze_structural_layout(
                raw_bytes=raw,
                extension=ext,
                delimiter_default=cfg.csv_delimiter_default,
                max_rows=cfg.structural_max_rows,
                max_columns=cfg.structural_max_columns,
            )
        except StructuralValidationError as exc:
            logger.warning("Falha estrutural no upload Cleide: %s (%s)", exc.message, exc.code)
            _emit_processing_event("failure", 0, exc.message)
            return _json_error(exc.message, code=exc.code)
        try:
            analytics = build_analytics_context(
                raw_bytes=raw,
                extension=ext,
                structural_context=structural,
                delimiter_default=cfg.csv_delimiter_default,
                max_rows=cfg.analytics_max_rows,
                max_group_items=cfg.analytics_group_limit,
            )
        except AnalyticsProcessingError as exc:
            logger.warning("Falha analytics no upload Cleide: %s (%s)", exc.message, exc.code)
            _emit_processing_event("failure", 0, exc.message)
            return _json_error(exc.message, code=exc.code)

        maybe_cleanup_expired_cleide_uploads(cfg.upload_ttl_minutes)
        previous_ref = get_cleide_upload_ref(session)
        previous_ext = get_upload_ref_extension(previous_ref) if previous_ref else None
        if previous_ref:
            clear_cleide_upload_file(previous_ref)
            clear_cleide_upload_ref(session)
            clear_cleide_dataset_context(session)

        upload_info = save_cleide_upload_file(file_storage=file, safe_filename=safe_name or f"upload{ext}")
        upload_ref = str(upload_info["upload_ref"])
        upload_ref_for_key = upload_ref or "sem_ref"
        set_cleide_upload_ref(session, upload_ref)
        set_cleide_dataset_context(
            session,
            {
                "dataset_validado": bool(structural.get("dataset_validado")),
                "dataset_tipo": structural.get("dataset_tipo"),
                "sheet_detectada": structural.get("sheet_detectada"),
                "linhas_detectadas": int(structural.get("linhas_detectadas") or 0),
                "colunas_detectadas": list(structural.get("colunas_detectadas") or []),
                "colunas_faltantes": list(structural.get("colunas_faltantes") or []),
                "aliases_resolvidos": dict(structural.get("aliases_resolvidos") or {}),
                "raw_headers": list(structural.get("raw_headers") or []),
                "canonical_headers": list(structural.get("canonical_headers") or []),
                "normalized_columns": list(structural.get("normalized_columns") or []),
                "colunas_duplicadas": list(structural.get("colunas_duplicadas") or []),
                "delimiter_detectado": structural.get("delimiter_detectado"),
                "detected_encoding": structural.get("detected_encoding"),
                "structural_validation": {
                    "required_columns": list(structural.get("aliases_resolvidos", {}).keys()),
                    "missing_required": list(structural.get("colunas_faltantes") or []),
                    "duplicated_columns": list(structural.get("colunas_duplicadas") or []),
                },
                "analytics_context": analytics,
                "operational_context": build_cleide_operational_context(
                    upload_ref=upload_ref,
                    dataset_validado=bool(structural.get("dataset_validado")),
                    analytics_ready=bool(analytics.get("analytics_ready")),
                    stale_upload=False,
                    dataset_summary=dict(analytics.get("dataset_summary") or {}),
                    kpis=dict(analytics.get("kpis") or {}),
                    aggregate_tables={
                        "transportadora": list(analytics.get("transportadora_stats") or []),
                        "uf_origem": list(analytics.get("uf_origem_stats") or []),
                        "uf_destino": list(analytics.get("uf_destino_stats") or []),
                        "temporal": list(analytics.get("temporal_stats") or []),
                        "pareto_fretes_zerados_uf_destino": list(analytics.get("pareto_fretes_zerados_uf_destino") or []),
                        "pareto_fretes_zerados_transportadora": list(
                            analytics.get("pareto_fretes_zerados_transportadora") or []
                        ),
                    },
                    aggregate_counts=dict(analytics.get("aggregate_counts") or {}),
                    active_filters={},
                    filter_mode="aggregate_approximation",
                    kpi_scope="global_session",
                    no_row_level_intersection=True,
                    multi_dimension_filters_are_approximate=True,
                    kpis_are_global_session_scope=True,
                ),
            },
        )
        replaced_previous = bool(previous_ref)
        _emit_processing_event(
            "success",
            _resolve_rows_processed_from_analytics(analytics),
            None,
        )

        return (
            jsonify(
                {
                    "success": True,
                    "upload_ref": upload_ref,
                    "filename": upload_info["safe_filename"],
                    "stored_filename": upload_info["stored_filename"],
                    "file_size_bytes": upload_info["file_size_bytes"],
                    "extension": ext,
                    "encoding": structural.get("detected_encoding"),
                    "replaced_previous_upload": replaced_previous,
                    "previous_extension": previous_ext,
                    "dataset_validado": bool(structural.get("dataset_validado")),
                    "colunas_detectadas": list(structural.get("colunas_detectadas") or []),
                    "colunas_faltantes": list(structural.get("colunas_faltantes") or []),
                    "linhas_detectadas": int(structural.get("linhas_detectadas") or 0),
                    "sheet_detectada": structural.get("sheet_detectada"),
                    "aliases_resolvidos": dict(structural.get("aliases_resolvidos") or {}),
                    "analytics_ready": bool(analytics.get("analytics_ready")),
                    "kpis": dict(analytics.get("kpis") or {}),
                    "dataset_summary": dict(analytics.get("dataset_summary") or {}),
                    "aggregate_counts": dict(analytics.get("aggregate_counts") or {}),
                    "transportadora_stats": list(analytics.get("transportadora_stats") or []),
                    "uf_origem_stats": list(analytics.get("uf_origem_stats") or []),
                    "uf_destino_stats": list(analytics.get("uf_destino_stats") or []),
                    "temporal_stats": list(analytics.get("temporal_stats") or []),
                    "pareto_fretes_zerados_uf_destino": list(analytics.get("pareto_fretes_zerados_uf_destino") or []),
                    "pareto_fretes_zerados_transportadora": list(
                        analytics.get("pareto_fretes_zerados_transportadora") or []
                    ),
                    "cleide_contexto_operacional": build_cleide_operational_context(
                        upload_ref=upload_ref,
                        dataset_validado=bool(structural.get("dataset_validado")),
                        analytics_ready=bool(analytics.get("analytics_ready")),
                        stale_upload=False,
                        dataset_summary=dict(analytics.get("dataset_summary") or {}),
                        kpis=dict(analytics.get("kpis") or {}),
                        aggregate_tables={
                            "transportadora": list(analytics.get("transportadora_stats") or []),
                            "uf_origem": list(analytics.get("uf_origem_stats") or []),
                            "uf_destino": list(analytics.get("uf_destino_stats") or []),
                            "temporal": list(analytics.get("temporal_stats") or []),
                            "pareto_fretes_zerados_uf_destino": list(
                                analytics.get("pareto_fretes_zerados_uf_destino") or []
                            ),
                            "pareto_fretes_zerados_transportadora": list(
                                analytics.get("pareto_fretes_zerados_transportadora") or []
                            ),
                        },
                        aggregate_counts=dict(analytics.get("aggregate_counts") or {}),
                        active_filters={},
                        filter_mode="aggregate_approximation",
                        kpi_scope="global_session",
                        no_row_level_intersection=True,
                        multi_dimension_filters_are_approximate=True,
                        kpis_are_global_session_scope=True,
                    ),
                    "upload_lock": lock_token,
                    "message": (
                        "Upload concluido com sucesso."
                        if not replaced_previous
                        else "Upload concluido com sucesso e upload anterior substituido."
                    ),
                }
            ),
            200,
        )
    except ValueError as exc:
        logger.warning("Falha de validacao no upload Cleide: %s", exc)
        _emit_processing_event("failure", 0, str(exc))
        return _json_error(str(exc), code="invalid_upload")
    finally:
        clear_cleide_upload_in_progress(session)


def clear_cleide_upload() -> tuple[Any, int]:
    lock_token = get_or_create_cleide_upload_lock(session)
    request_token = (request.headers.get("X-Cleide-Upload-Lock") or "").strip()
    if request_token and request_token != lock_token:
        return _json_error(
            "Conflito de upload na sessao atual. Atualize a tela e tente novamente.",
            code="invalid_upload_lock",
            status=409,
        )
    current_ref = get_cleide_upload_ref(session)
    if current_ref:
        clear_cleide_upload_file(current_ref)
    clear_cleide_upload_ref(session)
    clear_cleide_dataset_context(session)
    clear_cleide_upload_in_progress(session)
    return jsonify({"success": True, "message": "Upload da Cleide removido."}), 200


def get_cleide_upload_status() -> tuple[Any, int]:
    lock_token = get_cleide_upload_lock(session)
    current_ref = get_cleide_upload_ref(session)
    dataset_context = get_cleide_dataset_context(session) or {}
    if not current_ref:
        return jsonify(
            {
                "success": True,
                "upload_ativo": False,
                "upload_lock": lock_token,
                "dataset_validado": False,
                "colunas_detectadas": [],
                "colunas_faltantes": [],
                "linhas_detectadas": 0,
                "sheet_detectada": None,
                "aliases_resolvidos": {},
                "analytics_ready": False,
                "kpis": {},
                "dataset_summary": {},
                "aggregate_counts": {},
                "transportadora_stats": [],
                "uf_origem_stats": [],
                "uf_destino_stats": [],
                "temporal_stats": [],
                "pareto_fretes_zerados_uf_destino": [],
                "pareto_fretes_zerados_transportadora": [],
            }
        ), 200
    analytics_context = dataset_context.get("analytics_context") or {}
    operational_context = dataset_context.get("operational_context") or {}
    upload_file = resolve_cleide_upload_file(current_ref)
    if upload_file is None:
        stale_operational_context = (
            dict(operational_context)
            if isinstance(operational_context, dict) and operational_context
            else build_cleide_operational_context(
                upload_ref=current_ref,
                dataset_validado=bool(dataset_context.get("dataset_validado")),
                analytics_ready=bool(analytics_context.get("analytics_ready")),
                stale_upload=True,
                dataset_summary=dict(analytics_context.get("dataset_summary") or {}),
                kpis=dict(analytics_context.get("kpis") or {}),
                aggregate_tables={
                    "transportadora": list(analytics_context.get("transportadora_stats") or []),
                    "uf_origem": list(analytics_context.get("uf_origem_stats") or []),
                    "uf_destino": list(analytics_context.get("uf_destino_stats") or []),
                    "temporal": list(analytics_context.get("temporal_stats") or []),
                    "pareto_fretes_zerados_uf_destino": list(
                        analytics_context.get("pareto_fretes_zerados_uf_destino") or []
                    ),
                    "pareto_fretes_zerados_transportadora": list(
                        analytics_context.get("pareto_fretes_zerados_transportadora") or []
                    ),
                },
                aggregate_counts=dict(analytics_context.get("aggregate_counts") or {}),
                active_filters={},
                filter_mode="aggregate_approximation",
                kpi_scope="global_session",
                no_row_level_intersection=True,
                multi_dimension_filters_are_approximate=True,
                kpis_are_global_session_scope=True,
            )
        )
        if isinstance(stale_operational_context, dict):
            stale_operational_context["session_scope"] = dict(stale_operational_context.get("session_scope") or {})
            stale_operational_context["session_scope"]["stale_upload"] = True
        return (
            jsonify(
                {
                    "success": True,
                    "upload_ativo": False,
                    "stale_upload": True,
                    "upload_ref": current_ref,
                    "upload_lock": lock_token,
                    "dataset_validado": bool(dataset_context.get("dataset_validado")),
                    "colunas_detectadas": list(dataset_context.get("colunas_detectadas") or []),
                    "colunas_faltantes": list(dataset_context.get("colunas_faltantes") or []),
                    "linhas_detectadas": int(dataset_context.get("linhas_detectadas") or 0),
                    "sheet_detectada": dataset_context.get("sheet_detectada"),
                    "aliases_resolvidos": dict(dataset_context.get("aliases_resolvidos") or {}),
                    "analytics_ready": bool(analytics_context.get("analytics_ready")),
                    "kpis": dict(analytics_context.get("kpis") or {}),
                    "dataset_summary": dict(analytics_context.get("dataset_summary") or {}),
                    "aggregate_counts": dict(analytics_context.get("aggregate_counts") or {}),
                    "transportadora_stats": list(analytics_context.get("transportadora_stats") or []),
                    "uf_origem_stats": list(analytics_context.get("uf_origem_stats") or []),
                    "uf_destino_stats": list(analytics_context.get("uf_destino_stats") or []),
                    "temporal_stats": list(analytics_context.get("temporal_stats") or []),
                    "pareto_fretes_zerados_uf_destino": list(
                        analytics_context.get("pareto_fretes_zerados_uf_destino") or []
                    ),
                    "pareto_fretes_zerados_transportadora": list(
                        analytics_context.get("pareto_fretes_zerados_transportadora") or []
                    ),
                    "cleide_contexto_operacional": stale_operational_context,
                }
            ),
            200,
        )
    active_operational_context = (
        dict(operational_context)
        if isinstance(operational_context, dict) and operational_context
        else build_cleide_operational_context(
            upload_ref=current_ref,
            dataset_validado=bool(dataset_context.get("dataset_validado")),
            analytics_ready=bool(analytics_context.get("analytics_ready")),
            stale_upload=False,
            dataset_summary=dict(analytics_context.get("dataset_summary") or {}),
            kpis=dict(analytics_context.get("kpis") or {}),
            aggregate_tables={
                "transportadora": list(analytics_context.get("transportadora_stats") or []),
                "uf_origem": list(analytics_context.get("uf_origem_stats") or []),
                "uf_destino": list(analytics_context.get("uf_destino_stats") or []),
                "temporal": list(analytics_context.get("temporal_stats") or []),
                "pareto_fretes_zerados_uf_destino": list(analytics_context.get("pareto_fretes_zerados_uf_destino") or []),
                "pareto_fretes_zerados_transportadora": list(
                    analytics_context.get("pareto_fretes_zerados_transportadora") or []
                ),
            },
            aggregate_counts=dict(analytics_context.get("aggregate_counts") or {}),
            active_filters={},
            filter_mode="aggregate_approximation",
            kpi_scope="global_session",
            no_row_level_intersection=True,
            multi_dimension_filters_are_approximate=True,
            kpis_are_global_session_scope=True,
        )
    )
    return (
        jsonify(
            {
                "success": True,
                "upload_ativo": True,
                "upload_ref": current_ref,
                "filename": upload_file.name,
                "file_size_bytes": int(upload_file.stat().st_size),
                "extension": upload_file.suffix.lower(),
                "upload_lock": lock_token,
                "dataset_validado": bool(dataset_context.get("dataset_validado")),
                "colunas_detectadas": list(dataset_context.get("colunas_detectadas") or []),
                "colunas_faltantes": list(dataset_context.get("colunas_faltantes") or []),
                "linhas_detectadas": int(dataset_context.get("linhas_detectadas") or 0),
                "sheet_detectada": dataset_context.get("sheet_detectada"),
                "aliases_resolvidos": dict(dataset_context.get("aliases_resolvidos") or {}),
                "analytics_ready": bool(analytics_context.get("analytics_ready")),
                "kpis": dict(analytics_context.get("kpis") or {}),
                "dataset_summary": dict(analytics_context.get("dataset_summary") or {}),
                "aggregate_counts": dict(analytics_context.get("aggregate_counts") or {}),
                "transportadora_stats": list(analytics_context.get("transportadora_stats") or []),
                "uf_origem_stats": list(analytics_context.get("uf_origem_stats") or []),
                "uf_destino_stats": list(analytics_context.get("uf_destino_stats") or []),
                "temporal_stats": list(analytics_context.get("temporal_stats") or []),
                "pareto_fretes_zerados_uf_destino": list(analytics_context.get("pareto_fretes_zerados_uf_destino") or []),
                "pareto_fretes_zerados_transportadora": list(
                    analytics_context.get("pareto_fretes_zerados_transportadora") or []
                ),
                "cleide_contexto_operacional": active_operational_context,
            }
        ),
        200,
    )


def get_cleide_dashboard_filtered_analytics(filters: dict[str, Any] | None) -> tuple[Any, int]:
    started_ms = int(time.perf_counter() * 1000)
    current_ref = get_cleide_upload_ref(session)
    if not current_ref:
        logger.info(
            "Cleide dashboard filter failed missing_upload_ref elapsed_ms=%s",
            int(time.perf_counter() * 1000) - started_ms,
        )
        return _json_error(
            "Nenhum upload ativo na sessao da Cleide.",
            code="missing_upload_ref",
            status=400,
        )
    upload_file = resolve_cleide_upload_file(current_ref)
    if upload_file is None:
        logger.info(
            "Cleide dashboard filter failed stale_upload_ref upload_ref=%s elapsed_ms=%s",
            current_ref,
            int(time.perf_counter() * 1000) - started_ms,
        )
        return _json_error(
            "Upload ativo nao encontrado. Reenvie o arquivo para continuar.",
            code="stale_upload_ref",
            status=409,
        )

    dataset_context = get_cleide_dataset_context(session) or {}
    if not bool(dataset_context.get("dataset_validado")):
        logger.info(
            "Cleide dashboard filter failed invalid_dataset_context upload_ref=%s elapsed_ms=%s",
            current_ref,
            int(time.perf_counter() * 1000) - started_ms,
        )
        return _json_error(
            "Dataset da sessao nao esta validado para analytics.",
            code="invalid_dataset_context",
            status=400,
        )

    cfg = get_cleide_config()
    normalized_filters = _normalize_dashboard_filters(filters)
    active_filters = {key: value for key, value in normalized_filters.items() if value}
    logger.info(
        "Cleide dashboard filter start upload_ref=%s active_filters=%s",
        current_ref,
        active_filters,
    )
    raw = upload_file.read_bytes()
    structural_context = dict(dataset_context)
    if not structural_context.get("raw_headers"):
        structural_context["raw_headers"] = list(
            dataset_context.get("canonical_headers") or dataset_context.get("colunas_detectadas") or []
        )
    try:
        analytics = build_filtered_analytics_context(
            raw_bytes=raw,
            extension=upload_file.suffix.lower(),
            structural_context=structural_context,
            delimiter_default=cfg.csv_delimiter_default,
            max_rows=cfg.analytics_max_rows,
            max_group_items=cfg.analytics_group_limit,
            filters=normalized_filters,
        )
    except AnalyticsProcessingError as exc:
        logger.warning("Falha analytics no filtro dashboard Cleide: %s (%s)", exc.message, exc.code)
        logger.info(
            "Cleide dashboard filter failed analytics_error upload_ref=%s active_filters=%s elapsed_ms=%s",
            current_ref,
            active_filters,
            int(time.perf_counter() * 1000) - started_ms,
        )
        return _json_error(exc.message, code=exc.code)

    operational_context = build_cleide_operational_context(
        upload_ref=current_ref,
        dataset_validado=bool(dataset_context.get("dataset_validado")),
        analytics_ready=bool(analytics.get("analytics_ready")),
        stale_upload=False,
        dataset_summary=dict(analytics.get("dataset_summary") or {}),
        kpis=dict(analytics.get("kpis") or {}),
        aggregate_tables={
            "transportadora": list(analytics.get("transportadora_stats") or []),
            "uf_origem": list(analytics.get("uf_origem_stats") or []),
            "uf_destino": list(analytics.get("uf_destino_stats") or []),
            "temporal": list(analytics.get("temporal_stats") or []),
            "pareto_fretes_zerados_uf_destino": list(analytics.get("pareto_fretes_zerados_uf_destino") or []),
            "pareto_fretes_zerados_transportadora": list(analytics.get("pareto_fretes_zerados_transportadora") or []),
        },
        aggregate_counts=dict(analytics.get("aggregate_counts") or {}),
        active_filters=normalized_filters,
        filter_mode="row_level_intersection_backend",
        kpi_scope="filtered_session_intersection",
        no_row_level_intersection=False,
        multi_dimension_filters_are_approximate=False,
        kpis_are_global_session_scope=False,
    )
    dataset_context["operational_context"] = operational_context
    set_cleide_dataset_context(session, dataset_context)
    logger.info(
        "Cleide dashboard filter success upload_ref=%s active_filters=%s elapsed_ms=%s total_docs=%s aggregate_counts=%s",
        current_ref,
        active_filters,
        int(time.perf_counter() * 1000) - started_ms,
        int((analytics.get("kpis") or {}).get("total_documentos") or 0),
        dict(analytics.get("aggregate_counts") or {}),
    )

    return (
        jsonify(
            {
                "success": True,
                "upload_ref": current_ref,
                "analytics_ready": bool(analytics.get("analytics_ready")),
                "active_filters": normalized_filters,
                "kpis": dict(analytics.get("kpis") or {}),
                "dataset_summary": dict(analytics.get("dataset_summary") or {}),
                "aggregate_counts": dict(analytics.get("aggregate_counts") or {}),
                "transportadora_stats": list(analytics.get("transportadora_stats") or []),
                "uf_origem_stats": list(analytics.get("uf_origem_stats") or []),
                "uf_destino_stats": list(analytics.get("uf_destino_stats") or []),
                "temporal_stats": list(analytics.get("temporal_stats") or []),
                "pareto_fretes_zerados_uf_destino": list(analytics.get("pareto_fretes_zerados_uf_destino") or []),
                "pareto_fretes_zerados_transportadora": list(analytics.get("pareto_fretes_zerados_transportadora") or []),
                "cleide_contexto_operacional": operational_context,
            }
        ),
        200,
    )
