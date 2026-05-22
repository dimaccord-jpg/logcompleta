from __future__ import annotations

import csv
import io
import re
import unicodedata
import zipfile
from dataclasses import dataclass
from typing import Any

_CSV_DELIMITERS = (",", ";", "\t", "|")
_MAX_SNIFF_SAMPLE = 4096
_MAX_HEADER_COLUMNS = 1024

REQUIRED_COLUMNS = (
    "transportadora",
    "uf_origem",
    "uf_destino",
    "valor_frete",
    "peso",
    "data_emissao",
)

ALIASES_BY_REQUIRED: dict[str, tuple[str, ...]] = {
    "transportadora": (
        "transportadora",
        "transp",
        "transportador",
        "nome_transportadora",
        "nome transportadora",
        "carrier",
    ),
    "uf_origem": (
        "uf_origem",
        "uf origem",
        "origem_uf",
        "estado_origem",
        "estado de origem",
        "uf_remetente",
    ),
    "uf_destino": (
        "uf_destino",
        "uf destino",
        "destino_uf",
        "estado_destino",
        "estado de destino",
        "uf_destinatario",
    ),
    "valor_frete": (
        "valor_frete",
        "valor frete",
        "frete",
        "vl_frete",
        "valor do frete",
        "frete_valor",
    ),
    "peso": (
        "peso",
        "peso_kg",
        "peso kg",
        "peso_total",
        "kg",
        "peso_bruto",
    ),
    "data_emissao": (
        "data_emissao",
        "data emissao",
        "dt_emissao",
        "data de emissao",
        "data_documento",
        "data_nf",
    ),
}


@dataclass(frozen=True)
class StructuralValidationError(Exception):
    code: str
    message: str

    def __str__(self) -> str:
        return self.message


def canonicalize_header(value: Any) -> str:
    text = str(value or "").strip().lower()
    normalized = unicodedata.normalize("NFKD", text)
    normalized = "".join(ch for ch in normalized if not unicodedata.combining(ch))
    normalized = re.sub(r"[^a-z0-9]+", "_", normalized)
    normalized = normalized.strip("_")
    return normalized


def analyze_structural_layout(
    *,
    raw_bytes: bytes,
    extension: str,
    delimiter_default: str,
    max_rows: int,
    max_columns: int,
) -> dict[str, Any]:
    ext = (extension or "").strip().lower()
    if ext == ".csv":
        return _analyze_csv(
            raw_bytes=raw_bytes,
            delimiter_default=delimiter_default,
            max_rows=max_rows,
            max_columns=max_columns,
        )
    if ext == ".xlsx":
        return _analyze_xlsx(raw_bytes=raw_bytes, max_rows=max_rows, max_columns=max_columns)
    raise StructuralValidationError(code="invalid_extension", message="Formato invalido para analise.")


def _analyze_csv(
    *,
    raw_bytes: bytes,
    delimiter_default: str,
    max_rows: int,
    max_columns: int,
) -> dict[str, Any]:
    decoded_text, encoding = _decode_csv(raw_bytes)
    if not decoded_text.strip():
        raise StructuralValidationError(code="invalid_csv", message="Arquivo CSV vazio.")
    delimiter = _detect_delimiter(decoded_text, delimiter_default)
    try:
        rows = csv.reader(io.StringIO(decoded_text), delimiter=delimiter)
        base = _extract_tabular_structure(rows=rows, max_rows=max_rows, max_columns=max_columns)
    except csv.Error as exc:
        raise StructuralValidationError(code="invalid_csv", message=f"CSV invalido: {exc}") from exc
    except StructuralValidationError as exc:
        if exc.code == "invalid_header":
            raise StructuralValidationError(
                code="invalid_csv",
                message="CSV invalido: cabecalho contem colunas vazias ou invalidas.",
            ) from exc
        raise
    if len(base["canonical_headers"]) < 2:
        raise StructuralValidationError(
            code="invalid_csv",
            message="CSV invalido: cabecalho deve conter ao menos 2 colunas.",
        )
    if base["linhas_detectadas"] <= 0:
        raise StructuralValidationError(
            code="invalid_csv",
            message="CSV invalido: necessario ao menos 1 linha de dados.",
        )
    base["dataset_tipo"] = "csv"
    base["sheet_detectada"] = "csv"
    base["detected_encoding"] = encoding
    base["delimiter_detectado"] = delimiter
    return base


def _analyze_xlsx(*, raw_bytes: bytes, max_rows: int, max_columns: int) -> dict[str, Any]:
    try:
        import openpyxl
    except ImportError as exc:
        raise StructuralValidationError(code="invalid_xlsx", message="Servico XLSX indisponivel.") from exc
    try:
        workbook = openpyxl.load_workbook(io.BytesIO(raw_bytes), read_only=True, data_only=True)
    except (OSError, ValueError, KeyError, RuntimeError, zipfile.BadZipFile) as exc:
        raise StructuralValidationError(code="invalid_xlsx", message="Arquivo XLSX invalido ou corrompido.") from exc

    first_error: StructuralValidationError | None = None
    try:
        for name in workbook.sheetnames:
            sheet = workbook[name]
            try:
                base = _extract_tabular_structure(
                    rows=sheet.iter_rows(values_only=True),
                    max_rows=max_rows,
                    max_columns=max_columns,
                )
            except StructuralValidationError as exc:
                if first_error is None:
                    first_error = exc
                continue
            if len(base["canonical_headers"]) < 2:
                first_error = first_error or StructuralValidationError(
                    code="invalid_xlsx",
                    message="Planilha XLSX sem cabecalho tabular valido.",
                )
                continue
            if base["linhas_detectadas"] <= 0:
                first_error = first_error or StructuralValidationError(
                    code="invalid_xlsx",
                    message="Planilha XLSX sem linhas de dados.",
                )
                continue
            base["dataset_tipo"] = "xlsx"
            base["sheet_detectada"] = name
            base["detected_encoding"] = None
            return base
    finally:
        workbook.close()

    if first_error is not None:
        if first_error.code in {"layout_too_many_rows", "layout_too_many_columns"}:
            raise first_error
        raise StructuralValidationError(code="invalid_xlsx", message=first_error.message)
    raise StructuralValidationError(code="invalid_xlsx", message="XLSX sem planilha valida.")


def _decode_csv(raw_bytes: bytes) -> tuple[str, str]:
    last_exc: UnicodeDecodeError | None = None
    for encoding in ("utf-8-sig", "utf-8", "latin1"):
        try:
            text = raw_bytes.decode(encoding)
            if encoding == "utf-8-sig" and not raw_bytes.startswith(b"\xef\xbb\xbf"):
                return text, "utf-8"
            return text, encoding
        except UnicodeDecodeError as exc:
            last_exc = exc
    if last_exc is not None:
        raise StructuralValidationError(
            code="invalid_csv",
            message="Nao foi possivel decodificar CSV (utf-8/utf-8-sig/latin1).",
        ) from last_exc
    raise StructuralValidationError(code="invalid_csv", message="CSV invalido.")


def _detect_delimiter(text: str, fallback: str) -> str:
    sample = text[:_MAX_SNIFF_SAMPLE]
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=_CSV_DELIMITERS)
        if dialect.delimiter in _CSV_DELIMITERS:
            return dialect.delimiter
    except csv.Error:
        pass
    return fallback if fallback in _CSV_DELIMITERS else ","


def _extract_tabular_structure(*, rows, max_rows: int, max_columns: int) -> dict[str, Any]:
    safe_max_rows = max(100, int(max_rows))
    safe_max_columns = min(max(10, int(max_columns)), _MAX_HEADER_COLUMNS)

    header_raw: list[str] | None = None
    canonical_headers: list[str] = []
    rows_detected = 0

    for row in rows:
        normalized_row = [str(cell).strip() if cell is not None else "" for cell in row]
        while normalized_row and not normalized_row[-1]:
            normalized_row.pop()
        if len(normalized_row) > safe_max_columns:
            raise StructuralValidationError(
                code="layout_too_many_columns",
                message="Layout invalido: quantidade de colunas acima do limite permitido.",
            )
        if not _row_has_data(normalized_row):
            continue
        if header_raw is None:
            header_raw = normalized_row
            canonical_headers = [canonicalize_header(col) for col in header_raw]
            if not any(canonical_headers):
                raise StructuralValidationError(code="invalid_header", message="Cabecalho invalido.")
            if any(not col for col in canonical_headers):
                raise StructuralValidationError(code="invalid_header", message="Cabecalho com colunas vazias.")
            continue
        rows_detected += 1
        if rows_detected > safe_max_rows:
            raise StructuralValidationError(
                code="layout_too_many_rows",
                message="Layout invalido: quantidade de linhas acima do limite permitido.",
            )

    if header_raw is None:
        raise StructuralValidationError(code="invalid_header", message="Cabecalho inexistente.")

    duplicates = _find_duplicates(canonical_headers)
    resolved_aliases = _resolve_aliases(header_raw, canonical_headers)
    missing_required = [name for name in REQUIRED_COLUMNS if name not in resolved_aliases]

    detected_columns = []
    for name in canonical_headers:
        if name and name not in detected_columns:
            detected_columns.append(name)

    return {
        "dataset_validado": not missing_required and not duplicates and rows_detected > 0,
        "colunas_detectadas": detected_columns,
        "colunas_faltantes": missing_required,
        "colunas_duplicadas": duplicates,
        "linhas_detectadas": rows_detected,
        "aliases_resolvidos": resolved_aliases,
        "canonical_headers": canonical_headers,
        "normalized_columns": canonical_headers,
        "raw_headers": header_raw,
    }


def _resolve_aliases(raw_headers: list[str], canonical_headers: list[str]) -> dict[str, str]:
    alias_sets = {
        required: {canonicalize_header(alias) for alias in aliases}
        for required, aliases in ALIASES_BY_REQUIRED.items()
    }
    resolved: dict[str, str] = {}
    for required in REQUIRED_COLUMNS:
        options = alias_sets[required]
        for idx, canonical in enumerate(canonical_headers):
            if canonical and canonical in options:
                resolved[required] = raw_headers[idx]
                break
    return resolved


def _find_duplicates(headers: list[str]) -> list[str]:
    counts: dict[str, int] = {}
    for value in headers:
        if not value:
            continue
        counts[value] = counts.get(value, 0) + 1
    return sorted([name for name, qty in counts.items() if qty > 1])


def _row_has_data(row: list[str]) -> bool:
    return any(cell.strip() for cell in row)
