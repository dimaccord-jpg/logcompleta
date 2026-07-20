"""
Wrapper documental fino do domínio Agente Compara (Fase 1).

Reaproveita preparação, store e configuração governados do Cleiton.
Opera exclusivamente sobre session keys `agente_compara_*`; não toca sessão da Júlia.
Sem rotas, chat, IA ou billing.
"""
from __future__ import annotations

import csv
import copy
import io
import json
import logging
import re
import time
import unicodedata
import zipfile
from datetime import UTC, datetime, timedelta
from decimal import Decimal, InvalidOperation, ROUND_CEILING, ROUND_HALF_UP
from uuid import uuid4

from flask import has_request_context, request, session
from werkzeug.utils import secure_filename

from app.cleiton_doc_contracts import (
    CONTEXT_KIND_GEMINI_FILE,
    CONTEXT_KIND_PLACEHOLDER,
    CONTEXT_KIND_TEXT,
    DOC_TYPE_CSV,
    DOC_TYPE_DOCX,
    DOC_TYPE_PDF,
    DOC_TYPE_TXT,
    DOC_TYPE_XLSX,
    DOC_TYPE_XML,
    ERROR_DOC_NOT_FOUND,
    ERROR_GEMINI_FILE_UPLOAD,
    ERROR_INVALID_SIZE,
    ERROR_MAX_FILES,
    ERROR_SESSION_BYTES,
    FIELD_CHAR_COUNT,
    FIELD_COLUMN_COUNT,
    FIELD_CONTEXT_KIND,
    FIELD_CONTEXT_REF,
    FIELD_CREATED_AT,
    FIELD_DISPLAY_NAME,
    FIELD_DOC_ID,
    FIELD_DOC_TYPE,
    FIELD_ERROR_CODE,
    FIELD_EXPIRES_AT,
    FIELD_EXTENSION,
    FIELD_GEMINI_FILE_NAME,
    FIELD_GEMINI_FILE_STATE,
    FIELD_GEMINI_FILE_URI,
    FIELD_GEMINI_MIME_TYPE,
    FIELD_GEMINI_UPLOADED_AT,
    FIELD_MAX_DEPTH,
    FIELD_MIME_TYPE,
    FIELD_NODE_COUNT,
    FIELD_PAGE_COUNT,
    FIELD_PDF_CONTEXT_READY,
    FIELD_PREPARED_CONTEXT,
    FIELD_ROW_COUNT,
    FIELD_SAFE_NAME,
    FIELD_SESSION_KEY,
    FIELD_SIZE_BYTES,
    FIELD_SOURCE_AGENT,
    FIELD_STATUS,
    FIELD_TRUNCATED,
    FIELD_WARNINGS,
    STATUS_ACTIVE,
    STATUS_ERROR,
)
from app.cleiton_doc_gemini_files import (
    pdf_context_ready_from_record,
    upload_pdf_to_gemini_files_api,
)
from app.cleiton_doc_prepare import prepare_document
from app.cleiton_doc_service import CleitonDocSessionError, maybe_cleanup_expired_cleiton_docs
from app.cleiton_doc_store import (
    document_record_matches_domain_scope,
    get_cleiton_doc_tmp_dir,
    load_document_record,
    peek_document_record,
    remove_document_record,
    save_document_record,
)
from app.services.cleiton_doc_config_service import get_cleiton_doc_config
from app.services.agente_compara_config_service import (
    get_active_calculation_base_by_id,
    get_active_calculation_bases_for_runtime,
    get_agente_compara_config,
    normalize_calculation_base_unit,
    resolve_audited_file_limits,
    resolve_calculation_base_status,
)
from app.agente_compara_correction_service import build_audit_correction_suggestions

logger = logging.getLogger(__name__)

AGENTE_COMPARA_DOMAIN = "agente_compara"

AGENTE_COMPARA_DOC_IDS_SESSION_KEY = "agente_compara_doc_ids"
AGENTE_COMPARA_DOC_CONTEXT_SESSION_KEY = "agente_compara_doc_context"
AGENTE_COMPARA_CHAT_HISTORY_SESSION_KEY = "agente_compara_chat_history"
AGENTE_COMPARA_UPLOAD_LOCK_SESSION_KEY = "agente_compara_upload_lock"
AGENTE_COMPARA_LAST_REQUEST_ID_SESSION_KEY = "agente_compara_last_request_id"
AGENTE_COMPARA_UPLOAD_IN_PROGRESS_SESSION_KEY = "agente_compara_upload_in_progress"
AGENTE_COMPARA_TEMP_TABLE_ID_SESSION_KEY = "agente_compara_temp_table_id"
AGENTE_COMPARA_TEMP_TABLE_SOURCE_DOCS_SESSION_KEY = "agente_compara_temp_table_source_doc_ids"

TEMP_TABLE_STATUS_PROCESSING = "processing"
TEMP_TABLE_STATUS_AWAITING_VALIDATION = "awaiting_validation"
TEMP_TABLE_STATUS_VALIDATED = "validated"
TEMP_TABLE_STATUS_NEEDS_REVIEW = "needs_review"
TEMP_TABLE_STATUS_FAILED = "failed"
TEMP_TABLE_STATUS_EXPIRED = "expired"
TEMP_TABLE_STATUS_DISCARDED = "discarded"

TEMP_TABLE_VERSION_MARKER = "agente_compara_temp_table_v1"
TEMP_TABLE_OPERATIONAL_OWNER = "cleiton"
TEMP_TABLE_UI_DISPLAY_NAME = "Tabela temporária extraída"
TEMP_TABLE_JSON_BEGIN = "---AGENTE_COMPARA_TEMP_TABLE---"
TEMP_TABLE_JSON_END = "---END_AGENTE_COMPARA_TEMP_TABLE---"

TEMP_TABLE_SAVE_MAX_PAYLOAD_BYTES = 512 * 1024
TEMP_TABLE_REVIEW_ACTION_SAVE_AND_ADVANCE = "save_and_advance"
TEMP_TABLE_REVIEW_ACTION_SAVE_DRAFT = "save_draft"
HUMAN_REVIEW_STATUS_REVIEWED = "reviewed"
HUMAN_REVIEW_STATUS_EDITED = "edited"
UNMAPPED_CALCULATION_BASIS_LABEL = "não mapeado / revisar"
ERROR_ACCESSORIAL_CALCULATION_BASE_MESSAGE = "Selecione uma base de cálculo ou exclua a linha."
ERROR_ACCESSORIAL_VALUE_MESSAGE = "Preencha um valor válido para esta taxa ou exclua a linha."
ERROR_ACCESSORIAL_UNIT_MESSAGE = "A unidade não é compatível com a base selecionada."
ERROR_ACCESSORIAL_OPERATION_MESSAGE = "Revise a operação da base de cálculo selecionada."
ERROR_ACCESSORIAL_MINIMUM_LINK_MESSAGE = (
    "Esta regra mínima não possui uma taxa principal válida vinculada. Corrija ou exclua a regra antes de continuar."
)
ERROR_ACCESSORIAL_ADVANCE_MESSAGE = "Revise as generalidades antes de avançar."
ERROR_ACCESSORIAL_RATE_CONFLICT_MESSAGE_TEMPLATE = (
    "Há informações contraditórias na regra {name}. O valor informado é {structured_percent}%, "
    "mas a observação indica {described_percent}%. Corrija uma das informações ou exclua a regra antes de continuar."
)

ERROR_TEMP_TABLE_NOT_FOUND = "agente_compara_temp_table_not_found"
ERROR_TEMP_TABLE_ID_MISMATCH = "agente_compara_temp_table_id_mismatch"
ERROR_TEMP_TABLE_EXPIRED = "agente_compara_temp_table_expired"
ERROR_TEMP_TABLE_INVALID_PAYLOAD = "agente_compara_temp_table_invalid_payload"
ERROR_TEMP_TABLE_INVALID_ACCESSORIAL_FEES = "invalid_accessorial_fees"
ERROR_TEMP_TABLE_PAYLOAD_TOO_LARGE = "agente_compara_temp_table_payload_too_large"
ERROR_TEMP_TABLE_SCOPE_MISMATCH = "agente_compara_temp_table_scope_mismatch"

COVERAGE_TABLE_COLUMNS = ["UF destino", "Cidade destino", "Região de frete"]
COVERAGE_TABLE_STATUS_NEEDS_REVIEW = "needs_review"
COVERAGE_UPLOAD_MAX_BYTES = 512 * 1024
COVERAGE_UPLOAD_MAX_ROWS = 10000

ICMS_INTERSTATE_SOURCE_NAME = "Resolução Senado Federal nº 22/1989"
ICMS_INTERMUNICIPAL_SOURCE_NAME = "Cadastro estadual/manual"
UF_NORTH = frozenset({"AC", "AP", "AM", "PA", "RO", "RR", "TO"})
UF_NORTHEAST = frozenset({"AL", "BA", "CE", "MA", "PB", "PE", "PI", "RN", "SE"})
UF_CENTER_WEST = frozenset({"DF", "GO", "MT", "MS"})
UF_SOUTHEAST = frozenset({"ES", "MG", "RJ", "SP"})
UF_SOUTH = frozenset({"PR", "RS", "SC"})
BRAZILIAN_UFS = UF_NORTH | UF_NORTHEAST | UF_CENTER_WEST | UF_SOUTHEAST | UF_SOUTH
ICMS_7_PERCENT_ORIGIN_UFS = UF_SOUTH | UF_SOUTHEAST
ICMS_7_PERCENT_DESTINATION_UFS = UF_NORTH | UF_NORTHEAST | UF_CENTER_WEST | {"ES"}

ERROR_COVERAGE_NO_TEMP_TABLE = "agente_compara_coverage_no_temp_table"
ERROR_COVERAGE_INVALID_FORMAT = "agente_compara_coverage_invalid_format"
ERROR_COVERAGE_EMPTY_FILE = "agente_compara_coverage_empty_file"
ERROR_COVERAGE_PAYLOAD_TOO_LARGE = "agente_compara_coverage_payload_too_large"
ERROR_COVERAGE_PARSE_FAILED = "agente_compara_coverage_parse_failed"
ERROR_COVERAGE_INVALID_PAYLOAD = "agente_compara_coverage_invalid_payload"
ERROR_COVERAGE_EXPIRED = "agente_compara_coverage_expired"
ERROR_COVERAGE_SCOPE_MISMATCH = "agente_compara_coverage_scope_mismatch"

AUDIT_BATCH_SHEET_NAME = "Modelo AgenteCompara"
AUDIT_INPUT_SCHEMA_VERSION = "agente_compara_input_v1"
AUDIT_BATCH_STATUS_UPLOADED = "uploaded"
AUDIT_BATCH_STATUS_PROCESSED = "processed"
AGENTE_COMPARA_TEMPLATE_FILENAME = "template_agente_compara.xlsx"

ERROR_AUDIT_NO_TEMP_TABLE = "agente_compara_audit_no_temp_table"
ERROR_AUDIT_INVALID_FORMAT = "agente_compara_audit_invalid_format"
ERROR_AUDIT_EMPTY_FILE = "agente_compara_audit_empty_file"
ERROR_AUDIT_PAYLOAD_TOO_LARGE = "agente_compara_audit_payload_too_large"
ERROR_AUDIT_PARSE_FAILED = "agente_compara_audit_parse_failed"
ERROR_AUDIT_MISSING_COLUMNS = "agente_compara_audit_missing_columns"
ERROR_AUDIT_TOO_MANY_ROWS = "agente_compara_audit_too_many_rows"
ERROR_AUDIT_EXPIRED = "agente_compara_audit_expired"
ERROR_AUDIT_SCOPE_MISMATCH = "agente_compara_audit_scope_mismatch"
ERROR_AUDIT_INVALID_SHEET = "agente_compara_audit_invalid_sheet"
ERROR_AUDIT_EMPTY_ROWS = "agente_compara_audit_empty_rows"
ERROR_AUDIT_BATCH_NOT_FOUND = "agente_compara_audit_batch_not_found"
ERROR_AUDIT_BATCH_EMPTY = "agente_compara_audit_batch_empty"

AUDIT_STATUS_OK = "ok"
AUDIT_STATUS_DIVERGENT = "divergent"
AUDIT_STATUS_MISSING_COVERAGE = "missing_coverage_mapping"
AUDIT_STATUS_AMBIGUOUS_COVERAGE = "ambiguous_coverage_mapping"
AUDIT_STATUS_MISSING_FREIGHT_RULE = "missing_freight_rule"
AUDIT_STATUS_INVALID_WEIGHT = "invalid_weight"
AUDIT_STATUS_INVALID_CHARGED_FREIGHT = "invalid_charged_freight"
AUDIT_STATUS_INVALID_INVOICE_VALUE = "invalid_invoice_value"
AUDIT_STATUS_UNSUPPORTED_PRICING = "unsupported_pricing_model"
AUDIT_DIAGNOSTIC_PRICING_DIMENSION_MISMATCH = "pricing_dimension_mismatch"
AUDIT_TRANSFORMATION_SELECT_PRICING_DIMENSION = "select_pricing_dimension"

AUDIT_BI_DATASET_VERSION = "agente_compara_bi_v1"
AUDIT_BI_SOURCE = "audit_batch"
AUDIT_BI_FILTER_MODE = "frontend_row_level"
AUDIT_BI_CHARTS_SUPPORTED = (
    "transportadora",
    "uf_destino",
    "temporal",
    "pareto_transportadora",
)
AUDIT_BI_PUBLIC_ROW_FIELDS = (
    "row_index",
    "carrier",
    "origin_uf",
    "destination_uf",
    "issue_date",
    "audited_weight",
    "charged_freight",
    "expected_freight",
    "divergence_value",
    "status",
)
AUDIT_BI_FIELD_PRESENCE_FIELDS = (
    "carrier",
    "origin_uf",
    "destination_uf",
    "issue_date",
    "audited_weight",
    "charged_freight",
    "expected_freight",
    "divergence_value",
    "status",
)
AUDIT_BI_FORBIDDEN_PUBLIC_FIELDS = frozenset(
    {
        "document_number",
        "invoice_value",
        "origin_city",
        "destination_city",
        "source_file_name",
        "calculation_details",
        "calculation_basis",
        "calculation_components",
        "numero_documento",
        "freight_region",
    }
)
AUDIT_BI_NOT_READY_MESSAGE = (
    "Ainda não há dados auditados disponíveis para gráficos. "
    "Envie o arquivo auditado para habilitar esta área."
)

TAX_CALCULATION_VERSION = "agente_compara_tax_v2"
TAX_CALCULATION_MODE_INSIDE = "inside"
TAX_CALCULATION_MODE_OUTSIDE = "outside"
PRICING_RULE_PARSER_VERSION = "agente_compara_pricing_matrix_v2"
_WEIGHT_KG_HEADER_RE = re.compile(r"\bkg\b|\bkgs\b|\bpeso\b")
ISS_SOURCE_NAME = "Cadastro municipal/manual"
ISS_SOURCE_TYPE = "manual"
AUDIT_BATCH_STALE_TAX_CONFIG_REASON = "tax_config_changed"
AUDIT_BATCH_STALE_TAX_CONFIG_ALERT = (
    "Configuração fiscal alterada após processamento da auditoria. "
    "Reprocesse o lote para atualizar os resultados."
)
AUDIT_BATCH_STALE_PRICING_RULE_REASON = "pricing_rule_changed"
AUDIT_BATCH_STALE_PRICING_RULE_ALERT = (
    "Tabela tarifária, rotas ou generalidades alteradas após processamento da auditoria. "
    "Reprocesse o lote para atualizar os resultados."
)
AUDIT_BATCH_STALE_FISCAL_OUTDATED_REASON = "fiscal_calculation_outdated"
AUDIT_BATCH_STALE_FISCAL_OUTDATED_ALERT = (
    "O lote foi calculado com regra fiscal anterior. "
    "Reprocesse o lote para aplicar a metodologia fiscal atual."
)
AUDIT_REASON_ROUTE_TOLL_APPLIED = "route_toll_applied"
AUDIT_REASON_ROUTE_TOLL_DUPLICATE_IGNORED = "route_toll_duplicate_ignored"

AUDIT_REASON_ACCESSORIAL_PERCENTAGE_CALCULATED = "accessorial_percentage_calculated"
AUDIT_REASON_DUPLICATE_INVOICE_PERCENTAGE_FEE_IGNORED = "duplicate_invoice_percentage_fee_ignored"
AUDIT_REASON_AMBIGUOUS_ACCESSORIAL_PERCENTAGE = "ambiguous_accessorial_percentage"
AUDIT_REASON_UNSUPPORTED_ACCESSORIAL_CONDITION = "unsupported_accessorial_condition"
AUDIT_REASON_ACCESSORIAL_FEE_NOT_APPLIED = "accessorial_fee_not_applied"
AUDIT_REASON_ACCESSORIAL_MINIMUM_WITHOUT_BASE_IGNORED = "accessorial_minimum_without_base_ignored"
AUDIT_REASON_CONFIGURED_ACCESSORIAL_CALCULATED = "configured_accessorial_fee_calculated"
AUDIT_REASON_LEGACY_CLASSIFIER_NOT_CALCULATED = "legacy_classifier_not_calculated"
AUDIT_REASON_NOT_CONFIGURED_CALCULATION_BASE = "not_configured_calculation_base"
AUDIT_REASON_UNSUPPORTED_CONFIGURED_OPERATION = "unsupported_operation"
AUDIT_REASON_MISSING_AUDIT_VARIABLE = "missing_audit_variable"
AUDIT_REASON_INVALID_CONFIGURED_AMOUNT = "invalid_amount"
AUDIT_REASON_CONDITIONS_PRESENT = "conditions_present"
AUDIT_REASON_UNSUPPORTED_REASON_PRESENT = "unsupported_reason_present"
AUDIT_REASON_CLASSIFICATION_WARNING_PRESENT = "classification_warning_present"

_AUDIT_REQUIRED_FIELDS = (
    "destination_city",
    "destination_uf",
    "charged_freight",
    "audited_weight",
)
_AUDIT_OPTIONAL_FIELDS = (
    "carrier",
    "document_number",
    "origin_city",
    "origin_uf",
    "invoice_value",
    "modal",
    "issue_date",
    "delivery_date",
)
_AUDIT_FIELD_LABELS = {
    "destination_city": "cidade_destino",
    "destination_uf": "uf_destino",
    "charged_freight": "valor_frete",
    "audited_weight": "peso",
    "carrier": "transportadora",
    "document_number": "numero_documento",
    "origin_city": "cidade_origem",
    "origin_uf": "uf_origem",
    "invoice_value": "valor_nf",
    "modal": "modal",
    "issue_date": "data_emissao",
    "delivery_date": "data_entrega",
}

_HTML_TAG_RE = re.compile(r"<[^>]+>")
_BR_UFS = frozenset(
    {
        "AC", "AL", "AP", "AM", "BA", "CE", "DF", "ES", "GO", "MA", "MT", "MS", "MG",
        "PA", "PB", "PR", "PE", "PI", "RJ", "RN", "RS", "RO", "RR", "SC", "SP", "SE", "TO",
    }
)
_COVERAGE_REJECT_GENERIC_HEADERS = frozenset(
    {
        "destino",
        "grupo",
        "tipo",
        "obs",
        "observacao",
        "observacoes",
        "nota",
        "notas",
        "comentario",
        "comentarios",
        "info",
        "informacao",
        "informacoes",
    }
)
_COVERAGE_FIELD_LABELS = {
    "destination_uf": "UF destino",
    "destination_city": "Cidade destino",
    "freight_region": "Região de frete",
}
_COVERAGE_FIELD_HINTS = {
    "destination_uf": "Você pode usar nomes como UF, Estado ou Unidade Federativa.",
    "destination_city": "Você pode usar nomes como Cidade, Município, Localidade ou Cidade atendida.",
    "freight_region": "Você pode usar nomes como Praça, Região, Rota, Itinerário, Área ou Zona.",
}
_COVERAGE_REQUIRED_FIELDS = ("destination_uf", "destination_city", "freight_region")
_COVERAGE_HEADER_ALIASES: dict[str, str] = {}

AGENTE_COMPARA_DOCUMENT_UPLOAD_FLOW_TYPE = "agente_compara_document_upload"
AGENTE_COMPARA_DOCUMENT_PREPARE_FLOW_TYPE = "agente_compara_document_prepare"
AGENTE_COMPARA_CHAT_FLOW_TYPE = "agente_compara_chat"
AGENTE_COMPARA_INSIGHTS_CHAT_FLOW_TYPE = "agente_compara_insights_chat"
AGENTE_COMPARA_TEMP_TABLE_EXTRACTION_FLOW_TYPE = "agente_compara_temp_table_extraction"
AGENTE_COMPARA_COVERAGE_UPLOAD_FLOW_TYPE = "agente_compara_coverage_upload"
AGENTE_COMPARA_BATCH_UPLOAD_FLOW_TYPE = "agente_compara_batch_upload"
AGENTE_COMPARA_BATCH_PROCESSED_FLOW_TYPE = "agente_compara_batch_processed"

SOURCE_AGENT_AGENTE_COMPARA = "agente_compara"


class AgenteComparaTempTableError(ValueError):
    def __init__(self, error_code: str, message: str, *, errors: list[dict] | None = None):
        super().__init__(message)
        self.error_code = error_code
        self.message = message
        self.errors = list(errors or [])


class AgenteComparaCoverageError(ValueError):
    def __init__(self, error_code: str, message: str):
        super().__init__(message)
        self.error_code = error_code
        self.message = message


class AgenteComparaBatchError(ValueError):
    def __init__(self, error_code: str, message: str):
        super().__init__(message)
        self.error_code = error_code
        self.message = message


def agente_compara_upload_idempotency_key(request_id: str) -> str:
    return f"agente-compara-upload:{(request_id or '').strip()}"


def agente_compara_upload_doc_idempotency_key(doc_id: str) -> str:
    return f"agente-compara-upload-doc:{(doc_id or '').strip()}"


def agente_compara_chat_idempotency_key(request_id: str) -> str:
    return f"agente-compara-chat:{(request_id or '').strip()}"


def agente_compara_insights_chat_idempotency_key(request_id: str, batch_scope: str = "") -> str:
    scope = (batch_scope or "").strip() or "unknown"
    return f"agente-compara-insights-chat:{scope}:{(request_id or '').strip()}"


def agente_compara_temp_table_extraction_idempotency_key(source_doc_ids: list[str]) -> str:
    normalized = _normalize_source_doc_ids(source_doc_ids)
    joined = ":".join(normalized)
    return f"agente-compara-temp-table:{TEMP_TABLE_VERSION_MARKER}:{joined}"


def _resolve_agente_compara_execution_id() -> str:
    if not has_request_context():
        return str(uuid4())
    execution_id = (request.headers.get("X-Execution-ID") or "").strip()
    if not execution_id:
        execution_id = (request.form.get("execution_id") or "").strip()
    if not execution_id:
        execution_id = str(uuid4())
    return execution_id[:120]


def agente_compara_coverage_upload_idempotency_key(session_id: str, coverage_version: str) -> str:
    return f"agente-compara-coverage-upload:{(session_id or '').strip()}:{(coverage_version or '').strip()}"


def agente_compara_batch_upload_idempotency_key(session_id: str, audit_batch_id: str) -> str:
    return f"agente-compara-batch-upload:{(session_id or '').strip()}:{(audit_batch_id or '').strip()}"


def agente_compara_batch_run_idempotency_key(session_id: str, audit_batch_id: str, run_version: str) -> str:
    return (
        f"agente-compara-batch-run:{(session_id or '').strip()}:"
        f"{(audit_batch_id or '').strip()}:{(run_version or '').strip()}"
    )


def _emit_agente_compara_operational_billing(
    *,
    emitted: list[bool],
    started_at: float,
    flow_type: str,
    idempotency_key: str,
    rows_processed: int,
    status: str = "success",
    error_summary: str | None = None,
    execution_id: str | None = None,
) -> None:
    if emitted[0]:
        return
    emitted[0] = True
    elapsed_ms = int((time.perf_counter() - started_at) * 1000)
    try:
        from app.services.cleiton_upload_billing_service import apropriar_billing_agente_compara_operational_flow

        apropriar_billing_agente_compara_operational_flow(
            flow_type=flow_type,
            idempotency_key=idempotency_key,
            rows_processed=max(0, int(rows_processed)),
            processing_time_ms=max(0, elapsed_ms),
            status=status,
            error_summary=error_summary,
            execution_id=execution_id,
        )
    except Exception:
        logger.exception("Falha ao apropriar billing operacional da Auditoria Agente Compara (%s).", flow_type)
        try:
            from app.run_cleiton_processing_governance import cleiton_register_processing_event

            cleiton_register_processing_event(
                agent="agente_compara",
                flow_type=flow_type,
                processing_type="non_llm",
                rows_processed=max(0, int(rows_processed)),
                processing_time_ms=max(0, elapsed_ms),
                status=status,
                error_summary=error_summary,
                execution_id=execution_id,
                apply_operational_motor=False,
            )
        except Exception:
            logger.exception(
                "Falha no fallback de ProcessingEvent operacional da Auditoria Agente Compara (%s).",
                flow_type,
            )


def _utcnow() -> datetime:
    return datetime.now(UTC).replace(tzinfo=None)


def _require_session() -> None:
    if not has_request_context():
        raise RuntimeError("Sessão documental da Agente Compara requer request context Flask.")


def _mark_session_modified() -> None:
    session.modified = True


def get_agente_compara_doc_ids(session_obj) -> list[str]:
    raw = session_obj.get(AGENTE_COMPARA_DOC_IDS_SESSION_KEY)
    if not isinstance(raw, list):
        return []
    ids: list[str] = []
    for item in raw:
        if isinstance(item, str):
            ref = item.strip()
            if ref:
                ids.append(ref)
    return ids


def set_agente_compara_doc_ids(session_obj, doc_ids: list[str]) -> None:
    cleaned: list[str] = []
    seen: set[str] = set()
    for item in doc_ids or []:
        if not isinstance(item, str):
            continue
        ref = item.strip()
        if not ref or ref in seen:
            continue
        seen.add(ref)
        cleaned.append(ref)
    session_obj[AGENTE_COMPARA_DOC_IDS_SESSION_KEY] = cleaned


def clear_agente_compara_doc_ids(session_obj) -> None:
    session_obj.pop(AGENTE_COMPARA_DOC_IDS_SESSION_KEY, None)


def append_agente_compara_doc_id(session_obj, doc_id: str) -> None:
    ref = (doc_id or "").strip()
    if not ref:
        raise ValueError("doc_id inválido para sessão documental da Agente Compara.")
    ids = get_agente_compara_doc_ids(session_obj)
    if ref not in ids:
        ids.append(ref)
    set_agente_compara_doc_ids(session_obj, ids)


def remove_agente_compara_doc_id(session_obj, doc_id: str) -> None:
    ref = (doc_id or "").strip()
    if not ref:
        return
    ids = [item for item in get_agente_compara_doc_ids(session_obj) if item != ref]
    if ids:
        set_agente_compara_doc_ids(session_obj, ids)
    else:
        clear_agente_compara_doc_ids(session_obj)


def _parse_size_bytes(size_bytes) -> int:
    if size_bytes is None:
        raise CleitonDocSessionError(
            ERROR_INVALID_SIZE,
            "size_bytes ausente ou inválido para documento Agente Compara.",
        )
    if isinstance(size_bytes, bool):
        raise CleitonDocSessionError(
            ERROR_INVALID_SIZE,
            "size_bytes não numérico para documento Agente Compara.",
        )
    try:
        parsed = int(size_bytes)
    except (TypeError, ValueError):
        raise CleitonDocSessionError(
            ERROR_INVALID_SIZE,
            "size_bytes não numérico para documento Agente Compara.",
        ) from None
    if parsed <= 0:
        raise CleitonDocSessionError(
            ERROR_INVALID_SIZE,
            "size_bytes deve ser maior que zero para documento Agente Compara.",
        )
    return parsed


def _normalize_extension(extension: str) -> str:
    raw = (extension or "").strip().lower()
    if not raw:
        return ""
    return raw if raw.startswith(".") else f".{raw}"


def _build_safe_display_name(display_name: str, extension: str) -> tuple[str, str]:
    raw = (display_name or "").strip() or "documento"
    safe = secure_filename(raw) or "documento"
    ext = _normalize_extension(extension)
    if ext and not safe.lower().endswith(ext):
        safe_name = f"{safe}{ext}"
    else:
        safe_name = safe
        if not ext and "." in safe_name:
            ext = "." + safe_name.rsplit(".", 1)[-1].lower()
    return raw, safe_name


def _context_ref_for_kind(context_kind: str, doc_id: str) -> str:
    kind = (context_kind or CONTEXT_KIND_PLACEHOLDER).strip() or CONTEXT_KIND_PLACEHOLDER
    if kind == CONTEXT_KIND_TEXT:
        return f"text:{doc_id}"
    if kind == CONTEXT_KIND_GEMINI_FILE:
        return f"gemini_file:{doc_id}"
    return f"placeholder:{doc_id}"


def _public_record(record: dict) -> dict:
    pdf_ready = pdf_context_ready_from_record(record)
    return {
        FIELD_DOC_ID: record.get(FIELD_DOC_ID),
        FIELD_DOC_TYPE: record.get(FIELD_DOC_TYPE),
        FIELD_DISPLAY_NAME: record.get(FIELD_DISPLAY_NAME),
        FIELD_SAFE_NAME: record.get(FIELD_SAFE_NAME),
        FIELD_EXTENSION: record.get(FIELD_EXTENSION),
        FIELD_MIME_TYPE: record.get(FIELD_MIME_TYPE),
        FIELD_SIZE_BYTES: record.get(FIELD_SIZE_BYTES),
        FIELD_CREATED_AT: record.get(FIELD_CREATED_AT),
        FIELD_EXPIRES_AT: record.get(FIELD_EXPIRES_AT),
        FIELD_STATUS: record.get(FIELD_STATUS),
        FIELD_TRUNCATED: record.get(FIELD_TRUNCATED),
        FIELD_CONTEXT_KIND: record.get(FIELD_CONTEXT_KIND),
        FIELD_CONTEXT_REF: record.get(FIELD_CONTEXT_REF),
        FIELD_CHAR_COUNT: record.get(FIELD_CHAR_COUNT),
        FIELD_ROW_COUNT: record.get(FIELD_ROW_COUNT),
        FIELD_COLUMN_COUNT: record.get(FIELD_COLUMN_COUNT),
        FIELD_PAGE_COUNT: record.get(FIELD_PAGE_COUNT),
        FIELD_NODE_COUNT: record.get(FIELD_NODE_COUNT),
        FIELD_MAX_DEPTH: record.get(FIELD_MAX_DEPTH),
        FIELD_WARNINGS: record.get(FIELD_WARNINGS) or [],
        FIELD_SESSION_KEY: record.get(FIELD_SESSION_KEY),
        FIELD_ERROR_CODE: record.get(FIELD_ERROR_CODE),
        FIELD_PDF_CONTEXT_READY: pdf_ready,
    }


def get_allowed_document_formats() -> list[dict]:
    """Retorna formatos habilitados a partir da config central do Cleiton."""
    cfg = get_cleiton_doc_config()
    catalog = [
        (DOC_TYPE_TXT, ".txt", cfg.txt_enabled),
        (DOC_TYPE_XML, ".xml", cfg.xml_enabled),
        (DOC_TYPE_CSV, ".csv", cfg.csv_enabled),
        (DOC_TYPE_XLSX, ".xlsx", cfg.excel_enabled),
        (DOC_TYPE_DOCX, ".docx", cfg.docx_enabled),
        (DOC_TYPE_PDF, ".pdf", cfg.pdf_enabled),
    ]
    return [
        {"doc_type": doc_type, "extension": extension, "enabled": bool(enabled)}
        for doc_type, extension, enabled in catalog
        if enabled
    ]


def cleanup_expired_documents_for_session() -> int:
    _require_session()
    cfg = get_cleiton_doc_config()
    removed = 0
    stale_ids: list[str] = []

    for doc_id in get_agente_compara_doc_ids(session):
        record = load_document_record(doc_id, ttl_hours=cfg.upload_ttl_hours)
        if record is None:
            stale_ids.append(doc_id)
            removed += 1

    for doc_id in stale_ids:
        remove_document_record(doc_id)
        remove_agente_compara_doc_id(session, doc_id)

    if stale_ids:
        _mark_session_modified()
    return removed


def get_active_documents_for_session() -> list[dict]:
    _require_session()
    cfg = get_cleiton_doc_config()
    active: list[dict] = []
    stale_ids: list[str] = []

    for doc_id in get_agente_compara_doc_ids(session):
        record = load_document_record(doc_id, ttl_hours=cfg.upload_ttl_hours)
        if record is None:
            stale_ids.append(doc_id)
            continue
        active.append(_public_record(record))

    if stale_ids:
        for doc_id in stale_ids:
            remove_agente_compara_doc_id(session, doc_id)
        _mark_session_modified()

    return active


def get_document_session_totals() -> dict:
    _require_session()
    cfg = get_cleiton_doc_config()
    cleanup_expired_documents_for_session()
    active = get_active_documents_for_session()
    total_bytes = sum(int(item.get(FIELD_SIZE_BYTES) or 0) for item in active)
    active_count = len(active)
    max_files = int(cfg.max_files_per_session)
    max_bytes = int(cfg.session_max_bytes)
    return {
        "active_count": active_count,
        "total_bytes": total_bytes,
        "max_files_per_session": max_files,
        "session_max_bytes": max_bytes,
        "remaining_files": max(0, max_files - active_count),
        "remaining_bytes": max(0, max_bytes - total_bytes),
    }


def assert_session_can_accept_document(size_bytes) -> None:
    _require_session()
    cfg = get_cleiton_doc_config()
    incoming = _parse_size_bytes(size_bytes)

    if incoming > int(cfg.session_max_bytes):
        raise CleitonDocSessionError(
            ERROR_INVALID_SIZE,
            "size_bytes excede o limite total configurado da sessão documental.",
        )

    totals = get_document_session_totals()

    if totals["active_count"] >= int(cfg.max_files_per_session):
        raise CleitonDocSessionError(
            ERROR_MAX_FILES,
            "Limite de arquivos documentais por sessão atingido.",
        )

    projected = int(totals["total_bytes"]) + incoming
    if projected > int(cfg.session_max_bytes):
        raise CleitonDocSessionError(
            ERROR_SESSION_BYTES,
            "Limite total de bytes documentais da sessão excedido.",
        )


def _register_document_record(
    *,
    display_name: str,
    extension: str,
    mime_type: str,
    size_bytes: int,
    context_kind: str = CONTEXT_KIND_PLACEHOLDER,
    context_ref: str | None = None,
    truncated: bool = False,
    doc_type: str | None = None,
    prepared_context: str | None = None,
    char_count: int | None = None,
    row_count: int | None = None,
    column_count: int | None = None,
    page_count: int | None = None,
    node_count: int | None = None,
    max_depth: int | None = None,
    warnings: list[str] | None = None,
    status: str = STATUS_ACTIVE,
    error_code: str | None = None,
    gemini_file_name: str | None = None,
    gemini_file_uri: str | None = None,
    gemini_mime_type: str | None = None,
    gemini_file_state: str | None = None,
    gemini_uploaded_at: str | None = None,
) -> dict:
    _require_session()
    cfg = get_cleiton_doc_config()
    validated_size = _parse_size_bytes(size_bytes)
    assert_session_can_accept_document(validated_size)

    doc_id = uuid4().hex
    display, safe_name = _build_safe_display_name(display_name, extension)
    ext = _normalize_extension(extension) or (
        f".{safe_name.rsplit('.', 1)[-1].lower()}" if "." in safe_name else ""
    )
    created_at = _utcnow()
    expires_at = created_at + timedelta(hours=max(1, int(cfg.upload_ttl_hours)))
    resolved_kind = (context_kind or CONTEXT_KIND_PLACEHOLDER).strip() or CONTEXT_KIND_PLACEHOLDER

    record = {
        FIELD_DOC_ID: doc_id,
        FIELD_DOC_TYPE: doc_type,
        FIELD_DISPLAY_NAME: display,
        FIELD_SAFE_NAME: safe_name,
        FIELD_EXTENSION: ext,
        FIELD_MIME_TYPE: (mime_type or "application/octet-stream").strip(),
        FIELD_SIZE_BYTES: validated_size,
        FIELD_CREATED_AT: created_at.isoformat(),
        FIELD_EXPIRES_AT: expires_at.isoformat(),
        FIELD_STATUS: (status or STATUS_ACTIVE).strip() or STATUS_ACTIVE,
        FIELD_TRUNCATED: bool(truncated),
        FIELD_CONTEXT_KIND: resolved_kind,
        FIELD_CONTEXT_REF: context_ref or _context_ref_for_kind(resolved_kind, doc_id),
        FIELD_PREPARED_CONTEXT: prepared_context,
        FIELD_CHAR_COUNT: char_count,
        FIELD_ROW_COUNT: row_count,
        FIELD_COLUMN_COUNT: column_count,
        FIELD_PAGE_COUNT: page_count,
        FIELD_NODE_COUNT: node_count,
        FIELD_MAX_DEPTH: max_depth,
        FIELD_WARNINGS: list(warnings or []),
        FIELD_SOURCE_AGENT: SOURCE_AGENT_AGENTE_COMPARA,
        FIELD_SESSION_KEY: AGENTE_COMPARA_DOC_IDS_SESSION_KEY,
        FIELD_ERROR_CODE: error_code,
        FIELD_GEMINI_FILE_NAME: gemini_file_name,
        FIELD_GEMINI_FILE_URI: gemini_file_uri,
        FIELD_GEMINI_MIME_TYPE: gemini_mime_type,
        FIELD_GEMINI_FILE_STATE: gemini_file_state,
        FIELD_GEMINI_UPLOADED_AT: gemini_uploaded_at,
    }

    save_document_record(record)
    append_agente_compara_doc_id(session, doc_id)
    _mark_session_modified()
    return _public_record(record)


def prepare_and_register_document(
    *,
    display_name: str,
    file_bytes: bytes,
    mime_type: str | None = None,
    extension: str | None = None,
) -> dict:
    """
    Valida, prepara e registra documento na sessão Agente Compara após sucesso.

    Delega validação/preparação ao Cleiton; persiste IDs em `agente_compara_doc_ids`.
    """
    prepared = prepare_document(
        display_name=display_name,
        file_bytes=file_bytes,
        mime_type=mime_type,
        extension=extension,
    )

    register_kwargs = {
        "display_name": prepared["display_name"],
        "extension": prepared[FIELD_EXTENSION],
        "mime_type": prepared[FIELD_MIME_TYPE],
        "size_bytes": prepared[FIELD_SIZE_BYTES],
        "context_kind": prepared[FIELD_CONTEXT_KIND],
        "truncated": prepared[FIELD_TRUNCATED],
        "doc_type": prepared[FIELD_DOC_TYPE],
        "prepared_context": prepared[FIELD_PREPARED_CONTEXT],
        "char_count": prepared[FIELD_CHAR_COUNT],
        "row_count": prepared[FIELD_ROW_COUNT],
        "column_count": prepared[FIELD_COLUMN_COUNT],
        "page_count": prepared[FIELD_PAGE_COUNT],
        "node_count": prepared[FIELD_NODE_COUNT],
        "max_depth": prepared[FIELD_MAX_DEPTH],
        "warnings": prepared[FIELD_WARNINGS],
    }

    if prepared[FIELD_CONTEXT_KIND] == CONTEXT_KIND_GEMINI_FILE:
        cfg = get_cleiton_doc_config()
        upload_result = upload_pdf_to_gemini_files_api(
            file_bytes=file_bytes,
            mime_type=prepared[FIELD_MIME_TYPE],
            display_name=prepared["display_name"],
            page_count=prepared[FIELD_PAGE_COUNT],
            max_pages=int(cfg.pdf_max_pages),
        )
        register_kwargs["prepared_context"] = upload_result.prepared_context or prepared[FIELD_PREPARED_CONTEXT]
        register_kwargs["warnings"] = list(prepared[FIELD_WARNINGS]) + list(upload_result.warnings or [])
        if upload_result.ok:
            register_kwargs.update(
                {
                    "status": STATUS_ACTIVE,
                    "gemini_file_name": upload_result.gemini_file_name,
                    "gemini_file_uri": upload_result.gemini_file_uri,
                    "gemini_mime_type": upload_result.gemini_mime_type,
                    "gemini_file_state": upload_result.gemini_file_state,
                    "gemini_uploaded_at": upload_result.gemini_uploaded_at,
                }
            )
        else:
            register_kwargs.update(
                {
                    "status": STATUS_ERROR,
                    "error_code": ERROR_GEMINI_FILE_UPLOAD,
                    "gemini_file_name": upload_result.gemini_file_name,
                    "gemini_file_uri": upload_result.gemini_file_uri,
                    "gemini_mime_type": upload_result.gemini_mime_type,
                    "gemini_file_state": upload_result.gemini_file_state,
                }
            )
            logger.warning(
                "Agente Compara audit doc: upload Gemini Files API falhou para PDF (summary=%s).",
                upload_result.error_summary,
            )

    return _register_document_record(**register_kwargs)


def _agente_compara_document_owned_by_session(doc_id: str) -> bool:
    ref = (doc_id or "").strip()
    if not ref or ref not in get_agente_compara_doc_ids(session):
        return False
    record = peek_document_record(ref)
    return document_record_matches_domain_scope(
        record,
        expected_source_agent=SOURCE_AGENT_AGENTE_COMPARA,
        expected_session_key=AGENTE_COMPARA_DOC_IDS_SESSION_KEY,
    )


def remove_document_from_session(doc_id: str) -> dict:
    _require_session()
    ref = (doc_id or "").strip()
    if not ref:
        return {
            "ok": False,
            "doc_id": doc_id,
            "removed_from_store": False,
            "removed_from_session": False,
            "error_code": ERROR_DOC_NOT_FOUND,
        }

    if not _agente_compara_document_owned_by_session(ref):
        return {
            "ok": True,
            "doc_id": ref,
            "removed_from_store": False,
            "removed_from_session": False,
            "error_code": ERROR_DOC_NOT_FOUND,
        }

    store_result = remove_document_record(ref)
    remove_agente_compara_doc_id(session, ref)
    _mark_session_modified()

    invalidate_temp_table_if_source_changed(
        reason=TEMP_TABLE_STATUS_DISCARDED,
        removed_doc_id=ref,
    )

    return {
        "ok": True,
        "doc_id": ref,
        "removed_from_store": bool(store_result.get("removed")),
        "removed_from_session": True,
        "error_code": None,
    }


def clear_documents_for_session() -> dict:
    _require_session()
    ids = list(get_agente_compara_doc_ids(session))
    removed_store = 0
    removed_session = 0
    for doc_id in ids:
        if not _agente_compara_document_owned_by_session(doc_id):
            continue
        result = remove_document_record(doc_id)
        if result.get("removed"):
            removed_store += 1
        remove_agente_compara_doc_id(session, doc_id)
        removed_session += 1
    if not get_agente_compara_doc_ids(session):
        clear_agente_compara_doc_ids(session)
    if removed_session:
        invalidate_temp_table_for_session(reason=TEMP_TABLE_STATUS_DISCARDED)
        _mark_session_modified()
    return {
        "ok": True,
        "requested": len(ids),
        "removed_from_store": removed_store,
        "removed_from_session": removed_session,
    }


def _temp_table_filename(temp_table_id: str) -> str:
    ref = (temp_table_id or "").strip()
    if not ref:
        raise ValueError("temp_table_id inválido.")
    return f"tt_{ref}.json"


def _temp_table_path(temp_table_id: str):
    from pathlib import Path

    safe_name = _temp_table_filename(temp_table_id)
    base = Path(get_cleiton_doc_tmp_dir()).resolve()
    candidate = (base / safe_name).resolve()
    if candidate.parent != base:
        raise ValueError("temp_table path inválido.")
    return candidate


def _write_temp_table_atomic(path, payload: dict) -> None:
    from app.cleiton_doc_store import _write_json_atomic

    _write_json_atomic(path, payload)


def _normalize_source_doc_ids(doc_ids: list[str] | None) -> list[str]:
    cleaned: list[str] = []
    seen: set[str] = set()
    for item in doc_ids or []:
        if not isinstance(item, str):
            continue
        ref = item.strip()
        if not ref or ref in seen:
            continue
        seen.add(ref)
        cleaned.append(ref)
    return cleaned


def get_temp_table_id(session_obj) -> str | None:
    raw = session_obj.get(AGENTE_COMPARA_TEMP_TABLE_ID_SESSION_KEY)
    if not isinstance(raw, str):
        return None
    ref = raw.strip()
    return ref or None


def set_temp_table_id(session_obj, temp_table_id: str | None) -> None:
    ref = (temp_table_id or "").strip()
    if ref:
        session_obj[AGENTE_COMPARA_TEMP_TABLE_ID_SESSION_KEY] = ref
    else:
        session_obj.pop(AGENTE_COMPARA_TEMP_TABLE_ID_SESSION_KEY, None)


def get_temp_table_source_doc_ids(session_obj) -> list[str]:
    raw = session_obj.get(AGENTE_COMPARA_TEMP_TABLE_SOURCE_DOCS_SESSION_KEY)
    if not isinstance(raw, list):
        return []
    return _normalize_source_doc_ids(raw)


def set_temp_table_source_doc_ids(session_obj, doc_ids: list[str]) -> None:
    session_obj[AGENTE_COMPARA_TEMP_TABLE_SOURCE_DOCS_SESSION_KEY] = _normalize_source_doc_ids(doc_ids)


def clear_temp_table_session_refs(session_obj) -> None:
    session_obj.pop(AGENTE_COMPARA_TEMP_TABLE_ID_SESSION_KEY, None)
    session_obj.pop(AGENTE_COMPARA_TEMP_TABLE_SOURCE_DOCS_SESSION_KEY, None)


def _normalize_coverage_header(value) -> str:
    if value is None:
        return ""
    text = str(value).strip()
    text = unicodedata.normalize("NFKD", text)
    text = "".join(ch for ch in text if not unicodedata.combining(ch))
    text = text.lower()
    text = re.sub(r"[_\-\/.:]+", " ", text)
    text = re.sub(r"[^a-z0-9\s]", " ", text)
    text = re.sub(r"\s+", " ", text)
    return text.strip()


def _coverage_alias_entries() -> tuple[tuple[str, tuple[str, ...]], ...]:
    return (
        (
            "destination_uf",
            (
                "uf",
                "uf destino",
                "uf de destino",
                "uf dest",
                "uf destinatario",
                "uf destinatário",
                "uf destin",
                "uf destno",
                "estado",
                "estado destino",
                "estado de destino",
                "estado dest",
                "estado destin",
                "estado detino",
                "sigla uf",
                "sigla estado",
                "unidade federativa",
                "uf entrega",
                "uf de entrega",
                "uf_destino",
            ),
        ),
        (
            "destination_city",
            (
                "cidade",
                "cidade destino",
                "cidade de destino",
                "cidade dest",
                "cidade destinatario",
                "cidade destinatário",
                "cidade destin",
                "cidade detino",
                "cid destino",
                "municipio",
                "município",
                "municipio destino",
                "município destino",
                "municipio de destino",
                "municipo",
                "municípo",
                "municip",
                "mun destino",
                "localidade",
                "localidade destino",
                "cidade atendida",
                "cidade de atendimento",
                "cidade entrega",
                "cidade de entrega",
                "destino cidade",
                "cidade_destino",
            ),
        ),
        (
            "freight_region",
            (
                "praca",
                "praça",
                "praca destino",
                "praça destino",
                "praca dest",
                "prasa",
                "regiao",
                "região",
                "regiao de frete",
                "região de frete",
                "regiao frete",
                "região frete",
                "regiao fret",
                "regiao fretee",
                "regiao de fret",
                "regiao de fretee",
                "regiao de atendimento",
                "região de atendimento",
                "regiao atendida",
                "região atendida",
                "regiao destino",
                "região destino",
                "area",
                "área",
                "area de atendimento",
                "área de atendimento",
                "zona",
                "zona de entrega",
                "setor",
                "setor de entrega",
                "rota",
                "rota destino",
                "rota de entrega",
                "codigo regiao",
                "código região",
                "cod regiao",
                "cód regiao",
                "itinerario",
                "itinerário",
                "itinerario entrega",
                "itinerário entrega",
                "itinerarioo",
                "regional",
                "regional destino",
                "faixa regional",
                "grupo destino",
                "grupo de destino",
                "regiao_frete",
            ),
        ),
    )


def _build_coverage_header_aliases() -> dict[str, str]:
    aliases: dict[str, str] = {}
    for field_name, labels in _coverage_alias_entries():
        for label in labels:
            key = _normalize_coverage_header(label)
            if not key or key in _COVERAGE_REJECT_GENERIC_HEADERS:
                continue
            aliases.setdefault(key, field_name)
    return aliases


def _ensure_coverage_header_aliases() -> dict[str, str]:
    global _COVERAGE_HEADER_ALIASES
    if not _COVERAGE_HEADER_ALIASES:
        _COVERAGE_HEADER_ALIASES = _build_coverage_header_aliases()
    return _COVERAGE_HEADER_ALIASES


def _coverage_edit_distance(left: str, right: str) -> int:
    if left == right:
        return 0
    if not left:
        return len(right)
    if not right:
        return len(left)
    prev = list(range(len(right) + 1))
    for i, left_char in enumerate(left, start=1):
        current = [i]
        for j, right_char in enumerate(right, start=1):
            cost = 0 if left_char == right_char else 1
            current.append(
                min(
                    current[-1] + 1,
                    prev[j] + 1,
                    prev[j - 1] + cost,
                )
            )
        prev = current
    return prev[-1]


def _coverage_fuzzy_max_distance(alias_key: str) -> int:
    if len(alias_key) <= 4:
        return 1
    if len(alias_key) <= 7:
        return 1
    return 2


def _resolve_coverage_field(header: str) -> str | None:
    normalized = _normalize_coverage_header(header)
    if not normalized:
        return None
    aliases = _ensure_coverage_header_aliases()
    exact = aliases.get(normalized)
    if exact:
        return exact
    if normalized in _COVERAGE_REJECT_GENERIC_HEADERS:
        return None

    best_field: str | None = None
    best_distance = 999
    second_best_distance = 999
    for alias_key, field_name in aliases.items():
        max_distance = _coverage_fuzzy_max_distance(alias_key)
        distance = _coverage_edit_distance(normalized, alias_key)
        if distance > max_distance:
            continue
        if distance < best_distance:
            second_best_distance = best_distance
            best_distance = distance
            best_field = field_name
        elif distance == best_distance and field_name != best_field:
            second_best_distance = distance
    if best_field is None:
        return None
    if second_best_distance == best_distance:
        return None
    return best_field


def _coverage_column_sample_values(raw_rows: list[list], col_index: int, *, max_rows: int = 50) -> list[str]:
    values: list[str] = []
    for raw_row in raw_rows[1 : max_rows + 1]:
        if not isinstance(raw_row, list) or col_index >= len(raw_row):
            continue
        value = str(raw_row[col_index] or "").strip()
        if value:
            values.append(value)
    return values


def _score_coverage_column_as_uf(values: list[str]) -> float:
    if not values:
        return 0.0
    hits = sum(1 for value in values if _normalize_destination_uf(value) in _BR_UFS)
    return hits / len(values)


def _score_coverage_column_as_city(values: list[str]) -> float:
    if not values:
        return 0.0
    hits = 0
    for value in values:
        if _normalize_destination_uf(value) in _BR_UFS:
            continue
        cleaned = _sanitize_cell_string(value)
        if not cleaned or len(cleaned) < 3:
            continue
        alpha_ratio = sum(ch.isalpha() or ch.isspace() for ch in cleaned) / len(cleaned)
        if alpha_ratio >= 0.7:
            hits += 1
    return hits / len(values)


_REGION_NAME_HINTS = frozenset(
    {
        "norte",
        "nordeste",
        "sul",
        "sudeste",
        "centro oeste",
        "centro-oeste",
        "interior",
        "capital",
        "metropolitana",
        "litoral",
        "fluvial",
        "fluvias",
    }
)
_REGION_PATTERN_RE = re.compile(
    r"^[A-Za-z]{2}\s*[-/]\s*.+|.+\s+(interior|capital|metropolitana|litoral)\b.*",
    re.IGNORECASE,
)


def _score_coverage_column_as_region(values: list[str]) -> float:
    if not values:
        return 0.0
    hits = 0
    unique_values = {value.casefold() for value in values}
    repetition_bonus = 0.0
    if len(values) >= 3:
        repetition_bonus = min(0.25, (1 - (len(unique_values) / len(values))) * 0.5)
    for value in values:
        cleaned = _sanitize_cell_string(value)
        if not cleaned:
            continue
        normalized = _normalize_coverage_header(cleaned)
        if _REGION_PATTERN_RE.match(cleaned):
            hits += 1
            continue
        if any(hint in normalized for hint in _REGION_NAME_HINTS):
            hits += 1
            continue
        if re.search(r"\d", cleaned) and re.search(r"[A-Za-z]{2}", cleaned):
            hits += 0.75
    base = hits / len(values)
    return min(1.0, base + repetition_bonus)


def _infer_coverage_columns_from_content(
    raw_rows: list[list],
    field_indexes: dict[str, int],
    header_row: list,
) -> dict[str, int]:
    used_indexes = set(field_indexes.values())
    candidate_indexes = [
        index for index in range(len(header_row)) if index not in used_indexes
    ]
    if not candidate_indexes:
        return {}

    missing_fields = [field for field in _COVERAGE_REQUIRED_FIELDS if field not in field_indexes]
    if not missing_fields:
        return {}

    scores: dict[str, dict[int, float]] = {field: {} for field in missing_fields}
    for col_index in candidate_indexes:
        values = _coverage_column_sample_values(raw_rows, col_index)
        if not values:
            continue
        if "destination_uf" in missing_fields:
            scores["destination_uf"][col_index] = _score_coverage_column_as_uf(values)
        if "destination_city" in missing_fields:
            scores["destination_city"][col_index] = _score_coverage_column_as_city(values)
        if "freight_region" in missing_fields:
            scores["freight_region"][col_index] = _score_coverage_column_as_region(values)

    inferred: dict[str, int] = {}
    assigned: set[int] = set()
    thresholds = {
        "destination_uf": 0.75,
        "destination_city": 0.55,
        "freight_region": 0.45,
    }

    while True:
        best_field: str | None = None
        best_col: int | None = None
        best_score = -1.0
        for field in missing_fields:
            if field in inferred:
                continue
            threshold = thresholds[field]
            for col_index, score in scores.get(field, {}).items():
                if col_index in assigned or score < threshold:
                    continue
                if score > best_score:
                    best_score = score
                    best_field = field
                    best_col = col_index
        if best_field is None or best_col is None:
            break
        inferred[best_field] = best_col
        assigned.add(best_col)

    return inferred


def _resolve_coverage_field_indexes(raw_rows: list[list]) -> dict[str, int]:
    header_row = raw_rows[0]
    field_indexes: dict[str, int] = {}
    for index, header in enumerate(header_row):
        field_name = _resolve_coverage_field(str(header or ""))
        if field_name and field_name not in field_indexes:
            field_indexes[field_name] = index

    missing_fields = [field for field in _COVERAGE_REQUIRED_FIELDS if field not in field_indexes]
    if missing_fields:
        inferred = _infer_coverage_columns_from_content(raw_rows, field_indexes, header_row)
        for field_name, index in inferred.items():
            if field_name not in field_indexes:
                field_indexes[field_name] = index

    return field_indexes


def _format_coverage_missing_columns_error(missing_fields: list[str]) -> str:
    labels = [_COVERAGE_FIELD_LABELS[field] for field in missing_fields if field in _COVERAGE_FIELD_LABELS]
    hints = [
        _COVERAGE_FIELD_HINTS[field]
        for field in missing_fields
        if field in _COVERAGE_FIELD_HINTS
    ]
    message = "Colunas obrigatórias ausentes: " + ", ".join(labels) + "."
    if hints:
        message += " " + " ".join(hints)
    return message


def _normalize_destination_uf(value) -> str | None:
    cleaned = _sanitize_cell_string(value)
    if not cleaned:
        return None
    candidate = re.sub(r"[^A-Za-z]", "", cleaned).upper()
    if len(candidate) != 2:
        return None
    return candidate


def _empty_coverage_table_shell(*, uploaded_at: str | None = None) -> dict:
    return {
        "status": COVERAGE_TABLE_STATUS_NEEDS_REVIEW,
        "columns": list(COVERAGE_TABLE_COLUMNS),
        "rows": [],
        "validation_warnings": [],
        "human_review_status": None,
        "human_edited_at": None,
        "human_edited_by_user_id": None,
        "edit_version": 0,
        "uploaded_at": uploaded_at or _utcnow().isoformat(),
    }


def _public_coverage_table(coverage) -> dict | None:
    if not isinstance(coverage, dict):
        return None
    rows = coverage.get("rows")
    return {
        "status": coverage.get("status"),
        "columns": list(coverage.get("columns") or COVERAGE_TABLE_COLUMNS),
        "rows": list(rows) if isinstance(rows, list) else [],
        "validation_warnings": list(coverage.get("validation_warnings") or []),
        "human_review_status": coverage.get("human_review_status"),
        "human_edited_at": coverage.get("human_edited_at"),
        "human_edited_by_user_id": coverage.get("human_edited_by_user_id"),
        "edit_version": coverage.get("edit_version"),
        "uploaded_at": coverage.get("uploaded_at"),
    }


def _decode_coverage_csv_bytes(file_bytes: bytes) -> str:
    for encoding in ("utf-8-sig", "utf-8", "cp1252", "latin-1"):
        try:
            return (file_bytes or b"").decode(encoding)
        except UnicodeDecodeError:
            continue
    raise AgenteComparaCoverageError(
        ERROR_COVERAGE_PARSE_FAILED,
        "Não foi possível decodificar o arquivo CSV de cobertura.",
    )


def _parse_coverage_tabular_rows(raw_rows: list[list], *, source_file_name: str) -> tuple[list[dict], list[str]]:
    if not raw_rows:
        raise AgenteComparaCoverageError(
            ERROR_COVERAGE_EMPTY_FILE,
            "O arquivo de cobertura está vazio.",
        )
    header_row = raw_rows[0]
    if not isinstance(header_row, list):
        raise AgenteComparaCoverageError(
            ERROR_COVERAGE_PARSE_FAILED,
            "Cabeçalho do arquivo de cobertura inválido.",
        )
    field_indexes = _resolve_coverage_field_indexes(raw_rows)
    missing_fields = [
        field for field in _COVERAGE_REQUIRED_FIELDS if field not in field_indexes
    ]
    if missing_fields:
        raise AgenteComparaCoverageError(
            ERROR_COVERAGE_PARSE_FAILED,
            _format_coverage_missing_columns_error(missing_fields),
        )

    parsed_rows: list[dict] = []
    warnings: list[str] = []
    row_index = 0
    for raw_row in raw_rows[1:]:
        if not isinstance(raw_row, list):
            continue
        if not any(str(cell or "").strip() for cell in raw_row):
            continue
        row_index += 1
        if row_index > COVERAGE_UPLOAD_MAX_ROWS:
            raise AgenteComparaCoverageError(
                ERROR_COVERAGE_PAYLOAD_TOO_LARGE,
                "O arquivo de cobertura excede o limite de linhas permitido.",
            )

        def _cell(field: str) -> str:
            idx = field_indexes[field]
            if idx >= len(raw_row):
                return ""
            return str(raw_row[idx] or "").strip()

        destination_uf = _normalize_destination_uf(_cell("destination_uf"))
        destination_city = _sanitize_cell_string(_cell("destination_city"))
        freight_region = _sanitize_cell_string(_cell("freight_region"))
        row_errors: list[str] = []
        if not destination_uf:
            row_errors.append("UF inválida ou ausente")
        if not destination_city:
            row_errors.append("cidade ausente")
        if not freight_region:
            row_errors.append("região de frete ausente")
        if row_errors:
            warnings.append(f"Linha {row_index} ignorada: {', '.join(row_errors)}.")
            continue
        parsed_rows.append(
            {
                "destination_uf": destination_uf,
                "destination_city": destination_city,
                "freight_region": freight_region,
                "row_index": row_index,
                "source_file_name": source_file_name,
                "confidence": None,
                "evidence_ref": None,
                "notes": "",
            }
        )
    if not parsed_rows and not warnings:
        raise AgenteComparaCoverageError(
            ERROR_COVERAGE_EMPTY_FILE,
            "Nenhuma linha válida encontrada no arquivo de cobertura.",
        )
    return parsed_rows, warnings


def _parse_coverage_csv_bytes(file_bytes: bytes, *, source_file_name: str) -> tuple[list[dict], list[str]]:
    text = _decode_coverage_csv_bytes(file_bytes)
    if not text.strip():
        raise AgenteComparaCoverageError(
            ERROR_COVERAGE_EMPTY_FILE,
            "O arquivo de cobertura está vazio.",
        )
    reader = csv.reader(io.StringIO(text))
    raw_rows = [row for row in reader]
    return _parse_coverage_tabular_rows(raw_rows, source_file_name=source_file_name)


def _parse_coverage_xlsx_bytes(file_bytes: bytes, *, source_file_name: str) -> tuple[list[dict], list[str]]:
    if not file_bytes:
        raise AgenteComparaCoverageError(
            ERROR_COVERAGE_EMPTY_FILE,
            "O arquivo de cobertura está vazio.",
        )
    if len(file_bytes) > COVERAGE_UPLOAD_MAX_BYTES:
        raise AgenteComparaCoverageError(
            ERROR_COVERAGE_PAYLOAD_TOO_LARGE,
            "O arquivo de cobertura excede o limite de tamanho permitido.",
        )
    try:
        if zipfile.is_zipfile(io.BytesIO(file_bytes)):
            with zipfile.ZipFile(io.BytesIO(file_bytes)) as archive:
                if archive.testzip() is not None:
                    raise AgenteComparaCoverageError(
                        ERROR_COVERAGE_PARSE_FAILED,
                        "Arquivo XLSX de cobertura corrompido.",
                    )
        from openpyxl import load_workbook

        workbook = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    except AgenteComparaCoverageError:
        raise
    except Exception as exc:
        raise AgenteComparaCoverageError(
            ERROR_COVERAGE_PARSE_FAILED,
            "Não foi possível ler o arquivo XLSX de cobertura.",
        ) from exc

    raw_rows: list[list] = []
    try:
        sheet = workbook.active
        for row in sheet.iter_rows(values_only=True):
            raw_rows.append(["" if cell is None else str(cell) for cell in row])
    finally:
        workbook.close()
    return _parse_coverage_tabular_rows(raw_rows, source_file_name=source_file_name)


def _validate_coverage_row_for_save(item, *, row_index: int) -> dict:
    if not isinstance(item, dict):
        raise AgenteComparaTempTableError(
            ERROR_TEMP_TABLE_INVALID_PAYLOAD,
            "Cada linha de coverage_table deve ser um objeto.",
        )
    destination_uf = _normalize_destination_uf(item.get("destination_uf"))
    destination_city = _sanitize_cell_string(item.get("destination_city"))
    freight_region = _sanitize_cell_string(item.get("freight_region"))
    notes = _sanitize_cell_string(item.get("notes")) or ""
    if not destination_uf:
        raise AgenteComparaTempTableError(
            ERROR_TEMP_TABLE_INVALID_PAYLOAD,
            f"Linha {row_index}: UF destino inválida ou ausente.",
        )
    if not destination_city:
        raise AgenteComparaTempTableError(
            ERROR_TEMP_TABLE_INVALID_PAYLOAD,
            f"Linha {row_index}: cidade destino é obrigatória.",
        )
    if not freight_region:
        raise AgenteComparaTempTableError(
            ERROR_TEMP_TABLE_INVALID_PAYLOAD,
            f"Linha {row_index}: região de frete é obrigatória.",
        )
    return {
        "destination_uf": destination_uf,
        "destination_city": destination_city,
        "freight_region": freight_region,
        "row_index": row_index,
        "source_file_name": _sanitize_cell_string(item.get("source_file_name")),
        "confidence": item.get("confidence"),
        "evidence_ref": _sanitize_cell_string(item.get("evidence_ref")),
        "notes": notes,
    }


def _validate_coverage_table_for_save(raw_coverage) -> dict | None:
    if raw_coverage is None:
        return None
    if not isinstance(raw_coverage, dict):
        raise AgenteComparaTempTableError(
            ERROR_TEMP_TABLE_INVALID_PAYLOAD,
            "coverage_table deve ser um objeto.",
        )
    raw_rows = raw_coverage.get("rows")
    if raw_rows is None:
        return {"rows": []}
    if not isinstance(raw_rows, list):
        raise AgenteComparaTempTableError(
            ERROR_TEMP_TABLE_INVALID_PAYLOAD,
            "coverage_table.rows deve ser uma lista.",
        )
    normalized_rows: list[dict] = []
    for index, row in enumerate(raw_rows, start=1):
        normalized_rows.append(_validate_coverage_row_for_save(row, row_index=index))
    return {"rows": normalized_rows}


def upload_coverage_table_from_file(
    *,
    display_name: str,
    file_bytes: bytes,
    extension: str | None,
    user_scope=None,
    franquia_scope=None,
) -> dict:
    """
    Upload complementar determinístico de coverage_table no tt_*.json ativo.

    Não registra documento principal, não chama Gemini e não dispara extração de frete.
    """
    started_at = time.perf_counter()
    emitted_processing_event = [False]
    execution_id = _resolve_agente_compara_execution_id()
    _require_session()
    if not file_bytes:
        raise AgenteComparaCoverageError(
            ERROR_COVERAGE_EMPTY_FILE,
            "O arquivo de cobertura está vazio.",
        )
    if len(file_bytes) > COVERAGE_UPLOAD_MAX_BYTES:
        raise AgenteComparaCoverageError(
            ERROR_COVERAGE_PAYLOAD_TOO_LARGE,
            "O arquivo de cobertura excede o limite de tamanho permitido.",
        )

    ext = (extension or "").strip().lower()
    if not ext.startswith("."):
        ext = f".{ext}" if ext else ""
    if ext == ".pdf":
        raise AgenteComparaCoverageError(
            ERROR_COVERAGE_INVALID_FORMAT,
            "Upload de cobertura aceita apenas CSV e XLSX nesta fase.",
        )
    if ext not in {".csv", ".xlsx"}:
        raise AgenteComparaCoverageError(
            ERROR_COVERAGE_INVALID_FORMAT,
            "Upload de cobertura aceita apenas CSV e XLSX nesta fase.",
        )

    sync_temp_table_with_session_documents()
    active_id = get_temp_table_id(session)
    if not active_id:
        raise AgenteComparaCoverageError(
            ERROR_COVERAGE_NO_TEMP_TABLE,
            "Nenhuma tabela temporária ativa nesta sessão.",
        )

    cfg = get_cleiton_doc_config()
    record = load_temp_table_record(active_id, ttl_hours=cfg.upload_ttl_hours)
    if record is None:
        clear_temp_table_session_refs(session)
        _mark_session_modified()
        raise AgenteComparaCoverageError(
            ERROR_COVERAGE_NO_TEMP_TABLE,
            "Tabela temporária ativa não encontrada.",
        )
    status = (record.get("status") or "").strip().lower()
    if status == TEMP_TABLE_STATUS_EXPIRED:
        raise AgenteComparaCoverageError(
            ERROR_COVERAGE_EXPIRED,
            "A tabela temporária desta sessão expirou.",
        )
    if status in {TEMP_TABLE_STATUS_DISCARDED, TEMP_TABLE_STATUS_PROCESSING}:
        raise AgenteComparaCoverageError(
            ERROR_COVERAGE_NO_TEMP_TABLE,
            "Tabela temporária indisponível para upload complementar.",
        )
    _assert_temp_table_scope(record, user_scope=user_scope, franquia_scope=franquia_scope)

    safe_name = secure_filename(display_name or "coverage") or "coverage"
    if ext == ".csv":
        rows, warnings = _parse_coverage_csv_bytes(file_bytes, source_file_name=safe_name)
    else:
        rows, warnings = _parse_coverage_xlsx_bytes(file_bytes, source_file_name=safe_name)

    now = _utcnow().isoformat()
    preserved_expires_at = record.get("expires_at")
    updated = dict(record)
    coverage = _empty_coverage_table_shell(uploaded_at=now)
    existing_coverage = record.get("coverage_table")
    if isinstance(existing_coverage, dict):
        coverage["human_review_status"] = existing_coverage.get("human_review_status")
        coverage["human_edited_at"] = existing_coverage.get("human_edited_at")
        coverage["human_edited_by_user_id"] = existing_coverage.get("human_edited_by_user_id")
        coverage["edit_version"] = existing_coverage.get("edit_version") or 0
    coverage["rows"] = rows
    coverage["validation_warnings"] = warnings
    updated["coverage_table"] = coverage
    updated["updated_at"] = now
    updated["expires_at"] = preserved_expires_at

    saved = save_temp_table_record(updated)
    logger.info(
        "Agente Compara coverage upload: temp_table_id=%s user_id=%s rows=%s warnings=%s",
        saved.get("temp_table_id"),
        user_scope,
        len(rows),
        len(warnings),
    )
    public = _public_temp_table(saved)
    if public is None:
        raise AgenteComparaCoverageError(
            ERROR_COVERAGE_NO_TEMP_TABLE,
            "Não foi possível retornar a tabela temporária atualizada.",
        )
    _emit_agente_compara_operational_billing(
        emitted=emitted_processing_event,
        started_at=started_at,
        flow_type=AGENTE_COMPARA_COVERAGE_UPLOAD_FLOW_TYPE,
        idempotency_key=agente_compara_coverage_upload_idempotency_key(active_id, execution_id),
        rows_processed=len(rows),
        status="success",
        execution_id=execution_id,
    )
    return public


def _audit_header_aliases() -> dict[str, tuple[str, ...]]:
    return {
        "destination_city": (
            "cidade destino",
            "cidade de destino",
            "cidade_destino",
            "destino cidade",
        ),
        "destination_uf": (
            "uf destino",
            "uf de destino",
            "uf_destino",
            "destino uf",
            "estado destino",
        ),
        "charged_freight": (
            "valor frete",
            "valor do frete",
            "valor_frete",
            "valor frete cobrado",
            "valor do frete cobrado",
            "valor_frete_cobrado",
            "frete cobrado",
            "frete",
        ),
        "audited_weight": (
            "peso",
            "peso auditado",
            "peso_auditado",
            "peso cobrado",
            "peso kg",
        ),
        "carrier": ("transportadora", "nome transportadora", "empresa transportadora"),
        "document_number": (
            "numero documento",
            "numero do documento",
            "numero_documento",
            "documento",
            "nf",
            "nota fiscal",
        ),
        "origin_city": ("cidade origem", "cidade de origem", "cidade_origem"),
        "origin_uf": ("uf origem", "uf de origem", "uf_origem", "estado origem"),
        "invoice_value": ("valor nf", "valor da nf", "valor_nf", "valor nota fiscal"),
        "modal": ("modal", "tipo modal", "modalidade"),
        "issue_date": ("data emissao", "data de emissao", "data_emissao", "emissao"),
        "delivery_date": ("data entrega", "data de entrega", "data_entrega", "entrega"),
    }


def _resolve_audit_field(header_value) -> str | None:
    normalized = _normalize_coverage_header(header_value)
    if not normalized:
        return None
    for field_name, aliases in _audit_header_aliases().items():
        if normalized in aliases:
            return field_name
    return None


def _resolve_audit_field_indexes(header_row: list) -> tuple[dict[str, int], dict[str, str]]:
    field_indexes: dict[str, int] = {}
    header_map: dict[str, str] = {}
    for index, header in enumerate(header_row):
        field_name = _resolve_audit_field(header)
        if field_name and field_name not in field_indexes:
            field_indexes[field_name] = index
            source_header = _sanitize_cell_string(header) or str(header or "").strip()
            header_map[source_header] = field_name
    return field_indexes, header_map


def _format_audit_missing_columns_error(missing_fields: list[str]) -> str:
    labels = [_AUDIT_FIELD_LABELS.get(field, field) for field in missing_fields]
    return "Colunas obrigatórias ausentes: " + ", ".join(labels) + "."


def _parse_audit_numeric(value) -> float | None:
    cleaned = _sanitize_cell_string(value)
    if cleaned is None:
        return None
    text = cleaned.replace(" ", "")
    if not text:
        return None
    if "," in text and "." in text:
        text = text.replace(".", "").replace(",", ".")
    elif "," in text:
        text = text.replace(",", ".")
    try:
        parsed = float(text)
    except (TypeError, ValueError):
        return None
    if parsed < 0:
        return None
    return parsed


def _audit_row_is_empty(raw_row: list, field_indexes: dict[str, int]) -> bool:
    for index in field_indexes.values():
        if index < len(raw_row):
            cell = _sanitize_cell_string(raw_row[index])
            if cell:
                return False
    return True


def _normalize_audit_row(
    raw_row: list,
    *,
    row_index: int,
    field_indexes: dict[str, int],
    source_file_name: str,
) -> dict | None:
    if _audit_row_is_empty(raw_row, field_indexes):
        return None

    destination_city = _sanitize_cell_string(
        raw_row[field_indexes["destination_city"]]
        if field_indexes["destination_city"] < len(raw_row)
        else None
    )
    destination_uf = _normalize_destination_uf(
        raw_row[field_indexes["destination_uf"]]
        if field_indexes["destination_uf"] < len(raw_row)
        else None
    )
    charged_freight = _parse_audit_numeric(
        raw_row[field_indexes["charged_freight"]]
        if field_indexes["charged_freight"] < len(raw_row)
        else None
    )
    audited_weight = _parse_audit_numeric(
        raw_row[field_indexes["audited_weight"]]
        if field_indexes["audited_weight"] < len(raw_row)
        else None
    )

    if not destination_city or not destination_uf or charged_freight is None or audited_weight is None:
        raise AgenteComparaBatchError(
            ERROR_AUDIT_PARSE_FAILED,
            f"Linha {row_index}: dados obrigatórios inválidos ou ausentes.",
        )

    normalized: dict = {
        "row_index": row_index,
        "destination_city": destination_city,
        "destination_uf": destination_uf,
        "charged_freight": charged_freight,
        "audited_weight": audited_weight,
        "source_file_name": source_file_name,
    }

    for field_name in _AUDIT_OPTIONAL_FIELDS:
        if field_name not in field_indexes:
            continue
        index = field_indexes[field_name]
        raw_value = raw_row[index] if index < len(raw_row) else None
        if field_name in {"invoice_value"}:
            parsed_value = _parse_audit_numeric(raw_value)
            if parsed_value is not None:
                normalized[field_name] = parsed_value
            continue
        if field_name == "origin_uf":
            parsed_uf = _normalize_destination_uf(raw_value)
            if parsed_uf:
                normalized[field_name] = parsed_uf
            continue
        cleaned = _sanitize_cell_string(raw_value)
        if cleaned:
            normalized[field_name] = cleaned

    return normalized


def _parse_audit_tabular_rows(
    raw_rows: list[list],
    *,
    source_file_name: str,
    max_rows: int,
) -> tuple[list[dict], dict[str, str], str | None]:
    if not raw_rows:
        raise AgenteComparaBatchError(
            ERROR_AUDIT_EMPTY_FILE,
            "O arquivo auditado está vazio.",
        )

    header_row = raw_rows[0]
    field_indexes, header_map = _resolve_audit_field_indexes(header_row)
    missing_fields = [
        field for field in _AUDIT_REQUIRED_FIELDS if field not in field_indexes
    ]
    if missing_fields:
        raise AgenteComparaBatchError(
            ERROR_AUDIT_MISSING_COLUMNS,
            _format_audit_missing_columns_error(missing_fields),
        )

    normalized_rows: list[dict] = []
    data_row_index = 0
    for raw_row in raw_rows[1:]:
        if _audit_row_is_empty(raw_row, field_indexes):
            continue
        data_row_index += 1
        if data_row_index > max_rows:
            raise AgenteComparaBatchError(
                ERROR_AUDIT_TOO_MANY_ROWS,
                f"O arquivo excede o limite de {max_rows} linhas configurado para auditoria.",
            )
        normalized = _normalize_audit_row(
            raw_row,
            row_index=data_row_index,
            field_indexes=field_indexes,
            source_file_name=source_file_name,
        )
        if normalized is not None:
            normalized_rows.append(normalized)

    if not normalized_rows:
        raise AgenteComparaBatchError(
            ERROR_AUDIT_EMPTY_ROWS,
            "Nenhuma linha válida encontrada no arquivo auditado.",
        )

    public_header_map = {
        _AUDIT_FIELD_LABELS.get(field, field): field
        for field in field_indexes
        if field in _AUDIT_FIELD_LABELS
    }
    public_header_map.update(
        {
            source: target
            for source, target in header_map.items()
            if source and target
        }
    )
    return normalized_rows, public_header_map, None


def _parse_audit_csv_bytes(
    file_bytes: bytes,
    *,
    source_file_name: str,
    max_rows: int,
) -> tuple[list[dict], dict[str, str], str | None]:
    text = _decode_coverage_csv_bytes(file_bytes)
    reader = csv.reader(io.StringIO(text))
    raw_rows = [[cell for cell in row] for row in reader if row]
    rows, header_map, _ = _parse_audit_tabular_rows(
        raw_rows,
        source_file_name=source_file_name,
        max_rows=max_rows,
    )
    return rows, header_map, None


def _parse_audit_xlsx_bytes(
    file_bytes: bytes,
    *,
    source_file_name: str,
    max_bytes: int,
    max_rows: int,
) -> tuple[list[dict], dict[str, str], str]:
    if not file_bytes:
        raise AgenteComparaBatchError(
            ERROR_AUDIT_EMPTY_FILE,
            "O arquivo auditado está vazio.",
        )
    if len(file_bytes) > int(max_bytes):
        raise AgenteComparaBatchError(
            ERROR_AUDIT_PAYLOAD_TOO_LARGE,
            "O arquivo auditado excede o limite de tamanho permitido.",
        )
    try:
        if zipfile.is_zipfile(io.BytesIO(file_bytes)):
            with zipfile.ZipFile(io.BytesIO(file_bytes)) as archive:
                if archive.testzip() is not None:
                    raise AgenteComparaBatchError(
                        ERROR_AUDIT_PARSE_FAILED,
                        "Arquivo XLSX auditado corrompido.",
                    )
        from openpyxl import load_workbook

        workbook = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    except AgenteComparaBatchError:
        raise
    except Exception as exc:
        raise AgenteComparaBatchError(
            ERROR_AUDIT_PARSE_FAILED,
            "Não foi possível ler o arquivo XLSX auditado.",
        ) from exc

    sheet = None
    try:
        if AUDIT_BATCH_SHEET_NAME in workbook.sheetnames:
            sheet = workbook[AUDIT_BATCH_SHEET_NAME]
        else:
            raise AgenteComparaBatchError(
                ERROR_AUDIT_INVALID_SHEET,
                f"A aba '{AUDIT_BATCH_SHEET_NAME}' é obrigatória no arquivo XLSX auditado.",
            )
        raw_rows: list[list] = []
        for row in sheet.iter_rows(values_only=True):
            raw_rows.append(["" if cell is None else cell for cell in row])
    finally:
        workbook.close()

    rows, header_map, _ = _parse_audit_tabular_rows(
        raw_rows,
        source_file_name=source_file_name,
        max_rows=max_rows,
    )
    return rows, header_map, AUDIT_BATCH_SHEET_NAME


def _empty_audit_batch_shell(*, uploaded_at: str | None = None) -> dict:
    return {
        "status": AUDIT_BATCH_STATUS_UPLOADED,
        "audit_batch_id": None,
        "temp_table_id": None,
        "source_file_name": None,
        "sheet_name": None,
        "uploaded_at": uploaded_at,
        "created_at": None,
        "updated_at": None,
        "expires_at": None,
        "row_count": 0,
        "max_rows": None,
        "input_schema_version": AUDIT_INPUT_SCHEMA_VERSION,
        "header_map": {},
        "normalized_rows": [],
        "results": [],
        "summary": None,
        "audit_diagnostics": None,
    }


def _normalize_audit_lookup_text(value) -> str:
    cleaned = _sanitize_cell_string(value)
    if not cleaned:
        return ""
    text = unicodedata.normalize("NFKD", str(cleaned))
    text = "".join(ch for ch in text if not unicodedata.combining(ch))
    text = text.upper().strip()
    text = re.sub(r"[_\-\/.,:;]+", " ", text)
    text = re.sub(r"[^A-Z0-9\s]", " ", text)
    text = re.sub(r"\s+", " ", text)
    return text.strip()


def _coverage_lookup_key(destination_uf, destination_city) -> str | None:
    uf = _normalize_destination_uf(destination_uf)
    city = _normalize_audit_lookup_text(destination_city)
    if not uf or not city:
        return None
    return f"{uf}|{city}"


def build_coverage_index(coverage_table) -> dict:
    rows = []
    if isinstance(coverage_table, dict):
        rows = coverage_table.get("rows") or []
    elif isinstance(coverage_table, list):
        rows = coverage_table
    if not isinstance(rows, list):
        return {}

    grouped: dict[str, set[str]] = {}
    for row in rows:
        if not isinstance(row, dict):
            continue
        key = _coverage_lookup_key(row.get("destination_uf"), row.get("destination_city"))
        region = _sanitize_cell_string(row.get("freight_region"))
        if not key or not region:
            continue
        grouped.setdefault(key, set()).add(region)

    index: dict[str, object] = {}
    for key, regions in grouped.items():
        ordered = sorted(regions)
        if len(ordered) == 1:
            index[key] = ordered[0]
        else:
            index[key] = {
                "reason_code": AUDIT_STATUS_AMBIGUOUS_COVERAGE,
                "regions": ordered,
            }
    return index


def _parse_brazilian_money(value) -> float | None:
    cleaned = _sanitize_cell_string(value)
    if cleaned is None:
        return None
    text = cleaned.strip()
    if not text:
        return None
    text = re.sub(r"(?i)\bR\$\b|R\$", "", text)
    text = re.sub(r"[^0-9,\.\-]", "", text)
    if not text or text in {"-", ".", ","}:
        return None
    if "," in text and "." in text:
        text = text.replace(".", "").replace(",", ".")
    elif "," in text:
        text = text.replace(",", ".")
    try:
        parsed = float(text)
    except (TypeError, ValueError):
        return None
    if parsed < 0:
        return None
    return parsed


def _parse_weight_number(value) -> float | None:
    cleaned = _sanitize_cell_string(value)
    if cleaned is None:
        return None
    text = cleaned.strip()
    if not text:
        return None
    text = re.sub(r"[^0-9,\.\-]", "", text)
    if "," in text and "." in text:
        text = text.replace(".", "").replace(",", ".")
    elif "," in text:
        text = text.replace(",", ".")
    try:
        parsed = float(text)
    except (TypeError, ValueError):
        return None
    if parsed < 0:
        return None
    return parsed


def _parse_range_from_label(label) -> tuple[float | None, float | None] | None:
    if _is_excess_column(label):
        return None
    normalized = _normalize_coverage_header(label)
    if not normalized or "acima" in normalized:
        return None
    numbers = [float(item.replace(",", ".")) for item in re.findall(r"\d+(?:[,.]\d+)?", normalized)]
    if not numbers:
        return None
    has_weight_unit = bool(_WEIGHT_KG_HEADER_RE.search(normalized))
    if re.search(r"\bate\b", normalized) and has_weight_unit:
        return (0.0, numbers[-1])
    if normalized.startswith("ate ") or normalized.startswith("ate"):
        return (0.0, numbers[-1])
    if len(numbers) >= 2 and (
        " a " in f" {normalized} "
        or normalized.startswith("de ")
        or " ate " in f" {normalized} "
    ):
        return (numbers[0], numbers[1])
    if len(numbers) == 1 and has_weight_unit:
        return (0.0, numbers[0])
    return None


def _normalize_brackets(brackets: list[dict]) -> list[dict]:
    cleaned = [
        bracket
        for bracket in brackets
        if isinstance(bracket.get("max_kg"), (int, float))
        and isinstance(bracket.get("value"), (int, float))
    ]
    cleaned.sort(key=lambda item: (float(item.get("max_kg")), float(item.get("min_kg") or 0)))
    previous_max = 0.0
    normalized: list[dict] = []
    for index, bracket in enumerate(cleaned):
        max_kg = float(bracket["max_kg"])
        if max_kg < previous_max:
            continue
        min_kg = 0.0 if index == 0 else previous_max
        normalized.append(
            {
                "min_kg": min_kg,
                "max_kg": max_kg,
                "value": round(float(bracket["value"]), 2),
                "label": bracket.get("label") or f"Faixa até {max_kg:g} kg",
            }
        )
        previous_max = max_kg
    return normalized


def _is_region_column(column_name) -> bool:
    normalized = _normalize_coverage_header(column_name)
    return normalized in {
        "uf cidades",
        "uf cidade",
        "regiao",
        "regiao de frete",
        "praca",
        "rota",
        "itinerario",
        "destino",
        "destino frete",
        "cidade",
        "cidades",
        "cidade destino",
        "municipio",
        "municipio destino",
    }


def _is_destination_uf_column(column_name) -> bool:
    normalized = _normalize_coverage_header(column_name)
    return normalized in {
        "uf",
        "uf destino",
        "uf de destino",
        "uf dest",
        "uf entrega",
        "uf de entrega",
    }


def _is_city_destination_column(column_name) -> bool:
    normalized = _normalize_coverage_header(column_name)
    return normalized in {
        "uf cidades",
        "uf cidade",
        "cidade",
        "cidades",
        "cidade destino",
        "municipio",
        "municipio destino",
        "destino",
    }


def _is_value_column(column_name) -> bool:
    normalized = _normalize_coverage_header(column_name)
    if "pedagio" in normalized or "gris" in normalized or "seguro" in normalized:
        return False
    if "tso" in normalized or "tas" in normalized or "frete valor" in normalized:
        return False
    return normalized in {"frete", "valor", "valor frete", "frete peso", "tarifa"}


def _is_excess_column(column_name) -> bool:
    normalized = _normalize_coverage_header(column_name)
    if not normalized:
        return False
    if "excedente" in normalized or "excesso" in normalized:
        return True
    if "adicional por kg" in normalized or "kg excedente" in normalized:
        return True
    if "acima" in normalized and _WEIGHT_KG_HEADER_RE.search(normalized):
        return True
    return False


def _is_toll_column(column_name) -> bool:
    normalized = _normalize_coverage_header(column_name)
    return normalized == "pedagio" or normalized.startswith("pedagio ")


def _extract_route_toll_from_row(columns: list, row: dict) -> dict | None:
    for column in columns:
        if not _is_toll_column(column):
            continue
        source_value = _sanitize_cell_string(row.get(column))
        rate = _parse_brazilian_money(source_value)
        if rate is None:
            continue
        return {
            "rate_per_fraction": float(rate),
            "fraction_size_kg": 100.0,
            "source_column": column,
            "source_value": source_value,
        }
    return None


def _is_direct_kg_column(column_name) -> bool:
    normalized = _normalize_coverage_header(column_name)
    if "excedente" in normalized or "frete valor" in normalized:
        return False
    return (
        normalized in {"kg", "por kg", "valor kg", "frete kg", "frete peso kg", "frete peso"}
        or "r kg" in normalized
        or "rs kg" in normalized
    )


def _is_direct_ton_column(column_name) -> bool:
    normalized = _normalize_coverage_header(column_name)
    return (
        "tonelada" in normalized
        or normalized in {"ton", "por ton", "valor ton", "frete ton"}
        or "r ton" in normalized
        or "rs ton" in normalized
    )


def _is_freight_value_percent_column(column_name) -> bool:
    raw = _sanitize_cell_string(column_name) or ""
    normalized = _normalize_coverage_header(raw)
    if not normalized or "seguro" in normalized:
        return False

    has_percent_marker = "%" in raw or "pct" in normalized or "percentual" in normalized or "perc" in normalized
    if normalized in {
        "adv",
        "adv %",
        "adv percent",
        "adv percentual",
        "ad val",
        "ad valorem",
        "frete valor",
        "frete valor %",
        "frete valor percent",
        "frete valor percentual",
    }:
        return True
    if normalized in {
        "sobre nf",
        "sob nf",
        "s nf",
        "percentual nf",
        "perc nf",
    }:
        return True
    if normalized in {
        "nf",
        "nota",
        "nota fiscal",
        "valor nf",
        "frete valor",
        "freight value pct",
        "sobre nf",
        "fv",
        "f v",
        "ad valorem",
    }:
        return has_percent_marker
    return False


def _parse_decimal_number(value) -> Decimal | None:
    cleaned = _sanitize_cell_string(value)
    if cleaned is None:
        return None
    text = cleaned.strip()
    if not text:
        return None
    text = re.sub(r"(?i)\bR\$\b|R\$", "", text)
    text = re.sub(r"[^0-9,\.\-]", "", text)
    if not text or text in {"-", ".", ","}:
        return None
    if "," in text and "." in text:
        text = text.replace(".", "").replace(",", ".")
    elif "," in text:
        text = text.replace(",", ".")
    try:
        parsed = Decimal(text)
    except (InvalidOperation, ValueError):
        return None
    if parsed < 0:
        return None
    return parsed


def _parse_percentage_rate(value) -> Decimal | None:
    parsed = _parse_decimal_number(value)
    if parsed is None:
        return None
    return parsed / Decimal("100")


def _parse_freight_value_rate(value) -> Decimal | None:
    return _parse_percentage_rate(value)


def _extract_freight_value_from_row(columns: list, row: dict) -> dict | None:
    for column in columns:
        if not _is_freight_value_percent_column(column):
            continue
        source_value = _sanitize_cell_string(row.get(column))
        rate = _parse_freight_value_rate(source_value)
        if rate is None:
            continue
        return {
            "rate": float(rate),
            "source_column": column,
            "source_value": source_value,
            "calculation_base": "invoice_value",
        }
    return None


def _calculate_route_toll_amount(route_toll: dict, weight_kg) -> tuple[Decimal, dict] | None:
    if not isinstance(route_toll, dict):
        return None
    weight = _parse_weight_number(weight_kg)
    if weight is None:
        return None
    rate = _decimal_money(route_toll.get("rate_per_fraction"))
    fraction_size = _parse_decimal_number(route_toll.get("fraction_size_kg")) or Decimal("100")
    if rate is None or rate <= 0 or fraction_size <= 0:
        return None
    weight_decimal = Decimal(str(weight))
    fractions = int((weight_decimal / fraction_size).to_integral_value(rounding=ROUND_CEILING))
    amount = rate * Decimal(fractions)
    component = {
        "amount": _round_money(amount),
        "rate_per_fraction": _round_money(rate),
        "fraction_size_kg": float(fraction_size),
        "fractions": fractions,
        "source_column": route_toll.get("source_column"),
        "source_value": route_toll.get("source_value"),
        "details": (
            f"ceil({_format_brazilian_decimal(weight_decimal)} / "
            f"{_format_brazilian_decimal(fraction_size)}) x "
            f"{_format_brazilian_decimal(rate)} = {_format_brazilian_decimal(amount)}"
        ),
        "reason_code": AUDIT_REASON_ROUTE_TOLL_APPLIED,
    }
    return amount, component


_ACCESSORIAL_NOT_APPLIED_ALIASES = {
    "gris",
    "gerenciamento de risco",
    "taxa de risco",
    "adicional de risco",
    "risco",
    "seguro",
    "seguro carga",
    "seguro rctr c",
    "seguro ad valorem",
    "frete valor seguro",
}

_ACCESSORIAL_AMBIGUOUS_ALIASES = {
    "valor frete",
    "taxa",
    "outras taxas",
    "generalidades",
    "valor",
    "taxa administrativa",
    "minimo",
    "minimo nf",
    "minimo por conhecimento",
}

_ACCESSORIAL_BASE_ALIASES = {
    "nf",
    "nota",
    "nota fiscal",
    "valor nf",
    "valor nota",
    "valor da nota",
    "valor nota fiscal",
    "valor da nota fiscal",
    "sobre nf",
    "sob nf",
    "s nf",
    "sobre nota",
    "sobre nota fiscal",
    "sobre valor da nota fiscal",
    "percentual nf",
    "perc nf",
}

_ACCESSORIAL_PERCENTUAL_NF_ALIASES = {
    "nf",
    "nota",
    "nota fiscal",
    "valor nf",
    "sobre nf",
    "sob nf",
    "s nf",
    "percentual nf",
    "perc nf",
    "taxa sobre nf",
    "taxa valor nf",
}


def _has_percent_marker(*values) -> bool:
    for value in values:
        raw = _sanitize_cell_string(value)
        if not raw:
            continue
        normalized = _normalize_coverage_header(raw)
        if "%" in raw or "pct" in normalized or "percentual" in normalized or "perc" in normalized:
            return True
    return False


def _is_invoice_value_basis(value) -> bool:
    normalized = _normalize_coverage_header(value)
    return not normalized or normalized in _ACCESSORIAL_BASE_ALIASES


def _classify_accessorial_invoice_percentage_fee(fee: dict) -> tuple[str | None, str | None]:
    label = _sanitize_cell_string(fee.get("name")) or ""
    normalized = _normalize_coverage_header(label)
    if not normalized:
        return None, AUDIT_REASON_AMBIGUOUS_ACCESSORIAL_PERCENTAGE
    if normalized in _ACCESSORIAL_NOT_APPLIED_ALIASES:
        return normalized.replace(" ", "_"), AUDIT_REASON_ACCESSORIAL_FEE_NOT_APPLIED
    if normalized in _ACCESSORIAL_AMBIGUOUS_ALIASES:
        return None, AUDIT_REASON_AMBIGUOUS_ACCESSORIAL_PERCENTAGE
    if not _is_invoice_value_basis(fee.get("calculation_basis")):
        return None, AUDIT_REASON_UNSUPPORTED_ACCESSORIAL_CONDITION

    has_percent_marker = _has_percent_marker(label, fee.get("value"), fee.get("unit"))
    if normalized in {"ad valorem", "advalorem", "ad val"}:
        if has_percent_marker:
            return "ad_valorem", None
        return "ad_valorem", AUDIT_REASON_AMBIGUOUS_ACCESSORIAL_PERCENTAGE
    if normalized in {"frete valor", "fv", "f v"}:
        if has_percent_marker:
            return "freight_value", None
        return "freight_value", AUDIT_REASON_AMBIGUOUS_ACCESSORIAL_PERCENTAGE
    if normalized in _ACCESSORIAL_PERCENTUAL_NF_ALIASES:
        return "invoice_percentage", None
    return None, AUDIT_REASON_AMBIGUOUS_ACCESSORIAL_PERCENTAGE


def _accessorial_ignored_component(fee: dict, canonical_name: str | None, reason_code: str) -> dict:
    canonical_component = fee.get("canonical_component") or canonical_name
    component = {
        "label": _sanitize_cell_string(fee.get("name")),
        "canonical_name": canonical_name,
        "canonical_component": canonical_component,
        "component_group": fee.get("component_group"),
        "calculation_type": fee.get("calculation_type"),
        "source_value": _sanitize_cell_string(fee.get("value")),
        "source_unit": _sanitize_cell_string(fee.get("unit")),
        "source_block": "accessorial_fees",
        "reason_code": reason_code,
    }
    return {key: value for key, value in component.items() if value is not None}


def _accessorial_runtime_rate(fee: dict) -> Decimal | None:
    rate = _parse_decimal_number(fee.get("rate"))
    if rate is None:
        rate = _parse_percentage_rate(fee.get("value"))
    if rate is None or rate <= 0:
        return None
    return rate


def _accessorial_runtime_minimum_amount(fee: dict) -> Decimal | None:
    amount = _decimal_money(fee.get("minimum_amount"))
    if amount is None or amount <= 0:
        return None
    return amount


def _accessorial_uses_invoice_value_base(fee: dict) -> bool:
    calculation_base = _normalize_coverage_header(fee.get("calculation_base"))
    if calculation_base in {"invoice value", "valor invoice", "valor da invoice"}:
        return True
    if calculation_base and calculation_base not in {"invoice", "nota", "nota fiscal", "valor nf"}:
        return False
    return _is_invoice_value_basis(fee.get("calculation_basis"))


def _accessorial_fee_component_ref(fee: dict) -> set[str]:
    refs = {
        fee.get("component_group"),
        fee.get("canonical_component"),
        fee.get("related_to"),
    }
    return {str(ref) for ref in refs if ref}


def _accessorial_is_duplicate_of_tariff_freight_value(fee: dict, has_tariff_freight_value: bool) -> bool:
    if not has_tariff_freight_value:
        return False
    refs = _accessorial_fee_component_ref(fee)
    return bool(refs & {"freight_value", "ad_valorem"})


def _accessorial_is_calculable_invoice_percentage(fee: dict) -> bool:
    return (
        fee.get("calculation_type") == "invoice_percentage"
        and fee.get("status") == "calculable"
        and fee.get("classification_confidence") == "high"
        and not fee.get("conditions")
        and not fee.get("unsupported_reason")
        and _accessorial_uses_invoice_value_base(fee)
        and _accessorial_runtime_rate(fee) is not None
    )


def _accessorial_runtime_ignored_reason(fee: dict) -> str:
    if fee.get("conditions") or fee.get("unsupported_reason"):
        return AUDIT_REASON_UNSUPPORTED_ACCESSORIAL_CONDITION
    if fee.get("calculation_type") in {"invoice_percentage", "unknown"}:
        return AUDIT_REASON_AMBIGUOUS_ACCESSORIAL_PERCENTAGE
    return AUDIT_REASON_UNSUPPORTED_ACCESSORIAL_CONDITION


_CONFIGURED_ACCESSORIAL_OPERATIONS = {
    "percentage_of_variable",
    "fixed_amount",
    "ceil_fraction",
}


def _accessorial_is_configured_calculation_base(fee: dict) -> bool:
    return (
        fee.get("classification_source") in {
            "configured_calculation_base",
            "manual_configured_calculation_base",
        }
        and bool(fee.get("calculation_base_id"))
    )


def _configured_accessorial_ignored_reason(fee: dict) -> str | None:
    if not _accessorial_is_configured_calculation_base(fee):
        return AUDIT_REASON_NOT_CONFIGURED_CALCULATION_BASE
    if fee.get("conditions"):
        return AUDIT_REASON_CONDITIONS_PRESENT
    if fee.get("unsupported_reason"):
        return AUDIT_REASON_UNSUPPORTED_REASON_PRESENT
    if fee.get("classification_warning"):
        return AUDIT_REASON_CLASSIFICATION_WARNING_PRESENT
    if fee.get("status") != "calculable" or fee.get("classification_confidence") != "high":
        return AUDIT_REASON_LEGACY_CLASSIFIER_NOT_CALCULATED
    if fee.get("operation") not in _CONFIGURED_ACCESSORIAL_OPERATIONS:
        return AUDIT_REASON_UNSUPPORTED_CONFIGURED_OPERATION
    return None


def _accessorial_configured_ignored_component(fee: dict, reason_code: str) -> dict:
    component = _accessorial_ignored_component(
        fee,
        fee.get("canonical_component"),
        reason_code,
    )
    component.update(
        {
            "calculation_base_id": fee.get("calculation_base_id"),
            "calculation_base_label": fee.get("calculation_base_label"),
            "operation": fee.get("operation"),
            "audit_variable": fee.get("audit_variable"),
            "classification_source": fee.get("classification_source"),
        }
    )
    return {key: value for key, value in component.items() if value is not None}


def _accessorial_runtime_amount(fee: dict) -> Decimal | None:
    amount = _decimal_money(fee.get("amount"))
    if amount is not None and amount > 0:
        return amount
    parsed = _parse_single_accessorial_money(
        fee.get("value"),
        allow_bare_number=bool(fee.get("unit")),
    )
    if parsed is not None and parsed > 0:
        return parsed
    return None


def _audit_variable_decimal(audit_variables: dict[str, Decimal | None], variable: str | None) -> Decimal | None:
    if not variable:
        return None
    value = audit_variables.get(variable)
    if value is None or value < 0:
        return None
    return value


def _configured_accessorial_common_component(fee: dict, *, amount: Decimal) -> dict:
    return {
        "label": _sanitize_cell_string(fee.get("name")),
        "canonical_name": fee.get("canonical_component"),
        "canonical_component": fee.get("canonical_component"),
        "component_group": fee.get("component_group"),
        "calculation_type": fee.get("calculation_type"),
        "calculation_base_id": fee.get("calculation_base_id"),
        "calculation_base_label": fee.get("calculation_base_label"),
        "operation": fee.get("operation"),
        "audit_variable": fee.get("audit_variable"),
        "value": _sanitize_cell_string(fee.get("value")),
        "amount": _round_money(amount),
        "source_block": "accessorial_fees",
        "classification_source": "configured_calculation_base",
        "reason_code": AUDIT_REASON_CONFIGURED_ACCESSORIAL_CALCULATED,
    }


def _configured_percentage_component(
    fee: dict,
    audit_variables: dict[str, Decimal | None],
) -> tuple[dict | None, dict | None, Decimal]:
    variable_name = fee.get("audit_variable")
    variable_value = _audit_variable_decimal(audit_variables, variable_name)
    if variable_value is None:
        return None, _accessorial_configured_ignored_component(
            fee,
            AUDIT_REASON_MISSING_AUDIT_VARIABLE,
        ), Decimal("0")

    rate = _accessorial_runtime_rate(fee)
    if rate is None:
        return None, _accessorial_configured_ignored_component(
            fee,
            AUDIT_REASON_INVALID_CONFIGURED_AMOUNT,
        ), Decimal("0")

    amount = variable_value * rate
    component = _configured_accessorial_common_component(fee, amount=amount)
    component.update(
        {
            "rate": float(rate),
            "variable_value": _round_money(variable_value),
            "details": (
                f"{variable_name}: {_format_brazilian_decimal(variable_value)} x "
                f"{_format_brazilian_decimal(rate * Decimal('100'))}% = "
                f"{_format_brazilian_decimal(amount)}"
            ),
        }
    )
    if variable_name == "valor_nf":
        component["invoice_value"] = _round_money(variable_value)
    return component, None, amount


def _configured_fixed_amount_component(fee: dict) -> tuple[dict | None, dict | None, Decimal]:
    amount = _accessorial_runtime_amount(fee)
    if amount is None:
        return None, _accessorial_configured_ignored_component(
            fee,
            AUDIT_REASON_INVALID_CONFIGURED_AMOUNT,
        ), Decimal("0")
    component = _configured_accessorial_common_component(fee, amount=amount)
    component["details"] = f"valor fixo = {_format_brazilian_decimal(amount)}"
    return component, None, amount


def _configured_ceil_fraction_component(
    fee: dict,
    audit_variables: dict[str, Decimal | None],
) -> tuple[dict | None, dict | None, Decimal]:
    variable_name = fee.get("audit_variable")
    variable_value = _audit_variable_decimal(audit_variables, variable_name)
    if variable_value is None:
        return None, _accessorial_configured_ignored_component(
            fee,
            AUDIT_REASON_MISSING_AUDIT_VARIABLE,
        ), Decimal("0")

    base_amount = _accessorial_runtime_amount(fee)
    parameters = fee.get("operation_parameters") if isinstance(fee.get("operation_parameters"), dict) else {}
    fraction_size = _parse_decimal_number(parameters.get("fraction_size"))
    if base_amount is None or fraction_size is None or fraction_size <= 0:
        return None, _accessorial_configured_ignored_component(
            fee,
            AUDIT_REASON_INVALID_CONFIGURED_AMOUNT,
        ), Decimal("0")

    fractions = int((variable_value / fraction_size).to_integral_value(rounding=ROUND_CEILING))
    amount = base_amount * Decimal(fractions)
    component = _configured_accessorial_common_component(fee, amount=amount)
    component.update(
        {
            "fraction_size": _round_money(fraction_size),
            "weight": _round_money(variable_value) if variable_name == "peso" else _round_money(variable_value),
            "base_amount": _round_money(base_amount),
            "fractions": fractions,
            "details": (
                f"{variable_name}: ceil({_format_brazilian_decimal(variable_value)} / "
                f"{_format_brazilian_decimal(fraction_size)}) x "
                f"{_format_brazilian_decimal(base_amount)} = {_format_brazilian_decimal(amount)}"
            ),
        }
    )
    return component, None, amount


def _build_configured_accessorial_fee_component(
    fee: dict,
    audit_variables: dict[str, Decimal | None],
) -> tuple[dict | None, dict | None, Decimal]:
    reason = _configured_accessorial_ignored_reason(fee)
    if reason is not None:
        return None, _accessorial_configured_ignored_component(fee, reason), Decimal("0")

    operation = fee.get("operation")
    if operation == "percentage_of_variable":
        return _configured_percentage_component(fee, audit_variables)
    if operation == "fixed_amount":
        return _configured_fixed_amount_component(fee)
    if operation == "ceil_fraction":
        return _configured_ceil_fraction_component(fee, audit_variables)
    return None, _accessorial_configured_ignored_component(
        fee,
        AUDIT_REASON_UNSUPPORTED_CONFIGURED_OPERATION,
    ), Decimal("0")


def _accessorial_find_linked_minimum(fee: dict, minimum_fees: list[dict]) -> Decimal | None:
    refs = _accessorial_fee_component_ref(fee)
    if not refs:
        return None
    matches: list[Decimal] = []
    for minimum_fee in minimum_fees:
        minimum_refs = _accessorial_fee_component_ref(minimum_fee)
        if not refs & minimum_refs:
            continue
        amount = _accessorial_runtime_minimum_amount(minimum_fee)
        if amount is not None:
            matches.append(amount)
    if len(matches) != 1:
        return None
    return matches[0]


def _accessorial_resolve_linked_minimum(
    fee: dict,
    calculated_amount: Decimal,
    minimum_fees: list[dict],
    consumed_minimum_ids: set[int],
) -> tuple[Decimal, Decimal | None, bool]:
    minimum_amount = _accessorial_find_linked_minimum(fee, minimum_fees)
    minimum_applied = minimum_amount is not None and minimum_amount > calculated_amount
    final_amount = minimum_amount if minimum_applied else calculated_amount
    if minimum_amount is not None:
        for minimum_fee in minimum_fees:
            if _accessorial_runtime_minimum_amount(minimum_fee) == minimum_amount and (
                _accessorial_fee_component_ref(fee) & _accessorial_fee_component_ref(minimum_fee)
            ):
                consumed_minimum_ids.add(id(minimum_fee))
                break
    return final_amount, minimum_amount, minimum_applied


def _accessorial_append_minimum_details(
    details: str,
    minimum_amount: Decimal,
    minimum_applied: bool,
) -> str:
    suffix = "mínimo aplicado" if minimum_applied else "mínimo não aplicado"
    return f"{details}; {suffix} = {_format_brazilian_decimal(minimum_amount)}"


def _accessorial_component_details(
    fee: dict,
    *,
    invoice_value: Decimal,
    rate: Decimal,
    calculated_amount: Decimal,
    minimum_amount: Decimal | None,
    amount: Decimal,
    minimum_applied: bool,
) -> str:
    label = _sanitize_cell_string(fee.get("name")) or "Generalidade"
    details = (
        f"{label}: {_format_brazilian_decimal(invoice_value)} x "
        f"{_format_brazilian_decimal(rate * Decimal('100'))}% = "
        f"{_format_brazilian_decimal(calculated_amount)}"
    )
    if minimum_amount is not None:
        suffix = "mínimo aplicado" if minimum_applied else "mínimo não aplicado"
        details = f"{details}; {suffix} = {_format_brazilian_decimal(minimum_amount)}"
    return details


def _build_accessorial_percent_fee_components(
    accessorial_fees,
    *,
    invoice_value: Decimal | None,
    audit_variables: dict[str, Decimal | None] | None = None,
    has_tariff_freight_value: bool,
    has_route_toll: bool = False,
) -> tuple[list[dict], list[dict], Decimal]:
    calculated: list[dict] = []
    ignored: list[dict] = []
    total = Decimal("0")
    audit_variables = audit_variables or {}

    if not isinstance(accessorial_fees, list):
        return calculated, ignored, total

    normalized_fees = _normalize_accessorial_fees(accessorial_fees)
    minimum_fee_items = [
        fee
        for fee in normalized_fees
        if fee.get("calculation_type") == "minimum_amount"
        and fee.get("modifier_type") == "minimum_amount"
    ]
    minimum_fees = [
        fee
        for fee in minimum_fee_items
        if fee.get("status") == "calculable"
        and _accessorial_runtime_minimum_amount(fee) is not None
    ]
    consumed_minimum_ids: set[int] = set()

    for fee in normalized_fees:
        canonical_component = fee.get("canonical_component")
        if fee.get("calculation_type") == "minimum_amount":
            continue
        if has_route_toll and canonical_component == "toll":
            ignored.append(
                _accessorial_ignored_component(
                    fee,
                    canonical_component,
                    AUDIT_REASON_ROUTE_TOLL_DUPLICATE_IGNORED,
                )
            )
            continue
        if _accessorial_is_configured_calculation_base(fee):
            component, ignored_component, amount = _build_configured_accessorial_fee_component(
                fee,
                audit_variables,
            )
            if component is not None:
                calculated_amount = amount
                final_amount, minimum_amount, minimum_applied = _accessorial_resolve_linked_minimum(
                    fee,
                    calculated_amount,
                    minimum_fees,
                    consumed_minimum_ids,
                )
                component["calculated_amount"] = _round_money(calculated_amount)
                component["minimum_amount"] = (
                    _round_money(minimum_amount) if minimum_amount is not None else None
                )
                component["minimum_applied"] = minimum_applied
                component["amount"] = _round_money(final_amount)
                if minimum_amount is not None and component.get("details"):
                    component["details"] = _accessorial_append_minimum_details(
                        component["details"],
                        minimum_amount,
                        minimum_applied,
                    )
                calculated.append(component)
                total += final_amount
            elif ignored_component is not None:
                ignored.append(ignored_component)
            continue
        if not _accessorial_is_calculable_invoice_percentage(fee):
            ignored.append(
                _accessorial_ignored_component(
                    fee,
                    canonical_component,
                    _accessorial_runtime_ignored_reason(fee),
                )
            )
            continue
        if _accessorial_is_duplicate_of_tariff_freight_value(fee, has_tariff_freight_value):
            ignored.append(
                _accessorial_ignored_component(
                    fee,
                    canonical_component,
                    AUDIT_REASON_DUPLICATE_INVOICE_PERCENTAGE_FEE_IGNORED,
                )
            )
            continue

        rate = _accessorial_runtime_rate(fee)
        if rate is None:
            continue
        if invoice_value is None:
            ignored.append(
                _accessorial_ignored_component(
                    fee,
                    canonical_component,
                    AUDIT_STATUS_INVALID_INVOICE_VALUE,
                )
            )
            continue

        calculated_amount = invoice_value * rate
        final_amount, minimum_amount, minimum_applied = _accessorial_resolve_linked_minimum(
            fee,
            calculated_amount,
            minimum_fees,
            consumed_minimum_ids,
        )
        total += final_amount
        calculated.append(
            {
                "label": _sanitize_cell_string(fee.get("name")),
                "canonical_name": canonical_component,
                "canonical_component": canonical_component,
                "component_group": fee.get("component_group"),
                "calculation_type": "invoice_percentage",
                "calculated_amount": _round_money(calculated_amount),
                "minimum_amount": _round_money(minimum_amount) if minimum_amount is not None else None,
                "minimum_applied": minimum_applied,
                "amount": _round_money(final_amount),
                "rate": float(rate),
                "source_value": _sanitize_cell_string(fee.get("value")),
                "invoice_value": _round_money(invoice_value),
                "details": _accessorial_component_details(
                    fee,
                    invoice_value=invoice_value,
                    rate=rate,
                    calculated_amount=calculated_amount,
                    minimum_amount=minimum_amount,
                    amount=final_amount,
                    minimum_applied=minimum_applied,
                ),
                "source_block": "accessorial_fees",
                "reason_code": AUDIT_REASON_ACCESSORIAL_PERCENTAGE_CALCULATED,
            }
        )

    for minimum_fee in minimum_fee_items:
        if id(minimum_fee) not in consumed_minimum_ids:
            ignored.append(
                _accessorial_ignored_component(
                    minimum_fee,
                    minimum_fee.get("canonical_component"),
                    AUDIT_REASON_ACCESSORIAL_MINIMUM_WITHOUT_BASE_IGNORED,
                )
            )

    return calculated, ignored, total


def _decimal_money(value) -> Decimal | None:
    if isinstance(value, Decimal):
        return value if value >= 0 else None
    parsed = _parse_brazilian_money(value)
    if parsed is None:
        return None
    try:
        return Decimal(str(parsed))
    except (InvalidOperation, ValueError):
        return None


def _round_money(value: Decimal) -> float:
    return float(value.quantize(Decimal("0.01"), rounding=ROUND_HALF_UP))


def _format_brazilian_decimal(value: Decimal) -> str:
    text = f"{value.normalize():f}"
    if "." in text:
        text = text.rstrip("0").rstrip(".")
    return text.replace(".", ",") if text else "0"


def _region_from_table_context(table: dict) -> str | None:
    context = table.get("context") if isinstance(table.get("context"), dict) else {}
    for key in ("route_label", "destination", "region", "freight_region", "praca", "rota"):
        candidate = _sanitize_cell_string(context.get(key))
        if candidate:
            return candidate
    return _sanitize_cell_string(table.get("table_title"))


def _make_unsupported_rule(region: str, source_title: str | None, note: str) -> dict:
    return {
        "pricing_type": AUDIT_STATUS_UNSUPPORTED_PRICING,
        "region": region,
        "source_table_title": source_title,
        "brackets": [],
        "excess": None,
        "unit": "kg",
        "normalization_notes": [note],
    }


def _register_pricing_rule(index: dict, region: str | None, rule: dict) -> None:
    if not region:
        return
    if region in index:
        index[region] = _make_unsupported_rule(
            region,
            rule.get("source_table_title"),
            "Mais de uma regra de frete para a mesma região.",
        )
        return
    index[region] = rule


def _pricing_rule_keys_for_row(
    region: str,
    destination_uf: str | None = None,
    *,
    include_normalized_region: bool = False,
) -> list[str]:
    keys = [region]
    normalized_region = _normalize_audit_lookup_text(region)
    if include_normalized_region and normalized_region:
        keys.append(normalized_region)
    uf = _normalize_destination_uf(destination_uf)
    if uf and normalized_region:
        keys.append(f"{uf}|{normalized_region}")
    return list(dict.fromkeys(keys))


def _region_uf_from_composite_route_destination(destination: str) -> tuple[str, str] | None:
    cleaned = _sanitize_cell_string(destination)
    if not cleaned:
        return None
    for separator in (" - ", " – ", " — "):
        parts = [part.strip() for part in cleaned.split(separator) if part.strip()]
        if len(parts) != 2:
            continue
        left, right = parts
        left_uf = _normalize_destination_uf(left)
        right_uf = _normalize_destination_uf(right)
        if left_uf and left_uf in _BR_UFS and not right_uf:
            return right, left_uf
        if right_uf and right_uf in _BR_UFS and not left_uf:
            return left, right_uf
    return None


_FREIGHT_ROUTE_REGION_NOTE_HEADERS = frozenset(
    {
        "regiao",
        "regiao de frete",
        "praca",
        "rota",
        "itinerario",
        "zona",
        "area",
    }
)


def _freight_route_has_weight_brackets(route: dict) -> bool:
    for limit in (10, 20, 30, 50, 70, 100):
        if _parse_brazilian_money(route.get(f"weight_{limit}") or route.get(f"weight_{limit}kg")) is not None:
            return True
    return False


def _region_label_from_freight_route_classification_notes(tail: str) -> str | None:
    text = _sanitize_cell_string(tail)
    if not text:
        return None
    for quote in ("'", '"', "\u2018", "\u2019", "\u201c", "\u201d"):
        if not text.startswith(quote):
            continue
        end = text.find(quote, 1)
        if end > 1:
            candidate = _sanitize_cell_string(text[1:end])
            if candidate:
                return candidate
    stripped = text.rstrip(".,; ")
    return _sanitize_cell_string(stripped) if stripped else None


def _region_label_from_freight_route_notes(notes) -> str | None:
    cleaned = _sanitize_cell_string(notes)
    if not cleaned:
        return None
    for separator in (":", " - ", " – ", " — "):
        if separator not in cleaned:
            continue
        parts = [part.strip() for part in cleaned.split(separator) if part.strip()]
        if len(parts) < 2:
            continue
        header = _normalize_coverage_header(parts[0])
        if header not in _FREIGHT_ROUTE_REGION_NOTE_HEADERS and not header.startswith("regiao"):
            continue
        label = parts[1] if len(parts) == 2 else " ".join(parts[1:])
        candidate = _sanitize_cell_string(label)
        if candidate:
            return candidate
    lowered = cleaned.lower()
    for marker in ("classificada como", "classificado como"):
        if marker not in lowered:
            continue
        idx = lowered.index(marker)
        candidate = _region_label_from_freight_route_classification_notes(cleaned[idx + len(marker) :])
        if candidate:
            return candidate
    return None


def _resolve_freight_route_region_uf_table(route: dict) -> tuple[str, str] | None:
    if not isinstance(route, dict):
        return None
    origin = _sanitize_cell_string(route.get("origin"))
    destination = _sanitize_cell_string(route.get("destination"))
    if not destination:
        return None
    if _region_uf_from_composite_route_destination(destination):
        return None
    destination_uf = _normalize_destination_uf(destination)
    if not destination_uf or destination_uf not in _BR_UFS:
        return None
    if not _freight_route_has_weight_brackets(route):
        return None
    notes_region = _region_label_from_freight_route_notes(route.get("notes"))
    if not origin:
        if not notes_region:
            return None
        return notes_region, destination_uf
    origin_uf = _normalize_destination_uf(origin)
    if origin_uf and origin_uf in _BR_UFS:
        return None
    region_label = notes_region or origin
    if not region_label:
        return None
    if notes_region and origin:
        if _normalize_audit_lookup_text(origin) != _normalize_audit_lookup_text(notes_region):
            return None
    return region_label, destination_uf


def _freight_route_region_label(route: dict) -> str | None:
    region_uf = _resolve_freight_route_region_uf_table(route)
    if region_uf:
        return region_uf[0]
    return (
        _sanitize_cell_string(
            route.get("destination")
            or route.get("freight_region")
            or route.get("region")
            or route.get("route")
        )
        or None
    )


def _pricing_rule_keys_for_freight_route(route: dict, region: str) -> list[str]:
    keys = [region]
    parsed = _region_uf_from_composite_route_destination(region)
    if parsed:
        region_label, uf = parsed
        normalized_region = _normalize_audit_lookup_text(region_label)
        if uf and normalized_region:
            keys.append(f"{uf}|{normalized_region}")
    region_uf = _resolve_freight_route_region_uf_table(route)
    if region_uf:
        region_label, uf = region_uf
        keys.extend(_pricing_rule_keys_for_row(region_label, uf))
    return list(dict.fromkeys(keys))


def _build_rule_from_row_range_table(table: dict) -> dict | None:
    rows = table.get("rows") if isinstance(table.get("rows"), list) else []
    columns = table.get("columns") if isinstance(table.get("columns"), list) else []
    if not rows or not columns:
        return None
    range_col = next((col for col in columns if _parse_range_from_label(col) is None and "peso" in _normalize_coverage_header(col)), None)
    value_col = next((col for col in columns if _is_value_column(col)), None)
    if not range_col or not value_col:
        return None
    region = _region_from_table_context(table)
    if not region:
        return None
    brackets = []
    freight_value = None
    for row in rows:
        if not isinstance(row, dict):
            continue
        parsed_range = _parse_range_from_label(row.get(range_col))
        value = _parse_brazilian_money(row.get(value_col))
        if parsed_range and value is not None:
            row_freight_value = _extract_freight_value_from_row(columns, row)
            if freight_value is None and row_freight_value is not None:
                freight_value = row_freight_value
            brackets.append(
                {
                    "min_kg": parsed_range[0] or 0.0,
                    "max_kg": parsed_range[1],
                    "value": value,
                    "label": _sanitize_cell_string(row.get(range_col)),
                }
            )
    brackets = _normalize_brackets(brackets)
    if not brackets:
        return _make_unsupported_rule(region, table.get("table_title"), "Faixas por linha sem valor calculável.")
    excess_col = next((col for col in columns if _is_excess_column(col)), None)
    excess_rate = None
    if excess_col:
        for row in rows:
            excess_rate = _parse_brazilian_money(row.get(excess_col))
            if excess_rate is not None:
                break
    route_toll = None
    for row in rows:
        if not isinstance(row, dict):
            continue
        route_toll = _extract_route_toll_from_row(columns, row)
        if route_toll is not None:
            break
    rule = {
        "pricing_type": "range_plus_excess_per_kg" if excess_rate is not None else "fixed_range",
        "region": region,
        "source_table_title": table.get("table_title"),
        "brackets": brackets,
        "excess": {"rate_per_kg": excess_rate} if excess_rate is not None else None,
        "freight_value": freight_value,
        "route_toll": route_toll,
        "unit": "kg",
        "normalization_notes": [],
    }
    return rule


def _build_rules_from_matrix_table(table: dict) -> list[tuple[str, dict]]:
    rows = table.get("rows") if isinstance(table.get("rows"), list) else []
    columns = table.get("columns") if isinstance(table.get("columns"), list) else []
    if not rows or not columns:
        return []

    selected_dimension_col = _sanitize_cell_string(table.get("_selected_pricing_dimension_column"))
    region_col = selected_dimension_col if selected_dimension_col in columns else None
    if region_col is None:
        region_col = next((col for col in columns if _is_region_column(col)), None)
    uf_col = next((col for col in columns if _is_destination_uf_column(col)), None)
    range_cols = [
        (col, parsed)
        for col in columns
        if not _is_excess_column(col)
        for parsed in [_parse_range_from_label(col)]
        if parsed is not None
    ]
    direct_kg_col = next((col for col in columns if _is_direct_kg_column(col)), None)
    direct_ton_col = next((col for col in columns if _is_direct_ton_column(col)), None)
    excess_col = next((col for col in columns if _is_excess_column(col)), None)
    context_region = _region_from_table_context(table)
    region_is_city_destination = bool(region_col and _is_city_destination_column(region_col))
    rules: list[tuple[str, dict]] = []

    for row in rows:
        if not isinstance(row, dict):
            continue
        region = _sanitize_cell_string(row.get(region_col)) if region_col else context_region
        destination_uf = _sanitize_cell_string(row.get(uf_col)) if uf_col else None
        if not region:
            continue
        if direct_ton_col:
            value = _parse_brazilian_money(row.get(direct_ton_col))
            if value is not None:
                freight_value = _extract_freight_value_from_row(columns, row)
                route_toll = _extract_route_toll_from_row(columns, row)
                rule = {
                    "pricing_type": "direct_weight_rate",
                    "region": region,
                    "source_table_title": table.get("table_title"),
                    "brackets": [],
                    "excess": None,
                    "unit": "ton",
                    "value_per_ton": value,
                    "freight_value": freight_value,
                    "route_toll": route_toll,
                    "normalization_notes": [],
                }
                rules.extend(
                    (key, rule)
                    for key in _pricing_rule_keys_for_row(
                        region,
                        destination_uf,
                        include_normalized_region=region_is_city_destination,
                    )
                )
                continue
        if direct_kg_col:
            value = _parse_brazilian_money(row.get(direct_kg_col))
            if value is not None:
                freight_value = _extract_freight_value_from_row(columns, row)
                route_toll = _extract_route_toll_from_row(columns, row)
                rule = {
                    "pricing_type": "direct_weight_rate",
                    "region": region,
                    "source_table_title": table.get("table_title"),
                    "brackets": [],
                    "excess": None,
                    "unit": "kg",
                    "value_per_kg": value,
                    "freight_value": freight_value,
                    "route_toll": route_toll,
                    "normalization_notes": [],
                }
                rules.extend(
                    (key, rule)
                    for key in _pricing_rule_keys_for_row(
                        region,
                        destination_uf,
                        include_normalized_region=region_is_city_destination,
                    )
                )
                continue
        brackets = []
        for col, parsed_range in range_cols:
            value = _parse_brazilian_money(row.get(col))
            if value is None:
                continue
            brackets.append(
                {
                    "min_kg": parsed_range[0] or 0.0,
                    "max_kg": parsed_range[1],
                    "value": value,
                    "label": col,
                }
            )
        brackets = _normalize_brackets(brackets)
        if brackets:
            excess_rate = _parse_brazilian_money(row.get(excess_col)) if excess_col else None
            freight_value = _extract_freight_value_from_row(columns, row)
            route_toll = _extract_route_toll_from_row(columns, row)
            rule = {
                "pricing_type": "range_plus_excess_per_kg" if excess_rate is not None else "fixed_range",
                "region": region,
                "source_table_title": table.get("table_title"),
                "brackets": brackets,
                "excess": {"rate_per_kg": excess_rate} if excess_rate is not None else None,
                "freight_value": freight_value,
                "route_toll": route_toll,
                "unit": "kg",
                "normalization_notes": [],
            }
            rules.extend(
                (key, rule)
                for key in _pricing_rule_keys_for_row(
                    region,
                    destination_uf,
                    include_normalized_region=region_is_city_destination,
                )
            )
        elif region_col:
            rule = _make_unsupported_rule(
                region,
                table.get("table_title"),
                "Linha de destino sem modelo de peso/faixa reconhecido.",
            )
            rules.extend(
                (key, rule)
                for key in _pricing_rule_keys_for_row(
                    region,
                    destination_uf,
                    include_normalized_region=region_is_city_destination,
                )
            )
    return rules


def _build_rule_from_freight_route(route: dict) -> tuple[str, dict] | None:
    region = _freight_route_region_label(route)
    if not region:
        return None
    brackets = []
    for limit in (10, 20, 30, 50, 70, 100):
        value = _parse_brazilian_money(
            route.get(f"weight_{limit}") or route.get(f"weight_{limit}kg")
        )
        if value is not None:
            brackets.append(
                {
                    "min_kg": 0.0,
                    "max_kg": float(limit),
                    "value": value,
                    "label": f"Até {limit} kg",
                }
            )
    brackets = _normalize_brackets(brackets)
    excess_rate = _parse_brazilian_money(route.get("freight_weight_kg") or route.get("frete_peso_kg"))
    route_toll_rate = _parse_brazilian_money(route.get("pedagio") or route.get("pedagio_valor"))
    route_toll = None
    if route_toll_rate is not None:
        route_toll = {
            "rate_per_fraction": float(route_toll_rate),
            "fraction_size_kg": 100.0,
            "source_column": "pedagio",
            "source_value": _sanitize_cell_string(route.get("pedagio") or route.get("pedagio_valor")),
        }
    if brackets:
        return (
            region,
            {
                "pricing_type": "range_plus_excess_per_kg" if excess_rate is not None else "fixed_range",
                "region": region,
                "source_table_title": "freight_routes",
                "brackets": brackets,
                "excess": {"rate_per_kg": excess_rate} if excess_rate is not None else None,
                "route_toll": route_toll,
                "unit": "kg",
                "normalization_notes": [],
            },
        )
    if excess_rate is not None:
        return (
            region,
            {
                "pricing_type": "direct_weight_rate",
                "region": region,
                "source_table_title": "freight_routes",
                "brackets": [],
                "excess": None,
                "unit": "kg",
                "value_per_kg": excess_rate,
                "normalization_notes": [],
            },
        )
    return (region, _make_unsupported_rule(region, "freight_routes", "Rota sem faixa de peso calculável."))


def build_freight_pricing_index(temp_table) -> dict:
    if not isinstance(temp_table, dict):
        return {}
    index: dict[str, dict] = {}

    for table in temp_table.get("freight_tables") or []:
        if not isinstance(table, dict):
            continue
        row_range_rule = _build_rule_from_row_range_table(table)
        if row_range_rule is not None:
            _register_pricing_rule(index, row_range_rule.get("region"), row_range_rule)
            continue
        matrix_rules = _build_rules_from_matrix_table(table)
        if matrix_rules:
            for region, rule in matrix_rules:
                _register_pricing_rule(index, region, rule)
            continue
        region = _region_from_table_context(table)
        if region:
            _register_pricing_rule(
                index,
                region,
                _make_unsupported_rule(region, table.get("table_title"), "Tabela sem modelo de peso/faixa reconhecido."),
            )

    for route in temp_table.get("freight_routes") or []:
        if not isinstance(route, dict):
            continue
        route_rule = _build_rule_from_freight_route(route)
        if route_rule is not None:
            region, rule = route_rule
            for key in _pricing_rule_keys_for_freight_route(route, region):
                _register_pricing_rule(index, key, rule)

    return index


def calculate_weight_freight(weight_kg, pricing_rule) -> dict | None:
    weight = _parse_weight_number(weight_kg)
    if weight is None or not isinstance(pricing_rule, dict):
        return None
    pricing_type = pricing_rule.get("pricing_type")
    if pricing_type == "fixed_range":
        for bracket in pricing_rule.get("brackets") or []:
            min_kg = float(bracket.get("min_kg") or 0)
            max_kg = bracket.get("max_kg")
            value = bracket.get("value")
            if max_kg is None or value is None:
                continue
            max_kg = float(max_kg)
            if (min_kg <= 0 and 0 <= weight <= max_kg) or (min_kg < weight <= max_kg):
                return {
                    "expected_freight": round(float(value), 2),
                    "calculation_basis": "fixed_range",
                    "calculation_details": bracket.get("label") or f"Faixa até {max_kg:g} kg",
                }
        return None
    if pricing_type == "range_plus_excess_per_kg":
        brackets = pricing_rule.get("brackets") or []
        for bracket in brackets:
            min_kg = float(bracket.get("min_kg") or 0)
            max_kg = bracket.get("max_kg")
            value = bracket.get("value")
            if max_kg is None or value is None:
                continue
            max_kg = float(max_kg)
            if (min_kg <= 0 and 0 <= weight <= max_kg) or (min_kg < weight <= max_kg):
                return {
                    "expected_freight": round(float(value), 2),
                    "calculation_basis": "range_plus_excess_per_kg",
                    "calculation_details": bracket.get("label") or f"Faixa até {max_kg:g} kg",
                }
        if not brackets:
            return None
        last = max(brackets, key=lambda item: float(item.get("max_kg") or 0))
        last_max = float(last.get("max_kg") or 0)
        last_value = last.get("value")
        excess = pricing_rule.get("excess") if isinstance(pricing_rule.get("excess"), dict) else {}
        excess_rate = excess.get("rate_per_kg")
        if last_value is None or excess_rate is None or weight <= last_max:
            return None
        expected = float(last_value) + float(excess_rate) * (weight - last_max)
        return {
            "expected_freight": round(expected, 2),
            "calculation_basis": "range_plus_excess_per_kg",
            "calculation_details": f"Faixa até {last_max:g} kg + excedente por kg",
        }
    if pricing_type == "direct_weight_rate":
        unit = str(pricing_rule.get("unit") or "kg").strip().lower()
        if unit in {"ton", "tonelada", "toneladas", "t"}:
            value = pricing_rule.get("value_per_ton")
            if value is None:
                return None
            expected = (weight / 1000.0) * float(value)
            return {
                "expected_freight": round(expected, 2),
                "calculation_basis": "direct_weight_rate",
                "calculation_details": "Peso em toneladas x valor por tonelada",
            }
        value = pricing_rule.get("value_per_kg")
        if value is None:
            return None
        expected = weight * float(value)
        return {
            "expected_freight": round(expected, 2),
            "calculation_basis": "direct_weight_rate",
            "calculation_details": "Peso em kg x valor por kg",
        }
    return None


def compare_charged_vs_expected(charged_freight, expected_freight) -> dict:
    charged = round(float(charged_freight), 2)
    expected = round(float(expected_freight), 2)
    divergence = round(charged - expected, 2)
    return {
        "charged_freight": charged,
        "expected_freight": expected,
        "divergence_value": divergence,
        "status": AUDIT_STATUS_OK if divergence == 0 else AUDIT_STATUS_DIVERGENT,
    }


def _format_brazilian_money_display(value) -> str:
    money = _decimal_money(value)
    if money is None:
        return "R$ 0,00"
    quantized = money.quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
    return f"R$ {f'{quantized:.2f}'.replace('.', ',')}"


def _format_brazilian_percent_display(rate: float) -> str:
    quantized = Decimal(str(rate)).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
    return f"{f'{quantized:.2f}'.replace('.', ',')}%"


def _tax_config_fingerprint(tax_config) -> str | None:
    if not isinstance(tax_config, dict):
        return None
    try:
        return json.dumps(tax_config, sort_keys=True, ensure_ascii=False, default=str)
    except (TypeError, ValueError):
        return None


def _serialize_pricing_rule_for_fingerprint(rule: dict) -> dict:
    brackets = []
    for bracket in rule.get("brackets") or []:
        if not isinstance(bracket, dict):
            continue
        brackets.append(
            {
                "min_kg": bracket.get("min_kg"),
                "max_kg": bracket.get("max_kg"),
                "value": bracket.get("value"),
            }
        )
    return {
        "pricing_type": rule.get("pricing_type"),
        "brackets": brackets,
        "excess": rule.get("excess"),
        "value_per_kg": rule.get("value_per_kg"),
        "value_per_ton": rule.get("value_per_ton"),
        "unit": rule.get("unit"),
        "freight_value": rule.get("freight_value"),
        "route_toll": rule.get("route_toll"),
    }


def _pricing_rule_fingerprint(temp_table) -> str | None:
    if not isinstance(temp_table, dict):
        return None
    index = build_freight_pricing_index(temp_table)
    try:
        payload = {
            "parser_version": PRICING_RULE_PARSER_VERSION,
            "rules": {
                key: _serialize_pricing_rule_for_fingerprint(rule)
                for key, rule in sorted(index.items())
                if isinstance(rule, dict)
            },
        }
        return json.dumps(payload, sort_keys=True, ensure_ascii=False, default=str)
    except (TypeError, ValueError):
        return None


def _cities_match_reliably(origin_city, destination_city) -> bool:
    origin = _normalize_audit_lookup_text(origin_city)
    destination = _normalize_audit_lookup_text(destination_city)
    return bool(origin and destination and origin == destination)


def _resolve_applicable_tax_type(row: dict, tax_config: dict) -> str | None:
    origin_uf = _normalize_uf(tax_config.get("origin_uf"))
    destination_uf = _normalize_destination_uf(row.get("destination_uf"))
    if not origin_uf or not destination_uf:
        return None
    if origin_uf != destination_uf:
        return "icms"
    if _cities_match_reliably(tax_config.get("origin_city"), row.get("destination_city")):
        iss_rate = tax_config.get("iss_rate")
        if iss_rate is not None:
            try:
                if float(iss_rate) > 0:
                    return "iss"
            except (TypeError, ValueError):
                return None
        return None
    return "icms"


def _find_active_icms_rate(tax_config: dict, destination_uf) -> dict | None:
    destination = _normalize_destination_uf(destination_uf)
    if not destination:
        return None
    for item in tax_config.get("icms_rates") or []:
        if not isinstance(item, dict):
            continue
        if _normalize_uf(item.get("destination_uf")) != destination:
            continue
        if not item.get("is_active"):
            continue
        rate = item.get("applied_rate")
        if rate is None:
            continue
        try:
            rate_float = float(rate)
        except (TypeError, ValueError):
            continue
        if rate_float <= 0:
            continue
        return item
    return None


def _build_pre_tax_subtotal_info(calculated: dict) -> dict:
    components = calculated.get("calculation_components") if isinstance(calculated.get("calculation_components"), dict) else {}
    ignored = components.get("ignored_accessorial_fees") if isinstance(components.get("ignored_accessorial_fees"), list) else []
    subtotal = _decimal_money(calculated.get("expected_freight"))
    return {
        "subtotal_before_taxes": subtotal,
        "is_partial_base": bool(ignored),
        "ignored_components_count": len(ignored),
    }


def _append_tax_memory_to_details(existing_details: str | None, memory_lines: list[str]) -> str:
    if not memory_lines:
        return existing_details or ""
    tax_details = " | ".join(memory_lines)
    if existing_details:
        return f"{existing_details} | {tax_details}"
    return tax_details


def _calculate_inside_tax_amounts(subtotal: Decimal, rate: float) -> tuple[Decimal, Decimal]:
    rate_decimal = Decimal(str(rate)) / Decimal("100")
    if rate_decimal <= 0 or rate_decimal >= 1:
        return Decimal("0"), subtotal
    total_with_tax = (subtotal / (Decimal("1") - rate_decimal)).quantize(
        Decimal("0.01"), rounding=ROUND_HALF_UP
    )
    tax_amount = (total_with_tax - subtotal).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
    return tax_amount, total_with_tax


def _build_icms_tax_component(
    *,
    icms_rate: dict,
    row: dict,
    tax_config: dict,
    subtotal: Decimal,
    tax_amount: Decimal,
) -> dict:
    rate = float(icms_rate["applied_rate"])
    return {
        "tax_type": "ICMS",
        "base_amount": _round_money(subtotal),
        "rate": rate,
        "amount": _round_money(tax_amount),
        "calculation_mode": TAX_CALCULATION_MODE_INSIDE,
        "source_name": _sanitize_cell_string(icms_rate.get("source_name")),
        "source_type": _sanitize_cell_string(icms_rate.get("source_type")),
        "user_edited": bool(icms_rate.get("user_edited")),
        "suggested_rate": icms_rate.get("suggested_rate"),
        "applied": True,
        "ignored_reason": None,
        "origin_uf": _normalize_uf(tax_config.get("origin_uf")),
        "destination_uf": _normalize_destination_uf(row.get("destination_uf")),
    }


def _build_iss_tax_component(
    *,
    row: dict,
    tax_config: dict,
    subtotal: Decimal,
    tax_amount: Decimal,
) -> dict:
    rate = float(tax_config["iss_rate"])
    return {
        "tax_type": "ISS",
        "base_amount": _round_money(subtotal),
        "rate": rate,
        "amount": _round_money(tax_amount),
        "calculation_mode": TAX_CALCULATION_MODE_INSIDE,
        "source_name": ISS_SOURCE_NAME,
        "source_type": ISS_SOURCE_TYPE,
        "user_edited": False,
        "suggested_rate": None,
        "applied": True,
        "ignored_reason": None,
        "origin_uf": _normalize_uf(tax_config.get("origin_uf")),
        "destination_uf": _normalize_destination_uf(row.get("destination_uf")),
    }


def _build_unapplied_tax_component(
    *,
    tax_type: str,
    row: dict,
    tax_config: dict,
    ignored_reason: str,
) -> dict:
    return {
        "tax_type": tax_type,
        "base_amount": None,
        "rate": None,
        "amount": None,
        "calculation_mode": TAX_CALCULATION_MODE_INSIDE,
        "source_name": None,
        "source_type": None,
        "user_edited": False,
        "suggested_rate": None,
        "applied": False,
        "ignored_reason": ignored_reason,
        "origin_uf": _normalize_uf(tax_config.get("origin_uf")),
        "destination_uf": _normalize_destination_uf(row.get("destination_uf")),
    }


def _apply_row_tax_components(
    calculated: dict,
    row: dict,
    tax_config: dict | None,
) -> dict:
    if not isinstance(tax_config, dict) or tax_config.get("include_taxes") is not True:
        return calculated

    subtotal_info = _build_pre_tax_subtotal_info(calculated)
    subtotal = subtotal_info["subtotal_before_taxes"]
    if subtotal is None or subtotal <= 0:
        return calculated

    components = dict(calculated.get("calculation_components") or {})
    components["subtotal_before_taxes"] = _round_money(subtotal)
    components["expected_freight_before_taxes"] = _round_money(subtotal)

    tax_type = _resolve_applicable_tax_type(row, tax_config)
    tax_components: list[dict] = []
    memory_lines: list[str] = []
    tax_amount = Decimal("0")
    total_with_tax = None

    memory_lines.append(f"Subtotal antes dos impostos: {_format_brazilian_money_display(subtotal)}")
    if subtotal_info["is_partial_base"]:
        memory_lines.append(
            "Base parcial: imposto calculado sobre componentes disponíveis "
            "(alguns componentes foram ignorados)."
        )

    iss_rate_configured = tax_config.get("iss_rate") is not None
    try:
        iss_rate_positive = iss_rate_configured and float(tax_config.get("iss_rate")) > 0
    except (TypeError, ValueError):
        iss_rate_positive = False

    if tax_type == "icms":
        icms_rate = _find_active_icms_rate(tax_config, row.get("destination_uf"))
        if icms_rate:
            rate = float(icms_rate["applied_rate"])
            tax_amount, total_with_tax = _calculate_inside_tax_amounts(subtotal, rate)
            tax_components.append(
                _build_icms_tax_component(
                    icms_rate=icms_rate,
                    row=row,
                    tax_config=tax_config,
                    subtotal=subtotal,
                    tax_amount=tax_amount,
                )
            )
            memory_lines.append(
                f"ICMS por dentro: {_format_brazilian_percent_display(rate)} — base "
                f"{_format_brazilian_money_display(subtotal)}, imposto "
                f"{_format_brazilian_money_display(tax_amount)}, total "
                f"{_format_brazilian_money_display(total_with_tax)}"
            )
            source_name = _sanitize_cell_string(icms_rate.get("source_name"))
            if source_name:
                memory_lines.append(f"Fonte: {source_name}")
            if icms_rate.get("user_edited"):
                memory_lines.append("Alíquota editada pelo usuário.")
        else:
            tax_components.append(
                _build_unapplied_tax_component(
                    tax_type="ICMS",
                    row=row,
                    tax_config=tax_config,
                    ignored_reason="Alíquota ICMS não disponível, inativa ou ausente para a UF destino.",
                )
            )
        if iss_rate_positive:
            memory_lines.append(
                "ISS não aplicado nesta linha: transporte não identificado como municipal."
            )
    elif tax_type == "iss":
        rate = float(tax_config["iss_rate"])
        tax_amount, total_with_tax = _calculate_inside_tax_amounts(subtotal, rate)
        tax_components.append(
            _build_iss_tax_component(
                row=row,
                tax_config=tax_config,
                subtotal=subtotal,
                tax_amount=tax_amount,
            )
        )
        memory_lines.append(
            f"ISS por dentro: {_format_brazilian_percent_display(rate)} — base "
            f"{_format_brazilian_money_display(subtotal)}, imposto "
            f"{_format_brazilian_money_display(tax_amount)}, total "
            f"{_format_brazilian_money_display(total_with_tax)}"
        )
        memory_lines.append(f"Fonte: {ISS_SOURCE_NAME}")
    elif iss_rate_positive:
        memory_lines.append(
            "ISS não aplicado nesta linha: transporte não identificado como municipal."
        )

    if tax_amount > 0 and total_with_tax is not None:
        calculated["expected_freight"] = _round_money(total_with_tax)
        components["tax_total"] = _round_money(tax_amount)
        memory_lines.append(
            f"Total esperado com impostos: {_format_brazilian_money_display(total_with_tax)}"
        )
    else:
        components["tax_total"] = 0.0

    components["tax_components"] = tax_components
    calculated["calculation_components"] = components
    calculated["calculation_details"] = _append_tax_memory_to_details(
        calculated.get("calculation_details"),
        memory_lines,
    )
    return calculated


def _build_tax_fiscal_snapshot(tax_config, results: list[dict]) -> dict:
    snapshot = {
        "tax_applied": False,
        "tax_calculation_version": TAX_CALCULATION_VERSION,
        "tax_calculation_mode": TAX_CALCULATION_MODE_INSIDE,
    }
    if not isinstance(tax_config, dict) or tax_config.get("include_taxes") is not True:
        return snapshot

    snapshot["tax_config_snapshot"] = copy.deepcopy(tax_config)
    rows_with_tax = 0
    total_tax = Decimal("0")
    icms_rows = 0
    iss_rows = 0
    for result in results:
        if not isinstance(result, dict):
            continue
        tax_items = (result.get("calculation_components") or {}).get("tax_components") or []
        applied_item = next(
            (
                item
                for item in tax_items
                if isinstance(item, dict) and item.get("applied") and item.get("amount")
            ),
            None,
        )
        if applied_item is None:
            continue
        rows_with_tax += 1
        total_tax += _decimal_money(applied_item.get("amount")) or Decimal("0")
        if applied_item.get("tax_type") == "ICMS":
            icms_rows += 1
        elif applied_item.get("tax_type") == "ISS":
            iss_rows += 1

    snapshot["tax_applied"] = rows_with_tax > 0
    snapshot["tax_summary"] = {
        "rows_with_tax": rows_with_tax,
        "total_tax_amount": _round_money(total_tax),
        "icms_rows": icms_rows,
        "iss_rows": iss_rows,
    }
    return snapshot


def _apply_tax_fiscal_snapshot_to_audit_batch(audit_batch: dict, fiscal_snapshot: dict) -> dict:
    updated = dict(audit_batch)
    for key in (
        "tax_config_snapshot",
        "tax_applied",
        "tax_calculation_version",
        "tax_calculation_mode",
        "tax_summary",
    ):
        if key in fiscal_snapshot:
            updated[key] = fiscal_snapshot[key]
    updated.pop("needs_reprocess", None)
    updated.pop("stale_reason", None)
    updated.pop("stale_at", None)
    return updated


def _audit_batch_is_fiscally_outdated(audit_batch) -> bool:
    if not isinstance(audit_batch, dict):
        return False
    if audit_batch.get("status") != AUDIT_BATCH_STATUS_PROCESSED:
        return False
    if not audit_batch.get("tax_applied"):
        return False
    version = audit_batch.get("tax_calculation_version")
    mode = audit_batch.get("tax_calculation_mode")
    return version != TAX_CALCULATION_VERSION or mode != TAX_CALCULATION_MODE_INSIDE


def _audit_batch_is_pricing_rule_parser_outdated(audit_batch) -> bool:
    if not isinstance(audit_batch, dict):
        return False
    if audit_batch.get("status") != AUDIT_BATCH_STATUS_PROCESSED:
        return False
    return audit_batch.get("pricing_rule_parser_version") != PRICING_RULE_PARSER_VERSION


def _audit_batch_effective_needs_reprocess(audit_batch) -> bool:
    if not isinstance(audit_batch, dict):
        return False
    if audit_batch.get("needs_reprocess"):
        return True
    if _audit_batch_is_fiscally_outdated(audit_batch):
        return True
    return _audit_batch_is_pricing_rule_parser_outdated(audit_batch)


def _audit_batch_should_bill_operational_run(audit_batch) -> bool:
    """
    Linhas da planilha auditada são cobradas no upload (agente_compara_batch_upload).
    O processamento só gera débito operacional em reprocessamento explícito.
    """
    if not isinstance(audit_batch, dict):
        return False
    if audit_batch.get("status") != AUDIT_BATCH_STATUS_PROCESSED:
        return False
    return _audit_batch_effective_needs_reprocess(audit_batch)


def _mark_audit_batch_stale_if_processed(record: dict, *, reason: str, alert: str) -> dict:
    audit_batch = record.get("audit_batch")
    if not isinstance(audit_batch, dict):
        return record
    if audit_batch.get("status") != AUDIT_BATCH_STATUS_PROCESSED:
        return record

    updated = dict(record)
    updated_batch = dict(audit_batch)
    updated_batch["needs_reprocess"] = True
    updated_batch["stale_reason"] = reason
    updated_batch["stale_at"] = _utcnow().isoformat()
    updated["audit_batch"] = updated_batch

    alerts = list(updated.get("reading_alerts") or [])
    if alert not in alerts:
        alerts.append(alert)
    updated["reading_alerts"] = alerts
    return updated


def _audit_bi_methodology(audit_batch) -> dict:
    if not isinstance(audit_batch, dict):
        return {
            "taxes_included": False,
            "tax_calculation_mode": None,
            "tax_calculation_version": None,
            "needs_reprocess": False,
        }
    return {
        "taxes_included": bool(audit_batch.get("tax_applied")),
        "tax_calculation_mode": audit_batch.get("tax_calculation_mode"),
        "tax_calculation_version": audit_batch.get("tax_calculation_version"),
        "needs_reprocess": _audit_batch_effective_needs_reprocess(audit_batch),
    }


def _apply_freight_value_component(
    calculated: dict,
    row: dict,
    pricing_rule: dict,
    accessorial_fees=None,
) -> dict | None:
    weight_freight = _decimal_money(calculated.get("expected_freight"))
    if weight_freight is None:
        return None

    components = {
        "weight_freight": {
            "amount": _round_money(weight_freight),
            "basis": calculated.get("calculation_basis"),
            "details": calculated.get("calculation_details"),
        }
    }
    result = {
        **calculated,
        "weight_freight": _round_money(weight_freight),
        "freight_value_amount": None,
        "route_toll_amount": None,
        "accessorial_fees_amount": None,
        "accessorial_percent_fees_amount": None,
        "calculation_components": components,
    }

    freight_value = pricing_rule.get("freight_value")
    has_tariff_freight_value = isinstance(freight_value, dict)
    route_toll = pricing_rule.get("route_toll")
    has_route_toll = isinstance(route_toll, dict)
    invoice_value = _decimal_money(row.get("invoice_value"))
    audited_weight = _parse_weight_number(row.get("audited_weight"))
    audit_variables = {
        "valor_nf": invoice_value,
        "peso": Decimal(str(audited_weight)) if audited_weight is not None else None,
    }
    expected = weight_freight

    if has_tariff_freight_value:
        if invoice_value is None:
            return None
        rate = _parse_decimal_number(freight_value.get("rate"))
        if rate is not None:
            freight_value_amount = invoice_value * rate
            expected += freight_value_amount
            result["freight_value_amount"] = _round_money(freight_value_amount)
            result["expected_freight"] = _round_money(expected)
            tariff_component = {
                "amount": _round_money(freight_value_amount),
                "rate": float(rate),
                "source_column": freight_value.get("source_column"),
                "source_value": freight_value.get("source_value"),
                "invoice_value": _round_money(invoice_value),
                "details": (
                    f"Valor NF {_format_brazilian_decimal(invoice_value)} x "
                    f"{_format_brazilian_decimal(rate * Decimal('100'))}%"
                ),
            }
            components["freight_value"] = tariff_component
            components["tariff_freight_value"] = dict(tariff_component)

    if has_route_toll:
        toll_result = _calculate_route_toll_amount(route_toll, row.get("audited_weight"))
        if toll_result is not None:
            toll_amount, toll_component = toll_result
            expected += toll_amount
            result["route_toll_amount"] = _round_money(toll_amount)
            result["expected_freight"] = _round_money(expected)
            components["route_toll"] = toll_component
            components["tariff_route_toll"] = dict(toll_component)

    accessorial_components, ignored_accessorial_fees, accessorial_total = _build_accessorial_percent_fee_components(
        accessorial_fees,
        invoice_value=invoice_value,
        audit_variables=audit_variables,
        has_tariff_freight_value=has_tariff_freight_value,
        has_route_toll=has_route_toll and result.get("route_toll_amount") is not None,
    )
    accessorial_percent_total = sum(
        (
            _decimal_money(item.get("amount")) or Decimal("0")
            for item in accessorial_components
            if item.get("operation") == "percentage_of_variable"
            or item.get("calculation_type") == "invoice_percentage"
        ),
        Decimal("0"),
    )
    accessorial_percent_components = [
        item
        for item in accessorial_components
        if item.get("operation") == "percentage_of_variable"
        or item.get("calculation_type") == "invoice_percentage"
    ]
    if accessorial_components:
        expected += accessorial_total
        result["accessorial_fees_amount"] = _round_money(accessorial_total)
        result["accessorial_percent_fees_amount"] = (
            _round_money(accessorial_percent_total)
            if accessorial_percent_components
            else None
        )
        result["expected_freight"] = _round_money(expected)
    elif any(
        item.get("reason_code") == AUDIT_STATUS_INVALID_INVOICE_VALUE
        for item in ignored_accessorial_fees
    ):
        result["accessorial_fees_amount"] = _round_money(accessorial_total)
        result["accessorial_percent_fees_amount"] = _round_money(accessorial_total)
    components["accessorial_fees"] = accessorial_components
    components["accessorial_percent_fees"] = accessorial_percent_components
    components["ignored_accessorial_fees"] = ignored_accessorial_fees
    return result


def _resolve_pricing_index_entry(
    pricing_index: dict,
    lookup_key: str,
) -> tuple[dict, str] | None:
    if lookup_key in pricing_index:
        return pricing_index[lookup_key], lookup_key
    wanted = _normalize_audit_lookup_text(lookup_key)
    matches = [
        (region, rule)
        for region, rule in pricing_index.items()
        if _normalize_audit_lookup_text(region) == wanted
    ]
    if len(matches) == 1:
        matched_key, matched_rule = matches[0]
        return matched_rule, matched_key
    return None


def _pricing_rule_lookup_candidates(
    freight_region: str | None,
    destination_uf: str | None = None,
    destination_city: str | None = None,
) -> list[tuple[str, str]]:
    candidates: list[tuple[str, str]] = []
    if freight_region:
        candidates.append(("freight_region", freight_region))
        uf = _normalize_destination_uf(destination_uf)
        if uf:
            composed_key = f"{uf}|{_normalize_audit_lookup_text(freight_region)}"
            if composed_key:
                candidates.append(("freight_region", composed_key))
    uf_city_key = _coverage_lookup_key(destination_uf, destination_city)
    if uf_city_key:
        candidates.append(("destination_uf_city", uf_city_key))
    city_key = _normalize_audit_lookup_text(destination_city)
    if city_key:
        candidates.append(("destination_city", city_key))

    ordered: list[tuple[str, str]] = []
    seen: set[str] = set()
    for lookup_kind, lookup_key in candidates:
        if not lookup_key or lookup_key in seen:
            continue
        seen.add(lookup_key)
        ordered.append((lookup_kind, lookup_key))
    return ordered


def _find_pricing_rule_match(
    pricing_index: dict,
    freight_region: str | None,
    destination_uf: str | None = None,
    destination_city: str | None = None,
) -> tuple[dict, str, str] | None:
    fallback_unsupported: tuple[dict, str, str] | None = None
    for lookup_kind, lookup_key in _pricing_rule_lookup_candidates(
        freight_region,
        destination_uf,
        destination_city,
    ):
        resolved = _resolve_pricing_index_entry(pricing_index, lookup_key)
        if resolved is None:
            continue
        rule, matched_key = resolved
        if rule.get("pricing_type") == AUDIT_STATUS_UNSUPPORTED_PRICING:
            if fallback_unsupported is None:
                fallback_unsupported = (rule, lookup_kind, matched_key)
            continue
        return rule, lookup_kind, matched_key
    if fallback_unsupported is not None:
        return fallback_unsupported
    return None


def _find_pricing_rule(
    pricing_index: dict,
    freight_region: str | None,
    destination_uf: str | None = None,
    destination_city: str | None = None,
) -> dict | None:
    match = _find_pricing_rule_match(pricing_index, freight_region, destination_uf, destination_city)
    return match[0] if match else None


def _resolve_region_without_coverage(row: dict, pricing_index: dict) -> tuple[str | None, str | None]:
    city = _normalize_audit_lookup_text(row.get("destination_city"))
    uf = _normalize_destination_uf(row.get("destination_uf"))
    if not city and not uf:
        return None, AUDIT_STATUS_MISSING_COVERAGE
    candidates = []
    for region in pricing_index:
        normalized_region = _normalize_audit_lookup_text(region)
        if city and normalized_region == city:
            candidates.append(region)
        elif uf and normalized_region == uf:
            candidates.append(region)
        elif city and uf and normalized_region in {f"{uf} {city}", f"{city} {uf}", f"{uf}|{city}"}:
            candidates.append(region)
    unique = sorted(set(candidates))
    if len(unique) == 1:
        return unique[0], None
    return None, AUDIT_STATUS_MISSING_COVERAGE


def _base_audit_result(row: dict) -> dict:
    return {
        "row_index": row.get("row_index"),
        "numero_documento": row.get("document_number"),
        "destination_uf": row.get("destination_uf"),
        "destination_city": row.get("destination_city"),
        "freight_region": None,
        "audited_weight": row.get("audited_weight"),
        "charged_freight": row.get("charged_freight"),
        "expected_freight": None,
        "weight_freight": None,
        "freight_value_amount": None,
        "route_toll_amount": None,
        "accessorial_fees_amount": None,
        "accessorial_percent_fees_amount": None,
        "divergence_value": None,
        "status": None,
        "reason_code": None,
        "calculation_basis": None,
        "calculation_details": None,
        "calculation_components": {},
    }


def _limit_diagnostic_text(value, max_chars: int = 500) -> str | None:
    cleaned = _sanitize_cell_string(value)
    if not cleaned:
        return None
    if len(cleaned) <= max_chars:
        return cleaned
    return f"{cleaned[: max_chars - 1].rstrip()}…"


def _audit_search_context(row: dict, freight_region: str | None = None) -> dict:
    return {
        "destination_uf": _sanitize_cell_string(row.get("destination_uf")),
        "destination_city": _sanitize_cell_string(row.get("destination_city")),
        "coverage_classification": _sanitize_cell_string(freight_region),
    }


def _sanitize_attempted_keys(keys) -> list[str]:
    sanitized: list[str] = []
    seen: set[str] = set()
    for key in keys or []:
        cleaned = _sanitize_cell_string(key)
        if not cleaned or cleaned in seen:
            continue
        seen.add(cleaned)
        sanitized.append(cleaned)
        if len(sanitized) >= 8:
            break
    return sanitized


def _audit_failure_diagnostic(
    row: dict,
    *,
    failure_stage: str,
    freight_region: str | None = None,
    attempted_keys=None,
    message: str | None = None,
    diagnostic_group_code: str | None = None,
) -> dict:
    return {
        "failure_stage": _limit_diagnostic_text(failure_stage, 80),
        "diagnostic_group_code": _limit_diagnostic_text(diagnostic_group_code, 120),
        "search_context": _audit_search_context(row, freight_region),
        "attempted_keys": _sanitize_attempted_keys(attempted_keys),
        "message": _limit_diagnostic_text(message or "A linha não pôde ser calculada com os dados registrados."),
    }


def _status_result(
    row: dict,
    status: str,
    *,
    freight_region: str | None = None,
    diagnostic: dict | None = None,
) -> dict:
    result = _base_audit_result(row)
    result["freight_region"] = freight_region
    result["status"] = status
    result["reason_code"] = status
    if isinstance(diagnostic, dict):
        result["diagnostic"] = diagnostic
    return result


def _audit_single_row(
    row: dict,
    *,
    coverage_index: dict,
    pricing_index: dict,
    has_coverage: bool,
    accessorial_fees=None,
    tax_config=None,
) -> dict:
    weight = _parse_weight_number(row.get("audited_weight"))
    if weight is None:
        return _status_result(
            row,
            AUDIT_STATUS_INVALID_WEIGHT,
            diagnostic=_audit_failure_diagnostic(
                row,
                failure_stage="input_validation",
                message="Peso inválido ou ausente no arquivo auditado.",
            ),
        )
    charged = _parse_brazilian_money(row.get("charged_freight"))
    if charged is None:
        return _status_result(
            row,
            AUDIT_STATUS_INVALID_CHARGED_FREIGHT,
            diagnostic=_audit_failure_diagnostic(
                row,
                failure_stage="input_validation",
                message="Frete cobrado inválido ou ausente no arquivo auditado.",
            ),
        )

    freight_region = None
    if has_coverage:
        key = _coverage_lookup_key(row.get("destination_uf"), row.get("destination_city"))
        match = coverage_index.get(key) if key else None
        if isinstance(match, dict):
            return _status_result(
                row,
                AUDIT_STATUS_AMBIGUOUS_COVERAGE,
                diagnostic=_audit_failure_diagnostic(
                    row,
                    failure_stage="coverage_mapping",
                    message="Mais de uma classificação de cobertura foi encontrada para este destino.",
                ),
            )
        if isinstance(match, str) and match.strip():
            freight_region = match
        else:
            return _status_result(
                row,
                AUDIT_STATUS_MISSING_COVERAGE,
                diagnostic=_audit_failure_diagnostic(
                    row,
                    failure_stage="coverage_mapping",
                    message="Nenhuma classificação de cobertura foi encontrada para este destino.",
                ),
            )
    else:
        freight_region, reason = _resolve_region_without_coverage(row, pricing_index)
        if reason:
            return _status_result(
                row,
                reason,
                freight_region=freight_region,
                diagnostic=_audit_failure_diagnostic(
                    row,
                    failure_stage="coverage_mapping",
                    freight_region=freight_region,
                    message="Nenhuma classificação de cobertura foi encontrada para este destino.",
                ),
            )

    attempted_pricing_keys = [
        lookup_key
        for _lookup_kind, lookup_key in _pricing_rule_lookup_candidates(
            freight_region,
            row.get("destination_uf"),
            row.get("destination_city"),
        )
    ]
    rule_match = _find_pricing_rule_match(
        pricing_index,
        freight_region,
        row.get("destination_uf"),
        row.get("destination_city"),
    )
    if rule_match is None:
        uf = _normalize_destination_uf(row.get("destination_uf"))
        coverage = _sanitize_cell_string(freight_region)
        message = (
            f"Nenhuma regra compatível foi encontrada para {uf} / {coverage}."
            if uf and coverage
            else "Nenhuma regra compatível foi encontrada para a classificação de cobertura da linha."
        )
        return _status_result(
            row,
            AUDIT_STATUS_MISSING_FREIGHT_RULE,
            freight_region=freight_region,
            diagnostic=_audit_failure_diagnostic(
                row,
                failure_stage="pricing_rule_match",
                freight_region=freight_region,
                attempted_keys=attempted_pricing_keys,
                message=message,
            ),
        )
    rule, lookup_kind, lookup_key = rule_match
    if rule.get("pricing_type") == AUDIT_STATUS_UNSUPPORTED_PRICING:
        return _status_result(
            row,
            AUDIT_STATUS_UNSUPPORTED_PRICING,
            freight_region=freight_region,
            diagnostic=_audit_failure_diagnostic(
                row,
                failure_stage="pricing_rule_match",
                freight_region=freight_region,
                attempted_keys=attempted_pricing_keys,
                message="A regra localizada usa um modelo de tarifa ainda não suportado pela auditoria.",
            ),
        )

    calculated = calculate_weight_freight(weight, rule)
    if calculated is None:
        return _status_result(
            row,
            AUDIT_STATUS_UNSUPPORTED_PRICING,
            freight_region=freight_region,
            diagnostic=_audit_failure_diagnostic(
                row,
                failure_stage="pricing_calculation",
                freight_region=freight_region,
                attempted_keys=attempted_pricing_keys,
                message="A regra localizada não possui dados suficientes para calcular o frete esperado.",
            ),
        )
    calculated = _apply_freight_value_component(calculated, row, rule, accessorial_fees=accessorial_fees)
    if calculated is None:
        return _status_result(
            row,
            AUDIT_STATUS_INVALID_INVOICE_VALUE,
            freight_region=freight_region,
            diagnostic=_audit_failure_diagnostic(
                row,
                failure_stage="pricing_calculation",
                freight_region=freight_region,
                attempted_keys=attempted_pricing_keys,
                message="Valor de nota fiscal inválido ou ausente para uma regra que depende desse campo.",
            ),
        )

    calculated = _apply_row_tax_components(calculated, row, tax_config)

    comparison = compare_charged_vs_expected(charged, calculated["expected_freight"])
    result = _base_audit_result(row)
    result.update(comparison)
    result["freight_region"] = freight_region
    result["audited_weight"] = weight
    result["weight_freight"] = calculated.get("weight_freight")
    result["freight_value_amount"] = calculated.get("freight_value_amount")
    result["route_toll_amount"] = calculated.get("route_toll_amount")
    result["accessorial_fees_amount"] = calculated.get("accessorial_fees_amount")
    result["accessorial_percent_fees_amount"] = calculated.get("accessorial_percent_fees_amount")
    result["reason_code"] = None if comparison["status"] == AUDIT_STATUS_OK else AUDIT_STATUS_DIVERGENT
    result["calculation_basis"] = calculated["calculation_basis"]
    result["calculation_details"] = calculated["calculation_details"]
    result["calculation_components"] = calculated.get("calculation_components") or {}
    if lookup_kind != "freight_region":
        result["calculation_details"] = (
            f"{result['calculation_details']} | regra localizada por cidade/destino: {lookup_key}"
        )
    return result


def _build_audit_summary(results: list[dict], total_rows: int) -> dict:
    summary = {
        "total_rows": total_rows,
        "processed_rows": 0,
        "ok": 0,
        "divergent": 0,
        "missing_coverage_mapping": 0,
        "ambiguous_coverage_mapping": 0,
        "missing_freight_rule": 0,
        "invalid_rows": 0,
        "unsupported_pricing_model": 0,
    }
    for result in results:
        status = result.get("status")
        if status == AUDIT_STATUS_OK:
            summary["ok"] += 1
            summary["processed_rows"] += 1
        elif status == AUDIT_STATUS_DIVERGENT:
            summary["divergent"] += 1
            summary["processed_rows"] += 1
        elif status == AUDIT_STATUS_MISSING_COVERAGE:
            summary["missing_coverage_mapping"] += 1
        elif status == AUDIT_STATUS_AMBIGUOUS_COVERAGE:
            summary["ambiguous_coverage_mapping"] += 1
        elif status == AUDIT_STATUS_MISSING_FREIGHT_RULE:
            summary["missing_freight_rule"] += 1
        elif status in {
            AUDIT_STATUS_INVALID_WEIGHT,
            AUDIT_STATUS_INVALID_CHARGED_FREIGHT,
            AUDIT_STATUS_INVALID_INVOICE_VALUE,
        }:
            summary["invalid_rows"] += 1
        elif status == AUDIT_STATUS_UNSUPPORTED_PRICING:
            summary["unsupported_pricing_model"] += 1
    return summary


def _ordered_non_empty_strings(values) -> list[str]:
    ordered: list[str] = []
    seen: set[str] = set()
    for value in values or []:
        cleaned = _sanitize_cell_string(value)
        normalized = _normalize_audit_lookup_text(cleaned)
        if not cleaned or not normalized or normalized in seen:
            continue
        seen.add(normalized)
        ordered.append(cleaned)
    return ordered


def _pricing_index_available_values(pricing_index: dict) -> list[str]:
    if not isinstance(pricing_index, dict):
        return []
    # Composite keys are lookup helpers, not the tariff dimension shown to the user.
    return _ordered_non_empty_strings(key for key in pricing_index if "|" not in str(key))


def _candidate_column_for_requested_values(temp_table: dict, requested_values: list[str]) -> dict | None:
    requested_norm = {_normalize_audit_lookup_text(value) for value in requested_values}
    requested_norm.discard("")
    if not requested_norm or not isinstance(temp_table, dict):
        return None

    matches: list[dict] = []
    for table_index, table in enumerate(temp_table.get("freight_tables") or []):
        if not isinstance(table, dict):
            continue
        columns = table.get("columns") if isinstance(table.get("columns"), list) else []
        rows = table.get("rows") if isinstance(table.get("rows"), list) else []
        if not columns or not rows:
            continue
        current_column = next((col for col in columns if _is_region_column(col)), None)
        for column in columns:
            if _is_region_column(column) or _is_destination_uf_column(column):
                continue
            candidate_values = _ordered_non_empty_strings(
                row.get(column) for row in rows if isinstance(row, dict)
            )
            candidate_norm = {_normalize_audit_lookup_text(value) for value in candidate_values}
            candidate_norm.discard("")
            if candidate_norm == requested_norm:
                matches.append(
                    {
                        "candidate_column": _sanitize_cell_string(column),
                        "candidate_values": candidate_values,
                        "current_column": _sanitize_cell_string(current_column),
                        "table_refs": [f"freight_tables[{table_index}]"],
                        "exact_candidate_match": True,
                    }
                )
    if len(matches) != 1:
        return None
    return matches[0]


def _build_audit_diagnostics(
    temp_table: dict,
    *,
    results: list[dict],
    pricing_index: dict,
    generated_at: str,
) -> dict:
    error_statuses = {
        AUDIT_STATUS_MISSING_COVERAGE,
        AUDIT_STATUS_AMBIGUOUS_COVERAGE,
        AUDIT_STATUS_MISSING_FREIGHT_RULE,
        AUDIT_STATUS_INVALID_WEIGHT,
        AUDIT_STATUS_INVALID_CHARGED_FREIGHT,
        AUDIT_STATUS_INVALID_INVOICE_VALUE,
        AUDIT_STATUS_UNSUPPORTED_PRICING,
    }
    error_results = [
        result
        for result in results
        if isinstance(result, dict) and result.get("status") in error_statuses
    ]
    missing_rule_results = [
        result
        for result in error_results
        if result.get("status") == AUDIT_STATUS_MISSING_FREIGHT_RULE
        and _sanitize_cell_string(result.get("freight_region"))
    ]

    groups: list[dict] = []
    requested_values = _ordered_non_empty_strings(
        result.get("freight_region") for result in missing_rule_results
    )
    available_values = _pricing_index_available_values(pricing_index)
    requested_norm = {_normalize_audit_lookup_text(value) for value in requested_values}
    available_norm = {_normalize_audit_lookup_text(value) for value in available_values}
    requested_norm.discard("")
    available_norm.discard("")
    candidate = _candidate_column_for_requested_values(temp_table, requested_values)

    if (
        missing_rule_results
        and requested_norm
        and available_norm
        and requested_norm.isdisjoint(available_norm)
        and candidate is not None
    ):
        groups.append(
            {
                "code": AUDIT_DIAGNOSTIC_PRICING_DIMENSION_MISMATCH,
                "title": "Dimensão tarifária incompatível",
                "failure_stage": "pricing_rule_match",
                "affected_rows": len(missing_rule_results),
                "sample_row_indexes": [
                    result.get("row_index")
                    for result in missing_rule_results[:5]
                    if result.get("row_index") is not None
                ],
                "requested_values": requested_values,
                "available_values": available_values,
                "table_refs": list(candidate.get("table_refs") or []),
                "current_column": candidate.get("current_column"),
                "candidate_column": candidate["candidate_column"],
                "candidate_values": candidate["candidate_values"],
                "confidence": "high",
                "exact_candidate_match": bool(candidate.get("exact_candidate_match")),
                "ambiguous": False,
                "message": (
                    "As regiões utilizadas na cobertura não coincidem com a dimensão usada nas tarifas."
                ),
                "evidence": [],
                "actionability": {
                    "can_review_registered_table": True,
                    "can_apply_automatically": False,
                    "can_fix_source_files": True,
                },
            }
        )

    diagnostics = {
        "has_errors": bool(error_results),
        "total_errors": len(error_results),
        "generated_at": generated_at,
        "groups": groups,
    }
    diagnostics["suggestions"] = build_audit_correction_suggestions(temp_table, diagnostics)
    return diagnostics


def _format_human_list_pt(values) -> str:
    items = _ordered_non_empty_strings(values)
    if not items:
        return ""
    if len(items) == 1:
        return items[0]
    return f"{', '.join(items[:-1])} e {items[-1]}"


def _pricing_dimension_mismatch_row_message(result: dict, group: dict) -> str:
    diagnostic = result.get("diagnostic") if isinstance(result.get("diagnostic"), dict) else {}
    context = diagnostic.get("search_context") if isinstance(diagnostic.get("search_context"), dict) else {}
    city = _sanitize_cell_string(context.get("destination_city")) or "A cidade"
    uf = _sanitize_cell_string(context.get("destination_uf"))
    coverage = _sanitize_cell_string(context.get("coverage_classification"))
    current_values = _format_human_list_pt(group.get("available_values"))
    candidate_column = _sanitize_cell_string(group.get("candidate_column"))
    candidate_values = _format_human_list_pt(group.get("candidate_values"))

    parts = []
    if city and coverage:
        parts.append(f"A cidade {city} foi classificada como {coverage}.")
    if uf and coverage:
        parts.append(f"Nenhuma regra compatível foi encontrada para {uf} + {coverage}.")
    if current_values:
        parts.append(f"A tabela cadastrada está organizada atualmente pelos valores {current_values}.")
    if candidate_column and candidate_values:
        parts.append(
            f"Foi encontrada a coluna {candidate_column} com os valores {candidate_values}."
        )
    return " ".join(parts) or (
        "As regiões utilizadas na cobertura não coincidem com a dimensão usada nas tarifas."
    )


def _attach_audit_diagnostic_groups(results: list[dict], audit_diagnostics: dict) -> None:
    groups = audit_diagnostics.get("groups") if isinstance(audit_diagnostics, dict) else []
    if not isinstance(groups, list):
        return

    pricing_mismatch_group = next(
        (
            group
            for group in groups
            if isinstance(group, dict)
            and group.get("code") == AUDIT_DIAGNOSTIC_PRICING_DIMENSION_MISMATCH
            and group.get("failure_stage") == "pricing_rule_match"
        ),
        None,
    )
    if not pricing_mismatch_group:
        return

    for result in results:
        if not isinstance(result, dict):
            continue
        diagnostic = result.get("diagnostic")
        if not isinstance(diagnostic, dict):
            continue
        if (
            result.get("status") != AUDIT_STATUS_MISSING_FREIGHT_RULE
            or diagnostic.get("failure_stage") != "pricing_rule_match"
        ):
            continue
        diagnostic["diagnostic_group_code"] = AUDIT_DIAGNOSTIC_PRICING_DIMENSION_MISMATCH
        diagnostic["message"] = _limit_diagnostic_text(
            _pricing_dimension_mismatch_row_message(result, pricing_mismatch_group)
        )


def compute_audit_outputs(record: dict, normalized_rows: list[dict]) -> dict:
    coverage_table = record.get("coverage_table") if isinstance(record.get("coverage_table"), dict) else None
    has_coverage = bool(coverage_table and isinstance(coverage_table.get("rows"), list) and coverage_table.get("rows"))
    coverage_index = build_coverage_index(coverage_table or {"rows": []})
    pricing_index = build_freight_pricing_index(record)
    accessorial_fees = record.get("accessorial_fees") if isinstance(record.get("accessorial_fees"), list) else []
    tax_config = record.get("tax_config") if isinstance(record.get("tax_config"), dict) else None
    results = [
        _audit_single_row(
            row if isinstance(row, dict) else {},
            coverage_index=coverage_index,
            pricing_index=pricing_index,
            has_coverage=has_coverage,
            accessorial_fees=accessorial_fees,
            tax_config=tax_config,
        )
        for row in normalized_rows
    ]
    now = _utcnow().isoformat()
    summary = _build_audit_summary(results, len(normalized_rows))
    audit_diagnostics = _build_audit_diagnostics(
        record,
        results=results,
        pricing_index=pricing_index,
        generated_at=now,
    )
    _attach_audit_diagnostic_groups(results, audit_diagnostics)
    fiscal_snapshot = _build_tax_fiscal_snapshot(tax_config, results)
    audit_batch = {
        "normalized_rows": normalized_rows,
        "results": results,
        "summary": summary,
        "audit_diagnostics": audit_diagnostics,
        **fiscal_snapshot,
    }
    return {
        "results": results,
        "summary": summary,
        "audit_diagnostics": audit_diagnostics,
        "audit_bi": _public_audit_bi(audit_batch),
        "pricing_index": pricing_index,
        "generated_at": now,
        "fiscal_snapshot": fiscal_snapshot,
    }


def run_audit_batch_for_session(*, user_scope=None, franquia_scope=None) -> dict:
    started_at = time.perf_counter()
    emitted_processing_event = [False]
    execution_id = _resolve_agente_compara_execution_id()
    _require_session()
    sync_temp_table_with_session_documents()
    active_id = get_temp_table_id(session)
    if not active_id:
        raise AgenteComparaBatchError(
            ERROR_AUDIT_NO_TEMP_TABLE,
            "Nenhuma tabela temporária ativa nesta sessão.",
        )

    cfg = get_cleiton_doc_config()
    record = load_temp_table_record(active_id, ttl_hours=cfg.upload_ttl_hours)
    if record is None:
        clear_temp_table_session_refs(session)
        _mark_session_modified()
        raise AgenteComparaBatchError(
            ERROR_AUDIT_NO_TEMP_TABLE,
            "Tabela temporária ativa não encontrada.",
        )
    status = (record.get("status") or "").strip().lower()
    if status == TEMP_TABLE_STATUS_EXPIRED:
        raise AgenteComparaBatchError(
            ERROR_AUDIT_EXPIRED,
            "A tabela temporária desta sessão expirou.",
        )
    if status in {TEMP_TABLE_STATUS_DISCARDED, TEMP_TABLE_STATUS_PROCESSING}:
        raise AgenteComparaBatchError(
            ERROR_AUDIT_NO_TEMP_TABLE,
            "Tabela temporária indisponível para processamento.",
        )
    _assert_temp_table_scope(record, user_scope=user_scope, franquia_scope=franquia_scope)

    audit_batch = record.get("audit_batch")
    if not isinstance(audit_batch, dict):
        raise AgenteComparaBatchError(
            ERROR_AUDIT_BATCH_NOT_FOUND,
            "Nenhum lote auditado foi enviado nesta sessão.",
        )
    normalized_rows = audit_batch.get("normalized_rows")
    if not isinstance(normalized_rows, list) or not normalized_rows:
        raise AgenteComparaBatchError(
            ERROR_AUDIT_BATCH_EMPTY,
            "O lote auditado não possui linhas normalizadas para processar.",
        )

    should_bill_operational_run = _audit_batch_should_bill_operational_run(audit_batch)

    outputs = compute_audit_outputs(record, normalized_rows)
    results = outputs["results"]
    now = outputs["generated_at"]
    summary = outputs["summary"]
    audit_diagnostics = outputs["audit_diagnostics"]
    fiscal_snapshot = outputs.get("fiscal_snapshot") or {}
    preserved_expires_at = record.get("expires_at")
    updated_batch = dict(audit_batch)
    updated_batch["status"] = AUDIT_BATCH_STATUS_PROCESSED
    updated_batch["results"] = results
    updated_batch["summary"] = summary
    updated_batch["audit_diagnostics"] = audit_diagnostics
    updated_batch["updated_at"] = now
    updated_batch["processed_at"] = now
    updated_batch["expires_at"] = audit_batch.get("expires_at") or preserved_expires_at
    updated_batch = _apply_tax_fiscal_snapshot_to_audit_batch(updated_batch, fiscal_snapshot)
    updated_batch["pricing_rule_parser_version"] = PRICING_RULE_PARSER_VERSION
    updated_batch["pricing_rule_fingerprint"] = _pricing_rule_fingerprint(record)

    updated = dict(record)
    updated["audit_batch"] = updated_batch
    updated["updated_at"] = now
    updated["expires_at"] = preserved_expires_at

    saved = save_temp_table_record(updated)
    public = _public_temp_table(saved)
    if public is None:
        raise AgenteComparaBatchError(
            ERROR_AUDIT_NO_TEMP_TABLE,
            "Não foi possível retornar a tabela temporária processada.",
        )
    rows_evaluated = len(normalized_rows)
    processed_rows = int((summary or {}).get("processed_rows") or 0)
    billing_summary = None
    if processed_rows != rows_evaluated:
        billing_summary = f"processed_rows={processed_rows}"
    audit_batch_id = str(audit_batch.get("audit_batch_id") or "")
    if should_bill_operational_run:
        _emit_agente_compara_operational_billing(
            emitted=emitted_processing_event,
            started_at=started_at,
            flow_type=AGENTE_COMPARA_BATCH_PROCESSED_FLOW_TYPE,
            idempotency_key=agente_compara_batch_run_idempotency_key(active_id, audit_batch_id, execution_id),
            rows_processed=rows_evaluated,
            status="success",
            error_summary=billing_summary,
            execution_id=execution_id,
        )
    return public


def _public_audit_diagnostics(audit_diagnostics) -> dict | None:
    if not isinstance(audit_diagnostics, dict):
        return None

    public_groups: list[dict] = []
    groups = audit_diagnostics.get("groups")
    if isinstance(groups, list):
        for group in groups:
            if not isinstance(group, dict):
                continue
            actionability = group.get("actionability")
            public_actionability = {
                "can_review_registered_table": False,
                "can_apply_automatically": False,
                "can_fix_source_files": False,
            }
            if isinstance(actionability, dict):
                public_actionability = {
                    "can_review_registered_table": bool(
                        actionability.get("can_review_registered_table")
                    ),
                    "can_apply_automatically": bool(actionability.get("can_apply_automatically")),
                    "can_fix_source_files": bool(actionability.get("can_fix_source_files")),
                }
            raw_evidence = group.get("evidence") if isinstance(group.get("evidence"), list) else []
            public_evidence = _ordered_non_empty_strings(raw_evidence)
            public_group = {
                "code": _sanitize_cell_string(group.get("code")),
                "title": _sanitize_cell_string(group.get("title")),
                "failure_stage": _sanitize_cell_string(group.get("failure_stage")),
                "affected_rows": group.get("affected_rows"),
                "sample_row_indexes": list(group.get("sample_row_indexes") or [])[:5],
                "requested_values": _ordered_non_empty_strings(group.get("requested_values")),
                "available_values": _ordered_non_empty_strings(group.get("available_values")),
                "candidate_column": _sanitize_cell_string(group.get("candidate_column")),
                "candidate_values": _ordered_non_empty_strings(group.get("candidate_values")),
                "confidence": _sanitize_cell_string(group.get("confidence")),
                "message": _sanitize_cell_string(group.get("message")),
                "evidence": public_evidence,
                "actionability": public_actionability,
            }
            public_table_refs = [
                _sanitize_cell_string(ref)
                for ref in list(group.get("table_refs") or [])[:10]
                if _sanitize_cell_string(ref)
            ]
            if public_table_refs:
                public_group["table_refs"] = public_table_refs
            public_current_column = _sanitize_cell_string(group.get("current_column"))
            if public_current_column:
                public_group["current_column"] = public_current_column
            public_groups.append(public_group)

    public_diagnostics = {
        "has_errors": bool(audit_diagnostics.get("has_errors")),
        "total_errors": audit_diagnostics.get("total_errors") or 0,
        "generated_at": _sanitize_cell_string(audit_diagnostics.get("generated_at")),
        "groups": public_groups,
    }
    suggestions = list(audit_diagnostics.get("suggestions") or [])
    if suggestions:
        public_diagnostics["suggestions"] = suggestions
    return public_diagnostics


def _public_audit_batch(audit_batch) -> dict | None:
    if not isinstance(audit_batch, dict):
        return None
    normalized_rows = audit_batch.get("normalized_rows")
    row_count = audit_batch.get("row_count")
    if row_count is None and isinstance(normalized_rows, list):
        row_count = len(normalized_rows)
    return {
        "status": audit_batch.get("status"),
        "audit_batch_id": audit_batch.get("audit_batch_id"),
        "temp_table_id": audit_batch.get("temp_table_id"),
        "source_file_name": audit_batch.get("source_file_name"),
        "sheet_name": audit_batch.get("sheet_name"),
        "uploaded_at": audit_batch.get("uploaded_at"),
        "created_at": audit_batch.get("created_at"),
        "updated_at": audit_batch.get("updated_at"),
        "expires_at": audit_batch.get("expires_at"),
        "row_count": row_count,
        "max_rows": audit_batch.get("max_rows"),
        "input_schema_version": audit_batch.get("input_schema_version"),
        "header_map": dict(audit_batch.get("header_map") or {}),
        "results": list(audit_batch.get("results") or []),
        "summary": audit_batch.get("summary"),
        "audit_diagnostics": _public_audit_diagnostics(audit_batch.get("audit_diagnostics")),
        "processed_at": audit_batch.get("processed_at"),
        "needs_reprocess": _audit_batch_effective_needs_reprocess(audit_batch),
        "stale_reason": audit_batch.get("stale_reason"),
        "tax_applied": audit_batch.get("tax_applied"),
        "tax_calculation_mode": audit_batch.get("tax_calculation_mode"),
        "tax_calculation_version": audit_batch.get("tax_calculation_version"),
        "tax_config_snapshot": audit_batch.get("tax_config_snapshot"),
        "tax_summary": audit_batch.get("tax_summary"),
    }


def get_agente_compara_template_path():
    from pathlib import Path

    from flask import current_app

    template_path = (
        Path(current_app.root_path)
        / "protected_files"
        / "templates"
        / AGENTE_COMPARA_TEMPLATE_FILENAME
    )
    return template_path


def upload_audit_batch_from_file(
    *,
    display_name: str,
    file_bytes: bytes,
    extension: str | None,
    user_scope=None,
    franquia_scope=None,
) -> dict:
    """
    Upload determinístico do arquivo auditado no tt_*.json ativo.

    Não registra documento principal, não chama Gemini e não executa cálculo de auditoria.
    """
    started_at = time.perf_counter()
    emitted_processing_event = [False]
    execution_id = _resolve_agente_compara_execution_id()
    _require_session()
    if not file_bytes:
        raise AgenteComparaBatchError(
            ERROR_AUDIT_EMPTY_FILE,
            "O arquivo auditado está vazio.",
        )
    cfg = get_cleiton_doc_config()
    audit_cfg = get_agente_compara_config()
    audit_limits = resolve_audited_file_limits(audit_cfg, global_cfg=cfg)
    max_bytes = int(audit_limits["effective_max_bytes"] or 0)
    if len(file_bytes) > max_bytes:
        raise AgenteComparaBatchError(
            ERROR_AUDIT_PAYLOAD_TOO_LARGE,
            "O arquivo auditado excede o limite de tamanho permitido.",
        )

    ext = (extension or "").strip().lower()
    if not ext.startswith("."):
        ext = f".{ext}" if ext else ""
    if ext == ".pdf":
        raise AgenteComparaBatchError(
            ERROR_AUDIT_INVALID_FORMAT,
            "Upload do arquivo auditado aceita apenas CSV e XLSX nesta fase.",
        )
    if ext not in {".csv", ".xlsx"}:
        raise AgenteComparaBatchError(
            ERROR_AUDIT_INVALID_FORMAT,
            "Upload do arquivo auditado aceita apenas CSV e XLSX nesta fase.",
        )

    sync_temp_table_with_session_documents()
    active_id = get_temp_table_id(session)
    if not active_id:
        raise AgenteComparaBatchError(
            ERROR_AUDIT_NO_TEMP_TABLE,
            "Nenhuma tabela temporária ativa nesta sessão.",
        )

    record = load_temp_table_record(active_id, ttl_hours=cfg.upload_ttl_hours)
    if record is None:
        clear_temp_table_session_refs(session)
        _mark_session_modified()
        raise AgenteComparaBatchError(
            ERROR_AUDIT_NO_TEMP_TABLE,
            "Tabela temporária ativa não encontrada.",
        )
    status = (record.get("status") or "").strip().lower()
    if status == TEMP_TABLE_STATUS_EXPIRED:
        raise AgenteComparaBatchError(
            ERROR_AUDIT_EXPIRED,
            "A tabela temporária desta sessão expirou.",
        )
    if status in {TEMP_TABLE_STATUS_DISCARDED, TEMP_TABLE_STATUS_PROCESSING}:
        raise AgenteComparaBatchError(
            ERROR_AUDIT_NO_TEMP_TABLE,
            "Tabela temporária indisponível para upload do arquivo auditado.",
        )
    _assert_temp_table_scope(record, user_scope=user_scope, franquia_scope=franquia_scope)

    max_rows = int(audit_limits["effective_max_rows"] or 0)
    safe_name = secure_filename(display_name or "auditado") or "auditado"
    if ext == ".csv":
        normalized_rows, header_map, sheet_name = _parse_audit_csv_bytes(
            file_bytes,
            source_file_name=safe_name,
            max_rows=max_rows,
        )
    else:
        normalized_rows, header_map, sheet_name = _parse_audit_xlsx_bytes(
            file_bytes,
            source_file_name=safe_name,
            max_bytes=max_bytes,
            max_rows=max_rows,
        )

    now = _utcnow().isoformat()
    preserved_expires_at = record.get("expires_at")
    batch_id = uuid4().hex
    audit_batch = {
        "status": AUDIT_BATCH_STATUS_UPLOADED,
        "audit_batch_id": batch_id,
        "temp_table_id": active_id,
        "source_file_name": safe_name,
        "sheet_name": sheet_name,
        "uploaded_at": now,
        "created_at": now,
        "updated_at": now,
        "expires_at": preserved_expires_at,
        "row_count": len(normalized_rows),
        "max_rows": max_rows,
        "input_schema_version": AUDIT_INPUT_SCHEMA_VERSION,
        "header_map": header_map,
        "normalized_rows": normalized_rows,
        "results": [],
        "summary": None,
        "audit_diagnostics": None,
    }

    updated = dict(record)
    updated["audit_batch"] = audit_batch
    updated["updated_at"] = now
    updated["expires_at"] = preserved_expires_at

    saved = save_temp_table_record(updated)
    logger.info(
        "Agente Compara audit batch upload: temp_table_id=%s user_id=%s rows=%s max_rows=%s",
        saved.get("temp_table_id"),
        user_scope,
        len(normalized_rows),
        max_rows,
    )
    public = _public_temp_table(saved)
    if public is None:
        raise AgenteComparaBatchError(
            ERROR_AUDIT_NO_TEMP_TABLE,
            "Não foi possível retornar a tabela temporária atualizada.",
        )
    _emit_agente_compara_operational_billing(
        emitted=emitted_processing_event,
        started_at=started_at,
        flow_type=AGENTE_COMPARA_BATCH_UPLOAD_FLOW_TYPE,
        idempotency_key=agente_compara_batch_upload_idempotency_key(active_id, batch_id),
        rows_processed=len(normalized_rows),
        status="success",
        execution_id=execution_id,
    )
    return public


def _audit_bi_field_has_value(value) -> bool:
    if value is None:
        return False
    if isinstance(value, str):
        return bool(value.strip())
    return True


def _audit_bi_results_by_row_index(results) -> dict[int, dict]:
    indexed: dict[int, dict] = {}
    if not isinstance(results, list):
        return indexed
    for result in results:
        if not isinstance(result, dict):
            continue
        row_index = result.get("row_index")
        if isinstance(row_index, bool) or row_index is None:
            continue
        try:
            indexed[int(row_index)] = result
        except (TypeError, ValueError):
            continue
    return indexed


def _audit_bi_public_row(normalized_row: dict, result: dict | None = None) -> dict:
    public_row = {
        "row_index": normalized_row.get("row_index"),
        "carrier": normalized_row.get("carrier"),
        "origin_uf": normalized_row.get("origin_uf"),
        "destination_uf": normalized_row.get("destination_uf"),
        "issue_date": normalized_row.get("issue_date"),
        "audited_weight": normalized_row.get("audited_weight"),
        "charged_freight": normalized_row.get("charged_freight"),
        "expected_freight": None,
        "divergence_value": None,
        "status": None,
    }
    if isinstance(result, dict):
        if result.get("expected_freight") is not None:
            public_row["expected_freight"] = result.get("expected_freight")
        if result.get("divergence_value") is not None:
            public_row["divergence_value"] = result.get("divergence_value")
        if result.get("status") is not None:
            public_row["status"] = result.get("status")
    return public_row


def _audit_bi_compute_field_presence(rows: list[dict]) -> dict[str, bool]:
    presence = {field: False for field in AUDIT_BI_FIELD_PRESENCE_FIELDS}
    for row in rows:
        if not isinstance(row, dict):
            continue
        for field in AUDIT_BI_FIELD_PRESENCE_FIELDS:
            if _audit_bi_field_has_value(row.get(field)):
                presence[field] = True
    return presence


def _public_audit_bi(audit_batch) -> dict:
    base = {
        "dataset_version": AUDIT_BI_DATASET_VERSION,
        "source": AUDIT_BI_SOURCE,
        "filter_mode": AUDIT_BI_FILTER_MODE,
        "charts_supported": list(AUDIT_BI_CHARTS_SUPPORTED),
    }
    normalized_rows = None
    results = None
    if isinstance(audit_batch, dict):
        raw_rows = audit_batch.get("normalized_rows")
        if isinstance(raw_rows, list):
            normalized_rows = raw_rows
        raw_results = audit_batch.get("results")
        if isinstance(raw_results, list):
            results = raw_results

    if not normalized_rows:
        return {
            **base,
            "ready": False,
            "row_count": 0,
            "rows": [],
            "field_presence": {field: False for field in AUDIT_BI_FIELD_PRESENCE_FIELDS},
            "methodology": _audit_bi_methodology(audit_batch if isinstance(audit_batch, dict) else None),
            "message": AUDIT_BI_NOT_READY_MESSAGE,
        }

    results_by_index = _audit_bi_results_by_row_index(results)
    public_rows: list[dict] = []
    for normalized_row in normalized_rows:
        if not isinstance(normalized_row, dict):
            continue
        row_index = normalized_row.get("row_index")
        result = None
        if row_index is not None and not isinstance(row_index, bool):
            try:
                result = results_by_index.get(int(row_index))
            except (TypeError, ValueError):
                result = None
        public_rows.append(_audit_bi_public_row(normalized_row, result))

    return {
        **base,
        "ready": True,
        "row_count": len(public_rows),
        "rows": public_rows,
        "field_presence": _audit_bi_compute_field_presence(public_rows),
        "methodology": _audit_bi_methodology(audit_batch if isinstance(audit_batch, dict) else None),
    }


def _public_temp_table(record: dict | None) -> dict | None:
    if not record:
        return None
    ui = record.get("ui_visibility") if isinstance(record.get("ui_visibility"), dict) else {}
    correction_history = record.get("correction_history") if isinstance(record.get("correction_history"), list) else []
    last_correction = correction_history[-1] if correction_history and isinstance(correction_history[-1], dict) else {}
    public = {
        "temp_table_id": record.get("temp_table_id"),
        "status": record.get("status"),
        "source_documents": list(record.get("source_documents") or []),
        "detected_carrier": record.get("detected_carrier"),
        "origins": list(record.get("origins") or []),
        "destinations": list(record.get("destinations") or []),
        "routes": list(record.get("routes") or []),
        "freight_tables": list(record.get("freight_tables") or []),
        "freight_routes": list(record.get("freight_routes") or []),
        "weight_ranges": list(record.get("weight_ranges") or []),
        "freight_values": list(record.get("freight_values") or []),
        "accessorial_fees": list(record.get("accessorial_fees") or []),
        "charge_type_detected": record.get("charge_type_detected"),
        "extracted_items": list(record.get("extracted_items") or []),
        "uncertain_fields": list(record.get("uncertain_fields") or []),
        "reading_alerts": list(record.get("reading_alerts") or []),
        "evidence_refs": list(record.get("evidence_refs") or []),
        "created_at": record.get("created_at"),
        "updated_at": record.get("updated_at"),
        "expires_at": record.get("expires_at"),
        "session_scope": record.get("session_scope"),
        "franquia_scope": record.get("franquia_scope"),
        "user_scope": record.get("user_scope"),
        "operational_owner": record.get("operational_owner"),
        "ui_visibility": {
            "display_name": ui.get("display_name") or TEMP_TABLE_UI_DISPLAY_NAME,
            "readonly": True,
        },
        "version_marker": record.get("version_marker"),
        "human_review_status": record.get("human_review_status"),
        "human_edited_at": record.get("human_edited_at"),
        "human_edited_by_user_id": record.get("human_edited_by_user_id"),
        "edit_version": record.get("edit_version"),
        "tax_config": copy.deepcopy(record.get("tax_config")) if isinstance(record.get("tax_config"), dict) else None,
        "audit_correction": {
            "can_undo": bool(correction_history),
            "last_application_id": record.get("last_correction_application_id") or last_correction.get("application_id"),
            "last_applied_at": record.get("last_correction_applied_at") or last_correction.get("applied_at"),
            "last_suggestion_id": last_correction.get("suggestion_id"),
        },
    }
    coverage = _public_coverage_table(record.get("coverage_table"))
    if coverage is not None:
        public["coverage_table"] = coverage
    audit_batch = _public_audit_batch(record.get("audit_batch"))
    if audit_batch is not None:
        public["audit_batch"] = audit_batch
    public["audit_bi"] = _public_audit_bi(record.get("audit_batch"))
    return public


def _sanitize_cell_string(value) -> str | None:
    if value is None:
        return None
    if isinstance(value, bool):
        return str(value)
    if not isinstance(value, str):
        value = str(value)
    cleaned = "".join(
        ch for ch in value if ch in {"\n", "\t"} or (ord(ch) >= 32 and ord(ch) != 127)
    )
    cleaned = _HTML_TAG_RE.sub("", cleaned)
    stripped = cleaned.strip()
    return stripped if stripped else ""


def _sanitize_freight_table_context(raw_context) -> dict:
    normalized = _normalize_freight_table_context(raw_context)
    return {
        key: _sanitize_cell_string(val) if val is not None else None
        for key, val in normalized.items()
    }


def _validate_freight_table_item_for_save(item) -> dict:
    if not isinstance(item, dict):
        raise AgenteComparaTempTableError(
            ERROR_TEMP_TABLE_INVALID_PAYLOAD,
            "Cada item de freight_tables deve ser um objeto.",
        )
    raw_columns = item.get("columns")
    columns: list[str] = []
    if raw_columns is not None:
        if not isinstance(raw_columns, list):
            raise AgenteComparaTempTableError(
                ERROR_TEMP_TABLE_INVALID_PAYLOAD,
                "freight_tables.columns deve ser uma lista.",
            )
        for col in raw_columns:
            if not isinstance(col, str):
                raise AgenteComparaTempTableError(
                    ERROR_TEMP_TABLE_INVALID_PAYLOAD,
                    "Nome de coluna inválido.",
                )
            candidate = _sanitize_cell_string(col)
            if not candidate:
                raise AgenteComparaTempTableError(
                    ERROR_TEMP_TABLE_INVALID_PAYLOAD,
                    "Coluna sem nome não é permitida.",
                )
            columns.append(candidate)
    raw_rows = item.get("rows")
    rows: list[dict] = []
    if raw_rows is not None:
        if not isinstance(raw_rows, list):
            raise AgenteComparaTempTableError(
                ERROR_TEMP_TABLE_INVALID_PAYLOAD,
                "freight_tables.rows deve ser uma lista.",
            )
        for row in raw_rows:
            if not isinstance(row, dict):
                raise AgenteComparaTempTableError(
                    ERROR_TEMP_TABLE_INVALID_PAYLOAD,
                    "Cada linha de freight_tables deve ser um objeto.",
                )
            normalized_row: dict = {}
            if columns:
                for col in columns:
                    val = row.get(col)
                    if val is None:
                        normalized_row[col] = None
                    else:
                        normalized_row[col] = _sanitize_cell_string(val)
            else:
                for key, val in row.items():
                    if not isinstance(key, str) or not key.strip():
                        continue
                    safe_key = _sanitize_cell_string(key)
                    if not safe_key:
                        continue
                    normalized_row[safe_key] = _sanitize_cell_string(val) if val is not None else None
            if normalized_row:
                rows.append(normalized_row)
    if not columns and not rows:
        raise AgenteComparaTempTableError(
            ERROR_TEMP_TABLE_INVALID_PAYLOAD,
            "Tabela principal não pode ficar completamente vazia.",
        )
    if columns and not rows:
        raise AgenteComparaTempTableError(
            ERROR_TEMP_TABLE_INVALID_PAYLOAD,
            "Tabela principal não pode ficar sem linhas.",
        )
    if rows and not columns:
        raise AgenteComparaTempTableError(
            ERROR_TEMP_TABLE_INVALID_PAYLOAD,
            "Tabela principal não pode ficar sem colunas.",
        )
    return {
        "table_title": _sanitize_cell_string(item.get("table_title")),
        "table_type": _sanitize_cell_string(item.get("table_type")),
        "context": _sanitize_freight_table_context(item.get("context")),
        "columns": columns,
        "rows": rows,
        "notes": _sanitize_cell_string(item.get("notes")) or "",
        "evidence_ref": _sanitize_cell_string(item.get("evidence_ref")),
        "confidence": _sanitize_cell_string(item.get("confidence")),
    }


def _validate_freight_tables_for_save(raw_tables) -> list[dict]:
    if raw_tables is None:
        return []
    if not isinstance(raw_tables, list):
        raise AgenteComparaTempTableError(
            ERROR_TEMP_TABLE_INVALID_PAYLOAD,
            "freight_tables deve ser uma lista.",
        )
    if not raw_tables:
        return []
    return [_validate_freight_table_item_for_save(item) for item in raw_tables]


def _normalize_uf(value) -> str | None:
    if value is None:
        return None
    text = str(value).strip().upper()
    if len(text) != 2 or text not in BRAZILIAN_UFS:
        return None
    return text


_BR_STATE_NAME_TO_UF = {
    "ACRE": "AC",
    "ALAGOAS": "AL",
    "AMAPA": "AP",
    "AMAZONAS": "AM",
    "BAHIA": "BA",
    "CEARA": "CE",
    "DISTRITO FEDERAL": "DF",
    "ESPIRITO SANTO": "ES",
    "GOIAS": "GO",
    "MARANHAO": "MA",
    "MATO GROSSO": "MT",
    "MATO GROSSO DO SUL": "MS",
    "MINAS GERAIS": "MG",
    "PARA": "PA",
    "PARAIBA": "PB",
    "PARANA": "PR",
    "PERNAMBUCO": "PE",
    "PIAUI": "PI",
    "RIO DE JANEIRO": "RJ",
    "RIO GRANDE DO NORTE": "RN",
    "RIO GRANDE DO SUL": "RS",
    "RONDONIA": "RO",
    "RORAIMA": "RR",
    "SANTA CATARINA": "SC",
    "SAO PAULO": "SP",
    "SERGIPE": "SE",
    "TOCANTINS": "TO",
}

_KNOWN_CITY_TO_UF = {
    "CAMPINAS": "SP",
    "SAO PAULO": "SP",
    "JOINVILLE": "SC",
}

_TAX_DESTINATION_SOURCE_PRIORITY = {
    "manual": 4,
    "automatic": 3,
    "inferred_state": 2,
    "inferred_city": 1,
}

_VALID_TAX_DESTINATION_SOURCES = frozenset(_TAX_DESTINATION_SOURCE_PRIORITY.keys())

_TAX_DESTINATION_RECORD_SOURCES = (
    "destinations",
    "routes",
    "freight_routes",
    "freight_tables",
    "extracted_items",
)


def _extract_uf_tokens(value) -> set[str]:
    if value is None:
        return set()
    text = str(value).upper()
    return {match.group(1) for match in re.finditer(r"\b([A-Z]{2})\b", text) if match.group(1) in BRAZILIAN_UFS}


def _normalize_state_name_key(value) -> str:
    return _normalize_audit_lookup_text(value)


def resolve_state_name_to_uf(value) -> str | None:
    key = _normalize_state_name_key(value)
    if not key:
        return None
    return _BR_STATE_NAME_TO_UF.get(key)


def _resolve_city_name_to_uf(value) -> str | None:
    city_key = _normalize_state_name_key(value)
    if not city_key:
        return None
    if city_key in _KNOWN_CITY_TO_UF:
        return _KNOWN_CITY_TO_UF[city_key]
    ufs = _lookup_city_ufs_from_base_localidades(value)
    if len(ufs) == 1:
        return ufs[0]
    return None


def _lookup_city_ufs_from_base_localidades(city_text: str) -> list[str]:
    city_key = str(city_text or "").strip().lower()
    if not city_key:
        return []
    city_key = unicodedata.normalize("NFKD", city_key)
    city_key = "".join(ch for ch in city_key if not unicodedata.combining(ch))
    city_key = re.sub(r"\s+", " ", city_key).strip()
    if not city_key:
        return []
    try:
        from app.models import BaseLocalidades

        rows = BaseLocalidades.query.filter(
            BaseLocalidades.chave_busca.like(f"{city_key}-%")
        ).all()
    except Exception:
        return []
    ufs: list[str] = []
    seen: set[str] = set()
    for row in rows:
        uf_sigla = resolve_state_name_to_uf(row.uf_nome)
        if uf_sigla and uf_sigla not in seen:
            seen.add(uf_sigla)
            ufs.append(uf_sigla)
    return sorted(ufs)


def _tax_destination_field_kind(key) -> str | None:
    normalized = _normalize_audit_lookup_text(key).lower()
    compact = normalized.replace(" ", "_")
    if compact in {"destination_uf", "uf_destino", "destino_uf"}:
        return "uf"
    if "uf_destino" in compact or "destination_uf" in compact:
        return "uf"
    if "estado" in compact and ("destino" in compact or "destination" in compact):
        return "state"
    if "cidade" in compact and ("destino" in compact or "destination" in compact):
        return "city"
    if "destino" in compact or "destination" in compact:
        if "estado" in compact:
            return "state"
        if "uf" in compact:
            return "uf"
        return "city"
    return None


def _destination_uf_field_score(key) -> int:
    field_kind = _tax_destination_field_kind(key)
    if field_kind == "uf":
        return 3
    if field_kind == "state":
        return 2
    if field_kind == "city":
        return 1
    return 0


def _resolve_tax_location_findings(value, field_kind: str) -> list[tuple[str, str, str]]:
    text = _sanitize_cell_string(value)
    if not text:
        return []

    findings: list[tuple[str, str, str]] = []
    seen_ufs: set[str] = set()

    def _append(uf: str | None, source: str) -> None:
        if not uf or uf in seen_ufs:
            return
        seen_ufs.add(uf)
        findings.append((uf, source, text))

    if field_kind == "uf":
        direct_uf = _normalize_uf(text)
        if direct_uf:
            _append(direct_uf, "automatic")
        else:
            for token in sorted(_extract_uf_tokens(text)):
                _append(token, "automatic")
        return findings

    if field_kind == "state":
        state_uf = resolve_state_name_to_uf(text)
        if state_uf:
            _append(state_uf, "inferred_state")
            return findings
        for token in sorted(_extract_uf_tokens(text)):
            _append(token, "automatic")
        return findings

    city_uf = _resolve_city_name_to_uf(text)
    if city_uf:
        _append(city_uf, "inferred_city")
        return findings

    state_uf = resolve_state_name_to_uf(text)
    if state_uf:
        _append(state_uf, "inferred_state")
        return findings

    for token in sorted(_extract_uf_tokens(text)):
        _append(token, "automatic")
    return findings


def _merge_tax_destination_entry(
    by_uf: dict[str, dict],
    *,
    uf: str,
    source: str,
    evidence: str | None = None,
    user_confirmed: bool = False,
) -> None:
    normalized_uf = _normalize_uf(uf)
    if not normalized_uf:
        return
    if source not in _VALID_TAX_DESTINATION_SOURCES:
        source = "manual"
    evidence_text = _sanitize_cell_string(evidence)
    if normalized_uf not in by_uf:
        by_uf[normalized_uf] = {
            "uf": normalized_uf,
            "source": source,
            "evidence": [evidence_text] if evidence_text else [],
            "user_confirmed": bool(user_confirmed),
        }
        return
    entry = by_uf[normalized_uf]
    if evidence_text and evidence_text not in entry["evidence"]:
        entry["evidence"].append(evidence_text)
    current_priority = _TAX_DESTINATION_SOURCE_PRIORITY.get(entry["source"], 0)
    incoming_priority = _TAX_DESTINATION_SOURCE_PRIORITY.get(source, 0)
    if incoming_priority > current_priority:
        entry["source"] = source
    if user_confirmed:
        entry["user_confirmed"] = True


def _collect_tax_destination_findings_from_value(
    value,
    findings: list[tuple[str, str, str]],
    *,
    key_hint: str | None = None,
) -> None:
    if value is None:
        return
    if isinstance(value, dict):
        for key, item in value.items():
            if not isinstance(key, str):
                continue
            field_kind = _tax_destination_field_kind(key)
            if field_kind:
                findings.extend(_resolve_tax_location_findings(item, field_kind))
            _collect_tax_destination_findings_from_value(item, findings, key_hint=key)
        return
    if isinstance(value, list):
        for item in value:
            _collect_tax_destination_findings_from_value(item, findings, key_hint=key_hint)
        return
    if key_hint:
        field_kind = _tax_destination_field_kind(key_hint)
        if field_kind:
            findings.extend(_resolve_tax_location_findings(value, field_kind))


def _extract_tax_destination_findings_from_record(record: dict | None) -> list[tuple[str, str, str]]:
    if not isinstance(record, dict):
        return []
    findings: list[tuple[str, str, str]] = []
    for key in _TAX_DESTINATION_RECORD_SOURCES:
        _collect_tax_destination_findings_from_value(record.get(key), findings)
    return findings


def _normalize_tax_destination_evidence(value) -> list[str]:
    if value is None:
        return []
    if isinstance(value, list):
        evidence: list[str] = []
        for item in value:
            text = _sanitize_cell_string(item)
            if text and text not in evidence:
                evidence.append(text)
        return evidence
    text = _sanitize_cell_string(value)
    return [text] if text else []


def _normalize_tax_destination_entry(raw_entry, *, default_source: str = "manual") -> dict | None:
    if not isinstance(raw_entry, dict):
        return None
    uf = _normalize_uf(raw_entry.get("uf"))
    if not uf:
        return None
    source = str(raw_entry.get("source") or default_source).strip()
    if source not in _VALID_TAX_DESTINATION_SOURCES:
        source = default_source
    user_confirmed = bool(raw_entry.get("user_confirmed"))
    if source == "manual":
        user_confirmed = True
    return {
        "uf": uf,
        "source": source,
        "evidence": _normalize_tax_destination_evidence(raw_entry.get("evidence")),
        "user_confirmed": user_confirmed,
    }


def consolidate_tax_destination_ufs(
    record: dict | None,
    *,
    submitted_destination_ufs: list | None = None,
) -> list[dict]:
    by_uf: dict[str, dict] = {}

    if submitted_destination_ufs is None:
        for uf, source, evidence in _extract_tax_destination_findings_from_record(record):
            _merge_tax_destination_entry(by_uf, uf=uf, source=source, evidence=evidence)
    else:
        if not isinstance(submitted_destination_ufs, list):
            return []
        for raw_entry in submitted_destination_ufs:
            normalized = _normalize_tax_destination_entry(raw_entry)
            if not normalized:
                continue
            evidence = normalized["evidence"]
            evidence_text = evidence[0] if evidence else None
            _merge_tax_destination_entry(
                by_uf,
                uf=normalized["uf"],
                source=normalized["source"],
                evidence=evidence_text,
                user_confirmed=normalized["user_confirmed"],
            )
            entry = by_uf[normalized["uf"]]
            for extra in evidence[1:]:
                if extra not in entry["evidence"]:
                    entry["evidence"].append(extra)

    return [by_uf[uf] for uf in sorted(by_uf)]


def _collect_destination_ufs_from_value(value, destination_ufs: set[str], *, key_hint: str | None = None) -> None:
    findings: list[tuple[str, str, str]] = []
    _collect_tax_destination_findings_from_value(value, findings, key_hint=key_hint)
    destination_ufs.update(uf for uf, _source, _evidence in findings)


def extract_tax_destination_ufs_from_temp_table(record: dict | None) -> list[str]:
    return [entry["uf"] for entry in consolidate_tax_destination_ufs(record)]


def suggested_icms_interstate_rate(origin_uf: str, destination_uf: str) -> float:
    origin = _normalize_uf(origin_uf)
    destination = _normalize_uf(destination_uf)
    if origin in ICMS_7_PERCENT_ORIGIN_UFS and destination in ICMS_7_PERCENT_DESTINATION_UFS:
        return 7.0
    return 12.0


def _parse_optional_rate(value) -> float | None:
    if value is None:
        return None
    if isinstance(value, bool):
        raise AgenteComparaTempTableError(
            ERROR_TEMP_TABLE_INVALID_PAYLOAD,
            "Alíquota fiscal inválida.",
        )
    text = str(value).strip()
    if not text:
        return None
    text = text.replace("%", "").replace(",", ".")
    try:
        parsed = Decimal(text)
    except (InvalidOperation, ValueError):
        raise AgenteComparaTempTableError(
            ERROR_TEMP_TABLE_INVALID_PAYLOAD,
            "Alíquota fiscal inválida.",
        ) from None
    if parsed < 0:
        raise AgenteComparaTempTableError(
            ERROR_TEMP_TABLE_INVALID_PAYLOAD,
            "Alíquota fiscal inválida.",
        )
    return float(parsed)


def _normalize_tax_rate_for_compare(value) -> Decimal | None:
    if value is None:
        return None
    try:
        return Decimal(str(value)).quantize(Decimal("0.000001"))
    except (InvalidOperation, ValueError):
        return None


def _validate_tax_config_for_save(raw_tax_config) -> dict:
    if not isinstance(raw_tax_config, dict):
        raise AgenteComparaTempTableError(
            ERROR_TEMP_TABLE_INVALID_PAYLOAD,
            "tax_config deve ser um objeto.",
        )
    include_taxes = raw_tax_config.get("include_taxes")
    if include_taxes is not True:
        return {"include_taxes": False}

    origin_uf = _normalize_uf(raw_tax_config.get("origin_uf"))
    if not origin_uf:
        raise AgenteComparaTempTableError(
            ERROR_TEMP_TABLE_INVALID_PAYLOAD,
            "UF origem é obrigatória para incluir impostos.",
        )
    origin_city = _sanitize_cell_string(raw_tax_config.get("origin_city"))
    iss_rate = _parse_optional_rate(raw_tax_config.get("iss_rate"))
    if iss_rate is not None and not origin_city:
        raise AgenteComparaTempTableError(
            ERROR_TEMP_TABLE_INVALID_PAYLOAD,
            "Cidade origem é obrigatória quando ISS estiver preenchido.",
        )

    raw_rates = raw_tax_config.get("icms_rates") or []
    if not isinstance(raw_rates, list):
        raise AgenteComparaTempTableError(
            ERROR_TEMP_TABLE_INVALID_PAYLOAD,
            "tax_config.icms_rates deve ser uma lista.",
        )
    incoming_rates: dict[str, float | None] = {}
    for item in raw_rates:
        if not isinstance(item, dict):
            raise AgenteComparaTempTableError(
                ERROR_TEMP_TABLE_INVALID_PAYLOAD,
                "Cada item de tax_config.icms_rates deve ser um objeto.",
            )
        destination_uf = _normalize_uf(item.get("destination_uf"))
        if not destination_uf:
            continue
        incoming_rates[destination_uf] = _parse_optional_rate(item.get("applied_rate"))

    raw_destination_ufs = raw_tax_config.get("destination_ufs")
    destination_ufs: list[dict] | None = None
    if raw_destination_ufs is not None:
        if not isinstance(raw_destination_ufs, list):
            raise AgenteComparaTempTableError(
                ERROR_TEMP_TABLE_INVALID_PAYLOAD,
                "tax_config.destination_ufs deve ser uma lista.",
            )
        destination_ufs = []
        seen_destination_ufs: set[str] = set()
        for item in raw_destination_ufs:
            normalized = _normalize_tax_destination_entry(item)
            if not normalized:
                raise AgenteComparaTempTableError(
                    ERROR_TEMP_TABLE_INVALID_PAYLOAD,
                    "Cada item de tax_config.destination_ufs deve conter uma UF válida.",
                )
            if normalized["uf"] in seen_destination_ufs:
                continue
            seen_destination_ufs.add(normalized["uf"])
            destination_ufs.append(normalized)

    return {
        "include_taxes": True,
        "origin_uf": origin_uf,
        "origin_city": origin_city or None,
        "iss_rate": iss_rate,
        "_incoming_icms_rates": incoming_rates,
        "_destination_ufs": destination_ufs,
    }


def build_tax_config_for_temp_table(record: dict, validated_tax_config: dict) -> dict:
    if not validated_tax_config.get("include_taxes"):
        return {"include_taxes": False}

    origin_uf = validated_tax_config["origin_uf"]
    incoming_rates = validated_tax_config.get("_incoming_icms_rates") or {}
    submitted_destination_ufs = validated_tax_config.get("_destination_ufs")
    destination_ufs = consolidate_tax_destination_ufs(
        record,
        submitted_destination_ufs=submitted_destination_ufs,
    )
    icms_rates: list[dict] = []
    for destination_entry in destination_ufs:
        destination_uf = destination_entry["uf"]
        if destination_uf == origin_uf:
            operation_type = "intermunicipal"
            suggested_rate = None
            source_name = ICMS_INTERMUNICIPAL_SOURCE_NAME
            source_type = "manual"
        else:
            operation_type = "interstate"
            suggested_rate = suggested_icms_interstate_rate(origin_uf, destination_uf)
            source_name = ICMS_INTERSTATE_SOURCE_NAME
            source_type = "official"

        has_incoming_rate = destination_uf in incoming_rates
        applied_rate = incoming_rates[destination_uf] if has_incoming_rate else suggested_rate
        suggested_cmp = _normalize_tax_rate_for_compare(suggested_rate)
        applied_cmp = _normalize_tax_rate_for_compare(applied_rate)
        user_edited = bool(has_incoming_rate and applied_cmp != suggested_cmp)
        icms_rates.append(
            {
                "destination_uf": destination_uf,
                "operation_type": operation_type,
                "suggested_rate": suggested_rate,
                "applied_rate": applied_rate,
                "source_name": source_name,
                "source_type": source_type,
                "user_edited": user_edited,
                "is_active": applied_rate is not None,
            }
        )

    return {
        "include_taxes": True,
        "origin_uf": origin_uf,
        "origin_city": validated_tax_config.get("origin_city"),
        "iss_rate": validated_tax_config.get("iss_rate"),
        "destination_ufs": destination_ufs,
        "icms_rates": icms_rates,
    }


def _validate_freight_routes_for_save(raw_routes) -> list[dict]:
    if raw_routes is None:
        return []
    if not isinstance(raw_routes, list):
        raise AgenteComparaTempTableError(
            ERROR_TEMP_TABLE_INVALID_PAYLOAD,
            "freight_routes deve ser uma lista.",
        )
    if not raw_routes:
        return []
    normalized = _normalize_freight_routes(raw_routes)
    if not normalized:
        raise AgenteComparaTempTableError(
            ERROR_TEMP_TABLE_INVALID_PAYLOAD,
            "freight_routes inválido.",
        )
    sanitized: list[dict] = []
    for route in normalized:
        sanitized.append(
            {
                key: _sanitize_cell_string(val) if val is not None else None
                for key, val in route.items()
            }
        )
    return sanitized


ACCESSORIAL_FEE_LEGACY_FIELDS = (
    "name",
    "value",
    "unit",
    "calculation_basis",
    "notes",
    "scope",
)
ACCESSORIAL_FEE_CALCULATION_TYPES = {
    "invoice_percentage",
    "fixed_amount",
    "weight_rate",
    "weight",
    "weight_fraction",
    "freight_percentage",
    "minimum_amount",
    "maximum_amount",
    "conditional",
    "unknown",
}
ACCESSORIAL_FEE_CANONICAL_COMPONENTS = {
    "freight_value",
    "ad_valorem",
    "risk_management",
    "insurance",
    "toll",
    "administrative_fee",
    "operational_fee",
    "generic_accessorial",
}
ACCESSORIAL_FEE_CLASSIFICATION_CONFIDENCES = {"high", "medium", "low"}
ACCESSORIAL_FEE_STATUSES = {"calculable", "needs_review", "unsupported", "unknown"}
ACCESSORIAL_FEE_MODIFIER_TYPES = {"base_fee", "minimum_amount", "maximum_amount"}
ACCESSORIAL_FEE_OPTIONAL_FIELDS = (
    "rate",
    "amount",
    "minimum_amount",
    "maximum_amount",
    "conditions",
    "unsupported_reason",
    "calculation_base",
    "calculation_base_id",
    "calculation_base_label",
    "audit_variable",
    "operation",
    "operation_parameters",
    "classification_source",
    "classification_warning",
    "note_classification",
    "raw_calculation_basis",
    "source_block",
    "original_text",
    "evidence_ref",
)
ACCESSORIAL_FEE_NULLABLE_OPTIONAL_FIELDS = {
    "calculation_base_id",
    "calculation_base_label",
    "audit_variable",
    "operation",
    "classification_warning",
}


_ACCESSORIAL_CONDITIONAL_MARKERS = (
    "somente",
    "apenas",
    "acima de",
    "abaixo de",
    "conforme operacao",
    "sob consulta",
    "entrega agendada",
    "regiao norte",
)


_ACCESSORIAL_OPERATIONAL_UNIT_MARKERS = (
    "por dia",
    "por diaria",
    "por pallet",
    "kg dia",
    "por veiculo",
    "por ajudante",
    "por hora",
)


_ACCESSORIAL_INVOICE_PERCENT_BASIS_MARKERS = {
    "% nf",
    "% valor nf",
    "% nota fiscal",
    "percentual nf",
    "perc nf",
    "sobre nf",
    "sob nf",
    "s nf",
    "sobre nota",
    "sobre nota fiscal",
    "sobre valor da nota fiscal",
    "valor nf",
    "nf",
    "nota",
    "nota fiscal",
    "valor da nota",
    "valor da mercadoria",
    "valor da carga",
    "mercadoria",
    "carga",
    "por nota",
}


_ACCESSORIAL_FREIGHT_VALUE_PERCENT_MARKERS = {
    "frete valor",
    "frete valor %",
    "fv",
    "fv %",
    "f v",
    "f v %",
}


_BRAZILIAN_UF_CODES = (
    "AC",
    "AL",
    "AM",
    "AP",
    "BA",
    "CE",
    "DF",
    "ES",
    "GO",
    "MA",
    "MG",
    "MS",
    "MT",
    "PA",
    "PB",
    "PE",
    "PI",
    "PR",
    "RJ",
    "RN",
    "RO",
    "RR",
    "RS",
    "SC",
    "SE",
    "SP",
    "TO",
)


def _accessorial_text_parts(item: dict) -> dict[str, str]:
    return {
        field: _sanitize_cell_string(item.get(field)) or ""
        for field in (
            "name",
            "value",
            "unit",
            "calculation_basis",
            "notes",
            "scope",
            "raw_calculation_basis",
            "original_text",
        )
    }


def _accessorial_joined_text(parts: dict[str, str], fields: tuple[str, ...] | None = None) -> str:
    selected = fields or tuple(parts.keys())
    return " ".join(parts.get(field, "") for field in selected if parts.get(field, "")).strip()


def _accessorial_normalized_text(parts: dict[str, str], fields: tuple[str, ...] | None = None) -> str:
    return _normalize_coverage_header(_accessorial_joined_text(parts, fields))


def _accessorial_contains_any(text: str, markers: tuple[str, ...] | set[str]) -> bool:
    normalized_markers = tuple(_normalize_coverage_header(marker) for marker in markers)
    return any(marker and marker in text for marker in normalized_markers)


def _accessorial_token_set(text: str) -> set[str]:
    return {token for token in text.split() if token}


def _accessorial_name_fragments(text: str) -> set[str]:
    fragments = set(_accessorial_token_set(text))
    for chunk in text.replace("_", "-").split("-"):
        chunk = chunk.strip()
        if chunk:
            fragments.add(chunk)
    return fragments


def _infer_accessorial_component_group(parts: dict[str, str]) -> str | None:
    text = _accessorial_normalized_text(parts, ("name", "notes", "original_text"))
    fragments = _accessorial_name_fragments(text)
    if "gris" in fragments:
        return "gris"
    if "trt" in fragments:
        return "trt"
    if "tde" in fragments:
        return "tde"
    if {"agendamento", "agendada", "agendado"} & fragments:
        return "agendamento"
    return None


def _infer_accessorial_modifier_type(calculation_type: str) -> str | None:
    if calculation_type == "minimum_amount":
        return "minimum_amount"
    if calculation_type == "maximum_amount":
        return "maximum_amount"
    if calculation_type in {
        "invoice_percentage",
        "freight_percentage",
        "fixed_amount",
        "weight_rate",
        "weight",
        "weight_fraction",
    }:
        return "base_fee"
    return None


def _accessorial_has_freight_value_alias(text: str) -> bool:
    if _accessorial_contains_any(text, _ACCESSORIAL_FREIGHT_VALUE_PERCENT_MARKERS):
        return True
    return bool(re.search(r"\bf\s*v\b", text))


def _accessorial_has_invoice_percent_basis(text: str) -> bool:
    if _accessorial_contains_any(
        text,
        {
            "percentual nf",
            "perc nf",
            "sobre nf",
            "sob nf",
            "s nf",
            "sobre nota",
            "sobre nota fiscal",
            "sobre valor da nota fiscal",
            "valor da nota",
            "valor da mercadoria",
            "valor da carga",
            "mercadoria",
            "carga",
            "por nota",
        },
    ):
        return True
    return bool(re.search(r"\b(?:nf|nota|nota fiscal|valor nf)\b", text))


def _classify_accessorial_canonical_component(parts: dict[str, str]) -> str:
    text = _accessorial_normalized_text(parts, ("name", "notes", "original_text"))
    if "gris" in text or "gerenciamento de risco" in text or "taxa de risco" in text or "risco" in text:
        return "risk_management"
    if "seguro" in text or "rctr c" in text or "rctrc" in text:
        return "insurance"
    if _accessorial_has_freight_value_alias(text):
        return "freight_value"
    if "ad valorem" in text or "advalorem" in text or "ad val" in text:
        return "ad_valorem"
    if "pedagio" in text:
        return "toll"
    if "tas" in text or "taxa administrativa" in text:
        return "administrative_fee"
    if "tso" in text or "taxa operacional" in text:
        return "operational_fee"
    return "generic_accessorial"


def _accessorial_has_percent_signal(parts: dict[str, str]) -> bool:
    if _has_percent_marker(
        parts.get("name"),
        parts.get("value"),
        parts.get("unit"),
        parts.get("calculation_basis"),
        parts.get("notes"),
        parts.get("original_text"),
    ):
        return True
    text = _accessorial_normalized_text(parts)
    return _accessorial_has_invoice_percent_basis(text) or _accessorial_has_freight_value_alias(text)


def _accessorial_has_money_signal(parts: dict[str, str]) -> bool:
    raw = _accessorial_joined_text(parts)
    normalized_unit = _normalize_coverage_header(parts.get("unit"))
    return "R$" in raw or normalized_unit in {"r", "rs", "brl"}


def _parse_single_accessorial_money(value, *, allow_bare_number: bool = False) -> Decimal | None:
    text = _sanitize_cell_string(value)
    if not text:
        return None
    money_match = re.search(r"(?i)R\$\s*\d{1,3}(?:\.\d{3})*(?:,\d{1,4})?|\d+(?:[,.]\d{1,4})?\s*(?:reais|real)\b", text)
    if money_match:
        return _decimal_money(money_match.group(0))
    if not allow_bare_number:
        return None
    numbers = re.findall(r"\d+(?:[,.]\d+)?", text)
    if len(numbers) != 1:
        return None
    return _decimal_money(numbers[0])


def _accessorial_first_money(parts: dict[str, str]) -> Decimal | None:
    value_allows_bare_number = bool(parts.get("unit")) or _accessorial_contains_any(
        _accessorial_normalized_text(parts, ("calculation_basis", "name")),
        {"conhecimento", "cte", "ct e", "nf", "nota fiscal", "entrega", "ocorrencia", "documento", "tde"},
    )
    parsed = _parse_single_accessorial_money(parts.get("value"), allow_bare_number=value_allows_bare_number)
    if parsed is not None:
        return parsed
    for field in ("original_text", "name"):
        parsed = _parse_single_accessorial_money(parts.get(field), allow_bare_number=True)
        if parsed is not None:
            return parsed
    return None


def _accessorial_first_rate(parts: dict[str, str]) -> Decimal | None:
    for field in ("value", "notes", "original_text", "name"):
        parsed = _parse_percentage_rate(parts.get(field))
        if parsed is not None:
            return parsed
    return None


def _accessorial_unit_for_configured_base(parts: dict[str, str]) -> str:
    unit = _sanitize_cell_string(parts.get("unit")) or ""
    if unit.strip():
        return unit
    if _has_percent_marker(parts.get("value")):
        return "%"
    return unit


def _accessorial_percent_values(parts: dict[str, str]) -> list[str]:
    text = _accessorial_joined_text(parts)
    matches = re.findall(r"\d+(?:[,.]\d+)?\s*%", text)
    return list(dict.fromkeys(_normalize_coverage_header(match) for match in matches))


def _accessorial_percent_decimals(parts: dict[str, str], fields: tuple[str, ...] | None = None) -> list[Decimal]:
    values = _accessorial_explicit_percent_values(_accessorial_joined_text(parts, fields))
    unique: list[Decimal] = []
    for value in values:
        if not any(value == existing for existing in unique):
            unique.append(value)
    return unique


def _accessorial_has_multiple_material_percentages(parts: dict[str, str]) -> bool:
    values = _accessorial_percent_decimals(parts, ("value", "notes", "original_text", "name"))
    if len(values) <= 1:
        return False
    first = values[0].quantize(Decimal("0.000001"))
    return any(value.quantize(Decimal("0.000001")) != first for value in values[1:])


def _accessorial_has_uf_condition(parts: dict[str, str]) -> bool:
    text = _accessorial_joined_text(parts)
    if not text:
        return False
    uf_pattern = "|".join(_BRAZILIAN_UF_CODES)
    return bool(re.search(rf"\b(?:{uf_pattern})(?:\s*/\s*(?:{uf_pattern}))?\b", text.upper()))


def _accessorial_has_operational_condition(parts: dict[str, str]) -> bool:
    condition_text = _accessorial_normalized_text(parts, ("notes", "scope", "original_text"))
    full_text = _accessorial_normalized_text(parts)
    return bool(
        _accessorial_contains_any(condition_text, _ACCESSORIAL_CONDITIONAL_MARKERS)
        or _accessorial_has_uf_condition(parts)
        or _accessorial_has_non_priced_day_condition(parts)
        or _accessorial_has_multiple_material_percentages(parts)
        or _accessorial_contains_any(
            full_text,
            {"mercadoria", "modalidade", "modal", "prazo", "origem especifica", "destino especifico"},
        )
    )


def _accessorial_note_classification(parts: dict[str, str]) -> str | None:
    if _accessorial_has_operational_condition(parts):
        return "unsupported_operational_condition"
    text = _accessorial_normalized_text(parts, ("notes", "original_text", "value", "name"))
    if "minimo" in text or "minima" in text or "cobranca minima" in text:
        return "structured_minimum_rule"
    if _accessorial_joined_text(parts, ("notes", "original_text")).strip():
        return "descriptive_notes"
    return None


def _accessorial_is_compound_rule(parts: dict[str, str]) -> bool:
    all_text = _accessorial_normalized_text(parts)
    has_percent = _accessorial_has_percent_signal(parts)
    percent_values = _accessorial_percent_values(parts)
    has_minimum_or_maximum = (
        "minimo" in all_text
        or "minima" in all_text
        or "cobranca minima" in all_text
        or "teto" in all_text
        or "maximo" in all_text
    )
    return bool(
        (has_percent and has_minimum_or_maximum)
        or _accessorial_has_multiple_material_percentages(parts)
        or (has_percent and _accessorial_has_uf_condition(parts))
    )


def _accessorial_has_non_priced_day_condition(parts: dict[str, str]) -> bool:
    text = _accessorial_normalized_text(parts, ("notes", "original_text"))
    return bool(re.search(r"\bapos\b.*\bdia\b", text) or re.search(r"\b\d+\s*o?\s*dia\b", text))


def _classify_accessorial_calculation_type(parts: dict[str, str]) -> tuple[str, str, str | None]:
    all_text = _accessorial_normalized_text(parts)
    basis_text = _accessorial_normalized_text(
        parts,
        ("name", "calculation_basis", "notes", "original_text", "value", "unit"),
    )
    name_text = _accessorial_normalized_text(parts, ("name",))
    percent_signal = _accessorial_has_percent_signal(parts)
    money_signal = _accessorial_has_money_signal(parts)

    has_condition = _accessorial_contains_any(all_text, _ACCESSORIAL_CONDITIONAL_MARKERS)
    minimum_signal = "minimo" in all_text or "minima" in all_text or "cobranca minima" in all_text
    maximum_signal = "teto" in all_text or "maximo" in all_text
    operational_signal = _accessorial_contains_any(all_text, _ACCESSORIAL_OPERATIONAL_UNIT_MARKERS)
    compound_signal = _accessorial_is_compound_rule(parts)

    if has_condition:
        return "conditional", "medium", "textual_condition"
    if _accessorial_has_non_priced_day_condition(parts) and not money_signal:
        return "conditional", "medium", "missing_monetary_amount"
    value_only_parts = {
        **parts,
        "name": "",
        "notes": "",
        "original_text": "",
        "calculation_basis": "",
        "raw_calculation_basis": "",
    }
    if minimum_signal and money_signal and not _accessorial_has_percent_signal(value_only_parts):
        confidence = "high" if _accessorial_first_money(parts) is not None else "medium"
        return "minimum_amount", confidence, None
    if compound_signal:
        if percent_signal and _accessorial_has_invoice_percent_basis(basis_text):
            return "invoice_percentage", "medium", "compound_accessorial_rule"
        return "conditional", "medium", "compound_accessorial_rule"
    if minimum_signal and maximum_signal:
        return "conditional", "medium", "combined_minimum_maximum"
    if minimum_signal:
        confidence = "high" if money_signal else "medium"
        return "minimum_amount", confidence, None
    if maximum_signal:
        confidence = "high" if money_signal else "medium"
        return "maximum_amount", confidence, None

    if percent_signal and _accessorial_has_invoice_percent_basis(basis_text):
        return "invoice_percentage", "high", None

    if percent_signal and _accessorial_has_freight_value_alias(name_text):
        return "invoice_percentage", "high", None

    if percent_signal and _accessorial_contains_any(
        basis_text,
        {"frete", "valor frete", "valor do frete", "sobre frete"},
    ):
        return "freight_percentage", "high", None

    if operational_signal and money_signal:
        return "fixed_amount", "medium", "operational_unit_rate"

    if money_signal and _accessorial_contains_any(basis_text, {"kg", "quilo", "quilograma"}):
        return "weight_rate", "high", None

    if money_signal and _accessorial_contains_any(
        basis_text,
        {"conhecimento", "cte", "ct e", "nf", "nota fiscal", "entrega", "ocorrencia", "documento"},
    ):
        return "fixed_amount", "high", None

    if _accessorial_first_money(parts) is not None and _accessorial_contains_any(name_text, {"tde"}):
        return "fixed_amount", "medium", "missing_application_basis"

    if percent_signal:
        if _accessorial_contains_any(name_text, {"gris", "seguro", "ad valorem", "advalorem"}) or _accessorial_has_freight_value_alias(name_text):
            return "invoice_percentage", "high", None
        if _accessorial_contains_any(name_text, {"tso", "taxa operacional"}):
            return "freight_percentage", "low", None
        return "unknown", "low", None

    if money_signal:
        if _accessorial_contains_any(name_text, {"tas", "taxa administrativa", "pedagio"}):
            return "fixed_amount", "low", None
        return "unknown", "low", None

    if operational_signal:
        return "conditional", "medium", "missing_monetary_amount"

    return "unknown", "low", None


def _configured_base_fields_from_base(base: dict, *, source: str) -> dict:
    label = _sanitize_cell_string(base.get("label"))
    calculation_type = _sanitize_cell_string(base.get("calculation_type")) or "unknown"
    return {
        "calculation_basis": label,
        "calculation_base_id": _sanitize_cell_string(base.get("id")),
        "calculation_base_label": label,
        "calculation_type": calculation_type,
        "audit_variable": _sanitize_cell_string(base.get("audit_variable")),
        "operation": _sanitize_cell_string(base.get("operation")),
        "operation_parameters": base.get("parameters") if isinstance(base.get("parameters"), dict) else {},
        "classification_source": source,
        "classification_confidence": "high",
        "modifier_type": _infer_accessorial_modifier_type(calculation_type),
    }


def _unmapped_calculation_base_fields(*, warning: str | None = None) -> dict:
    fields = {
        "calculation_basis": UNMAPPED_CALCULATION_BASIS_LABEL,
        "calculation_base_id": None,
        "calculation_base_label": None,
        "audit_variable": None,
        "operation": None,
        "operation_parameters": {},
        "classification_source": "unmapped_calculation_base",
    }
    if warning:
        fields["classification_warning"] = warning
    return fields


def _accessorial_modifier_link_refs(fee: dict) -> set[str]:
    refs: set[str] = set()
    for field in ("component_group", "canonical_component", "related_to"):
        value = fee.get(field)
        if value and value != "generic_accessorial":
            refs.add(str(value))
    return refs


def _accessorial_base_link_ref(fee: dict) -> str | None:
    group = fee.get("component_group")
    if group:
        return str(group)
    canonical = fee.get("canonical_component")
    if canonical and canonical != "generic_accessorial":
        return str(canonical)
    return None


def _finalize_accessorial_minimum_modifier(item: dict, parts: dict[str, str], derived: dict) -> None:
    modifier_type = derived.get("modifier_type")
    if modifier_type != "minimum_amount":
        return

    explicit_modifier = _normalize_accessorial_modifier_type(item.get("modifier_type"))
    derived["calculation_type"] = "minimum_amount"
    derived["modifier_type"] = "minimum_amount"
    for field in ("unsupported_reason", "conditions", "rate", "amount", "classification_warning"):
        derived.pop(field, None)

    amount = _decimal_money(item.get("minimum_amount"))
    if amount is None:
        amount = _accessorial_first_money(parts)
    if amount is not None:
        derived["minimum_amount"] = _round_money(amount)

    basis = str(parts.get("calculation_basis") or "").strip().lower()
    if basis == UNMAPPED_CALCULATION_BASIS_LABEL.lower():
        derived["calculation_basis"] = ""

    related_to = derived.get("related_to") or _normalize_accessorial_component_ref(item.get("related_to"))
    if related_to:
        derived["related_to"] = related_to

    if explicit_modifier != "minimum_amount":
        return

    if derived.get("minimum_amount") is not None:
        derived["classification_confidence"] = "high"
        derived["status"] = "calculable" if related_to else "needs_review"
    else:
        derived["status"] = "needs_review"


def _derive_accessorial_fee_fields(item: dict) -> dict:
    parts = _accessorial_text_parts(item)
    calculation_type, confidence, unsupported_reason = _classify_accessorial_calculation_type(parts)
    note_classification = _accessorial_note_classification(parts)
    canonical_component = _classify_accessorial_canonical_component(parts)
    explicit_group = _normalize_accessorial_component_ref(item.get("component_group"))
    component_group = explicit_group or _infer_accessorial_component_group(parts)
    if component_group is None and canonical_component not in {None, "", "generic_accessorial"}:
        component_group = canonical_component
    modifier_type = _normalize_accessorial_modifier_type(item.get("modifier_type"))
    if modifier_type is None:
        modifier_type = _infer_accessorial_modifier_type(calculation_type)
    explicit_related = _normalize_accessorial_component_ref(item.get("related_to"))
    derived = {
        "calculation_type": calculation_type,
        "canonical_component": canonical_component,
        "classification_confidence": confidence,
        "component_group": component_group,
        "modifier_type": modifier_type,
        "related_to": explicit_related,
    }
    if unsupported_reason:
        derived["unsupported_reason"] = unsupported_reason
    if note_classification:
        derived["note_classification"] = note_classification

    has_unmapped_basis = str(parts.get("calculation_basis") or "").strip().lower() == UNMAPPED_CALCULATION_BASIS_LABEL
    has_explicit_base_id = bool(item.get("_has_explicit_calculation_base_id")) and _optional_normalized_str(
        item.get("calculation_base_id")
    ) is not None
    try:
        calculation_bases = get_agente_compara_config().calculation_bases
        configured_base = (
            get_active_calculation_base_by_id(item.get("calculation_base_id"), calculation_bases)
            if has_explicit_base_id
            else None
        )
        configured_base_result = (
            {"status": "matched", "base": configured_base}
            if configured_base is not None
            else (
                {"status": "unmapped", "base": None}
                if has_unmapped_basis
                else (
                {"status": "invalid_id", "base": None}
                if has_explicit_base_id
                else resolve_calculation_base_status(
                    parts.get("calculation_basis"),
                    _accessorial_unit_for_configured_base(parts),
                    calculation_bases,
                )
                )
            )
        )
    except Exception:
        logger.exception("Falha ao resolver base de cálculo configurada da Agente Compara.")
        configured_base_result = {"status": "not_found", "base": None}

    configured_base = configured_base_result.get("base")
    if configured_base_result.get("status") == "matched" and isinstance(configured_base, dict):
        configured_calculation_type = str(configured_base.get("calculation_type") or calculation_type).strip()
        should_override_type = calculation_type not in {
            "minimum_amount",
            "maximum_amount",
            "conditional",
        }
        if should_override_type:
            calculation_type = configured_calculation_type
        source = (
            "manual_configured_calculation_base"
            if item.get("classification_source") == "manual_configured_calculation_base"
            else "configured_calculation_base"
        )
        derived.update(_configured_base_fields_from_base(configured_base, source=source))
        derived["calculation_type"] = calculation_type
        derived["modifier_type"] = _infer_accessorial_modifier_type(calculation_type)
        derived["classification_confidence"] = "high" if should_override_type else confidence
        if should_override_type:
            confidence = "high"
    elif has_explicit_base_id or has_unmapped_basis:
        if calculation_type in {"minimum_amount", "maximum_amount"}:
            derived.setdefault("classification_source", "legacy_classifier")
            if has_unmapped_basis:
                derived["calculation_basis"] = ""
            if has_explicit_base_id and configured_base_result.get("status") != "matched":
                derived["classification_warning"] = "invalid_calculation_base_id"
                derived["status"] = "needs_review"
        else:
            derived.update(
                _unmapped_calculation_base_fields(
                    warning="invalid_calculation_base_id" if has_explicit_base_id else None
                )
            )
            calculation_type = "unknown"
            confidence = "low"
            unsupported_reason = None
    else:
        derived["calculation_base_id"] = None
        derived["classification_source"] = "legacy_classifier"
        if configured_base_result.get("status") == "ambiguous":
            derived["classification_warning"] = "ambiguous_calculation_base"

    if calculation_type == "conditional":
        derived["status"] = "unsupported"
        condition_text = _accessorial_joined_text(parts, ("notes", "scope", "original_text", "value"))
        if condition_text:
            derived["conditions"] = condition_text
    elif calculation_type == "unknown":
        derived["status"] = "unknown"
    elif confidence == "high":
        derived["status"] = "calculable"
    else:
        derived["status"] = "needs_review"

    if unsupported_reason == "compound_accessorial_rule":
        amount = _accessorial_first_money(parts)
        normalized_text = _accessorial_normalized_text(parts)
        if amount is not None and ("minimo" in normalized_text or "minima" in normalized_text):
            derived["minimum_amount"] = _round_money(amount)
        if amount is not None and ("teto" in normalized_text or "maximo" in normalized_text):
            derived["maximum_amount"] = _round_money(amount)
        if note_classification == "unsupported_operational_condition":
            condition_text = _accessorial_joined_text(parts, ("notes", "scope", "original_text", "value"))
            if condition_text:
                derived["conditions"] = condition_text

    if calculation_type in {"invoice_percentage", "freight_percentage"}:
        rate = _accessorial_first_rate(parts)
        if rate is not None:
            derived["rate"] = float(rate)
        if unsupported_reason == "compound_accessorial_rule":
            amount = _accessorial_first_money(parts)
            if amount is not None and "minimo" in _accessorial_normalized_text(parts):
                derived["minimum_amount"] = _round_money(amount)
            if note_classification == "unsupported_operational_condition":
                condition_text = _accessorial_joined_text(parts, ("notes", "scope", "original_text", "value"))
                if condition_text:
                    derived["conditions"] = condition_text
        if (
            configured_base_result.get("status") == "matched"
            and note_classification != "unsupported_operational_condition"
            and rate is not None
            and str(derived.get("operation") or "").strip() in _SUPPORTED_ACCESSORIAL_ADVANCE_OPERATIONS
            and (
                str(derived.get("operation") or "").strip() != "percentage_of_variable"
                or str(derived.get("audit_variable") or "").strip()
            )
        ):
            derived.pop("unsupported_reason", None)
            derived.pop("conditions", None)
            derived["classification_confidence"] = "high"
            derived["status"] = "calculable"
    elif calculation_type == "fixed_amount":
        amount = _accessorial_first_money(parts)
        if amount is not None:
            derived["amount"] = _round_money(amount)
    elif calculation_type == "minimum_amount":
        amount = _accessorial_first_money(parts)
        if amount is not None:
            derived["minimum_amount"] = _round_money(amount)
    elif calculation_type == "maximum_amount":
        amount = _accessorial_first_money(parts)
        if amount is not None:
            derived["maximum_amount"] = _round_money(amount)

    _finalize_accessorial_minimum_modifier(item, parts, derived)
    return derived


def _accessorial_confidence_rank(value) -> int:
    return {"low": 1, "medium": 2, "high": 3}.get(str(value or "").lower(), 0)


ACCESSORIAL_FEE_DERIVED_FIELDS = (
    "calculation_type",
    "calculation_base_id",
    "calculation_base_label",
    "audit_variable",
    "operation",
    "operation_parameters",
    "classification_source",
    "classification_warning",
    "note_classification",
    "canonical_component",
    "classification_confidence",
    "status",
    "rate",
    "amount",
    "minimum_amount",
    "maximum_amount",
    "conditions",
    "unsupported_reason",
    "component_group",
    "modifier_type",
    "related_to",
)


def _apply_accessorial_fee_classification(item: dict) -> dict:
    classified = dict(item)
    derived = _derive_accessorial_fee_fields(classified)

    for field in ACCESSORIAL_FEE_DERIVED_FIELDS:
        classified.pop(field, None)
    for field in ACCESSORIAL_FEE_DERIVED_FIELDS:
        if field in derived:
            classified[field] = derived[field]
    if "calculation_basis" in derived:
        classified["calculation_basis"] = derived["calculation_basis"]
    classified.pop("_has_explicit_calculation_base_id", None)
    classified.setdefault("source_block", "accessorial_fees")
    return classified


def _normalize_accessorial_choice(value, allowed: set[str], default: str) -> str | None:
    candidate = _optional_normalized_str(value)
    if candidate is None:
        return default
    normalized = candidate.lower()
    return normalized if normalized in allowed else default


def _normalize_accessorial_extra_value(value):
    if value is None:
        return None
    if isinstance(value, str):
        return _sanitize_cell_string(value)
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        return value
    if isinstance(value, bool):
        return str(value)
    if isinstance(value, list):
        cleaned = [_normalize_accessorial_extra_value(item) for item in value]
        return [item for item in cleaned if item is not None]
    if isinstance(value, dict):
        cleaned: dict[str, object] = {}
        for key, item in value.items():
            safe_key = _sanitize_cell_string(key)
            if not safe_key:
                continue
            safe_value = _normalize_accessorial_extra_value(item)
            if safe_value is not None:
                cleaned[safe_key] = safe_value
        return cleaned
    return _sanitize_cell_string(value)


def _normalize_accessorial_modifier_type(value) -> str | None:
    candidate = _optional_normalized_str(value)
    if candidate is None:
        return None
    normalized = candidate.lower()
    return normalized if normalized in ACCESSORIAL_FEE_MODIFIER_TYPES else None


def _normalize_accessorial_component_ref(value) -> str | None:
    candidate = _optional_normalized_str(value)
    if candidate is None:
        return None
    normalized = _normalize_coverage_header(candidate).replace(" ", "_")
    return normalized or None


def _normalize_accessorial_fee_item(item) -> dict | None:
    if isinstance(item, str):
        text = _optional_normalized_str(item)
        if text is None:
            return None
        return {
            "name": text,
            "value": None,
            "unit": None,
            "calculation_basis": None,
            "notes": "",
            "scope": None,
        }
    if not isinstance(item, dict):
        return None
    normalized = {
        "name": _optional_normalized_str(item.get("name")),
        "value": _optional_normalized_str(item.get("value")),
        "unit": _optional_normalized_str(item.get("unit")),
        "calculation_basis": _optional_normalized_str(item.get("calculation_basis")),
        "notes": _optional_normalized_str(item.get("notes")) or "",
        "scope": _optional_normalized_str(item.get("scope")),
    }
    if "raw_calculation_basis" in item:
        normalized["raw_calculation_basis"] = _optional_normalized_str(item.get("raw_calculation_basis"))
    if "calculation_type" in item:
        normalized["calculation_type"] = _normalize_accessorial_choice(
            item.get("calculation_type"),
            ACCESSORIAL_FEE_CALCULATION_TYPES,
            "unknown",
        )
    if "canonical_component" in item:
        normalized["canonical_component"] = _normalize_accessorial_choice(
            item.get("canonical_component"),
            ACCESSORIAL_FEE_CANONICAL_COMPONENTS,
            "generic_accessorial",
        )
    if "classification_confidence" in item:
        normalized["classification_confidence"] = _normalize_accessorial_choice(
            item.get("classification_confidence"),
            ACCESSORIAL_FEE_CLASSIFICATION_CONFIDENCES,
            "low",
        )
    if "status" in item:
        normalized["status"] = _normalize_accessorial_choice(
            item.get("status"),
            ACCESSORIAL_FEE_STATUSES,
            "unknown",
        )
    if "component_group" in item:
        normalized["component_group"] = _normalize_accessorial_component_ref(item.get("component_group"))
    if "modifier_type" in item:
        normalized["modifier_type"] = _normalize_accessorial_modifier_type(item.get("modifier_type"))
    if "related_to" in item:
        normalized["related_to"] = _normalize_accessorial_component_ref(item.get("related_to"))
    for field in ACCESSORIAL_FEE_OPTIONAL_FIELDS:
        if field not in item:
            continue
        value = _normalize_accessorial_extra_value(item.get(field))
        if value is not None:
            normalized[field] = value
    if "calculation_base_id" in item:
        normalized["_has_explicit_calculation_base_id"] = True
    return _apply_accessorial_fee_classification(normalized)


def _link_accessorial_fee_modifiers(fees: list[dict]) -> list[dict]:
    base_fees = [fee for fee in fees if fee.get("modifier_type") == "base_fee"]
    base_ref_sets = [_accessorial_modifier_link_refs(fee) for fee in base_fees]
    all_base_refs: set[str] = set()
    for refs in base_ref_sets:
        all_base_refs.update(refs)

    linked: list[dict] = []
    for fee in fees:
        item = dict(fee)
        modifier_type = item.get("modifier_type")
        if modifier_type == "base_fee":
            item["related_to"] = None
            linked.append(item)
            continue
        if modifier_type not in {"minimum_amount", "maximum_amount"}:
            linked.append(item)
            continue

        explicit_related = item.get("related_to")
        item_refs = _accessorial_modifier_link_refs(item)
        compatible_indices = [
            idx for idx, base_refs in enumerate(base_ref_sets) if item_refs & base_refs
        ]

        if len(compatible_indices) == 1:
            base_fee = base_fees[compatible_indices[0]]
            shared_refs = item_refs & base_ref_sets[compatible_indices[0]]
            preferred = _accessorial_base_link_ref(base_fee)
            if preferred and preferred in shared_refs:
                item["related_to"] = preferred
            elif shared_refs:
                item["related_to"] = sorted(shared_refs)[0]
            elif preferred:
                item["related_to"] = preferred
            else:
                item["related_to"] = None
            if not item.get("component_group"):
                item["component_group"] = base_fee.get("component_group") or base_fee.get("canonical_component")
        elif explicit_related and explicit_related in all_base_refs:
            item["related_to"] = explicit_related
            if not item.get("component_group"):
                for idx, base_refs in enumerate(base_ref_sets):
                    if explicit_related in base_refs:
                        item["component_group"] = (
                            base_fees[idx].get("component_group") or explicit_related
                        )
                        break
        else:
            item["related_to"] = None
            if item.get("status") == "calculable":
                item["status"] = "needs_review"
        linked.append(item)
    return linked


def _normalize_accessorial_fees(raw_fees) -> list[dict]:
    if not isinstance(raw_fees, list):
        return []
    normalized: list[dict] = []
    for item in raw_fees:
        fee = _normalize_accessorial_fee_item(item)
        if fee is not None:
            normalized.append(fee)
    return _link_accessorial_fee_modifiers(normalized)


def _validate_accessorial_fees_for_save(raw_fees) -> list[dict]:
    if raw_fees is None:
        return []
    if not isinstance(raw_fees, list):
        raise AgenteComparaTempTableError(
            ERROR_TEMP_TABLE_INVALID_PAYLOAD,
            "accessorial_fees deve ser uma lista.",
        )
    if not raw_fees:
        return []
    normalized = _normalize_accessorial_fees(raw_fees)
    if len(normalized) != len(raw_fees):
        raise AgenteComparaTempTableError(
            ERROR_TEMP_TABLE_INVALID_PAYLOAD,
            "accessorial_fees inválido.",
        )
    sanitized: list[dict] = []
    for item in normalized:
        fee = {
            field: _sanitize_cell_string(item.get(field))
            for field in ACCESSORIAL_FEE_LEGACY_FIELDS
        }
        fee["notes"] = fee.get("notes") or ""
        for field, default, allowed in (
            ("calculation_type", "unknown", ACCESSORIAL_FEE_CALCULATION_TYPES),
            ("canonical_component", "generic_accessorial", ACCESSORIAL_FEE_CANONICAL_COMPONENTS),
            ("classification_confidence", "low", ACCESSORIAL_FEE_CLASSIFICATION_CONFIDENCES),
            ("status", "unknown", ACCESSORIAL_FEE_STATUSES),
        ):
            if field in item:
                fee[field] = _normalize_accessorial_choice(item.get(field), allowed, default)
        for field in ACCESSORIAL_FEE_OPTIONAL_FIELDS:
            if field not in item:
                continue
            value = _normalize_accessorial_extra_value(item.get(field))
            if value is None and field in ACCESSORIAL_FEE_NULLABLE_OPTIONAL_FIELDS:
                fee[field] = None
                continue
            if value is not None:
                fee[field] = value
        for field in ("component_group", "related_to"):
            if field in item:
                fee[field] = _normalize_accessorial_component_ref(item.get(field))
        if "modifier_type" in item:
            fee["modifier_type"] = _normalize_accessorial_modifier_type(item.get("modifier_type"))
        sanitized.append(fee)
    return sanitized


def _is_general_accessorial_fee_for_base_validation(fee: dict) -> bool:
    name = _normalize_coverage_header(fee.get("name"))
    if not name:
        return True
    return not (
        name.startswith("frete valor")
        or name.startswith("frete peso")
        or name.startswith("taxa embarque")
        or name.startswith("taxa de embarque")
    )


_SUPPORTED_ACCESSORIAL_ADVANCE_OPERATIONS = frozenset(
    {
        "fixed_amount",
        "percentage_of_variable",
        "multiply_by_variable",
        "ceil_fraction",
    }
)


def _accessorial_fee_has_required_value_for_operation(fee: dict) -> bool:
    operation = str(fee.get("operation") or "").strip()
    if operation == "percentage_of_variable":
        return _accessorial_runtime_rate(fee) is not None
    if operation in {"fixed_amount", "ceil_fraction", "multiply_by_variable"}:
        return _accessorial_runtime_amount(fee) is not None
    return False


def _accessorial_fee_unit_matches_base(fee: dict, active_bases_by_id: dict[str, dict]) -> bool:
    base_id = str(fee.get("calculation_base_id") or "").strip()
    base = active_bases_by_id.get(base_id)
    if not base:
        return False
    expected = normalize_calculation_base_unit(base.get("unit"))
    if not expected:
        return True
    return normalize_calculation_base_unit(fee.get("unit")) == expected


def _accessorial_fee_uses_new_base_contract(fee: dict) -> bool:
    base_id = str(fee.get("calculation_base_id") or "").strip()
    basis = str(fee.get("calculation_basis") or "").strip().lower()
    source = str(fee.get("classification_source") or "").strip()
    return bool(base_id) or basis == UNMAPPED_CALCULATION_BASIS_LABEL.lower() or source.startswith("manual_")


def _accessorial_fee_is_minimum_modifier(fee: dict) -> bool:
    modifier = str(fee.get("modifier_type") or "").strip()
    calculation_type = str(fee.get("calculation_type") or "").strip()
    return modifier == "minimum_amount" or calculation_type == "minimum_amount"


def _accessorial_fee_is_base_fee_for_minimum_link(fee: dict) -> bool:
    modifier = str(fee.get("modifier_type") or "").strip()
    calculation_type = str(fee.get("calculation_type") or "").strip()
    if modifier in {"minimum_amount", "maximum_amount"}:
        return False
    if calculation_type in {"minimum_amount", "maximum_amount"}:
        return False
    return True


def _find_accessorial_minimum_base_fee(
    minimum_fee: dict,
    accessorial_fees: list[dict],
    *,
    exclude_index: int | None = None,
) -> dict | None:
    related_to = str(minimum_fee.get("related_to") or "").strip()
    if not related_to:
        return None
    minimum_refs = _accessorial_modifier_link_refs(minimum_fee)
    minimum_refs.add(related_to)
    matches: list[dict] = []
    for idx, fee in enumerate(accessorial_fees):
        if exclude_index is not None and idx == exclude_index:
            continue
        if not isinstance(fee, dict) or not _accessorial_fee_is_base_fee_for_minimum_link(fee):
            continue
        base_refs = _accessorial_modifier_link_refs(fee)
        if not minimum_refs & base_refs or related_to not in base_refs:
            continue
        matches.append(fee)
    if len(matches) == 1:
        return matches[0]
    return None


def _accessorial_fee_has_valid_minimum_amount(fee: dict) -> bool:
    if _accessorial_runtime_minimum_amount(fee) is not None:
        return True
    return _accessorial_first_money(_accessorial_text_parts(fee)) is not None


def _accessorial_explicit_percent_values(text) -> list[Decimal]:
    source = _sanitize_cell_string(text) or ""
    values: list[Decimal] = []
    index = 0
    while index < len(source):
        percent_index = source.find("%", index)
        if percent_index < 0:
            break
        start = percent_index - 1
        while start >= 0 and source[start].isspace():
            start -= 1
        end = start + 1
        while start >= 0 and (
            source[start].isdigit()
            or source[start] in {".", ","}
        ):
            start -= 1
        token = source[start + 1:end]
        parsed = _parse_decimal_number(token)
        if parsed is not None:
            values.append(parsed)
        index = percent_index + 1
    unique: list[Decimal] = []
    for value in values:
        if not any(value == existing for existing in unique):
            unique.append(value)
    return unique


def _format_accessorial_percent(value: Decimal | None) -> str:
    if value is None:
        return ""
    return _format_brazilian_decimal(value)


def _accessorial_percent_materially_differs(left: Decimal, right: Decimal) -> bool:
    return left.quantize(Decimal("0.000001")) != right.quantize(Decimal("0.000001"))


def _accessorial_rate_conflict_for_advance(fee: dict, index: int) -> dict | None:
    unit = normalize_calculation_base_unit(fee.get("unit"))
    if unit != "%":
        return None
    structured_rate = _accessorial_runtime_rate(fee)
    if structured_rate is None:
        return None
    structured_percent = structured_rate * Decimal("100")
    described_values = _accessorial_explicit_percent_values(fee.get("notes"))
    if len(described_values) != 1:
        return None
    described_percent = described_values[0]
    if not _accessorial_percent_materially_differs(structured_percent, described_percent):
        return None
    name = _sanitize_cell_string(fee.get("name")) or f"Item {index + 1}"
    message = ERROR_ACCESSORIAL_RATE_CONFLICT_MESSAGE_TEMPLATE.format(
        name=name,
        structured_percent=_format_accessorial_percent(structured_percent),
        described_percent=_format_accessorial_percent(described_percent),
    )
    error = _accessorial_advance_validation_error(
        index=index,
        fee=fee,
        field="value",
        reason_code="accessorial_rate_conflict",
        message=message,
    )
    error["code"] = "accessorial_rate_conflict"
    error["severity"] = "blocking"
    error["structured_percent"] = float(structured_percent)
    error["described_percent"] = float(described_percent)
    error["related_fields"] = ["value", "notes"]
    return error


def _accessorial_fee_missing_calculation_base(
    fee: dict,
    active_bases_by_id: dict[str, dict],
) -> bool:
    base_id = str(fee.get("calculation_base_id") or "").strip()
    if not base_id or base_id not in active_bases_by_id:
        return True
    basis = str(fee.get("calculation_basis") or "").strip().lower()
    return basis == UNMAPPED_CALCULATION_BASIS_LABEL.lower()


def _accessorial_fee_operation_is_complete(fee: dict) -> bool:
    operation = str(fee.get("operation") or "").strip()
    if operation not in _SUPPORTED_ACCESSORIAL_ADVANCE_OPERATIONS:
        return False
    if operation in {"percentage_of_variable", "multiply_by_variable", "ceil_fraction"}:
        if not str(fee.get("audit_variable") or "").strip():
            return False
    if operation != "ceil_fraction":
        return True
    params = fee.get("operation_parameters")
    if not isinstance(params, dict):
        return False
    fraction_size = params.get("fraction_size")
    if fraction_size in (None, ""):
        return False
    try:
        size = Decimal(str(fraction_size).replace(",", "."))
    except Exception:
        return False
    return size > 0


def _accessorial_advance_validation_error(
    *,
    index: int,
    fee: dict,
    field: str,
    reason_code: str,
    message: str,
) -> dict:
    code_by_reason = {
        "missing_calculation_base": "unknown_calculation_basis",
        "unsupported_or_incomplete_operation": "unsupported_compound_rule",
        "invalid_accessorial_value": "invalid_accessorial_value",
        "incompatible_accessorial_unit": "incompatible_accessorial_unit",
        "missing_minimum_base_link": "minimum_without_base",
        "invalid_minimum_base_link": "minimum_without_base",
        "accessorial_rate_conflict": "accessorial_rate_conflict",
        "percentage_without_audit_variable": "percentage_without_audit_variable",
    }
    return {
        "code": code_by_reason.get(reason_code, reason_code),
        "section": "accessorial_fees",
        "index": index,
        "name": _sanitize_cell_string(fee.get("name")) or f"Item {index + 1}",
        "field": field,
        "reason_code": reason_code,
        "severity": "blocking",
        "message": message,
    }


def _validate_linked_minimum_amount_for_advance(
    fee: dict,
    index: int,
    accessorial_fees: list[dict],
    active_bases_by_id: dict[str, dict],
) -> dict | None:
    if not _accessorial_fee_has_valid_minimum_amount(fee):
        return _accessorial_advance_validation_error(
            index=index,
            fee=fee,
            field="value",
            reason_code="invalid_accessorial_value",
            message=ERROR_ACCESSORIAL_VALUE_MESSAGE,
        )
    related_to = str(fee.get("related_to") or "").strip()
    if not related_to:
        return _accessorial_advance_validation_error(
            index=index,
            fee=fee,
            field="related_to",
            reason_code="missing_minimum_base_link",
            message=ERROR_ACCESSORIAL_MINIMUM_LINK_MESSAGE,
        )
    base_fee = _find_accessorial_minimum_base_fee(fee, accessorial_fees, exclude_index=index)
    base_is_valid = False
    if base_fee is not None:
        base_error = None
        has_explicit_base_contract = bool(
            str(base_fee.get("operation") or "").strip()
            or str(base_fee.get("classification_source") or "").strip().startswith("manual_")
            or str(base_fee.get("classification_source") or "").strip() == "configured_calculation_base"
        )
        if has_explicit_base_contract and _accessorial_fee_uses_new_base_contract(base_fee):
            base_error = _validate_accessorial_fee_for_advance(
                base_fee,
                index,
                active_bases_by_id,
            )
        else:
            base_error = None if _accessorial_runtime_rate(base_fee) is not None else {}
        base_is_valid = base_error is None and _accessorial_rate_conflict_for_advance(base_fee, index) is None
    if not base_is_valid:
        return _accessorial_advance_validation_error(
            index=index,
            fee=fee,
            field="related_to",
            reason_code="invalid_minimum_base_link",
            message=ERROR_ACCESSORIAL_MINIMUM_LINK_MESSAGE,
        )
    return None


def _validate_accessorial_fee_for_advance(
    fee: dict,
    index: int,
    active_bases_by_id: dict[str, dict],
) -> dict | None:
    if _accessorial_fee_missing_calculation_base(fee, active_bases_by_id):
        return _accessorial_advance_validation_error(
            index=index,
            fee=fee,
            field="calculation_base_id",
            reason_code="missing_calculation_base",
            message=ERROR_ACCESSORIAL_CALCULATION_BASE_MESSAGE,
        )
    operation = str(fee.get("operation") or "").strip()
    if operation == "percentage_of_variable" and not str(fee.get("audit_variable") or "").strip():
        return _accessorial_advance_validation_error(
            index=index,
            fee=fee,
            field="calculation_base_id",
            reason_code="percentage_without_audit_variable",
            message=ERROR_ACCESSORIAL_OPERATION_MESSAGE,
        )
    if not _accessorial_fee_operation_is_complete(fee):
        return _accessorial_advance_validation_error(
            index=index,
            fee=fee,
            field="calculation_base_id",
            reason_code="unsupported_or_incomplete_operation",
            message=ERROR_ACCESSORIAL_OPERATION_MESSAGE,
        )
    if not _accessorial_fee_has_required_value_for_operation(fee):
        return _accessorial_advance_validation_error(
            index=index,
            fee=fee,
            field="value",
            reason_code="invalid_accessorial_value",
            message=ERROR_ACCESSORIAL_VALUE_MESSAGE,
        )
    if not _accessorial_fee_unit_matches_base(fee, active_bases_by_id):
        return _accessorial_advance_validation_error(
            index=index,
            fee=fee,
            field="unit",
            reason_code="incompatible_accessorial_unit",
            message=ERROR_ACCESSORIAL_UNIT_MESSAGE,
        )
    conflict = _accessorial_rate_conflict_for_advance(fee, index)
    if conflict is not None:
        return conflict
    return None


def _validate_accessorial_fees_ready_to_advance(accessorial_fees) -> None:
    if not isinstance(accessorial_fees, list):
        return
    active_bases = get_active_calculation_bases_for_runtime(
        get_agente_compara_config().calculation_bases
    )
    active_bases_by_id = {
        str(base.get("id") or "").strip(): base
        for base in active_bases
        if str(base.get("id") or "").strip()
    }
    errors: list[dict] = []
    for index, fee in enumerate(accessorial_fees):
        if not isinstance(fee, dict) or not _is_general_accessorial_fee_for_base_validation(fee):
            continue
        if _accessorial_fee_is_minimum_modifier(fee):
            error = _validate_linked_minimum_amount_for_advance(
                fee,
                index,
                accessorial_fees,
                active_bases_by_id,
            )
            if error is not None:
                errors.append(error)
            continue
        if not _accessorial_fee_uses_new_base_contract(fee):
            continue
        error = _validate_accessorial_fee_for_advance(fee, index, active_bases_by_id)
        if error is not None:
            errors.append(error)
    if errors:
        raise AgenteComparaTempTableError(
            ERROR_TEMP_TABLE_INVALID_ACCESSORIAL_FEES,
            ERROR_ACCESSORIAL_ADVANCE_MESSAGE,
            errors=errors,
        )


def _assert_temp_table_scope(record: dict, *, user_scope=None, franquia_scope=None) -> None:
    record_user = record.get("user_scope")
    if record_user is not None and user_scope is not None and record_user != user_scope:
        raise AgenteComparaTempTableError(
            ERROR_TEMP_TABLE_SCOPE_MISMATCH,
            "Escopo de usuário não autorizado para esta tabela temporária.",
        )
    record_franquia = record.get("franquia_scope")
    if (
        record_franquia is not None
        and franquia_scope is not None
        and record_franquia != franquia_scope
    ):
        raise AgenteComparaTempTableError(
            ERROR_TEMP_TABLE_SCOPE_MISMATCH,
            "Escopo de franquia não autorizado para esta tabela temporária.",
        )


def _validate_temp_table_save_payload(payload, *, content_length: int | None = None) -> dict:
    if content_length is not None and content_length > TEMP_TABLE_SAVE_MAX_PAYLOAD_BYTES:
        raise AgenteComparaTempTableError(
            ERROR_TEMP_TABLE_PAYLOAD_TOO_LARGE,
            "Payload de edição excede o limite permitido.",
        )
    if not isinstance(payload, dict):
        raise AgenteComparaTempTableError(
            ERROR_TEMP_TABLE_INVALID_PAYLOAD,
            "Payload deve ser um objeto JSON.",
        )
    try:
        serialized = json.dumps(payload, ensure_ascii=False)
    except (TypeError, ValueError):
        raise AgenteComparaTempTableError(
            ERROR_TEMP_TABLE_INVALID_PAYLOAD,
            "Payload JSON inválido.",
        ) from None
    if len(serialized.encode("utf-8")) > TEMP_TABLE_SAVE_MAX_PAYLOAD_BYTES:
        raise AgenteComparaTempTableError(
            ERROR_TEMP_TABLE_PAYLOAD_TOO_LARGE,
            "Payload de edição excede o limite permitido.",
        )
    temp_table_id = payload.get("temp_table_id")
    if not isinstance(temp_table_id, str) or not temp_table_id.strip():
        raise AgenteComparaTempTableError(
            ERROR_TEMP_TABLE_INVALID_PAYLOAD,
            "temp_table_id é obrigatório.",
        )
    edit_target = payload.get("edit_target")
    if not isinstance(edit_target, dict):
        raise AgenteComparaTempTableError(
            ERROR_TEMP_TABLE_INVALID_PAYLOAD,
            "edit_target é obrigatório.",
        )
    review_action = payload.get("review_action")
    if review_action is not None and review_action not in {
        TEMP_TABLE_REVIEW_ACTION_SAVE_AND_ADVANCE,
        TEMP_TABLE_REVIEW_ACTION_SAVE_DRAFT,
    }:
        raise AgenteComparaTempTableError(
            ERROR_TEMP_TABLE_INVALID_PAYLOAD,
            "review_action inválida.",
        )

    has_freight_tables_key = "freight_tables" in edit_target
    has_freight_routes_key = "freight_routes" in edit_target
    has_accessorial_fees_key = "accessorial_fees" in edit_target
    has_coverage_table_key = "coverage_table" in edit_target
    has_tax_config_key = "tax_config" in edit_target

    freight_tables = (
        _validate_freight_tables_for_save(edit_target.get("freight_tables"))
        if has_freight_tables_key
        else None
    )
    freight_routes = (
        _validate_freight_routes_for_save(edit_target.get("freight_routes"))
        if has_freight_routes_key
        else None
    )
    accessorial_fees = (
        _validate_accessorial_fees_for_save(edit_target.get("accessorial_fees"))
        if has_accessorial_fees_key
        else None
    )
    coverage_table = (
        _validate_coverage_table_for_save(edit_target.get("coverage_table"))
        if has_coverage_table_key
        else None
    )
    tax_config = (
        _validate_tax_config_for_save(edit_target.get("tax_config"))
        if has_tax_config_key
        else None
    )

    has_freight_structural_edit = bool(
        (freight_tables if has_freight_tables_key else [])
        or (freight_routes if has_freight_routes_key else [])
        or (accessorial_fees if has_accessorial_fees_key else [])
    )
    has_coverage_structural_edit = bool(
        coverage_table and coverage_table.get("rows") is not None
    )
    if not (
        has_freight_tables_key
        or has_freight_routes_key
        or has_accessorial_fees_key
        or has_coverage_table_key
        or has_tax_config_key
    ):
        raise AgenteComparaTempTableError(
            ERROR_TEMP_TABLE_INVALID_PAYLOAD,
            "edit_target deve conter ao menos uma seção editável.",
        )

    return {
        "temp_table_id": temp_table_id.strip(),
        "freight_tables": freight_tables,
        "freight_routes": freight_routes,
        "accessorial_fees": accessorial_fees,
        "coverage_table": coverage_table,
        "tax_config": tax_config,
        "has_freight_tables_key": has_freight_tables_key,
        "has_freight_routes_key": has_freight_routes_key,
        "has_accessorial_fees_key": has_accessorial_fees_key,
        "has_coverage_table_key": has_coverage_table_key,
        "has_tax_config_key": has_tax_config_key,
        "has_freight_structural_edit": has_freight_structural_edit,
        "has_coverage_structural_edit": has_coverage_structural_edit,
        "has_structural_edit": bool(has_freight_structural_edit or has_coverage_structural_edit or has_tax_config_key),
        "review_action": review_action or TEMP_TABLE_REVIEW_ACTION_SAVE_AND_ADVANCE,
    }


def save_temp_table_edit(
    payload: dict,
    *,
    user_scope=None,
    franquia_scope=None,
    content_length: int | None = None,
) -> dict:
    """
    Persiste revisão/edição humana no artefato temporário tt_*.json da sessão.

    Não cria novo artefato, não chama Gemini e não grava em banco relacional.
    """
    _require_session()
    validated = _validate_temp_table_save_payload(payload, content_length=content_length)
    sync_temp_table_with_session_documents()
    active_id = get_temp_table_id(session)
    if not active_id:
        raise AgenteComparaTempTableError(
            ERROR_TEMP_TABLE_NOT_FOUND,
            "Nenhuma tabela temporária ativa nesta sessão.",
        )
    if validated["temp_table_id"] != active_id:
        raise AgenteComparaTempTableError(
            ERROR_TEMP_TABLE_ID_MISMATCH,
            "temp_table_id não corresponde à tabela temporária ativa da sessão.",
        )
    cfg = get_cleiton_doc_config()
    record = load_temp_table_record(active_id, ttl_hours=cfg.upload_ttl_hours)
    if record is None:
        clear_temp_table_session_refs(session)
        _mark_session_modified()
        raise AgenteComparaTempTableError(
            ERROR_TEMP_TABLE_NOT_FOUND,
            "Tabela temporária ativa não encontrada.",
        )
    status = (record.get("status") or "").strip().lower()
    if status == TEMP_TABLE_STATUS_EXPIRED:
        raise AgenteComparaTempTableError(
            ERROR_TEMP_TABLE_EXPIRED,
            "A tabela temporária desta sessão expirou.",
        )
    if status in {TEMP_TABLE_STATUS_DISCARDED, TEMP_TABLE_STATUS_PROCESSING}:
        raise AgenteComparaTempTableError(
            ERROR_TEMP_TABLE_NOT_FOUND,
            "Tabela temporária indisponível para revisão.",
        )
    _assert_temp_table_scope(record, user_scope=user_scope, franquia_scope=franquia_scope)

    now = _utcnow().isoformat()
    preserved_expires_at = record.get("expires_at")
    previous_tax_fingerprint = _tax_config_fingerprint(record.get("tax_config"))
    updated = dict(record)
    if validated["freight_tables"]:
        updated["freight_tables"] = validated["freight_tables"]
    if validated["freight_routes"]:
        updated["freight_routes"] = validated["freight_routes"]
    if validated["has_accessorial_fees_key"] and validated["accessorial_fees"] is not None:
        updated["accessorial_fees"] = validated["accessorial_fees"]
    if validated["has_tax_config_key"] and validated["tax_config"] is not None:
        updated["tax_config"] = build_tax_config_for_temp_table(updated, validated["tax_config"])

    has_freight_edit = (
        validated["has_freight_tables_key"]
        or validated["has_freight_routes_key"]
        or validated["has_accessorial_fees_key"]
    )
    if has_freight_edit:
        updated["human_review_status"] = (
            HUMAN_REVIEW_STATUS_EDITED
            if validated["has_freight_structural_edit"]
            else HUMAN_REVIEW_STATUS_REVIEWED
        )
        updated["human_edited_at"] = now
        if user_scope is not None:
            updated["human_edited_by_user_id"] = user_scope
        current_edit_version = updated.get("edit_version")
        if isinstance(current_edit_version, int) and current_edit_version >= 0:
            updated["edit_version"] = current_edit_version + 1
        else:
            updated["edit_version"] = 1
        updated = _mark_audit_batch_stale_if_processed(
            updated,
            reason=AUDIT_BATCH_STALE_PRICING_RULE_REASON,
            alert=AUDIT_BATCH_STALE_PRICING_RULE_ALERT,
        )

    if validated["has_tax_config_key"] and not has_freight_edit:
        updated["human_edited_at"] = now
        if user_scope is not None:
            updated["human_edited_by_user_id"] = user_scope
        current_edit_version = updated.get("edit_version")
        if isinstance(current_edit_version, int) and current_edit_version >= 0:
            updated["edit_version"] = current_edit_version + 1
        else:
            updated["edit_version"] = 1

    if validated["has_coverage_table_key"] and validated["coverage_table"] is not None:
        existing_coverage = updated.get("coverage_table")
        if not isinstance(existing_coverage, dict):
            existing_coverage = _empty_coverage_table_shell()
        coverage = dict(existing_coverage)
        coverage["status"] = COVERAGE_TABLE_STATUS_NEEDS_REVIEW
        coverage["columns"] = list(COVERAGE_TABLE_COLUMNS)
        coverage["rows"] = validated["coverage_table"]["rows"]
        coverage["human_review_status"] = (
            HUMAN_REVIEW_STATUS_EDITED
            if validated["has_coverage_structural_edit"]
            else HUMAN_REVIEW_STATUS_REVIEWED
        )
        coverage["human_edited_at"] = now
        if user_scope is not None:
            coverage["human_edited_by_user_id"] = user_scope
        current_coverage_version = coverage.get("edit_version")
        if isinstance(current_coverage_version, int) and current_coverage_version >= 0:
            coverage["edit_version"] = current_coverage_version + 1
        else:
            coverage["edit_version"] = 1
        if "validation_warnings" not in coverage or not isinstance(coverage.get("validation_warnings"), list):
            coverage["validation_warnings"] = []
        if "uploaded_at" not in coverage:
            coverage["uploaded_at"] = now
        updated["coverage_table"] = coverage

    if validated["review_action"] == TEMP_TABLE_REVIEW_ACTION_SAVE_AND_ADVANCE:
        _validate_accessorial_fees_ready_to_advance(updated.get("accessorial_fees"))

    if validated["has_tax_config_key"]:
        new_tax_fingerprint = _tax_config_fingerprint(updated.get("tax_config"))
        if previous_tax_fingerprint != new_tax_fingerprint:
            updated = _mark_audit_batch_stale_if_processed(
                updated,
                reason=AUDIT_BATCH_STALE_TAX_CONFIG_REASON,
                alert=AUDIT_BATCH_STALE_TAX_CONFIG_ALERT,
            )

    updated["updated_at"] = now
    updated["expires_at"] = preserved_expires_at

    saved = save_temp_table_record(updated)
    logger.info(
        "Agente Compara temp_table save: temp_table_id=%s user_id=%s status=%s tables=%s routes=%s fees=%s coverage_rows=%s taxes=%s",
        saved.get("temp_table_id"),
        user_scope,
        updated.get("human_review_status"),
        len(saved.get("freight_tables") or []),
        len(saved.get("freight_routes") or []),
        len(saved.get("accessorial_fees") or []),
        len((saved.get("coverage_table") or {}).get("rows") or []),
        bool((saved.get("tax_config") or {}).get("include_taxes")),
    )
    public = _public_temp_table(saved)
    if public is None:
        raise AgenteComparaTempTableError(
            ERROR_TEMP_TABLE_NOT_FOUND,
            "Não foi possível retornar a tabela temporária atualizada.",
        )
    return public


def load_temp_table_record(temp_table_id: str, *, ttl_hours: int) -> dict | None:
    ref = (temp_table_id or "").strip()
    if not ref:
        return None
    try:
        path = _temp_table_path(ref)
    except ValueError:
        return None
    if not path.is_file():
        return None
    try:
        with open(path, "r", encoding="utf-8") as handle:
            payload = json.load(handle)
        if not isinstance(payload, dict):
            remove_temp_table_record(ref)
            return None
    except Exception:
        remove_temp_table_record(ref)
        return None

    expires_at = _parse_iso(payload.get(FIELD_EXPIRES_AT))
    if expires_at is None:
        created_at = _parse_iso(payload.get(FIELD_CREATED_AT))
        if created_at is None:
            remove_temp_table_record(ref)
            return None
        expires_at = created_at + timedelta(hours=max(1, int(ttl_hours)))
    if _utcnow() >= expires_at:
        payload["status"] = TEMP_TABLE_STATUS_EXPIRED
        payload["updated_at"] = _utcnow().isoformat()
        try:
            _write_temp_table_atomic(path, payload)
        except Exception:
            remove_temp_table_record(ref)
            return None
        return payload
    return payload


def _sync_outdated_fiscal_stale(record: dict) -> dict:
    if not isinstance(record, dict):
        return record
    audit_batch = record.get("audit_batch")
    if not _audit_batch_is_fiscally_outdated(audit_batch):
        return record
    if isinstance(audit_batch, dict) and audit_batch.get("needs_reprocess"):
        return record
    return _mark_audit_batch_stale_if_processed(
        record,
        reason=AUDIT_BATCH_STALE_FISCAL_OUTDATED_REASON,
        alert=AUDIT_BATCH_STALE_FISCAL_OUTDATED_ALERT,
    )


def _sync_outdated_pricing_rule_stale(record: dict) -> dict:
    if not isinstance(record, dict):
        return record
    audit_batch = record.get("audit_batch")
    if not _audit_batch_is_pricing_rule_parser_outdated(audit_batch):
        return record
    if isinstance(audit_batch, dict) and audit_batch.get("needs_reprocess"):
        return record
    return _mark_audit_batch_stale_if_processed(
        record,
        reason=AUDIT_BATCH_STALE_PRICING_RULE_REASON,
        alert=AUDIT_BATCH_STALE_PRICING_RULE_ALERT,
    )


def save_temp_table_record(record: dict) -> dict:
    _require_session()
    temp_table_id = (record.get("temp_table_id") or uuid4().hex).strip()
    record = _sync_outdated_fiscal_stale(dict(record))
    record = _sync_outdated_pricing_rule_stale(record)
    record["temp_table_id"] = temp_table_id
    path = _temp_table_path(temp_table_id)
    _write_temp_table_atomic(path, record)
    set_temp_table_id(session, temp_table_id)
    set_temp_table_source_doc_ids(session, list(record.get("source_documents") or []))
    _mark_session_modified()
    return record


def remove_temp_table_record(temp_table_id: str) -> bool:
    ref = (temp_table_id or "").strip()
    if not ref:
        return False
    try:
        path = _temp_table_path(ref)
    except ValueError:
        return False
    if path.is_file():
        try:
            path.unlink()
        except Exception:
            return False
    return True


def invalidate_temp_table_for_session(*, reason: str = TEMP_TABLE_STATUS_DISCARDED) -> None:
    _require_session()
    temp_table_id = get_temp_table_id(session)
    clear_temp_table_session_refs(session)
    if temp_table_id:
        remove_temp_table_record(temp_table_id)
    _mark_session_modified()


def invalidate_temp_table_if_source_changed(
    *,
    reason: str = TEMP_TABLE_STATUS_DISCARDED,
    removed_doc_id: str | None = None,
) -> None:
    _require_session()
    temp_table_id = get_temp_table_id(session)
    if not temp_table_id:
        return
    cfg = get_cleiton_doc_config()
    record = load_temp_table_record(temp_table_id, ttl_hours=cfg.upload_ttl_hours)
    if record is None:
        clear_temp_table_session_refs(session)
        _mark_session_modified()
        return
    source_docs = list(record.get("source_documents") or [])
    active_ids = set(get_agente_compara_doc_ids(session))
    if removed_doc_id and removed_doc_id in source_docs:
        invalidate_temp_table_for_session(reason=reason)
        return
    if source_docs and not all(doc_id in active_ids for doc_id in source_docs):
        invalidate_temp_table_for_session(reason=reason)


def _parse_iso(raw: str | None) -> datetime | None:
    if not raw:
        return None
    try:
        dt = datetime.fromisoformat(str(raw).replace("Z", "+00:00"))
        if dt.tzinfo:
            off = dt.utcoffset()
            dt = (dt.replace(tzinfo=None) - off) if off else dt.replace(tzinfo=None)
        return dt
    except Exception:
        return None


def _temp_table_expires_at(source_doc_ids: list[str]) -> str:
    cfg = get_cleiton_doc_config()
    latest: datetime | None = None
    for doc_id in source_doc_ids:
        record = load_document_record(doc_id, ttl_hours=cfg.upload_ttl_hours)
        if record is None:
            continue
        candidate = _parse_iso(record.get(FIELD_EXPIRES_AT))
        if candidate and (latest is None or candidate > latest):
            latest = candidate
    if latest is None:
        latest = _utcnow() + timedelta(hours=max(1, int(cfg.upload_ttl_hours)))
    return latest.isoformat()


def get_active_temp_table_for_session() -> dict | None:
    _require_session()
    sync_temp_table_with_session_documents()
    temp_table_id = get_temp_table_id(session)
    if not temp_table_id:
        return None
    cfg = get_cleiton_doc_config()
    record = load_temp_table_record(temp_table_id, ttl_hours=cfg.upload_ttl_hours)
    if record is None:
        clear_temp_table_session_refs(session)
        _mark_session_modified()
        return None
    status = (record.get("status") or "").strip()
    if status in {TEMP_TABLE_STATUS_DISCARDED, TEMP_TABLE_STATUS_EXPIRED}:
        return _public_temp_table(record)
    return _public_temp_table(record)


def sync_temp_table_with_session_documents() -> None:
    _require_session()
    temp_table_id = get_temp_table_id(session)
    if not temp_table_id:
        return
    cfg = get_cleiton_doc_config()
    record = load_temp_table_record(temp_table_id, ttl_hours=cfg.upload_ttl_hours)
    if record is None:
        clear_temp_table_session_refs(session)
        _mark_session_modified()
        return
    active_ids = set(get_agente_compara_doc_ids(session))
    source_docs = list(record.get("source_documents") or [])
    if source_docs and not all(doc_id in active_ids for doc_id in source_docs):
        invalidate_temp_table_for_session(reason=TEMP_TABLE_STATUS_DISCARDED)


def should_attempt_temp_table_extraction(session_obj, source_doc_ids: list[str]) -> bool:
    normalized = _normalize_source_doc_ids(source_doc_ids)
    if not normalized:
        return False
    sync_temp_table_with_session_documents()
    temp_table_id = get_temp_table_id(session_obj)
    if not temp_table_id:
        return True
    cfg = get_cleiton_doc_config()
    record = load_temp_table_record(temp_table_id, ttl_hours=cfg.upload_ttl_hours)
    if record is None:
        return True
    bound_sources = _normalize_source_doc_ids(list(record.get("source_documents") or []))
    if bound_sources != normalized:
        return True
    raw_status = record.get("status")
    status = raw_status.strip().lower() if isinstance(raw_status, str) else ""
    if status == TEMP_TABLE_STATUS_PROCESSING:
        return False
    return False


def mark_temp_table_processing(source_doc_ids: list[str], *, user_scope=None, franquia_scope=None) -> dict:
    _require_session()
    normalized = _normalize_source_doc_ids(source_doc_ids)
    now = _utcnow()
    temp_table_id = get_temp_table_id(session) or uuid4().hex
    record = {
        "temp_table_id": temp_table_id,
        "status": TEMP_TABLE_STATUS_PROCESSING,
        "source_documents": normalized,
        "detected_carrier": None,
        "origins": [],
        "destinations": [],
        "routes": [],
        "freight_tables": [],
        "freight_routes": [],
        "weight_ranges": [],
        "freight_values": [],
        "accessorial_fees": [],
        "charge_type_detected": None,
        "extracted_items": [],
        "uncertain_fields": [],
        "reading_alerts": [],
        "evidence_refs": [],
        "created_at": now.isoformat(),
        "updated_at": now.isoformat(),
        "expires_at": _temp_table_expires_at(normalized),
        "session_scope": AGENTE_COMPARA_DOC_IDS_SESSION_KEY,
        "franquia_scope": franquia_scope,
        "user_scope": user_scope,
        "operational_owner": TEMP_TABLE_OPERATIONAL_OWNER,
        "ui_visibility": {
            "display_name": TEMP_TABLE_UI_DISPLAY_NAME,
            "readonly": True,
        },
        "version_marker": TEMP_TABLE_VERSION_MARKER,
    }
    return save_temp_table_record(record)


def temp_table_status_message(status: str) -> str:
    mapping = {
        TEMP_TABLE_STATUS_PROCESSING: (
            "Recebi os anexos e iniciei a estruturação da tabela temporária de frete."
        ),
        TEMP_TABLE_STATUS_AWAITING_VALIDATION: (
            "A tabela temporária foi estruturada e está aguardando sua validação."
        ),
        TEMP_TABLE_STATUS_NEEDS_REVIEW: (
            "A tabela temporária foi gerada. Revise os dados antes de continuar."
        ),
        TEMP_TABLE_STATUS_FAILED: (
            "Não foi possível estruturar a tabela temporária a partir dos anexos enviados."
        ),
        TEMP_TABLE_STATUS_EXPIRED: (
            "A tabela temporária desta sessão expirou."
        ),
        TEMP_TABLE_STATUS_DISCARDED: (
            "Os documentos de origem foram alterados ou removidos, "
            "então a tabela temporária anterior foi invalidada."
        ),
    }
    return mapping.get((status or "").strip(), "")


def _normalize_temp_table_status(raw_status) -> str:
    allowed = {
        TEMP_TABLE_STATUS_AWAITING_VALIDATION,
        TEMP_TABLE_STATUS_NEEDS_REVIEW,
        TEMP_TABLE_STATUS_FAILED,
        TEMP_TABLE_STATUS_VALIDATED,
    }
    if not isinstance(raw_status, str):
        return TEMP_TABLE_STATUS_FAILED
    candidate = raw_status.strip().lower()
    if not candidate or candidate not in allowed:
        return TEMP_TABLE_STATUS_FAILED
    return candidate


def _list_field_from_raw(raw: dict, name: str) -> list:
    value = raw.get(name)
    return list(value) if isinstance(value, list) else []


def _is_useful_freight_value(item) -> bool:
    if isinstance(item, dict):
        label = item.get("label")
        if isinstance(label, str) and label.strip():
            return True
        return item.get("value") is not None
    if isinstance(item, str):
        return bool(item.strip())
    return False


def _is_useful_accessorial_fee(item) -> bool:
    if isinstance(item, dict):
        name = item.get("name")
        if isinstance(name, str) and name.strip():
            return True
        return item.get("value") is not None
    if isinstance(item, str):
        return bool(item.strip())
    return False


def _is_useful_weight_range(item) -> bool:
    if isinstance(item, dict):
        label = item.get("label")
        if isinstance(label, str) and label.strip():
            return True
        if item.get("min_weight") is not None or item.get("max_weight") is not None:
            return True
    if isinstance(item, str):
        return bool(item.strip())
    return False


def _optional_normalized_str(value) -> str | None:
    if value is None:
        return None
    if isinstance(value, str):
        candidate = value.strip()
        return candidate or None
    candidate = str(value).strip()
    return candidate or None


def _freight_route_field(item: dict, primary: str, *aliases: str) -> str | None:
    keys = (primary, *aliases)
    for key in keys:
        if key not in item:
            continue
        normalized = _optional_normalized_str(item.get(key))
        if normalized is not None:
            return normalized
    return None


def _freight_route_key_present(item: dict, primary: str, *aliases: str) -> bool:
    return any(key in item for key in (primary, *aliases))


FREIGHT_ROUTE_TECHNICAL_FIELDS: tuple[tuple[str, tuple[str, ...]], ...] = (
    ("weight_30", ("weight_30kg",)),
    ("weight_50", ("weight_50kg",)),
    ("weight_70", ("weight_70kg",)),
    ("weight_100", ("weight_100kg",)),
    ("boarding_fee", ("taxa_embarque_kg",)),
    ("freight_value_pct", ("frete_valor_pct",)),
    ("freight_weight_kg", ("frete_peso_kg",)),
    ("pedagio", ("toll", "pedagio_valor")),
)

FREIGHT_ROUTE_RESERVED_KEYS = frozenset(
    {
        "origin",
        "destination",
        "freight_type",
        "type",
        "notes",
        "observations",
        "observacoes",
        "evidence_ref",
        "confidence",
        "column_labels",
        *(alias for _, aliases in FREIGHT_ROUTE_TECHNICAL_FIELDS for alias in aliases),
        *(primary for primary, _ in FREIGHT_ROUTE_TECHNICAL_FIELDS),
    }
)

FREIGHT_ROUTE_DEFAULT_LABELS = {
    "origin": "Origem",
    "destination": "Destino",
    "freight_type": "Tipo",
    "weight_30": "Até 30 kg",
    "weight_50": "Até 50 kg",
    "weight_70": "Até 70 kg",
    "weight_100": "Até 100 kg",
    "boarding_fee": "Taxa embarque",
    "freight_value_pct": "Frete Valor %",
    "freight_weight_kg": "Excedente por kg",
    "pedagio": "Pedágio",
    "notes": "Observações",
}


def _normalize_column_labels(raw) -> dict[str, str]:
    if not isinstance(raw, dict):
        return {}
    labels: dict[str, str] = {}
    for key, val in raw.items():
        if not isinstance(key, str):
            continue
        normalized_key = key.strip()
        normalized_val = _optional_normalized_str(val)
        if normalized_key and normalized_val:
            labels[normalized_key] = normalized_val
    return labels


def _normalize_freight_table_column_meta(raw_meta, columns: list[str]) -> dict[str, dict]:
    if not isinstance(raw_meta, dict):
        return {}
    normalized: dict[str, dict] = {}
    for key, val in raw_meta.items():
        if not isinstance(key, str):
            continue
        col = key.strip()
        if col not in columns or not isinstance(val, dict):
            continue
        origin = _optional_normalized_str(val.get("origin"))
        if origin in {"observed", "inferred", "technical"}:
            normalized[col] = {"origin": origin}
    return normalized


def _should_keep_freight_table_column(
    col: str,
    rows: list[dict],
    column_meta: dict[str, dict],
) -> bool:
    if column_meta.get(col, {}).get("origin") == "observed":
        return True
    if not rows:
        return True
    return any(
        _optional_normalized_str(row.get(col)) is not None
        for row in rows
        if isinstance(row, dict)
    )


def _prune_freight_table_columns(table: dict) -> dict:
    columns = table.get("columns") if isinstance(table.get("columns"), list) else []
    rows = table.get("rows") if isinstance(table.get("rows"), list) else []
    column_meta = table.get("column_meta") if isinstance(table.get("column_meta"), dict) else {}
    if not columns:
        return table
    kept_columns = [
        col for col in columns if _should_keep_freight_table_column(col, rows, column_meta)
    ]
    if kept_columns == columns:
        return table
    pruned_rows: list[dict] = []
    for row in rows:
        if not isinstance(row, dict):
            continue
        pruned_rows.append({col: row.get(col) for col in kept_columns})
    pruned_meta = {col: column_meta[col] for col in kept_columns if col in column_meta}
    updated = dict(table)
    updated["columns"] = kept_columns
    updated["rows"] = pruned_rows
    if pruned_meta:
        updated["column_meta"] = pruned_meta
    elif "column_meta" in updated:
        updated.pop("column_meta", None)
    return updated


def _route_has_visible_value(route: dict, key: str) -> bool:
    return _optional_normalized_str(route.get(key)) is not None


def _collect_route_column_specs(routes: list[dict]) -> list[dict[str, str]]:
    if not routes:
        return []
    merged_labels: dict[str, str] = {}
    for route in routes:
        labels = route.get("column_labels")
        if isinstance(labels, dict):
            merged_labels.update(_normalize_column_labels(labels))

    ordered_keys = [
        "origin",
        "destination",
        "freight_type",
        "weight_30",
        "weight_50",
        "weight_70",
        "weight_100",
        "boarding_fee",
        "freight_value_pct",
        "freight_weight_kg",
        "pedagio",
    ]
    specs: list[dict[str, str]] = []
    seen: set[str] = set()
    for key in ordered_keys:
        if not any(_route_has_visible_value(route, key) for route in routes):
            continue
        label = merged_labels.get(key) or FREIGHT_ROUTE_DEFAULT_LABELS.get(key, key)
        specs.append({"key": key, "label": label})
        seen.add(key)

    for route in routes:
        for key, val in route.items():
            if key in seen or key in FREIGHT_ROUTE_RESERVED_KEYS or not isinstance(key, str):
                continue
            if _optional_normalized_str(val) is None:
                continue
            label = merged_labels.get(key) or key
            specs.append({"key": key, "label": label})
            seen.add(key)
    return specs


def _synthesize_freight_tables_from_routes(routes: list[dict]) -> list[dict]:
    if not routes:
        return []
    specs = _collect_route_column_specs(routes)
    if not specs:
        return []
    columns = [spec["label"] for spec in specs]
    rows: list[dict] = []
    for route in routes:
        row: dict = {}
        for spec in specs:
            value = route.get(spec["key"])
            row[spec["label"]] = value if _optional_normalized_str(value) is not None else None
        if _row_has_any_value(row):
            rows.append(row)
    if not rows:
        return []
    first = routes[0]
    return [
        {
            "table_title": None,
            "table_type": "route_matrix_synthesized",
            "context": {
                "route_label": None,
                "origin": first.get("origin"),
                "destination": None,
                "customer": None,
                "supplier": None,
                "valid_from": None,
                "valid_to": None,
                "delivery_deadline": None,
            },
            "columns": columns,
            "rows": rows,
            "notes": "",
            "evidence_ref": first.get("evidence_ref"),
            "confidence": first.get("confidence"),
        }
    ]


def _normalize_freight_route_item(item) -> dict | None:
    if not isinstance(item, dict):
        return None
    result: dict = {}
    has_useful_value = False

    for primary, aliases in (
        ("origin", ("origin",)),
        ("destination", ("destination",)),
        ("freight_type", ("freight_type", "type")),
    ):
        if not _freight_route_key_present(item, primary, *aliases):
            continue
        value = _freight_route_field(item, primary, *aliases)
        result[primary] = value
        if value is not None:
            has_useful_value = True

    for primary, aliases in FREIGHT_ROUTE_TECHNICAL_FIELDS:
        if not _freight_route_key_present(item, primary, *aliases):
            continue
        value = _freight_route_field(item, primary, *aliases)
        if value is None:
            continue
        result[primary] = value
        has_useful_value = True

    for key, val in item.items():
        if key in FREIGHT_ROUTE_RESERVED_KEYS or not isinstance(key, str):
            continue
        candidate = key.strip()
        if not candidate:
            continue
        normalized_val = _optional_normalized_str(val)
        if normalized_val is None:
            continue
        result[candidate] = normalized_val
        has_useful_value = True

    column_labels = _normalize_column_labels(item.get("column_labels"))
    if column_labels:
        result["column_labels"] = column_labels

    if _freight_route_key_present(item, "notes", "observations", "observacoes"):
        result["notes"] = _freight_route_field(item, "notes", "observations", "observacoes") or ""
    else:
        result["notes"] = ""

    for meta_key in ("evidence_ref", "confidence"):
        if _freight_route_key_present(item, meta_key):
            result[meta_key] = _freight_route_field(item, meta_key)

    if not has_useful_value and not result.get("notes") and not result.get("evidence_ref"):
        return None
    return result


def _is_useful_freight_route(item) -> bool:
    normalized = _normalize_freight_route_item(item) if isinstance(item, dict) else None
    if not normalized:
        return False
    useful_fields = (
        "origin",
        "destination",
        "freight_type",
        "weight_30",
        "weight_50",
        "weight_70",
        "weight_100",
        "boarding_fee",
        "freight_value_pct",
        "freight_weight_kg",
        "pedagio",
    )
    if any(normalized.get(field) is not None for field in useful_fields):
        return True
    return any(
        key not in FREIGHT_ROUTE_RESERVED_KEYS
        and _optional_normalized_str(normalized.get(key)) is not None
        for key in normalized
    )


def _normalize_freight_routes(raw_routes) -> list[dict]:
    if not isinstance(raw_routes, list):
        return []
    normalized: list[dict] = []
    for item in raw_routes:
        route = _normalize_freight_route_item(item)
        if route is not None:
            normalized.append(route)
    return normalized


def _normalize_freight_table_context(raw_context) -> dict:
    if not isinstance(raw_context, dict):
        return {
            "route_label": None,
            "origin": None,
            "destination": None,
            "customer": None,
            "supplier": None,
            "valid_from": None,
            "valid_to": None,
            "delivery_deadline": None,
        }
    return {
        "route_label": _optional_normalized_str(raw_context.get("route_label")),
        "origin": _optional_normalized_str(raw_context.get("origin")),
        "destination": _optional_normalized_str(raw_context.get("destination")),
        "customer": _optional_normalized_str(raw_context.get("customer")),
        "supplier": _optional_normalized_str(raw_context.get("supplier")),
        "valid_from": _optional_normalized_str(raw_context.get("valid_from")),
        "valid_to": _optional_normalized_str(raw_context.get("valid_to")),
        "delivery_deadline": _optional_normalized_str(raw_context.get("delivery_deadline")),
    }


def _normalize_freight_table_row(item, columns: list[str]) -> dict:
    if not isinstance(item, dict):
        return {}
    normalized: dict = {}
    for col in columns:
        if col in item:
            val = item.get(col)
            if val is None:
                normalized[col] = None
            elif isinstance(val, str):
                normalized[col] = val
            else:
                normalized[col] = str(val)
    for key, val in item.items():
        if key not in normalized and isinstance(key, str) and key.strip():
            if val is None:
                normalized[key] = None
            elif isinstance(val, str):
                normalized[key] = val
            else:
                normalized[key] = str(val)
    return normalized


def _normalize_freight_table_item(item) -> dict | None:
    if not isinstance(item, dict):
        return None
    raw_columns = item.get("columns")
    columns: list[str] = []
    if isinstance(raw_columns, list):
        for col in raw_columns:
            if isinstance(col, str):
                candidate = col.strip()
                if candidate:
                    columns.append(candidate)
    raw_rows = item.get("rows")
    rows: list[dict] = []
    if isinstance(raw_rows, list):
        for row in raw_rows:
            normalized_row = _normalize_freight_table_row(row, columns)
            if normalized_row:
                rows.append(normalized_row)
    column_meta = _normalize_freight_table_column_meta(item.get("column_meta"), columns)
    table = {
        "table_title": _optional_normalized_str(item.get("table_title")),
        "table_type": _optional_normalized_str(item.get("table_type")),
        "context": _normalize_freight_table_context(item.get("context")),
        "columns": columns,
        "rows": rows,
        "notes": _optional_normalized_str(item.get("notes")) or "",
        "evidence_ref": _optional_normalized_str(item.get("evidence_ref")),
        "confidence": _optional_normalized_str(item.get("confidence")),
    }
    if column_meta:
        table["column_meta"] = column_meta
    return _prune_freight_table_columns(table)


def _row_has_any_value(row: dict) -> bool:
    if not isinstance(row, dict):
        return False
    for val in row.values():
        if val is None:
            continue
        if isinstance(val, str) and not val.strip():
            continue
        return True
    return False


def _is_useful_freight_table(item) -> bool:
    normalized = _normalize_freight_table_item(item) if isinstance(item, dict) else None
    if not normalized:
        return False
    if normalized.get("table_title"):
        return True
    if normalized.get("columns"):
        return True
    for row in normalized.get("rows") or []:
        if _row_has_any_value(row):
            return True
    return False


def _normalize_freight_tables(raw_tables) -> list[dict]:
    if not isinstance(raw_tables, list):
        return []
    normalized: list[dict] = []
    for item in raw_tables:
        table = _normalize_freight_table_item(item)
        if table is not None:
            normalized.append(table)
    return normalized


def _has_useful_partial_extraction_data(raw: dict) -> bool:
    for item in _list_field_from_raw(raw, "freight_tables"):
        if _is_useful_freight_table(item):
            return True
    for item in _list_field_from_raw(raw, "freight_routes"):
        if _is_useful_freight_route(item):
            return True
    for item in _list_field_from_raw(raw, "freight_values"):
        if _is_useful_freight_value(item):
            return True
    for item in _list_field_from_raw(raw, "accessorial_fees"):
        if _is_useful_accessorial_fee(item):
            return True
    for item in _list_field_from_raw(raw, "weight_ranges"):
        if _is_useful_weight_range(item):
            return True
    return False


def _has_legacy_useful_extraction_data(raw: dict) -> bool:
    for name in ("origins", "destinations", "routes", "extracted_items", "uncertain_fields"):
        items = _list_field_from_raw(raw, name)
        if items:
            return True
    carrier = raw.get("detected_carrier")
    if isinstance(carrier, str) and carrier.strip():
        return True
    charge = raw.get("charge_type_detected")
    if isinstance(charge, str) and charge.strip():
        return True
    return False


def _resolve_extraction_status(raw: dict, expanded: dict) -> str:
    """Resolve status final após normalização partial-first."""
    has_partial = _has_useful_partial_extraction_data(expanded)
    has_legacy = _has_legacy_useful_extraction_data(expanded)
    candidate = _normalize_temp_table_status(raw.get("status"))

    if has_partial:
        return TEMP_TABLE_STATUS_NEEDS_REVIEW

    if candidate == TEMP_TABLE_STATUS_AWAITING_VALIDATION:
        return TEMP_TABLE_STATUS_AWAITING_VALIDATION

    if candidate == TEMP_TABLE_STATUS_VALIDATED:
        return TEMP_TABLE_STATUS_VALIDATED

    if has_legacy and candidate == TEMP_TABLE_STATUS_FAILED:
        return TEMP_TABLE_STATUS_NEEDS_REVIEW

    if has_legacy:
        return candidate

    if candidate == TEMP_TABLE_STATUS_NEEDS_REVIEW:
        return TEMP_TABLE_STATUS_FAILED

    return TEMP_TABLE_STATUS_FAILED


def normalize_partial_first_extraction_to_temp_table(raw: dict) -> dict:
    """
    Normaliza a resposta partial-first (Etapa A) para o contrato interno da temp_table.

    A Etapa A retorna apenas custos brutos detectados; o backend completa campos
    opcionais do contrato interno sem exigir fechamento de rotas ou transportadora.
    """
    alerts = [
        str(item).strip()
        for item in _list_field_from_raw(raw, "reading_alerts")
        if isinstance(item, str) and str(item).strip()
    ]
    expanded = {
        "status": raw.get("status"),
        "detected_carrier": raw.get("detected_carrier"),
        "origins": _list_field_from_raw(raw, "origins"),
        "destinations": _list_field_from_raw(raw, "destinations"),
        "routes": _list_field_from_raw(raw, "routes"),
        "freight_tables": _normalize_freight_tables(_list_field_from_raw(raw, "freight_tables")),
        "freight_routes": _normalize_freight_routes(_list_field_from_raw(raw, "freight_routes")),
        "weight_ranges": _list_field_from_raw(raw, "weight_ranges"),
        "freight_values": _list_field_from_raw(raw, "freight_values"),
        "accessorial_fees": _normalize_accessorial_fees(_list_field_from_raw(raw, "accessorial_fees")),
        "charge_type_detected": raw.get("charge_type_detected"),
        "extracted_items": _list_field_from_raw(raw, "extracted_items"),
        "uncertain_fields": _list_field_from_raw(raw, "uncertain_fields"),
        "reading_alerts": alerts,
        "evidence_refs": _list_field_from_raw(raw, "evidence_refs"),
        "franquia_scope": raw.get("franquia_scope"),
        "user_scope": raw.get("user_scope"),
    }
    if not expanded["freight_tables"] and expanded["freight_routes"]:
        synthesized = _synthesize_freight_tables_from_routes(expanded["freight_routes"])
        if synthesized:
            expanded["freight_tables"] = synthesized
    expanded["status"] = _resolve_extraction_status(raw, expanded)
    return expanded


def _coerce_temp_table_payload(raw: dict, *, source_doc_ids: list[str]) -> dict:
    now = _utcnow().isoformat()
    normalized = normalize_partial_first_extraction_to_temp_table(raw)
    status = _resolve_extraction_status(raw, normalized)
    uncertain = (
        normalized.get("uncertain_fields")
        if isinstance(normalized.get("uncertain_fields"), list)
        else []
    )
    alerts = [
        str(item).strip()
        for item in (
            normalized.get("reading_alerts")
            if isinstance(normalized.get("reading_alerts"), list)
            else []
        )
        if isinstance(item, str) and str(item).strip()
    ]

    if _has_useful_partial_extraction_data(normalized):
        status = TEMP_TABLE_STATUS_NEEDS_REVIEW
    elif status == TEMP_TABLE_STATUS_NEEDS_REVIEW and not _has_legacy_useful_extraction_data(
        normalized
    ):
        status = TEMP_TABLE_STATUS_FAILED

    if status == TEMP_TABLE_STATUS_AWAITING_VALIDATION and (uncertain or alerts):
        status = TEMP_TABLE_STATUS_NEEDS_REVIEW
    if status == TEMP_TABLE_STATUS_NEEDS_REVIEW and not alerts and not uncertain:
        alerts = [
            "A extração encontrou dados parciais e precisa de validação humana."
        ]
    if status == TEMP_TABLE_STATUS_FAILED and not alerts:
        alerts = [
            "Não foi possível estruturar a tabela temporária a partir dos anexos enviados."
        ]

    temp_table_id = get_temp_table_id(session) or uuid4().hex
    existing = load_temp_table_record(temp_table_id, ttl_hours=get_cleiton_doc_config().upload_ttl_hours)
    created_at = (existing or {}).get("created_at") or now
    preserved_coverage = None
    if existing and isinstance(existing.get("coverage_table"), dict):
        preserved_coverage = existing.get("coverage_table")
    preserved_audit_batch = None
    if existing and isinstance(existing.get("audit_batch"), dict):
        preserved_audit_batch = existing.get("audit_batch")
    record = {
        "temp_table_id": temp_table_id,
        "status": status,
        "source_documents": _normalize_source_doc_ids(source_doc_ids),
        "detected_carrier": normalized.get("detected_carrier"),
        "origins": _list_field_from_raw(normalized, "origins"),
        "destinations": _list_field_from_raw(normalized, "destinations"),
        "routes": _list_field_from_raw(normalized, "routes"),
        "freight_tables": _normalize_freight_tables(_list_field_from_raw(normalized, "freight_tables")),
        "freight_routes": _normalize_freight_routes(_list_field_from_raw(normalized, "freight_routes")),
        "weight_ranges": _list_field_from_raw(normalized, "weight_ranges"),
        "freight_values": _list_field_from_raw(normalized, "freight_values"),
        "accessorial_fees": _normalize_accessorial_fees(_list_field_from_raw(normalized, "accessorial_fees")),
        "charge_type_detected": normalized.get("charge_type_detected"),
        "extracted_items": _list_field_from_raw(normalized, "extracted_items"),
        "uncertain_fields": uncertain,
        "reading_alerts": alerts,
        "evidence_refs": _list_field_from_raw(normalized, "evidence_refs"),
        "created_at": created_at,
        "updated_at": now,
        "expires_at": _temp_table_expires_at(_normalize_source_doc_ids(source_doc_ids)),
        "session_scope": AGENTE_COMPARA_DOC_IDS_SESSION_KEY,
        "franquia_scope": normalized.get("franquia_scope"),
        "user_scope": normalized.get("user_scope"),
        "operational_owner": TEMP_TABLE_OPERATIONAL_OWNER,
        "ui_visibility": {
            "display_name": TEMP_TABLE_UI_DISPLAY_NAME,
            "readonly": True,
        },
        "version_marker": TEMP_TABLE_VERSION_MARKER,
    }
    if preserved_coverage is not None:
        record["coverage_table"] = preserved_coverage
    if preserved_audit_batch is not None:
        record["audit_batch"] = preserved_audit_batch
    return record


def _has_human_review_metadata(record: dict | None) -> bool:
    if not record or not isinstance(record, dict):
        return False
    if record.get("human_review_status"):
        return True
    if record.get("human_edited_at"):
        return True
    if record.get("human_edited_by_user_id") is not None:
        return True
    edit_version = record.get("edit_version")
    if isinstance(edit_version, int) and edit_version > 0:
        return True
    if isinstance(edit_version, str) and edit_version.strip().isdigit() and int(edit_version) > 0:
        return True
    return False


def _source_documents_match(existing_sources, incoming_sources: list[str]) -> bool:
    return _normalize_source_doc_ids(list(existing_sources or [])) == _normalize_source_doc_ids(
        incoming_sources
    )


def _should_skip_extraction_overwrite(
    existing: dict | None,
    *,
    source_doc_ids: list[str],
    force_overwrite: bool = False,
) -> bool:
    if force_overwrite or not existing:
        return False
    if not _has_human_review_metadata(existing):
        return False
    return _source_documents_match(existing.get("source_documents"), source_doc_ids)


def apply_temp_table_extraction_from_model_payload(
    payload: dict | None,
    *,
    source_doc_ids: list[str],
    force_overwrite: bool = False,
) -> dict | None:
    _require_session()
    normalized = _normalize_source_doc_ids(source_doc_ids)
    if not normalized:
        return None

    cfg = get_cleiton_doc_config()
    temp_table_id = get_temp_table_id(session)
    existing = None
    if temp_table_id:
        existing = load_temp_table_record(temp_table_id, ttl_hours=cfg.upload_ttl_hours)

    if _should_skip_extraction_overwrite(
        existing,
        source_doc_ids=normalized,
        force_overwrite=force_overwrite,
    ):
        logger.info(
            "Agente Compara temp_table extraction skipped because human-reviewed artifact already exists "
            "(temp_table_id=%s edit_version=%s human_review_status=%s).",
            existing.get("temp_table_id") if existing else None,
            (existing or {}).get("edit_version"),
            (existing or {}).get("human_review_status"),
        )
        return existing

    if not isinstance(payload, dict):
        record = _coerce_temp_table_payload({"status": TEMP_TABLE_STATUS_FAILED}, source_doc_ids=normalized)
        return save_temp_table_record(record)
    record = _coerce_temp_table_payload(payload, source_doc_ids=normalized)
    return save_temp_table_record(record)


def split_temp_table_block_from_answer(answer_text: str) -> tuple[str, dict | None]:
    text = answer_text or ""
    begin = text.find(TEMP_TABLE_JSON_BEGIN)
    if begin < 0:
        return text.strip(), None
    end = text.find(TEMP_TABLE_JSON_END, begin)
    if end < 0:
        return text.strip(), None
    json_chunk = text[begin + len(TEMP_TABLE_JSON_BEGIN) : end].strip()
    visible = (text[:begin] + text[end + len(TEMP_TABLE_JSON_END) :]).strip()
    if not json_chunk:
        return visible, None
    try:
        parsed = json.loads(json_chunk)
    except (TypeError, ValueError, json.JSONDecodeError):
        return visible, None
    return visible, parsed if isinstance(parsed, dict) else None


def build_document_status_metadata() -> dict:
    """Metadados básicos para futuro endpoint de status da Agente Compara."""
    _require_session()
    maybe_cleanup_expired_cleiton_docs()
    sync_temp_table_with_session_documents()
    totals = get_document_session_totals()
    temp_table = get_active_temp_table_for_session()
    return {
        "domain": AGENTE_COMPARA_DOMAIN,
        "flow_types": {
            "upload": AGENTE_COMPARA_DOCUMENT_UPLOAD_FLOW_TYPE,
            "prepare": AGENTE_COMPARA_DOCUMENT_PREPARE_FLOW_TYPE,
            "chat": AGENTE_COMPARA_CHAT_FLOW_TYPE,
            "temp_table_extraction": AGENTE_COMPARA_TEMP_TABLE_EXTRACTION_FLOW_TYPE,
        },
        "documents": get_active_documents_for_session(),
        "temp_table": temp_table,
        "calculation_bases": get_active_calculation_bases_for_runtime(
            get_agente_compara_config().calculation_bases
        ),
        "allowed_formats": get_allowed_document_formats(),
        "session": {
            "count": totals["active_count"],
            "max_files": totals["max_files_per_session"],
            "total_bytes": totals["total_bytes"],
            "session_max_bytes": totals["session_max_bytes"],
        },
    }
