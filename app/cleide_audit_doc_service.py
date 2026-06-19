"""
Wrapper documental fino do domínio Cleide Auditoria (Fase 1).

Reaproveita preparação, store e configuração governados do Cleiton.
Opera exclusivamente sobre session keys `cleide_audit_*`; não toca sessão da Júlia.
Sem rotas, chat, IA ou billing.
"""
from __future__ import annotations

import csv
import io
import json
import logging
import re
import unicodedata
import zipfile
from datetime import UTC, datetime, timedelta
from uuid import uuid4

from flask import has_request_context, session
from werkzeug.utils import secure_filename

from app.cleiton_doc_contracts import (
    CONTEXT_KIND_GEMINI_FILE,
    CONTEXT_KIND_PLACEHOLDER,
    CONTEXT_KIND_TEXT,
    DOC_TYPE_CSV,
    DOC_TYPE_DOCX,
    DOC_TYPE_PDF,
    DOC_TYPE_TXT,
    DOC_TYPE_XLSX,
    DOC_TYPE_XML,
    ERROR_DOC_NOT_FOUND,
    ERROR_GEMINI_FILE_UPLOAD,
    ERROR_INVALID_SIZE,
    ERROR_MAX_FILES,
    ERROR_SESSION_BYTES,
    FIELD_CHAR_COUNT,
    FIELD_COLUMN_COUNT,
    FIELD_CONTEXT_KIND,
    FIELD_CONTEXT_REF,
    FIELD_CREATED_AT,
    FIELD_DISPLAY_NAME,
    FIELD_DOC_ID,
    FIELD_DOC_TYPE,
    FIELD_ERROR_CODE,
    FIELD_EXPIRES_AT,
    FIELD_EXTENSION,
    FIELD_GEMINI_FILE_NAME,
    FIELD_GEMINI_FILE_STATE,
    FIELD_GEMINI_FILE_URI,
    FIELD_GEMINI_MIME_TYPE,
    FIELD_GEMINI_UPLOADED_AT,
    FIELD_MAX_DEPTH,
    FIELD_MIME_TYPE,
    FIELD_NODE_COUNT,
    FIELD_PAGE_COUNT,
    FIELD_PDF_CONTEXT_READY,
    FIELD_PREPARED_CONTEXT,
    FIELD_ROW_COUNT,
    FIELD_SAFE_NAME,
    FIELD_SESSION_KEY,
    FIELD_SIZE_BYTES,
    FIELD_SOURCE_AGENT,
    FIELD_STATUS,
    FIELD_TRUNCATED,
    FIELD_WARNINGS,
    STATUS_ACTIVE,
    STATUS_ERROR,
)
from app.cleiton_doc_gemini_files import (
    pdf_context_ready_from_record,
    upload_pdf_to_gemini_files_api,
)
from app.cleiton_doc_prepare import prepare_document
from app.cleiton_doc_service import CleitonDocSessionError, maybe_cleanup_expired_cleiton_docs
from app.cleiton_doc_store import (
    get_cleiton_doc_tmp_dir,
    load_document_record,
    remove_document_record,
    save_document_record,
)
from app.services.cleiton_doc_config_service import get_cleiton_doc_config
from app.services.cleide_audit_config_service import get_cleide_audit_config

logger = logging.getLogger(__name__)

CLEIDE_AUDIT_DOMAIN = "cleide_audit"

CLEIDE_AUDIT_DOC_IDS_SESSION_KEY = "cleide_audit_doc_ids"
CLEIDE_AUDIT_DOC_CONTEXT_SESSION_KEY = "cleide_audit_doc_context"
CLEIDE_AUDIT_CHAT_HISTORY_SESSION_KEY = "cleide_audit_chat_history"
CLEIDE_AUDIT_UPLOAD_LOCK_SESSION_KEY = "cleide_audit_upload_lock"
CLEIDE_AUDIT_LAST_REQUEST_ID_SESSION_KEY = "cleide_audit_last_request_id"
CLEIDE_AUDIT_UPLOAD_IN_PROGRESS_SESSION_KEY = "cleide_audit_upload_in_progress"
CLEIDE_AUDIT_TEMP_TABLE_ID_SESSION_KEY = "cleide_audit_temp_table_id"
CLEIDE_AUDIT_TEMP_TABLE_SOURCE_DOCS_SESSION_KEY = "cleide_audit_temp_table_source_doc_ids"

TEMP_TABLE_STATUS_PROCESSING = "processing"
TEMP_TABLE_STATUS_AWAITING_VALIDATION = "awaiting_validation"
TEMP_TABLE_STATUS_VALIDATED = "validated"
TEMP_TABLE_STATUS_NEEDS_REVIEW = "needs_review"
TEMP_TABLE_STATUS_FAILED = "failed"
TEMP_TABLE_STATUS_EXPIRED = "expired"
TEMP_TABLE_STATUS_DISCARDED = "discarded"

TEMP_TABLE_VERSION_MARKER = "cleide_audit_temp_table_v1"
TEMP_TABLE_OPERATIONAL_OWNER = "cleiton"
TEMP_TABLE_UI_DISPLAY_NAME = "Tabela temporária extraída"
TEMP_TABLE_JSON_BEGIN = "---CLEIDE_TEMP_TABLE---"
TEMP_TABLE_JSON_END = "---END_CLEIDE_TEMP_TABLE---"

TEMP_TABLE_SAVE_MAX_PAYLOAD_BYTES = 512 * 1024
TEMP_TABLE_REVIEW_ACTION_SAVE_AND_ADVANCE = "save_and_advance"
HUMAN_REVIEW_STATUS_REVIEWED = "reviewed"
HUMAN_REVIEW_STATUS_EDITED = "edited"

ERROR_TEMP_TABLE_NOT_FOUND = "cleide_audit_temp_table_not_found"
ERROR_TEMP_TABLE_ID_MISMATCH = "cleide_audit_temp_table_id_mismatch"
ERROR_TEMP_TABLE_EXPIRED = "cleide_audit_temp_table_expired"
ERROR_TEMP_TABLE_INVALID_PAYLOAD = "cleide_audit_temp_table_invalid_payload"
ERROR_TEMP_TABLE_PAYLOAD_TOO_LARGE = "cleide_audit_temp_table_payload_too_large"
ERROR_TEMP_TABLE_SCOPE_MISMATCH = "cleide_audit_temp_table_scope_mismatch"

COVERAGE_TABLE_COLUMNS = ["UF destino", "Cidade destino", "Região de frete"]
COVERAGE_TABLE_STATUS_NEEDS_REVIEW = "needs_review"
COVERAGE_UPLOAD_MAX_BYTES = 512 * 1024
COVERAGE_UPLOAD_MAX_ROWS = 10000

ERROR_COVERAGE_NO_TEMP_TABLE = "cleide_audit_coverage_no_temp_table"
ERROR_COVERAGE_INVALID_FORMAT = "cleide_audit_coverage_invalid_format"
ERROR_COVERAGE_EMPTY_FILE = "cleide_audit_coverage_empty_file"
ERROR_COVERAGE_PAYLOAD_TOO_LARGE = "cleide_audit_coverage_payload_too_large"
ERROR_COVERAGE_PARSE_FAILED = "cleide_audit_coverage_parse_failed"
ERROR_COVERAGE_INVALID_PAYLOAD = "cleide_audit_coverage_invalid_payload"
ERROR_COVERAGE_EXPIRED = "cleide_audit_coverage_expired"
ERROR_COVERAGE_SCOPE_MISMATCH = "cleide_audit_coverage_scope_mismatch"

AUDIT_BATCH_SHEET_NAME = "Modelo Cleide"
AUDIT_INPUT_SCHEMA_VERSION = "cleide_audit_input_v1"
AUDIT_BATCH_STATUS_UPLOADED = "uploaded"
AUDIT_BATCH_STATUS_PROCESSED = "processed"
AUDIT_UPLOAD_MAX_BYTES = 2 * 1024 * 1024
CLEIDE_AUDIT_TEMPLATE_FILENAME = "template_cleide_auditoria_frete.xlsx"

ERROR_AUDIT_NO_TEMP_TABLE = "cleide_audit_audit_no_temp_table"
ERROR_AUDIT_INVALID_FORMAT = "cleide_audit_audit_invalid_format"
ERROR_AUDIT_EMPTY_FILE = "cleide_audit_audit_empty_file"
ERROR_AUDIT_PAYLOAD_TOO_LARGE = "cleide_audit_audit_payload_too_large"
ERROR_AUDIT_PARSE_FAILED = "cleide_audit_audit_parse_failed"
ERROR_AUDIT_MISSING_COLUMNS = "cleide_audit_audit_missing_columns"
ERROR_AUDIT_TOO_MANY_ROWS = "cleide_audit_audit_too_many_rows"
ERROR_AUDIT_EXPIRED = "cleide_audit_audit_expired"
ERROR_AUDIT_SCOPE_MISMATCH = "cleide_audit_audit_scope_mismatch"
ERROR_AUDIT_INVALID_SHEET = "cleide_audit_audit_invalid_sheet"
ERROR_AUDIT_EMPTY_ROWS = "cleide_audit_audit_empty_rows"
ERROR_AUDIT_BATCH_NOT_FOUND = "cleide_audit_audit_batch_not_found"
ERROR_AUDIT_BATCH_EMPTY = "cleide_audit_audit_batch_empty"

AUDIT_STATUS_OK = "ok"
AUDIT_STATUS_DIVERGENT = "divergent"
AUDIT_STATUS_MISSING_COVERAGE = "missing_coverage_mapping"
AUDIT_STATUS_AMBIGUOUS_COVERAGE = "ambiguous_coverage_mapping"
AUDIT_STATUS_MISSING_FREIGHT_RULE = "missing_freight_rule"
AUDIT_STATUS_INVALID_WEIGHT = "invalid_weight"
AUDIT_STATUS_INVALID_CHARGED_FREIGHT = "invalid_charged_freight"
AUDIT_STATUS_UNSUPPORTED_PRICING = "unsupported_pricing_model"

_AUDIT_REQUIRED_FIELDS = (
    "destination_city",
    "destination_uf",
    "charged_freight",
    "audited_weight",
)
_AUDIT_OPTIONAL_FIELDS = (
    "carrier",
    "document_number",
    "origin_city",
    "origin_uf",
    "invoice_value",
    "modal",
    "issue_date",
    "delivery_date",
)
_AUDIT_FIELD_LABELS = {
    "destination_city": "cidade_destino",
    "destination_uf": "uf_destino",
    "charged_freight": "valor_frete",
    "audited_weight": "peso",
    "carrier": "transportadora",
    "document_number": "numero_documento",
    "origin_city": "cidade_origem",
    "origin_uf": "uf_origem",
    "invoice_value": "valor_nf",
    "modal": "modal",
    "issue_date": "data_emissao",
    "delivery_date": "data_entrega",
}

_HTML_TAG_RE = re.compile(r"<[^>]+>")
_BR_UFS = frozenset(
    {
        "AC", "AL", "AP", "AM", "BA", "CE", "DF", "ES", "GO", "MA", "MT", "MS", "MG",
        "PA", "PB", "PR", "PE", "PI", "RJ", "RN", "RS", "RO", "RR", "SC", "SP", "SE", "TO",
    }
)
_COVERAGE_REJECT_GENERIC_HEADERS = frozenset(
    {
        "destino",
        "grupo",
        "tipo",
        "obs",
        "observacao",
        "observacoes",
        "nota",
        "notas",
        "comentario",
        "comentarios",
        "info",
        "informacao",
        "informacoes",
    }
)
_COVERAGE_FIELD_LABELS = {
    "destination_uf": "UF destino",
    "destination_city": "Cidade destino",
    "freight_region": "Região de frete",
}
_COVERAGE_FIELD_HINTS = {
    "destination_uf": "Você pode usar nomes como UF, Estado ou Unidade Federativa.",
    "destination_city": "Você pode usar nomes como Cidade, Município, Localidade ou Cidade atendida.",
    "freight_region": "Você pode usar nomes como Praça, Região, Rota, Itinerário, Área ou Zona.",
}
_COVERAGE_REQUIRED_FIELDS = ("destination_uf", "destination_city", "freight_region")
_COVERAGE_HEADER_ALIASES: dict[str, str] = {}

CLEIDE_AUDIT_DOCUMENT_UPLOAD_FLOW_TYPE = "cleide_audit_document_upload"
CLEIDE_AUDIT_DOCUMENT_PREPARE_FLOW_TYPE = "cleide_audit_document_prepare"
CLEIDE_AUDIT_CHAT_FLOW_TYPE = "cleide_audit_chat"
CLEIDE_AUDIT_TEMP_TABLE_EXTRACTION_FLOW_TYPE = "cleide_audit_temp_table_extraction"

SOURCE_AGENT_CLEIDE_AUDIT = "cleide_audit"


class CleideAuditTempTableError(ValueError):
    def __init__(self, error_code: str, message: str):
        super().__init__(message)
        self.error_code = error_code
        self.message = message


class CleideAuditCoverageError(ValueError):
    def __init__(self, error_code: str, message: str):
        super().__init__(message)
        self.error_code = error_code
        self.message = message


class CleideAuditBatchError(ValueError):
    def __init__(self, error_code: str, message: str):
        super().__init__(message)
        self.error_code = error_code
        self.message = message


def cleide_audit_upload_idempotency_key(request_id: str) -> str:
    return f"cleide-audit-upload:{(request_id or '').strip()}"


def cleide_audit_upload_doc_idempotency_key(doc_id: str) -> str:
    return f"cleide-audit-upload-doc:{(doc_id or '').strip()}"


def cleide_audit_chat_idempotency_key(request_id: str) -> str:
    return f"cleide-audit-chat:{(request_id or '').strip()}"


def cleide_audit_temp_table_extraction_idempotency_key(source_doc_ids: list[str]) -> str:
    normalized = _normalize_source_doc_ids(source_doc_ids)
    joined = ":".join(normalized)
    return f"cleide-audit-temp-table:{TEMP_TABLE_VERSION_MARKER}:{joined}"


def _utcnow() -> datetime:
    return datetime.now(UTC).replace(tzinfo=None)


def _require_session() -> None:
    if not has_request_context():
        raise RuntimeError("Sessão documental da Cleide Auditoria requer request context Flask.")


def _mark_session_modified() -> None:
    session.modified = True


def get_cleide_audit_doc_ids(session_obj) -> list[str]:
    raw = session_obj.get(CLEIDE_AUDIT_DOC_IDS_SESSION_KEY)
    if not isinstance(raw, list):
        return []
    ids: list[str] = []
    for item in raw:
        if isinstance(item, str):
            ref = item.strip()
            if ref:
                ids.append(ref)
    return ids


def set_cleide_audit_doc_ids(session_obj, doc_ids: list[str]) -> None:
    cleaned: list[str] = []
    seen: set[str] = set()
    for item in doc_ids or []:
        if not isinstance(item, str):
            continue
        ref = item.strip()
        if not ref or ref in seen:
            continue
        seen.add(ref)
        cleaned.append(ref)
    session_obj[CLEIDE_AUDIT_DOC_IDS_SESSION_KEY] = cleaned


def clear_cleide_audit_doc_ids(session_obj) -> None:
    session_obj.pop(CLEIDE_AUDIT_DOC_IDS_SESSION_KEY, None)


def append_cleide_audit_doc_id(session_obj, doc_id: str) -> None:
    ref = (doc_id or "").strip()
    if not ref:
        raise ValueError("doc_id inválido para sessão documental da Cleide Auditoria.")
    ids = get_cleide_audit_doc_ids(session_obj)
    if ref not in ids:
        ids.append(ref)
    set_cleide_audit_doc_ids(session_obj, ids)


def remove_cleide_audit_doc_id(session_obj, doc_id: str) -> None:
    ref = (doc_id or "").strip()
    if not ref:
        return
    ids = [item for item in get_cleide_audit_doc_ids(session_obj) if item != ref]
    if ids:
        set_cleide_audit_doc_ids(session_obj, ids)
    else:
        clear_cleide_audit_doc_ids(session_obj)


def _parse_size_bytes(size_bytes) -> int:
    if size_bytes is None:
        raise CleitonDocSessionError(
            ERROR_INVALID_SIZE,
            "size_bytes ausente ou inválido para documento Cleide Auditoria.",
        )
    if isinstance(size_bytes, bool):
        raise CleitonDocSessionError(
            ERROR_INVALID_SIZE,
            "size_bytes não numérico para documento Cleide Auditoria.",
        )
    try:
        parsed = int(size_bytes)
    except (TypeError, ValueError):
        raise CleitonDocSessionError(
            ERROR_INVALID_SIZE,
            "size_bytes não numérico para documento Cleide Auditoria.",
        ) from None
    if parsed <= 0:
        raise CleitonDocSessionError(
            ERROR_INVALID_SIZE,
            "size_bytes deve ser maior que zero para documento Cleide Auditoria.",
        )
    return parsed


def _normalize_extension(extension: str) -> str:
    raw = (extension or "").strip().lower()
    if not raw:
        return ""
    return raw if raw.startswith(".") else f".{raw}"


def _build_safe_display_name(display_name: str, extension: str) -> tuple[str, str]:
    raw = (display_name or "").strip() or "documento"
    safe = secure_filename(raw) or "documento"
    ext = _normalize_extension(extension)
    if ext and not safe.lower().endswith(ext):
        safe_name = f"{safe}{ext}"
    else:
        safe_name = safe
        if not ext and "." in safe_name:
            ext = "." + safe_name.rsplit(".", 1)[-1].lower()
    return raw, safe_name


def _context_ref_for_kind(context_kind: str, doc_id: str) -> str:
    kind = (context_kind or CONTEXT_KIND_PLACEHOLDER).strip() or CONTEXT_KIND_PLACEHOLDER
    if kind == CONTEXT_KIND_TEXT:
        return f"text:{doc_id}"
    if kind == CONTEXT_KIND_GEMINI_FILE:
        return f"gemini_file:{doc_id}"
    return f"placeholder:{doc_id}"


def _public_record(record: dict) -> dict:
    pdf_ready = pdf_context_ready_from_record(record)
    return {
        FIELD_DOC_ID: record.get(FIELD_DOC_ID),
        FIELD_DOC_TYPE: record.get(FIELD_DOC_TYPE),
        FIELD_DISPLAY_NAME: record.get(FIELD_DISPLAY_NAME),
        FIELD_SAFE_NAME: record.get(FIELD_SAFE_NAME),
        FIELD_EXTENSION: record.get(FIELD_EXTENSION),
        FIELD_MIME_TYPE: record.get(FIELD_MIME_TYPE),
        FIELD_SIZE_BYTES: record.get(FIELD_SIZE_BYTES),
        FIELD_CREATED_AT: record.get(FIELD_CREATED_AT),
        FIELD_EXPIRES_AT: record.get(FIELD_EXPIRES_AT),
        FIELD_STATUS: record.get(FIELD_STATUS),
        FIELD_TRUNCATED: record.get(FIELD_TRUNCATED),
        FIELD_CONTEXT_KIND: record.get(FIELD_CONTEXT_KIND),
        FIELD_CONTEXT_REF: record.get(FIELD_CONTEXT_REF),
        FIELD_CHAR_COUNT: record.get(FIELD_CHAR_COUNT),
        FIELD_ROW_COUNT: record.get(FIELD_ROW_COUNT),
        FIELD_COLUMN_COUNT: record.get(FIELD_COLUMN_COUNT),
        FIELD_PAGE_COUNT: record.get(FIELD_PAGE_COUNT),
        FIELD_NODE_COUNT: record.get(FIELD_NODE_COUNT),
        FIELD_MAX_DEPTH: record.get(FIELD_MAX_DEPTH),
        FIELD_WARNINGS: record.get(FIELD_WARNINGS) or [],
        FIELD_SESSION_KEY: record.get(FIELD_SESSION_KEY),
        FIELD_ERROR_CODE: record.get(FIELD_ERROR_CODE),
        FIELD_PDF_CONTEXT_READY: pdf_ready,
    }


def get_allowed_document_formats() -> list[dict]:
    """Retorna formatos habilitados a partir da config central do Cleiton."""
    cfg = get_cleiton_doc_config()
    catalog = [
        (DOC_TYPE_TXT, ".txt", cfg.txt_enabled),
        (DOC_TYPE_XML, ".xml", cfg.xml_enabled),
        (DOC_TYPE_CSV, ".csv", cfg.csv_enabled),
        (DOC_TYPE_XLSX, ".xlsx", cfg.excel_enabled),
        (DOC_TYPE_DOCX, ".docx", cfg.docx_enabled),
        (DOC_TYPE_PDF, ".pdf", cfg.pdf_enabled),
    ]
    return [
        {"doc_type": doc_type, "extension": extension, "enabled": bool(enabled)}
        for doc_type, extension, enabled in catalog
        if enabled
    ]


def cleanup_expired_documents_for_session() -> int:
    _require_session()
    cfg = get_cleiton_doc_config()
    removed = 0
    stale_ids: list[str] = []

    for doc_id in get_cleide_audit_doc_ids(session):
        record = load_document_record(doc_id, ttl_hours=cfg.upload_ttl_hours)
        if record is None:
            stale_ids.append(doc_id)
            removed += 1

    for doc_id in stale_ids:
        remove_document_record(doc_id)
        remove_cleide_audit_doc_id(session, doc_id)

    if stale_ids:
        _mark_session_modified()
    return removed


def get_active_documents_for_session() -> list[dict]:
    _require_session()
    cfg = get_cleiton_doc_config()
    active: list[dict] = []
    stale_ids: list[str] = []

    for doc_id in get_cleide_audit_doc_ids(session):
        record = load_document_record(doc_id, ttl_hours=cfg.upload_ttl_hours)
        if record is None:
            stale_ids.append(doc_id)
            continue
        active.append(_public_record(record))

    if stale_ids:
        for doc_id in stale_ids:
            remove_cleide_audit_doc_id(session, doc_id)
        _mark_session_modified()

    return active


def get_document_session_totals() -> dict:
    _require_session()
    cfg = get_cleiton_doc_config()
    cleanup_expired_documents_for_session()
    active = get_active_documents_for_session()
    total_bytes = sum(int(item.get(FIELD_SIZE_BYTES) or 0) for item in active)
    active_count = len(active)
    max_files = int(cfg.max_files_per_session)
    max_bytes = int(cfg.session_max_bytes)
    return {
        "active_count": active_count,
        "total_bytes": total_bytes,
        "max_files_per_session": max_files,
        "session_max_bytes": max_bytes,
        "remaining_files": max(0, max_files - active_count),
        "remaining_bytes": max(0, max_bytes - total_bytes),
    }


def assert_session_can_accept_document(size_bytes) -> None:
    _require_session()
    cfg = get_cleiton_doc_config()
    incoming = _parse_size_bytes(size_bytes)

    if incoming > int(cfg.session_max_bytes):
        raise CleitonDocSessionError(
            ERROR_INVALID_SIZE,
            "size_bytes excede o limite total configurado da sessão documental.",
        )

    totals = get_document_session_totals()

    if totals["active_count"] >= int(cfg.max_files_per_session):
        raise CleitonDocSessionError(
            ERROR_MAX_FILES,
            "Limite de arquivos documentais por sessão atingido.",
        )

    projected = int(totals["total_bytes"]) + incoming
    if projected > int(cfg.session_max_bytes):
        raise CleitonDocSessionError(
            ERROR_SESSION_BYTES,
            "Limite total de bytes documentais da sessão excedido.",
        )


def _register_document_record(
    *,
    display_name: str,
    extension: str,
    mime_type: str,
    size_bytes: int,
    context_kind: str = CONTEXT_KIND_PLACEHOLDER,
    context_ref: str | None = None,
    truncated: bool = False,
    doc_type: str | None = None,
    prepared_context: str | None = None,
    char_count: int | None = None,
    row_count: int | None = None,
    column_count: int | None = None,
    page_count: int | None = None,
    node_count: int | None = None,
    max_depth: int | None = None,
    warnings: list[str] | None = None,
    status: str = STATUS_ACTIVE,
    error_code: str | None = None,
    gemini_file_name: str | None = None,
    gemini_file_uri: str | None = None,
    gemini_mime_type: str | None = None,
    gemini_file_state: str | None = None,
    gemini_uploaded_at: str | None = None,
) -> dict:
    _require_session()
    cfg = get_cleiton_doc_config()
    validated_size = _parse_size_bytes(size_bytes)
    assert_session_can_accept_document(validated_size)

    doc_id = uuid4().hex
    display, safe_name = _build_safe_display_name(display_name, extension)
    ext = _normalize_extension(extension) or (
        f".{safe_name.rsplit('.', 1)[-1].lower()}" if "." in safe_name else ""
    )
    created_at = _utcnow()
    expires_at = created_at + timedelta(hours=max(1, int(cfg.upload_ttl_hours)))
    resolved_kind = (context_kind or CONTEXT_KIND_PLACEHOLDER).strip() or CONTEXT_KIND_PLACEHOLDER

    record = {
        FIELD_DOC_ID: doc_id,
        FIELD_DOC_TYPE: doc_type,
        FIELD_DISPLAY_NAME: display,
        FIELD_SAFE_NAME: safe_name,
        FIELD_EXTENSION: ext,
        FIELD_MIME_TYPE: (mime_type or "application/octet-stream").strip(),
        FIELD_SIZE_BYTES: validated_size,
        FIELD_CREATED_AT: created_at.isoformat(),
        FIELD_EXPIRES_AT: expires_at.isoformat(),
        FIELD_STATUS: (status or STATUS_ACTIVE).strip() or STATUS_ACTIVE,
        FIELD_TRUNCATED: bool(truncated),
        FIELD_CONTEXT_KIND: resolved_kind,
        FIELD_CONTEXT_REF: context_ref or _context_ref_for_kind(resolved_kind, doc_id),
        FIELD_PREPARED_CONTEXT: prepared_context,
        FIELD_CHAR_COUNT: char_count,
        FIELD_ROW_COUNT: row_count,
        FIELD_COLUMN_COUNT: column_count,
        FIELD_PAGE_COUNT: page_count,
        FIELD_NODE_COUNT: node_count,
        FIELD_MAX_DEPTH: max_depth,
        FIELD_WARNINGS: list(warnings or []),
        FIELD_SOURCE_AGENT: SOURCE_AGENT_CLEIDE_AUDIT,
        FIELD_SESSION_KEY: CLEIDE_AUDIT_DOC_IDS_SESSION_KEY,
        FIELD_ERROR_CODE: error_code,
        FIELD_GEMINI_FILE_NAME: gemini_file_name,
        FIELD_GEMINI_FILE_URI: gemini_file_uri,
        FIELD_GEMINI_MIME_TYPE: gemini_mime_type,
        FIELD_GEMINI_FILE_STATE: gemini_file_state,
        FIELD_GEMINI_UPLOADED_AT: gemini_uploaded_at,
    }

    save_document_record(record)
    append_cleide_audit_doc_id(session, doc_id)
    _mark_session_modified()
    return _public_record(record)


def prepare_and_register_document(
    *,
    display_name: str,
    file_bytes: bytes,
    mime_type: str | None = None,
    extension: str | None = None,
) -> dict:
    """
    Valida, prepara e registra documento na sessão Cleide Auditoria após sucesso.

    Delega validação/preparação ao Cleiton; persiste IDs em `cleide_audit_doc_ids`.
    """
    prepared = prepare_document(
        display_name=display_name,
        file_bytes=file_bytes,
        mime_type=mime_type,
        extension=extension,
    )

    register_kwargs = {
        "display_name": prepared["display_name"],
        "extension": prepared[FIELD_EXTENSION],
        "mime_type": prepared[FIELD_MIME_TYPE],
        "size_bytes": prepared[FIELD_SIZE_BYTES],
        "context_kind": prepared[FIELD_CONTEXT_KIND],
        "truncated": prepared[FIELD_TRUNCATED],
        "doc_type": prepared[FIELD_DOC_TYPE],
        "prepared_context": prepared[FIELD_PREPARED_CONTEXT],
        "char_count": prepared[FIELD_CHAR_COUNT],
        "row_count": prepared[FIELD_ROW_COUNT],
        "column_count": prepared[FIELD_COLUMN_COUNT],
        "page_count": prepared[FIELD_PAGE_COUNT],
        "node_count": prepared[FIELD_NODE_COUNT],
        "max_depth": prepared[FIELD_MAX_DEPTH],
        "warnings": prepared[FIELD_WARNINGS],
    }

    if prepared[FIELD_CONTEXT_KIND] == CONTEXT_KIND_GEMINI_FILE:
        cfg = get_cleiton_doc_config()
        upload_result = upload_pdf_to_gemini_files_api(
            file_bytes=file_bytes,
            mime_type=prepared[FIELD_MIME_TYPE],
            display_name=prepared["display_name"],
            page_count=prepared[FIELD_PAGE_COUNT],
            max_pages=int(cfg.pdf_max_pages),
        )
        register_kwargs["prepared_context"] = upload_result.prepared_context or prepared[FIELD_PREPARED_CONTEXT]
        register_kwargs["warnings"] = list(prepared[FIELD_WARNINGS]) + list(upload_result.warnings or [])
        if upload_result.ok:
            register_kwargs.update(
                {
                    "status": STATUS_ACTIVE,
                    "gemini_file_name": upload_result.gemini_file_name,
                    "gemini_file_uri": upload_result.gemini_file_uri,
                    "gemini_mime_type": upload_result.gemini_mime_type,
                    "gemini_file_state": upload_result.gemini_file_state,
                    "gemini_uploaded_at": upload_result.gemini_uploaded_at,
                }
            )
        else:
            register_kwargs.update(
                {
                    "status": STATUS_ERROR,
                    "error_code": ERROR_GEMINI_FILE_UPLOAD,
                    "gemini_file_name": upload_result.gemini_file_name,
                    "gemini_file_uri": upload_result.gemini_file_uri,
                    "gemini_mime_type": upload_result.gemini_mime_type,
                    "gemini_file_state": upload_result.gemini_file_state,
                }
            )
            logger.warning(
                "Cleide audit doc: upload Gemini Files API falhou para PDF (summary=%s).",
                upload_result.error_summary,
            )

    return _register_document_record(**register_kwargs)


def remove_document_from_session(doc_id: str) -> dict:
    _require_session()
    ref = (doc_id or "").strip()
    if not ref:
        return {
            "ok": False,
            "doc_id": doc_id,
            "removed_from_store": False,
            "removed_from_session": False,
            "error_code": ERROR_DOC_NOT_FOUND,
        }

    store_result = remove_document_record(ref)
    had_session_ref = ref in get_cleide_audit_doc_ids(session)
    if had_session_ref:
        remove_cleide_audit_doc_id(session, ref)
        _mark_session_modified()

    invalidate_temp_table_if_source_changed(
        reason=TEMP_TABLE_STATUS_DISCARDED,
        removed_doc_id=ref,
    )

    return {
        "ok": True,
        "doc_id": ref,
        "removed_from_store": bool(store_result.get("removed")),
        "removed_from_session": had_session_ref,
        "error_code": None if had_session_ref or store_result.get("removed") else ERROR_DOC_NOT_FOUND,
    }


def clear_documents_for_session() -> dict:
    _require_session()
    ids = get_cleide_audit_doc_ids(session)
    removed_store = 0
    removed_session = 0
    for doc_id in ids:
        result = remove_document_record(doc_id)
        if result.get("removed"):
            removed_store += 1
        removed_session += 1
    clear_cleide_audit_doc_ids(session)
    invalidate_temp_table_for_session(reason=TEMP_TABLE_STATUS_DISCARDED)
    _mark_session_modified()
    return {
        "ok": True,
        "requested": len(ids),
        "removed_from_store": removed_store,
        "removed_from_session": removed_session,
    }


def _temp_table_filename(temp_table_id: str) -> str:
    ref = (temp_table_id or "").strip()
    if not ref:
        raise ValueError("temp_table_id inválido.")
    return f"tt_{ref}.json"


def _temp_table_path(temp_table_id: str):
    from pathlib import Path

    safe_name = _temp_table_filename(temp_table_id)
    base = Path(get_cleiton_doc_tmp_dir()).resolve()
    candidate = (base / safe_name).resolve()
    if candidate.parent != base:
        raise ValueError("temp_table path inválido.")
    return candidate


def _write_temp_table_atomic(path, payload: dict) -> None:
    from app.cleiton_doc_store import _write_json_atomic

    _write_json_atomic(path, payload)


def _normalize_source_doc_ids(doc_ids: list[str] | None) -> list[str]:
    cleaned: list[str] = []
    seen: set[str] = set()
    for item in doc_ids or []:
        if not isinstance(item, str):
            continue
        ref = item.strip()
        if not ref or ref in seen:
            continue
        seen.add(ref)
        cleaned.append(ref)
    return cleaned


def get_temp_table_id(session_obj) -> str | None:
    raw = session_obj.get(CLEIDE_AUDIT_TEMP_TABLE_ID_SESSION_KEY)
    if not isinstance(raw, str):
        return None
    ref = raw.strip()
    return ref or None


def set_temp_table_id(session_obj, temp_table_id: str | None) -> None:
    ref = (temp_table_id or "").strip()
    if ref:
        session_obj[CLEIDE_AUDIT_TEMP_TABLE_ID_SESSION_KEY] = ref
    else:
        session_obj.pop(CLEIDE_AUDIT_TEMP_TABLE_ID_SESSION_KEY, None)


def get_temp_table_source_doc_ids(session_obj) -> list[str]:
    raw = session_obj.get(CLEIDE_AUDIT_TEMP_TABLE_SOURCE_DOCS_SESSION_KEY)
    if not isinstance(raw, list):
        return []
    return _normalize_source_doc_ids(raw)


def set_temp_table_source_doc_ids(session_obj, doc_ids: list[str]) -> None:
    session_obj[CLEIDE_AUDIT_TEMP_TABLE_SOURCE_DOCS_SESSION_KEY] = _normalize_source_doc_ids(doc_ids)


def clear_temp_table_session_refs(session_obj) -> None:
    session_obj.pop(CLEIDE_AUDIT_TEMP_TABLE_ID_SESSION_KEY, None)
    session_obj.pop(CLEIDE_AUDIT_TEMP_TABLE_SOURCE_DOCS_SESSION_KEY, None)


def _normalize_coverage_header(value) -> str:
    if value is None:
        return ""
    text = str(value).strip()
    text = unicodedata.normalize("NFKD", text)
    text = "".join(ch for ch in text if not unicodedata.combining(ch))
    text = text.lower()
    text = re.sub(r"[_\-\/.:]+", " ", text)
    text = re.sub(r"[^a-z0-9\s]", " ", text)
    text = re.sub(r"\s+", " ", text)
    return text.strip()


def _coverage_alias_entries() -> tuple[tuple[str, tuple[str, ...]], ...]:
    return (
        (
            "destination_uf",
            (
                "uf",
                "uf destino",
                "uf de destino",
                "uf dest",
                "uf destinatario",
                "uf destinatário",
                "uf destin",
                "uf destno",
                "estado",
                "estado destino",
                "estado de destino",
                "estado dest",
                "estado destin",
                "estado detino",
                "sigla uf",
                "sigla estado",
                "unidade federativa",
                "uf entrega",
                "uf de entrega",
                "uf_destino",
            ),
        ),
        (
            "destination_city",
            (
                "cidade",
                "cidade destino",
                "cidade de destino",
                "cidade dest",
                "cidade destinatario",
                "cidade destinatário",
                "cidade destin",
                "cidade detino",
                "cid destino",
                "municipio",
                "município",
                "municipio destino",
                "município destino",
                "municipio de destino",
                "municipo",
                "municípo",
                "municip",
                "mun destino",
                "localidade",
                "localidade destino",
                "cidade atendida",
                "cidade de atendimento",
                "cidade entrega",
                "cidade de entrega",
                "destino cidade",
                "cidade_destino",
            ),
        ),
        (
            "freight_region",
            (
                "praca",
                "praça",
                "praca destino",
                "praça destino",
                "praca dest",
                "prasa",
                "regiao",
                "região",
                "regiao de frete",
                "região de frete",
                "regiao frete",
                "região frete",
                "regiao fret",
                "regiao fretee",
                "regiao de fret",
                "regiao de fretee",
                "regiao de atendimento",
                "região de atendimento",
                "regiao atendida",
                "região atendida",
                "regiao destino",
                "região destino",
                "area",
                "área",
                "area de atendimento",
                "área de atendimento",
                "zona",
                "zona de entrega",
                "setor",
                "setor de entrega",
                "rota",
                "rota destino",
                "rota de entrega",
                "codigo regiao",
                "código região",
                "cod regiao",
                "cód regiao",
                "itinerario",
                "itinerário",
                "itinerario entrega",
                "itinerário entrega",
                "itinerarioo",
                "regional",
                "regional destino",
                "faixa regional",
                "grupo destino",
                "grupo de destino",
                "regiao_frete",
            ),
        ),
    )


def _build_coverage_header_aliases() -> dict[str, str]:
    aliases: dict[str, str] = {}
    for field_name, labels in _coverage_alias_entries():
        for label in labels:
            key = _normalize_coverage_header(label)
            if not key or key in _COVERAGE_REJECT_GENERIC_HEADERS:
                continue
            aliases.setdefault(key, field_name)
    return aliases


def _ensure_coverage_header_aliases() -> dict[str, str]:
    global _COVERAGE_HEADER_ALIASES
    if not _COVERAGE_HEADER_ALIASES:
        _COVERAGE_HEADER_ALIASES = _build_coverage_header_aliases()
    return _COVERAGE_HEADER_ALIASES


def _coverage_edit_distance(left: str, right: str) -> int:
    if left == right:
        return 0
    if not left:
        return len(right)
    if not right:
        return len(left)
    prev = list(range(len(right) + 1))
    for i, left_char in enumerate(left, start=1):
        current = [i]
        for j, right_char in enumerate(right, start=1):
            cost = 0 if left_char == right_char else 1
            current.append(
                min(
                    current[-1] + 1,
                    prev[j] + 1,
                    prev[j - 1] + cost,
                )
            )
        prev = current
    return prev[-1]


def _coverage_fuzzy_max_distance(alias_key: str) -> int:
    if len(alias_key) <= 4:
        return 1
    if len(alias_key) <= 7:
        return 1
    return 2


def _resolve_coverage_field(header: str) -> str | None:
    normalized = _normalize_coverage_header(header)
    if not normalized:
        return None
    aliases = _ensure_coverage_header_aliases()
    exact = aliases.get(normalized)
    if exact:
        return exact
    if normalized in _COVERAGE_REJECT_GENERIC_HEADERS:
        return None

    best_field: str | None = None
    best_distance = 999
    second_best_distance = 999
    for alias_key, field_name in aliases.items():
        max_distance = _coverage_fuzzy_max_distance(alias_key)
        distance = _coverage_edit_distance(normalized, alias_key)
        if distance > max_distance:
            continue
        if distance < best_distance:
            second_best_distance = best_distance
            best_distance = distance
            best_field = field_name
        elif distance == best_distance and field_name != best_field:
            second_best_distance = distance
    if best_field is None:
        return None
    if second_best_distance == best_distance:
        return None
    return best_field


def _coverage_column_sample_values(raw_rows: list[list], col_index: int, *, max_rows: int = 50) -> list[str]:
    values: list[str] = []
    for raw_row in raw_rows[1 : max_rows + 1]:
        if not isinstance(raw_row, list) or col_index >= len(raw_row):
            continue
        value = str(raw_row[col_index] or "").strip()
        if value:
            values.append(value)
    return values


def _score_coverage_column_as_uf(values: list[str]) -> float:
    if not values:
        return 0.0
    hits = sum(1 for value in values if _normalize_destination_uf(value) in _BR_UFS)
    return hits / len(values)


def _score_coverage_column_as_city(values: list[str]) -> float:
    if not values:
        return 0.0
    hits = 0
    for value in values:
        if _normalize_destination_uf(value) in _BR_UFS:
            continue
        cleaned = _sanitize_cell_string(value)
        if not cleaned or len(cleaned) < 3:
            continue
        alpha_ratio = sum(ch.isalpha() or ch.isspace() for ch in cleaned) / len(cleaned)
        if alpha_ratio >= 0.7:
            hits += 1
    return hits / len(values)


_REGION_NAME_HINTS = frozenset(
    {
        "norte",
        "nordeste",
        "sul",
        "sudeste",
        "centro oeste",
        "centro-oeste",
        "interior",
        "capital",
        "metropolitana",
        "litoral",
        "fluvial",
        "fluvias",
    }
)
_REGION_PATTERN_RE = re.compile(
    r"^[A-Za-z]{2}\s*[-/]\s*.+|.+\s+(interior|capital|metropolitana|litoral)\b.*",
    re.IGNORECASE,
)


def _score_coverage_column_as_region(values: list[str]) -> float:
    if not values:
        return 0.0
    hits = 0
    unique_values = {value.casefold() for value in values}
    repetition_bonus = 0.0
    if len(values) >= 3:
        repetition_bonus = min(0.25, (1 - (len(unique_values) / len(values))) * 0.5)
    for value in values:
        cleaned = _sanitize_cell_string(value)
        if not cleaned:
            continue
        normalized = _normalize_coverage_header(cleaned)
        if _REGION_PATTERN_RE.match(cleaned):
            hits += 1
            continue
        if any(hint in normalized for hint in _REGION_NAME_HINTS):
            hits += 1
            continue
        if re.search(r"\d", cleaned) and re.search(r"[A-Za-z]{2}", cleaned):
            hits += 0.75
    base = hits / len(values)
    return min(1.0, base + repetition_bonus)


def _infer_coverage_columns_from_content(
    raw_rows: list[list],
    field_indexes: dict[str, int],
    header_row: list,
) -> dict[str, int]:
    used_indexes = set(field_indexes.values())
    candidate_indexes = [
        index for index in range(len(header_row)) if index not in used_indexes
    ]
    if not candidate_indexes:
        return {}

    missing_fields = [field for field in _COVERAGE_REQUIRED_FIELDS if field not in field_indexes]
    if not missing_fields:
        return {}

    scores: dict[str, dict[int, float]] = {field: {} for field in missing_fields}
    for col_index in candidate_indexes:
        values = _coverage_column_sample_values(raw_rows, col_index)
        if not values:
            continue
        if "destination_uf" in missing_fields:
            scores["destination_uf"][col_index] = _score_coverage_column_as_uf(values)
        if "destination_city" in missing_fields:
            scores["destination_city"][col_index] = _score_coverage_column_as_city(values)
        if "freight_region" in missing_fields:
            scores["freight_region"][col_index] = _score_coverage_column_as_region(values)

    inferred: dict[str, int] = {}
    assigned: set[int] = set()
    thresholds = {
        "destination_uf": 0.75,
        "destination_city": 0.55,
        "freight_region": 0.45,
    }

    while True:
        best_field: str | None = None
        best_col: int | None = None
        best_score = -1.0
        for field in missing_fields:
            if field in inferred:
                continue
            threshold = thresholds[field]
            for col_index, score in scores.get(field, {}).items():
                if col_index in assigned or score < threshold:
                    continue
                if score > best_score:
                    best_score = score
                    best_field = field
                    best_col = col_index
        if best_field is None or best_col is None:
            break
        inferred[best_field] = best_col
        assigned.add(best_col)

    return inferred


def _resolve_coverage_field_indexes(raw_rows: list[list]) -> dict[str, int]:
    header_row = raw_rows[0]
    field_indexes: dict[str, int] = {}
    for index, header in enumerate(header_row):
        field_name = _resolve_coverage_field(str(header or ""))
        if field_name and field_name not in field_indexes:
            field_indexes[field_name] = index

    missing_fields = [field for field in _COVERAGE_REQUIRED_FIELDS if field not in field_indexes]
    if missing_fields:
        inferred = _infer_coverage_columns_from_content(raw_rows, field_indexes, header_row)
        for field_name, index in inferred.items():
            if field_name not in field_indexes:
                field_indexes[field_name] = index

    return field_indexes


def _format_coverage_missing_columns_error(missing_fields: list[str]) -> str:
    labels = [_COVERAGE_FIELD_LABELS[field] for field in missing_fields if field in _COVERAGE_FIELD_LABELS]
    hints = [
        _COVERAGE_FIELD_HINTS[field]
        for field in missing_fields
        if field in _COVERAGE_FIELD_HINTS
    ]
    message = "Colunas obrigatórias ausentes: " + ", ".join(labels) + "."
    if hints:
        message += " " + " ".join(hints)
    return message


def _normalize_destination_uf(value) -> str | None:
    cleaned = _sanitize_cell_string(value)
    if not cleaned:
        return None
    candidate = re.sub(r"[^A-Za-z]", "", cleaned).upper()
    if len(candidate) != 2:
        return None
    return candidate


def _empty_coverage_table_shell(*, uploaded_at: str | None = None) -> dict:
    return {
        "status": COVERAGE_TABLE_STATUS_NEEDS_REVIEW,
        "columns": list(COVERAGE_TABLE_COLUMNS),
        "rows": [],
        "validation_warnings": [],
        "human_review_status": None,
        "human_edited_at": None,
        "human_edited_by_user_id": None,
        "edit_version": 0,
        "uploaded_at": uploaded_at or _utcnow().isoformat(),
    }


def _public_coverage_table(coverage) -> dict | None:
    if not isinstance(coverage, dict):
        return None
    rows = coverage.get("rows")
    return {
        "status": coverage.get("status"),
        "columns": list(coverage.get("columns") or COVERAGE_TABLE_COLUMNS),
        "rows": list(rows) if isinstance(rows, list) else [],
        "validation_warnings": list(coverage.get("validation_warnings") or []),
        "human_review_status": coverage.get("human_review_status"),
        "human_edited_at": coverage.get("human_edited_at"),
        "human_edited_by_user_id": coverage.get("human_edited_by_user_id"),
        "edit_version": coverage.get("edit_version"),
        "uploaded_at": coverage.get("uploaded_at"),
    }


def _decode_coverage_csv_bytes(file_bytes: bytes) -> str:
    for encoding in ("utf-8-sig", "utf-8", "cp1252", "latin-1"):
        try:
            return (file_bytes or b"").decode(encoding)
        except UnicodeDecodeError:
            continue
    raise CleideAuditCoverageError(
        ERROR_COVERAGE_PARSE_FAILED,
        "Não foi possível decodificar o arquivo CSV de cobertura.",
    )


def _parse_coverage_tabular_rows(raw_rows: list[list], *, source_file_name: str) -> tuple[list[dict], list[str]]:
    if not raw_rows:
        raise CleideAuditCoverageError(
            ERROR_COVERAGE_EMPTY_FILE,
            "O arquivo de cobertura está vazio.",
        )
    header_row = raw_rows[0]
    if not isinstance(header_row, list):
        raise CleideAuditCoverageError(
            ERROR_COVERAGE_PARSE_FAILED,
            "Cabeçalho do arquivo de cobertura inválido.",
        )
    field_indexes = _resolve_coverage_field_indexes(raw_rows)
    missing_fields = [
        field for field in _COVERAGE_REQUIRED_FIELDS if field not in field_indexes
    ]
    if missing_fields:
        raise CleideAuditCoverageError(
            ERROR_COVERAGE_PARSE_FAILED,
            _format_coverage_missing_columns_error(missing_fields),
        )

    parsed_rows: list[dict] = []
    warnings: list[str] = []
    row_index = 0
    for raw_row in raw_rows[1:]:
        if not isinstance(raw_row, list):
            continue
        if not any(str(cell or "").strip() for cell in raw_row):
            continue
        row_index += 1
        if row_index > COVERAGE_UPLOAD_MAX_ROWS:
            raise CleideAuditCoverageError(
                ERROR_COVERAGE_PAYLOAD_TOO_LARGE,
                "O arquivo de cobertura excede o limite de linhas permitido.",
            )

        def _cell(field: str) -> str:
            idx = field_indexes[field]
            if idx >= len(raw_row):
                return ""
            return str(raw_row[idx] or "").strip()

        destination_uf = _normalize_destination_uf(_cell("destination_uf"))
        destination_city = _sanitize_cell_string(_cell("destination_city"))
        freight_region = _sanitize_cell_string(_cell("freight_region"))
        row_errors: list[str] = []
        if not destination_uf:
            row_errors.append("UF inválida ou ausente")
        if not destination_city:
            row_errors.append("cidade ausente")
        if not freight_region:
            row_errors.append("região de frete ausente")
        if row_errors:
            warnings.append(f"Linha {row_index} ignorada: {', '.join(row_errors)}.")
            continue
        parsed_rows.append(
            {
                "destination_uf": destination_uf,
                "destination_city": destination_city,
                "freight_region": freight_region,
                "row_index": row_index,
                "source_file_name": source_file_name,
                "confidence": None,
                "evidence_ref": None,
                "notes": "",
            }
        )
    if not parsed_rows and not warnings:
        raise CleideAuditCoverageError(
            ERROR_COVERAGE_EMPTY_FILE,
            "Nenhuma linha válida encontrada no arquivo de cobertura.",
        )
    return parsed_rows, warnings


def _parse_coverage_csv_bytes(file_bytes: bytes, *, source_file_name: str) -> tuple[list[dict], list[str]]:
    text = _decode_coverage_csv_bytes(file_bytes)
    if not text.strip():
        raise CleideAuditCoverageError(
            ERROR_COVERAGE_EMPTY_FILE,
            "O arquivo de cobertura está vazio.",
        )
    reader = csv.reader(io.StringIO(text))
    raw_rows = [row for row in reader]
    return _parse_coverage_tabular_rows(raw_rows, source_file_name=source_file_name)


def _parse_coverage_xlsx_bytes(file_bytes: bytes, *, source_file_name: str) -> tuple[list[dict], list[str]]:
    if not file_bytes:
        raise CleideAuditCoverageError(
            ERROR_COVERAGE_EMPTY_FILE,
            "O arquivo de cobertura está vazio.",
        )
    if len(file_bytes) > COVERAGE_UPLOAD_MAX_BYTES:
        raise CleideAuditCoverageError(
            ERROR_COVERAGE_PAYLOAD_TOO_LARGE,
            "O arquivo de cobertura excede o limite de tamanho permitido.",
        )
    try:
        if zipfile.is_zipfile(io.BytesIO(file_bytes)):
            with zipfile.ZipFile(io.BytesIO(file_bytes)) as archive:
                if archive.testzip() is not None:
                    raise CleideAuditCoverageError(
                        ERROR_COVERAGE_PARSE_FAILED,
                        "Arquivo XLSX de cobertura corrompido.",
                    )
        from openpyxl import load_workbook

        workbook = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    except CleideAuditCoverageError:
        raise
    except Exception as exc:
        raise CleideAuditCoverageError(
            ERROR_COVERAGE_PARSE_FAILED,
            "Não foi possível ler o arquivo XLSX de cobertura.",
        ) from exc

    raw_rows: list[list] = []
    try:
        sheet = workbook.active
        for row in sheet.iter_rows(values_only=True):
            raw_rows.append(["" if cell is None else str(cell) for cell in row])
    finally:
        workbook.close()
    return _parse_coverage_tabular_rows(raw_rows, source_file_name=source_file_name)


def _validate_coverage_row_for_save(item, *, row_index: int) -> dict:
    if not isinstance(item, dict):
        raise CleideAuditTempTableError(
            ERROR_TEMP_TABLE_INVALID_PAYLOAD,
            "Cada linha de coverage_table deve ser um objeto.",
        )
    destination_uf = _normalize_destination_uf(item.get("destination_uf"))
    destination_city = _sanitize_cell_string(item.get("destination_city"))
    freight_region = _sanitize_cell_string(item.get("freight_region"))
    notes = _sanitize_cell_string(item.get("notes")) or ""
    if not destination_uf:
        raise CleideAuditTempTableError(
            ERROR_TEMP_TABLE_INVALID_PAYLOAD,
            f"Linha {row_index}: UF destino inválida ou ausente.",
        )
    if not destination_city:
        raise CleideAuditTempTableError(
            ERROR_TEMP_TABLE_INVALID_PAYLOAD,
            f"Linha {row_index}: cidade destino é obrigatória.",
        )
    if not freight_region:
        raise CleideAuditTempTableError(
            ERROR_TEMP_TABLE_INVALID_PAYLOAD,
            f"Linha {row_index}: região de frete é obrigatória.",
        )
    return {
        "destination_uf": destination_uf,
        "destination_city": destination_city,
        "freight_region": freight_region,
        "row_index": row_index,
        "source_file_name": _sanitize_cell_string(item.get("source_file_name")),
        "confidence": item.get("confidence"),
        "evidence_ref": _sanitize_cell_string(item.get("evidence_ref")),
        "notes": notes,
    }


def _validate_coverage_table_for_save(raw_coverage) -> dict | None:
    if raw_coverage is None:
        return None
    if not isinstance(raw_coverage, dict):
        raise CleideAuditTempTableError(
            ERROR_TEMP_TABLE_INVALID_PAYLOAD,
            "coverage_table deve ser um objeto.",
        )
    raw_rows = raw_coverage.get("rows")
    if raw_rows is None:
        return {"rows": []}
    if not isinstance(raw_rows, list):
        raise CleideAuditTempTableError(
            ERROR_TEMP_TABLE_INVALID_PAYLOAD,
            "coverage_table.rows deve ser uma lista.",
        )
    normalized_rows: list[dict] = []
    for index, row in enumerate(raw_rows, start=1):
        normalized_rows.append(_validate_coverage_row_for_save(row, row_index=index))
    return {"rows": normalized_rows}


def upload_coverage_table_from_file(
    *,
    display_name: str,
    file_bytes: bytes,
    extension: str | None,
    user_scope=None,
    franquia_scope=None,
) -> dict:
    """
    Upload complementar determinístico de coverage_table no tt_*.json ativo.

    Não registra documento principal, não chama Gemini e não dispara extração de frete.
    """
    _require_session()
    if not file_bytes:
        raise CleideAuditCoverageError(
            ERROR_COVERAGE_EMPTY_FILE,
            "O arquivo de cobertura está vazio.",
        )
    if len(file_bytes) > COVERAGE_UPLOAD_MAX_BYTES:
        raise CleideAuditCoverageError(
            ERROR_COVERAGE_PAYLOAD_TOO_LARGE,
            "O arquivo de cobertura excede o limite de tamanho permitido.",
        )

    ext = (extension or "").strip().lower()
    if not ext.startswith("."):
        ext = f".{ext}" if ext else ""
    if ext == ".pdf":
        raise CleideAuditCoverageError(
            ERROR_COVERAGE_INVALID_FORMAT,
            "Upload de cobertura aceita apenas CSV e XLSX nesta fase.",
        )
    if ext not in {".csv", ".xlsx"}:
        raise CleideAuditCoverageError(
            ERROR_COVERAGE_INVALID_FORMAT,
            "Upload de cobertura aceita apenas CSV e XLSX nesta fase.",
        )

    sync_temp_table_with_session_documents()
    active_id = get_temp_table_id(session)
    if not active_id:
        raise CleideAuditCoverageError(
            ERROR_COVERAGE_NO_TEMP_TABLE,
            "Nenhuma tabela temporária ativa nesta sessão.",
        )

    cfg = get_cleiton_doc_config()
    record = load_temp_table_record(active_id, ttl_hours=cfg.upload_ttl_hours)
    if record is None:
        clear_temp_table_session_refs(session)
        _mark_session_modified()
        raise CleideAuditCoverageError(
            ERROR_COVERAGE_NO_TEMP_TABLE,
            "Tabela temporária ativa não encontrada.",
        )
    status = (record.get("status") or "").strip().lower()
    if status == TEMP_TABLE_STATUS_EXPIRED:
        raise CleideAuditCoverageError(
            ERROR_COVERAGE_EXPIRED,
            "A tabela temporária desta sessão expirou.",
        )
    if status in {TEMP_TABLE_STATUS_DISCARDED, TEMP_TABLE_STATUS_PROCESSING}:
        raise CleideAuditCoverageError(
            ERROR_COVERAGE_NO_TEMP_TABLE,
            "Tabela temporária indisponível para upload complementar.",
        )
    _assert_temp_table_scope(record, user_scope=user_scope, franquia_scope=franquia_scope)

    safe_name = secure_filename(display_name or "coverage") or "coverage"
    if ext == ".csv":
        rows, warnings = _parse_coverage_csv_bytes(file_bytes, source_file_name=safe_name)
    else:
        rows, warnings = _parse_coverage_xlsx_bytes(file_bytes, source_file_name=safe_name)

    now = _utcnow().isoformat()
    preserved_expires_at = record.get("expires_at")
    updated = dict(record)
    coverage = _empty_coverage_table_shell(uploaded_at=now)
    existing_coverage = record.get("coverage_table")
    if isinstance(existing_coverage, dict):
        coverage["human_review_status"] = existing_coverage.get("human_review_status")
        coverage["human_edited_at"] = existing_coverage.get("human_edited_at")
        coverage["human_edited_by_user_id"] = existing_coverage.get("human_edited_by_user_id")
        coverage["edit_version"] = existing_coverage.get("edit_version") or 0
    coverage["rows"] = rows
    coverage["validation_warnings"] = warnings
    updated["coverage_table"] = coverage
    updated["updated_at"] = now
    updated["expires_at"] = preserved_expires_at

    saved = save_temp_table_record(updated)
    logger.info(
        "Cleide coverage upload: temp_table_id=%s user_id=%s rows=%s warnings=%s",
        saved.get("temp_table_id"),
        user_scope,
        len(rows),
        len(warnings),
    )
    public = _public_temp_table(saved)
    if public is None:
        raise CleideAuditCoverageError(
            ERROR_COVERAGE_NO_TEMP_TABLE,
            "Não foi possível retornar a tabela temporária atualizada.",
        )
    return public


def _audit_header_aliases() -> dict[str, tuple[str, ...]]:
    return {
        "destination_city": (
            "cidade destino",
            "cidade de destino",
            "cidade_destino",
            "destino cidade",
        ),
        "destination_uf": (
            "uf destino",
            "uf de destino",
            "uf_destino",
            "destino uf",
            "estado destino",
        ),
        "charged_freight": (
            "valor frete",
            "valor do frete",
            "valor_frete",
            "valor frete cobrado",
            "valor do frete cobrado",
            "valor_frete_cobrado",
            "frete cobrado",
            "frete",
        ),
        "audited_weight": (
            "peso",
            "peso auditado",
            "peso_auditado",
            "peso cobrado",
            "peso kg",
        ),
        "carrier": ("transportadora", "nome transportadora", "empresa transportadora"),
        "document_number": (
            "numero documento",
            "numero do documento",
            "numero_documento",
            "documento",
            "nf",
            "nota fiscal",
        ),
        "origin_city": ("cidade origem", "cidade de origem", "cidade_origem"),
        "origin_uf": ("uf origem", "uf de origem", "uf_origem", "estado origem"),
        "invoice_value": ("valor nf", "valor da nf", "valor_nf", "valor nota fiscal"),
        "modal": ("modal", "tipo modal", "modalidade"),
        "issue_date": ("data emissao", "data de emissao", "data_emissao", "emissao"),
        "delivery_date": ("data entrega", "data de entrega", "data_entrega", "entrega"),
    }


def _resolve_audit_field(header_value) -> str | None:
    normalized = _normalize_coverage_header(header_value)
    if not normalized:
        return None
    for field_name, aliases in _audit_header_aliases().items():
        if normalized in aliases:
            return field_name
    return None


def _resolve_audit_field_indexes(header_row: list) -> tuple[dict[str, int], dict[str, str]]:
    field_indexes: dict[str, int] = {}
    header_map: dict[str, str] = {}
    for index, header in enumerate(header_row):
        field_name = _resolve_audit_field(header)
        if field_name and field_name not in field_indexes:
            field_indexes[field_name] = index
            source_header = _sanitize_cell_string(header) or str(header or "").strip()
            header_map[source_header] = field_name
    return field_indexes, header_map


def _format_audit_missing_columns_error(missing_fields: list[str]) -> str:
    labels = [_AUDIT_FIELD_LABELS.get(field, field) for field in missing_fields]
    return "Colunas obrigatórias ausentes: " + ", ".join(labels) + "."


def _parse_audit_numeric(value) -> float | None:
    cleaned = _sanitize_cell_string(value)
    if cleaned is None:
        return None
    text = cleaned.replace(" ", "")
    if not text:
        return None
    if "," in text and "." in text:
        text = text.replace(".", "").replace(",", ".")
    elif "," in text:
        text = text.replace(",", ".")
    try:
        parsed = float(text)
    except (TypeError, ValueError):
        return None
    if parsed < 0:
        return None
    return parsed


def _audit_row_is_empty(raw_row: list, field_indexes: dict[str, int]) -> bool:
    for index in field_indexes.values():
        if index < len(raw_row):
            cell = _sanitize_cell_string(raw_row[index])
            if cell:
                return False
    return True


def _normalize_audit_row(
    raw_row: list,
    *,
    row_index: int,
    field_indexes: dict[str, int],
    source_file_name: str,
) -> dict | None:
    if _audit_row_is_empty(raw_row, field_indexes):
        return None

    destination_city = _sanitize_cell_string(
        raw_row[field_indexes["destination_city"]]
        if field_indexes["destination_city"] < len(raw_row)
        else None
    )
    destination_uf = _normalize_destination_uf(
        raw_row[field_indexes["destination_uf"]]
        if field_indexes["destination_uf"] < len(raw_row)
        else None
    )
    charged_freight = _parse_audit_numeric(
        raw_row[field_indexes["charged_freight"]]
        if field_indexes["charged_freight"] < len(raw_row)
        else None
    )
    audited_weight = _parse_audit_numeric(
        raw_row[field_indexes["audited_weight"]]
        if field_indexes["audited_weight"] < len(raw_row)
        else None
    )

    if not destination_city or not destination_uf or charged_freight is None or audited_weight is None:
        raise CleideAuditBatchError(
            ERROR_AUDIT_PARSE_FAILED,
            f"Linha {row_index}: dados obrigatórios inválidos ou ausentes.",
        )

    normalized: dict = {
        "row_index": row_index,
        "destination_city": destination_city,
        "destination_uf": destination_uf,
        "charged_freight": charged_freight,
        "audited_weight": audited_weight,
        "source_file_name": source_file_name,
    }

    for field_name in _AUDIT_OPTIONAL_FIELDS:
        if field_name not in field_indexes:
            continue
        index = field_indexes[field_name]
        raw_value = raw_row[index] if index < len(raw_row) else None
        if field_name in {"invoice_value"}:
            parsed_value = _parse_audit_numeric(raw_value)
            if parsed_value is not None:
                normalized[field_name] = parsed_value
            continue
        if field_name == "origin_uf":
            parsed_uf = _normalize_destination_uf(raw_value)
            if parsed_uf:
                normalized[field_name] = parsed_uf
            continue
        cleaned = _sanitize_cell_string(raw_value)
        if cleaned:
            normalized[field_name] = cleaned

    return normalized


def _parse_audit_tabular_rows(
    raw_rows: list[list],
    *,
    source_file_name: str,
    max_rows: int,
) -> tuple[list[dict], dict[str, str], str | None]:
    if not raw_rows:
        raise CleideAuditBatchError(
            ERROR_AUDIT_EMPTY_FILE,
            "O arquivo auditado está vazio.",
        )

    header_row = raw_rows[0]
    field_indexes, header_map = _resolve_audit_field_indexes(header_row)
    missing_fields = [
        field for field in _AUDIT_REQUIRED_FIELDS if field not in field_indexes
    ]
    if missing_fields:
        raise CleideAuditBatchError(
            ERROR_AUDIT_MISSING_COLUMNS,
            _format_audit_missing_columns_error(missing_fields),
        )

    normalized_rows: list[dict] = []
    data_row_index = 0
    for raw_row in raw_rows[1:]:
        if _audit_row_is_empty(raw_row, field_indexes):
            continue
        data_row_index += 1
        if data_row_index > max_rows:
            raise CleideAuditBatchError(
                ERROR_AUDIT_TOO_MANY_ROWS,
                f"O arquivo excede o limite de {max_rows} linhas configurado para auditoria.",
            )
        normalized = _normalize_audit_row(
            raw_row,
            row_index=data_row_index,
            field_indexes=field_indexes,
            source_file_name=source_file_name,
        )
        if normalized is not None:
            normalized_rows.append(normalized)

    if not normalized_rows:
        raise CleideAuditBatchError(
            ERROR_AUDIT_EMPTY_ROWS,
            "Nenhuma linha válida encontrada no arquivo auditado.",
        )

    public_header_map = {
        _AUDIT_FIELD_LABELS.get(field, field): field
        for field in field_indexes
        if field in _AUDIT_FIELD_LABELS
    }
    public_header_map.update(
        {
            source: target
            for source, target in header_map.items()
            if source and target
        }
    )
    return normalized_rows, public_header_map, None


def _parse_audit_csv_bytes(
    file_bytes: bytes,
    *,
    source_file_name: str,
    max_rows: int,
) -> tuple[list[dict], dict[str, str], str | None]:
    text = _decode_coverage_csv_bytes(file_bytes)
    reader = csv.reader(io.StringIO(text))
    raw_rows = [[cell for cell in row] for row in reader if row]
    rows, header_map, _ = _parse_audit_tabular_rows(
        raw_rows,
        source_file_name=source_file_name,
        max_rows=max_rows,
    )
    return rows, header_map, None


def _parse_audit_xlsx_bytes(
    file_bytes: bytes,
    *,
    source_file_name: str,
    max_rows: int,
) -> tuple[list[dict], dict[str, str], str]:
    if not file_bytes:
        raise CleideAuditBatchError(
            ERROR_AUDIT_EMPTY_FILE,
            "O arquivo auditado está vazio.",
        )
    if len(file_bytes) > AUDIT_UPLOAD_MAX_BYTES:
        raise CleideAuditBatchError(
            ERROR_AUDIT_PAYLOAD_TOO_LARGE,
            "O arquivo auditado excede o limite de tamanho permitido.",
        )
    try:
        if zipfile.is_zipfile(io.BytesIO(file_bytes)):
            with zipfile.ZipFile(io.BytesIO(file_bytes)) as archive:
                if archive.testzip() is not None:
                    raise CleideAuditBatchError(
                        ERROR_AUDIT_PARSE_FAILED,
                        "Arquivo XLSX auditado corrompido.",
                    )
        from openpyxl import load_workbook

        workbook = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    except CleideAuditBatchError:
        raise
    except Exception as exc:
        raise CleideAuditBatchError(
            ERROR_AUDIT_PARSE_FAILED,
            "Não foi possível ler o arquivo XLSX auditado.",
        ) from exc

    sheet = None
    try:
        if AUDIT_BATCH_SHEET_NAME in workbook.sheetnames:
            sheet = workbook[AUDIT_BATCH_SHEET_NAME]
        else:
            raise CleideAuditBatchError(
                ERROR_AUDIT_INVALID_SHEET,
                f"A aba '{AUDIT_BATCH_SHEET_NAME}' é obrigatória no arquivo XLSX auditado.",
            )
        raw_rows: list[list] = []
        for row in sheet.iter_rows(values_only=True):
            raw_rows.append(["" if cell is None else cell for cell in row])
    finally:
        workbook.close()

    rows, header_map, _ = _parse_audit_tabular_rows(
        raw_rows,
        source_file_name=source_file_name,
        max_rows=max_rows,
    )
    return rows, header_map, AUDIT_BATCH_SHEET_NAME


def _empty_audit_batch_shell(*, uploaded_at: str | None = None) -> dict:
    return {
        "status": AUDIT_BATCH_STATUS_UPLOADED,
        "audit_batch_id": None,
        "temp_table_id": None,
        "source_file_name": None,
        "sheet_name": None,
        "uploaded_at": uploaded_at,
        "created_at": None,
        "updated_at": None,
        "expires_at": None,
        "row_count": 0,
        "max_rows": None,
        "input_schema_version": AUDIT_INPUT_SCHEMA_VERSION,
        "header_map": {},
        "normalized_rows": [],
        "results": [],
        "summary": None,
    }


def _normalize_audit_lookup_text(value) -> str:
    cleaned = _sanitize_cell_string(value)
    if not cleaned:
        return ""
    text = unicodedata.normalize("NFKD", str(cleaned))
    text = "".join(ch for ch in text if not unicodedata.combining(ch))
    text = text.upper().strip()
    text = re.sub(r"[_\-\/.,:;]+", " ", text)
    text = re.sub(r"[^A-Z0-9\s]", " ", text)
    text = re.sub(r"\s+", " ", text)
    return text.strip()


def _coverage_lookup_key(destination_uf, destination_city) -> str | None:
    uf = _normalize_destination_uf(destination_uf)
    city = _normalize_audit_lookup_text(destination_city)
    if not uf or not city:
        return None
    return f"{uf}|{city}"


def build_coverage_index(coverage_table) -> dict:
    rows = []
    if isinstance(coverage_table, dict):
        rows = coverage_table.get("rows") or []
    elif isinstance(coverage_table, list):
        rows = coverage_table
    if not isinstance(rows, list):
        return {}

    grouped: dict[str, set[str]] = {}
    for row in rows:
        if not isinstance(row, dict):
            continue
        key = _coverage_lookup_key(row.get("destination_uf"), row.get("destination_city"))
        region = _sanitize_cell_string(row.get("freight_region"))
        if not key or not region:
            continue
        grouped.setdefault(key, set()).add(region)

    index: dict[str, object] = {}
    for key, regions in grouped.items():
        ordered = sorted(regions)
        if len(ordered) == 1:
            index[key] = ordered[0]
        else:
            index[key] = {
                "reason_code": AUDIT_STATUS_AMBIGUOUS_COVERAGE,
                "regions": ordered,
            }
    return index


def _parse_brazilian_money(value) -> float | None:
    cleaned = _sanitize_cell_string(value)
    if cleaned is None:
        return None
    text = cleaned.strip()
    if not text:
        return None
    text = re.sub(r"(?i)\bR\$\b|R\$", "", text)
    text = re.sub(r"[^0-9,\.\-]", "", text)
    if not text or text in {"-", ".", ","}:
        return None
    if "," in text and "." in text:
        text = text.replace(".", "").replace(",", ".")
    elif "," in text:
        text = text.replace(",", ".")
    try:
        parsed = float(text)
    except (TypeError, ValueError):
        return None
    if parsed < 0:
        return None
    return parsed


def _parse_weight_number(value) -> float | None:
    cleaned = _sanitize_cell_string(value)
    if cleaned is None:
        return None
    text = cleaned.strip()
    if not text:
        return None
    text = re.sub(r"[^0-9,\.\-]", "", text)
    if "," in text and "." in text:
        text = text.replace(".", "").replace(",", ".")
    elif "," in text:
        text = text.replace(",", ".")
    try:
        parsed = float(text)
    except (TypeError, ValueError):
        return None
    if parsed < 0:
        return None
    return parsed


def _parse_range_from_label(label) -> tuple[float | None, float | None] | None:
    normalized = _normalize_coverage_header(label)
    if not normalized:
        return None
    numbers = [float(item.replace(",", ".")) for item in re.findall(r"\d+(?:[,.]\d+)?", normalized)]
    if not numbers:
        return None
    if normalized.startswith("ate ") or normalized.startswith("ate"):
        return (0.0, numbers[-1])
    if len(numbers) >= 2 and (
        " a " in f" {normalized} "
        or normalized.startswith("de ")
        or " ate " in f" {normalized} "
    ):
        return (numbers[0], numbers[1])
    if len(numbers) == 1 and re.search(r"\bkg\b|\bpeso\b", normalized):
        return (0.0, numbers[0])
    return None


def _normalize_brackets(brackets: list[dict]) -> list[dict]:
    cleaned = [
        bracket
        for bracket in brackets
        if isinstance(bracket.get("max_kg"), (int, float))
        and isinstance(bracket.get("value"), (int, float))
    ]
    cleaned.sort(key=lambda item: (float(item.get("max_kg")), float(item.get("min_kg") or 0)))
    previous_max = 0.0
    normalized: list[dict] = []
    for index, bracket in enumerate(cleaned):
        max_kg = float(bracket["max_kg"])
        if max_kg < previous_max:
            continue
        min_kg = 0.0 if index == 0 else previous_max
        normalized.append(
            {
                "min_kg": min_kg,
                "max_kg": max_kg,
                "value": round(float(bracket["value"]), 2),
                "label": bracket.get("label") or f"Faixa até {max_kg:g} kg",
            }
        )
        previous_max = max_kg
    return normalized


def _is_region_column(column_name) -> bool:
    normalized = _normalize_coverage_header(column_name)
    return normalized in {
        "uf cidades",
        "uf cidade",
        "regiao",
        "regiao de frete",
        "praca",
        "rota",
        "itinerario",
        "destino",
        "destino frete",
        "cidade",
        "cidades",
        "cidade destino",
        "municipio",
        "municipio destino",
    }


def _is_destination_uf_column(column_name) -> bool:
    normalized = _normalize_coverage_header(column_name)
    return normalized in {
        "uf",
        "uf destino",
        "uf de destino",
        "uf dest",
        "uf entrega",
        "uf de entrega",
    }


def _is_city_destination_column(column_name) -> bool:
    normalized = _normalize_coverage_header(column_name)
    return normalized in {
        "uf cidades",
        "uf cidade",
        "cidade",
        "cidades",
        "cidade destino",
        "municipio",
        "municipio destino",
        "destino",
    }


def _is_value_column(column_name) -> bool:
    normalized = _normalize_coverage_header(column_name)
    if "pedagio" in normalized or "gris" in normalized or "seguro" in normalized:
        return False
    if "tso" in normalized or "tas" in normalized or "frete valor" in normalized:
        return False
    return normalized in {"frete", "valor", "valor frete", "frete peso", "tarifa"}


def _is_excess_column(column_name) -> bool:
    normalized = _normalize_coverage_header(column_name)
    return "excedente" in normalized or "excesso" in normalized


def _is_direct_kg_column(column_name) -> bool:
    normalized = _normalize_coverage_header(column_name)
    if "excedente" in normalized or "frete valor" in normalized:
        return False
    return (
        normalized in {"kg", "por kg", "valor kg", "frete kg", "frete peso kg", "frete peso"}
        or "r kg" in normalized
        or "rs kg" in normalized
    )


def _is_direct_ton_column(column_name) -> bool:
    normalized = _normalize_coverage_header(column_name)
    return (
        "tonelada" in normalized
        or normalized in {"ton", "por ton", "valor ton", "frete ton"}
        or "r ton" in normalized
        or "rs ton" in normalized
    )


def _region_from_table_context(table: dict) -> str | None:
    context = table.get("context") if isinstance(table.get("context"), dict) else {}
    for key in ("route_label", "destination", "region", "freight_region", "praca", "rota"):
        candidate = _sanitize_cell_string(context.get(key))
        if candidate:
            return candidate
    return _sanitize_cell_string(table.get("table_title"))


def _make_unsupported_rule(region: str, source_title: str | None, note: str) -> dict:
    return {
        "pricing_type": AUDIT_STATUS_UNSUPPORTED_PRICING,
        "region": region,
        "source_table_title": source_title,
        "brackets": [],
        "excess": None,
        "unit": "kg",
        "normalization_notes": [note],
    }


def _register_pricing_rule(index: dict, region: str | None, rule: dict) -> None:
    if not region:
        return
    if region in index:
        index[region] = _make_unsupported_rule(
            region,
            rule.get("source_table_title"),
            "Mais de uma regra de frete para a mesma região.",
        )
        return
    index[region] = rule


def _pricing_rule_keys_for_row(
    region: str,
    destination_uf: str | None = None,
    *,
    include_normalized_region: bool = False,
) -> list[str]:
    keys = [region]
    normalized_region = _normalize_audit_lookup_text(region)
    if include_normalized_region and normalized_region:
        keys.append(normalized_region)
    uf = _normalize_destination_uf(destination_uf)
    if uf and normalized_region:
        keys.append(f"{uf}|{normalized_region}")
    return list(dict.fromkeys(keys))


def _build_rule_from_row_range_table(table: dict) -> dict | None:
    rows = table.get("rows") if isinstance(table.get("rows"), list) else []
    columns = table.get("columns") if isinstance(table.get("columns"), list) else []
    if not rows or not columns:
        return None
    range_col = next((col for col in columns if _parse_range_from_label(col) is None and "peso" in _normalize_coverage_header(col)), None)
    value_col = next((col for col in columns if _is_value_column(col)), None)
    if not range_col or not value_col:
        return None
    region = _region_from_table_context(table)
    if not region:
        return None
    brackets = []
    for row in rows:
        if not isinstance(row, dict):
            continue
        parsed_range = _parse_range_from_label(row.get(range_col))
        value = _parse_brazilian_money(row.get(value_col))
        if parsed_range and value is not None:
            brackets.append(
                {
                    "min_kg": parsed_range[0] or 0.0,
                    "max_kg": parsed_range[1],
                    "value": value,
                    "label": _sanitize_cell_string(row.get(range_col)),
                }
            )
    brackets = _normalize_brackets(brackets)
    if not brackets:
        return _make_unsupported_rule(region, table.get("table_title"), "Faixas por linha sem valor calculável.")
    excess_col = next((col for col in columns if _is_excess_column(col)), None)
    excess_rate = None
    if excess_col:
        for row in rows:
            excess_rate = _parse_brazilian_money(row.get(excess_col))
            if excess_rate is not None:
                break
    return {
        "pricing_type": "range_plus_excess_per_kg" if excess_rate is not None else "fixed_range",
        "region": region,
        "source_table_title": table.get("table_title"),
        "brackets": brackets,
        "excess": {"rate_per_kg": excess_rate} if excess_rate is not None else None,
        "unit": "kg",
        "normalization_notes": [],
    }


def _build_rules_from_matrix_table(table: dict) -> list[tuple[str, dict]]:
    rows = table.get("rows") if isinstance(table.get("rows"), list) else []
    columns = table.get("columns") if isinstance(table.get("columns"), list) else []
    if not rows or not columns:
        return []

    region_col = next((col for col in columns if _is_region_column(col)), None)
    uf_col = next((col for col in columns if _is_destination_uf_column(col)), None)
    range_cols = [(col, _parse_range_from_label(col)) for col in columns]
    range_cols = [(col, parsed) for col, parsed in range_cols if parsed is not None]
    direct_kg_col = next((col for col in columns if _is_direct_kg_column(col)), None)
    direct_ton_col = next((col for col in columns if _is_direct_ton_column(col)), None)
    excess_col = next((col for col in columns if _is_excess_column(col)), None)
    context_region = _region_from_table_context(table)
    region_is_city_destination = bool(region_col and _is_city_destination_column(region_col))
    rules: list[tuple[str, dict]] = []

    for row in rows:
        if not isinstance(row, dict):
            continue
        region = _sanitize_cell_string(row.get(region_col)) if region_col else context_region
        destination_uf = _sanitize_cell_string(row.get(uf_col)) if uf_col else None
        if not region:
            continue
        if direct_ton_col:
            value = _parse_brazilian_money(row.get(direct_ton_col))
            if value is not None:
                rule = {
                    "pricing_type": "direct_weight_rate",
                    "region": region,
                    "source_table_title": table.get("table_title"),
                    "brackets": [],
                    "excess": None,
                    "unit": "ton",
                    "value_per_ton": value,
                    "normalization_notes": [],
                }
                rules.extend(
                    (key, rule)
                    for key in _pricing_rule_keys_for_row(
                        region,
                        destination_uf,
                        include_normalized_region=region_is_city_destination,
                    )
                )
                continue
        if direct_kg_col:
            value = _parse_brazilian_money(row.get(direct_kg_col))
            if value is not None:
                rule = {
                    "pricing_type": "direct_weight_rate",
                    "region": region,
                    "source_table_title": table.get("table_title"),
                    "brackets": [],
                    "excess": None,
                    "unit": "kg",
                    "value_per_kg": value,
                    "normalization_notes": [],
                }
                rules.extend(
                    (key, rule)
                    for key in _pricing_rule_keys_for_row(
                        region,
                        destination_uf,
                        include_normalized_region=region_is_city_destination,
                    )
                )
                continue
        brackets = []
        for col, parsed_range in range_cols:
            value = _parse_brazilian_money(row.get(col))
            if value is None:
                continue
            brackets.append(
                {
                    "min_kg": parsed_range[0] or 0.0,
                    "max_kg": parsed_range[1],
                    "value": value,
                    "label": col,
                }
            )
        brackets = _normalize_brackets(brackets)
        if brackets:
            excess_rate = _parse_brazilian_money(row.get(excess_col)) if excess_col else None
            rule = {
                "pricing_type": "range_plus_excess_per_kg" if excess_rate is not None else "fixed_range",
                "region": region,
                "source_table_title": table.get("table_title"),
                "brackets": brackets,
                "excess": {"rate_per_kg": excess_rate} if excess_rate is not None else None,
                "unit": "kg",
                "normalization_notes": [],
            }
            rules.extend(
                (key, rule)
                for key in _pricing_rule_keys_for_row(
                    region,
                    destination_uf,
                    include_normalized_region=region_is_city_destination,
                )
            )
        elif region_col:
            rule = _make_unsupported_rule(
                region,
                table.get("table_title"),
                "Linha de destino sem modelo de peso/faixa reconhecido.",
            )
            rules.extend(
                (key, rule)
                for key in _pricing_rule_keys_for_row(
                    region,
                    destination_uf,
                    include_normalized_region=region_is_city_destination,
                )
            )
    return rules


def _build_rule_from_freight_route(route: dict) -> tuple[str, dict] | None:
    region = _sanitize_cell_string(
        route.get("destination")
        or route.get("freight_region")
        or route.get("region")
        or route.get("route")
    )
    if not region:
        return None
    direct_kg = _parse_brazilian_money(route.get("freight_weight_kg") or route.get("frete_peso_kg"))
    if direct_kg is not None:
        return (
            region,
            {
                "pricing_type": "direct_weight_rate",
                "region": region,
                "source_table_title": "freight_routes",
                "brackets": [],
                "excess": None,
                "unit": "kg",
                "value_per_kg": direct_kg,
                "normalization_notes": [],
            },
        )
    brackets = []
    for limit in (10, 20, 30, 50, 70, 100):
        value = _parse_brazilian_money(
            route.get(f"weight_{limit}") or route.get(f"weight_{limit}kg")
        )
        if value is not None:
            brackets.append(
                {
                    "min_kg": 0.0,
                    "max_kg": float(limit),
                    "value": value,
                    "label": f"Até {limit} kg",
                }
            )
    brackets = _normalize_brackets(brackets)
    if not brackets:
        return (region, _make_unsupported_rule(region, "freight_routes", "Rota sem faixa de peso calculável."))
    return (
        region,
        {
            "pricing_type": "fixed_range",
            "region": region,
            "source_table_title": "freight_routes",
            "brackets": brackets,
            "excess": None,
            "unit": "kg",
            "normalization_notes": [],
        },
    )


def build_freight_pricing_index(temp_table) -> dict:
    if not isinstance(temp_table, dict):
        return {}
    index: dict[str, dict] = {}

    for table in temp_table.get("freight_tables") or []:
        if not isinstance(table, dict):
            continue
        row_range_rule = _build_rule_from_row_range_table(table)
        if row_range_rule is not None:
            _register_pricing_rule(index, row_range_rule.get("region"), row_range_rule)
            continue
        matrix_rules = _build_rules_from_matrix_table(table)
        if matrix_rules:
            for region, rule in matrix_rules:
                _register_pricing_rule(index, region, rule)
            continue
        region = _region_from_table_context(table)
        if region:
            _register_pricing_rule(
                index,
                region,
                _make_unsupported_rule(region, table.get("table_title"), "Tabela sem modelo de peso/faixa reconhecido."),
            )

    for route in temp_table.get("freight_routes") or []:
        if not isinstance(route, dict):
            continue
        route_rule = _build_rule_from_freight_route(route)
        if route_rule is not None:
            region, rule = route_rule
            _register_pricing_rule(index, region, rule)

    return index


def calculate_weight_freight(weight_kg, pricing_rule) -> dict | None:
    weight = _parse_weight_number(weight_kg)
    if weight is None or not isinstance(pricing_rule, dict):
        return None
    pricing_type = pricing_rule.get("pricing_type")
    if pricing_type == "fixed_range":
        for bracket in pricing_rule.get("brackets") or []:
            min_kg = float(bracket.get("min_kg") or 0)
            max_kg = bracket.get("max_kg")
            value = bracket.get("value")
            if max_kg is None or value is None:
                continue
            max_kg = float(max_kg)
            if (min_kg <= 0 and 0 <= weight <= max_kg) or (min_kg < weight <= max_kg):
                return {
                    "expected_freight": round(float(value), 2),
                    "calculation_basis": "fixed_range",
                    "calculation_details": bracket.get("label") or f"Faixa até {max_kg:g} kg",
                }
        return None
    if pricing_type == "range_plus_excess_per_kg":
        brackets = pricing_rule.get("brackets") or []
        for bracket in brackets:
            min_kg = float(bracket.get("min_kg") or 0)
            max_kg = bracket.get("max_kg")
            value = bracket.get("value")
            if max_kg is None or value is None:
                continue
            max_kg = float(max_kg)
            if (min_kg <= 0 and 0 <= weight <= max_kg) or (min_kg < weight <= max_kg):
                return {
                    "expected_freight": round(float(value), 2),
                    "calculation_basis": "range_plus_excess_per_kg",
                    "calculation_details": bracket.get("label") or f"Faixa até {max_kg:g} kg",
                }
        if not brackets:
            return None
        last = max(brackets, key=lambda item: float(item.get("max_kg") or 0))
        last_max = float(last.get("max_kg") or 0)
        last_value = last.get("value")
        excess = pricing_rule.get("excess") if isinstance(pricing_rule.get("excess"), dict) else {}
        excess_rate = excess.get("rate_per_kg")
        if last_value is None or excess_rate is None or weight <= last_max:
            return None
        expected = float(last_value) + float(excess_rate) * (weight - last_max)
        return {
            "expected_freight": round(expected, 2),
            "calculation_basis": "range_plus_excess_per_kg",
            "calculation_details": f"Faixa até {last_max:g} kg + excedente por kg",
        }
    if pricing_type == "direct_weight_rate":
        unit = str(pricing_rule.get("unit") or "kg").strip().lower()
        if unit in {"ton", "tonelada", "toneladas", "t"}:
            value = pricing_rule.get("value_per_ton")
            if value is None:
                return None
            expected = (weight / 1000.0) * float(value)
            return {
                "expected_freight": round(expected, 2),
                "calculation_basis": "direct_weight_rate",
                "calculation_details": "Peso em toneladas x valor por tonelada",
            }
        value = pricing_rule.get("value_per_kg")
        if value is None:
            return None
        expected = weight * float(value)
        return {
            "expected_freight": round(expected, 2),
            "calculation_basis": "direct_weight_rate",
            "calculation_details": "Peso em kg x valor por kg",
        }
    return None


def compare_charged_vs_expected(charged_freight, expected_freight) -> dict:
    charged = round(float(charged_freight), 2)
    expected = round(float(expected_freight), 2)
    divergence = round(charged - expected, 2)
    return {
        "charged_freight": charged,
        "expected_freight": expected,
        "divergence_value": divergence,
        "status": AUDIT_STATUS_OK if divergence == 0 else AUDIT_STATUS_DIVERGENT,
    }


def _find_pricing_rule_match(
    pricing_index: dict,
    freight_region: str | None,
    destination_uf: str | None = None,
    destination_city: str | None = None,
) -> tuple[dict, str, str] | None:
    candidates: list[tuple[str, str]] = []
    if freight_region:
        candidates.append(("freight_region", freight_region))
    uf_city_key = _coverage_lookup_key(destination_uf, destination_city)
    if uf_city_key:
        candidates.append(("destination_uf_city", uf_city_key))
    city_key = _normalize_audit_lookup_text(destination_city)
    if city_key:
        candidates.append(("destination_city", city_key))

    seen: set[str] = set()
    for lookup_kind, lookup_key in candidates:
        if not lookup_key or lookup_key in seen:
            continue
        seen.add(lookup_key)
        if lookup_key in pricing_index:
            return pricing_index[lookup_key], lookup_kind, lookup_key
        wanted = _normalize_audit_lookup_text(lookup_key)
        matches = [
            (region, rule)
            for region, rule in pricing_index.items()
            if _normalize_audit_lookup_text(region) == wanted
        ]
        if len(matches) == 1:
            matched_key, matched_rule = matches[0]
            return matched_rule, lookup_kind, matched_key
    return None


def _find_pricing_rule(
    pricing_index: dict,
    freight_region: str | None,
    destination_uf: str | None = None,
    destination_city: str | None = None,
) -> dict | None:
    match = _find_pricing_rule_match(pricing_index, freight_region, destination_uf, destination_city)
    return match[0] if match else None


def _resolve_region_without_coverage(row: dict, pricing_index: dict) -> tuple[str | None, str | None]:
    city = _normalize_audit_lookup_text(row.get("destination_city"))
    uf = _normalize_destination_uf(row.get("destination_uf"))
    if not city and not uf:
        return None, AUDIT_STATUS_MISSING_COVERAGE
    candidates = []
    for region in pricing_index:
        normalized_region = _normalize_audit_lookup_text(region)
        if city and normalized_region == city:
            candidates.append(region)
        elif uf and normalized_region == uf:
            candidates.append(region)
        elif city and uf and normalized_region in {f"{uf} {city}", f"{city} {uf}", f"{uf}|{city}"}:
            candidates.append(region)
    unique = sorted(set(candidates))
    if len(unique) == 1:
        return unique[0], None
    return None, AUDIT_STATUS_MISSING_COVERAGE


def _base_audit_result(row: dict) -> dict:
    return {
        "row_index": row.get("row_index"),
        "numero_documento": row.get("document_number"),
        "destination_uf": row.get("destination_uf"),
        "destination_city": row.get("destination_city"),
        "freight_region": None,
        "audited_weight": row.get("audited_weight"),
        "charged_freight": row.get("charged_freight"),
        "expected_freight": None,
        "divergence_value": None,
        "status": None,
        "reason_code": None,
        "calculation_basis": None,
        "calculation_details": None,
    }


def _status_result(row: dict, status: str, *, freight_region: str | None = None) -> dict:
    result = _base_audit_result(row)
    result["freight_region"] = freight_region
    result["status"] = status
    result["reason_code"] = status
    return result


def _audit_single_row(row: dict, *, coverage_index: dict, pricing_index: dict, has_coverage: bool) -> dict:
    weight = _parse_weight_number(row.get("audited_weight"))
    if weight is None:
        return _status_result(row, AUDIT_STATUS_INVALID_WEIGHT)
    charged = _parse_brazilian_money(row.get("charged_freight"))
    if charged is None:
        return _status_result(row, AUDIT_STATUS_INVALID_CHARGED_FREIGHT)

    freight_region = None
    if has_coverage:
        key = _coverage_lookup_key(row.get("destination_uf"), row.get("destination_city"))
        match = coverage_index.get(key) if key else None
        if isinstance(match, dict):
            return _status_result(row, AUDIT_STATUS_AMBIGUOUS_COVERAGE)
        if isinstance(match, str) and match.strip():
            freight_region = match
        else:
            return _status_result(row, AUDIT_STATUS_MISSING_COVERAGE)
    else:
        freight_region, reason = _resolve_region_without_coverage(row, pricing_index)
        if reason:
            return _status_result(row, reason)

    rule_match = _find_pricing_rule_match(
        pricing_index,
        freight_region,
        row.get("destination_uf"),
        row.get("destination_city"),
    )
    if rule_match is None:
        return _status_result(row, AUDIT_STATUS_MISSING_FREIGHT_RULE, freight_region=freight_region)
    rule, lookup_kind, lookup_key = rule_match
    if rule.get("pricing_type") == AUDIT_STATUS_UNSUPPORTED_PRICING:
        return _status_result(row, AUDIT_STATUS_UNSUPPORTED_PRICING, freight_region=freight_region)

    calculated = calculate_weight_freight(weight, rule)
    if calculated is None:
        return _status_result(row, AUDIT_STATUS_UNSUPPORTED_PRICING, freight_region=freight_region)

    comparison = compare_charged_vs_expected(charged, calculated["expected_freight"])
    result = _base_audit_result(row)
    result.update(comparison)
    result["freight_region"] = freight_region
    result["audited_weight"] = weight
    result["reason_code"] = None if comparison["status"] == AUDIT_STATUS_OK else AUDIT_STATUS_DIVERGENT
    result["calculation_basis"] = calculated["calculation_basis"]
    result["calculation_details"] = calculated["calculation_details"]
    if lookup_kind != "freight_region":
        result["calculation_details"] = (
            f"{result['calculation_details']} | regra localizada por cidade/destino: {lookup_key}"
        )
    return result


def _build_audit_summary(results: list[dict], total_rows: int) -> dict:
    summary = {
        "total_rows": total_rows,
        "processed_rows": 0,
        "ok": 0,
        "divergent": 0,
        "missing_coverage_mapping": 0,
        "ambiguous_coverage_mapping": 0,
        "missing_freight_rule": 0,
        "invalid_rows": 0,
        "unsupported_pricing_model": 0,
    }
    for result in results:
        status = result.get("status")
        if status == AUDIT_STATUS_OK:
            summary["ok"] += 1
            summary["processed_rows"] += 1
        elif status == AUDIT_STATUS_DIVERGENT:
            summary["divergent"] += 1
            summary["processed_rows"] += 1
        elif status == AUDIT_STATUS_MISSING_COVERAGE:
            summary["missing_coverage_mapping"] += 1
        elif status == AUDIT_STATUS_AMBIGUOUS_COVERAGE:
            summary["ambiguous_coverage_mapping"] += 1
        elif status == AUDIT_STATUS_MISSING_FREIGHT_RULE:
            summary["missing_freight_rule"] += 1
        elif status in {AUDIT_STATUS_INVALID_WEIGHT, AUDIT_STATUS_INVALID_CHARGED_FREIGHT}:
            summary["invalid_rows"] += 1
        elif status == AUDIT_STATUS_UNSUPPORTED_PRICING:
            summary["unsupported_pricing_model"] += 1
    return summary


def run_audit_batch_for_session(*, user_scope=None, franquia_scope=None) -> dict:
    _require_session()
    sync_temp_table_with_session_documents()
    active_id = get_temp_table_id(session)
    if not active_id:
        raise CleideAuditBatchError(
            ERROR_AUDIT_NO_TEMP_TABLE,
            "Nenhuma tabela temporária ativa nesta sessão.",
        )

    cfg = get_cleiton_doc_config()
    record = load_temp_table_record(active_id, ttl_hours=cfg.upload_ttl_hours)
    if record is None:
        clear_temp_table_session_refs(session)
        _mark_session_modified()
        raise CleideAuditBatchError(
            ERROR_AUDIT_NO_TEMP_TABLE,
            "Tabela temporária ativa não encontrada.",
        )
    status = (record.get("status") or "").strip().lower()
    if status == TEMP_TABLE_STATUS_EXPIRED:
        raise CleideAuditBatchError(
            ERROR_AUDIT_EXPIRED,
            "A tabela temporária desta sessão expirou.",
        )
    if status in {TEMP_TABLE_STATUS_DISCARDED, TEMP_TABLE_STATUS_PROCESSING}:
        raise CleideAuditBatchError(
            ERROR_AUDIT_NO_TEMP_TABLE,
            "Tabela temporária indisponível para processamento.",
        )
    _assert_temp_table_scope(record, user_scope=user_scope, franquia_scope=franquia_scope)

    audit_batch = record.get("audit_batch")
    if not isinstance(audit_batch, dict):
        raise CleideAuditBatchError(
            ERROR_AUDIT_BATCH_NOT_FOUND,
            "Nenhum lote auditado foi enviado nesta sessão.",
        )
    normalized_rows = audit_batch.get("normalized_rows")
    if not isinstance(normalized_rows, list) or not normalized_rows:
        raise CleideAuditBatchError(
            ERROR_AUDIT_BATCH_EMPTY,
            "O lote auditado não possui linhas normalizadas para processar.",
        )

    coverage_table = record.get("coverage_table") if isinstance(record.get("coverage_table"), dict) else None
    has_coverage = bool(coverage_table and isinstance(coverage_table.get("rows"), list) and coverage_table.get("rows"))
    coverage_index = build_coverage_index(coverage_table or {"rows": []})
    pricing_index = build_freight_pricing_index(record)
    results = [
        _audit_single_row(
            row if isinstance(row, dict) else {},
            coverage_index=coverage_index,
            pricing_index=pricing_index,
            has_coverage=has_coverage,
        )
        for row in normalized_rows
    ]
    summary = _build_audit_summary(results, len(normalized_rows))

    now = _utcnow().isoformat()
    preserved_expires_at = record.get("expires_at")
    updated_batch = dict(audit_batch)
    updated_batch["status"] = AUDIT_BATCH_STATUS_PROCESSED
    updated_batch["results"] = results
    updated_batch["summary"] = summary
    updated_batch["updated_at"] = now
    updated_batch["processed_at"] = now
    updated_batch["expires_at"] = audit_batch.get("expires_at") or preserved_expires_at

    updated = dict(record)
    updated["audit_batch"] = updated_batch
    updated["updated_at"] = now
    updated["expires_at"] = preserved_expires_at

    saved = save_temp_table_record(updated)
    public = _public_temp_table(saved)
    if public is None:
        raise CleideAuditBatchError(
            ERROR_AUDIT_NO_TEMP_TABLE,
            "Não foi possível retornar a tabela temporária processada.",
        )
    return public


def _public_audit_batch(audit_batch) -> dict | None:
    if not isinstance(audit_batch, dict):
        return None
    normalized_rows = audit_batch.get("normalized_rows")
    row_count = audit_batch.get("row_count")
    if row_count is None and isinstance(normalized_rows, list):
        row_count = len(normalized_rows)
    return {
        "status": audit_batch.get("status"),
        "audit_batch_id": audit_batch.get("audit_batch_id"),
        "temp_table_id": audit_batch.get("temp_table_id"),
        "source_file_name": audit_batch.get("source_file_name"),
        "sheet_name": audit_batch.get("sheet_name"),
        "uploaded_at": audit_batch.get("uploaded_at"),
        "created_at": audit_batch.get("created_at"),
        "updated_at": audit_batch.get("updated_at"),
        "expires_at": audit_batch.get("expires_at"),
        "row_count": row_count,
        "max_rows": audit_batch.get("max_rows"),
        "input_schema_version": audit_batch.get("input_schema_version"),
        "header_map": dict(audit_batch.get("header_map") or {}),
        "results": list(audit_batch.get("results") or []),
        "summary": audit_batch.get("summary"),
        "processed_at": audit_batch.get("processed_at"),
    }


def get_cleide_audit_template_path():
    from pathlib import Path

    from flask import current_app

    template_path = (
        Path(current_app.root_path)
        / "protected_files"
        / "templates"
        / CLEIDE_AUDIT_TEMPLATE_FILENAME
    )
    return template_path


def upload_audit_batch_from_file(
    *,
    display_name: str,
    file_bytes: bytes,
    extension: str | None,
    user_scope=None,
    franquia_scope=None,
) -> dict:
    """
    Upload determinístico do arquivo auditado no tt_*.json ativo.

    Não registra documento principal, não chama Gemini e não executa cálculo de auditoria.
    """
    _require_session()
    if not file_bytes:
        raise CleideAuditBatchError(
            ERROR_AUDIT_EMPTY_FILE,
            "O arquivo auditado está vazio.",
        )
    if len(file_bytes) > AUDIT_UPLOAD_MAX_BYTES:
        raise CleideAuditBatchError(
            ERROR_AUDIT_PAYLOAD_TOO_LARGE,
            "O arquivo auditado excede o limite de tamanho permitido.",
        )

    ext = (extension or "").strip().lower()
    if not ext.startswith("."):
        ext = f".{ext}" if ext else ""
    if ext == ".pdf":
        raise CleideAuditBatchError(
            ERROR_AUDIT_INVALID_FORMAT,
            "Upload do arquivo auditado aceita apenas CSV e XLSX nesta fase.",
        )
    if ext not in {".csv", ".xlsx"}:
        raise CleideAuditBatchError(
            ERROR_AUDIT_INVALID_FORMAT,
            "Upload do arquivo auditado aceita apenas CSV e XLSX nesta fase.",
        )

    sync_temp_table_with_session_documents()
    active_id = get_temp_table_id(session)
    if not active_id:
        raise CleideAuditBatchError(
            ERROR_AUDIT_NO_TEMP_TABLE,
            "Nenhuma tabela temporária ativa nesta sessão.",
        )

    cfg = get_cleiton_doc_config()
    record = load_temp_table_record(active_id, ttl_hours=cfg.upload_ttl_hours)
    if record is None:
        clear_temp_table_session_refs(session)
        _mark_session_modified()
        raise CleideAuditBatchError(
            ERROR_AUDIT_NO_TEMP_TABLE,
            "Tabela temporária ativa não encontrada.",
        )
    status = (record.get("status") or "").strip().lower()
    if status == TEMP_TABLE_STATUS_EXPIRED:
        raise CleideAuditBatchError(
            ERROR_AUDIT_EXPIRED,
            "A tabela temporária desta sessão expirou.",
        )
    if status in {TEMP_TABLE_STATUS_DISCARDED, TEMP_TABLE_STATUS_PROCESSING}:
        raise CleideAuditBatchError(
            ERROR_AUDIT_NO_TEMP_TABLE,
            "Tabela temporária indisponível para upload do arquivo auditado.",
        )
    _assert_temp_table_scope(record, user_scope=user_scope, franquia_scope=franquia_scope)

    audit_cfg = get_cleide_audit_config()
    max_rows = int(audit_cfg.audited_file_max_rows)
    safe_name = secure_filename(display_name or "auditado") or "auditado"
    if ext == ".csv":
        normalized_rows, header_map, sheet_name = _parse_audit_csv_bytes(
            file_bytes,
            source_file_name=safe_name,
            max_rows=max_rows,
        )
    else:
        normalized_rows, header_map, sheet_name = _parse_audit_xlsx_bytes(
            file_bytes,
            source_file_name=safe_name,
            max_rows=max_rows,
        )

    now = _utcnow().isoformat()
    preserved_expires_at = record.get("expires_at")
    batch_id = uuid4().hex
    audit_batch = {
        "status": AUDIT_BATCH_STATUS_UPLOADED,
        "audit_batch_id": batch_id,
        "temp_table_id": active_id,
        "source_file_name": safe_name,
        "sheet_name": sheet_name,
        "uploaded_at": now,
        "created_at": now,
        "updated_at": now,
        "expires_at": preserved_expires_at,
        "row_count": len(normalized_rows),
        "max_rows": max_rows,
        "input_schema_version": AUDIT_INPUT_SCHEMA_VERSION,
        "header_map": header_map,
        "normalized_rows": normalized_rows,
        "results": [],
        "summary": None,
    }

    updated = dict(record)
    updated["audit_batch"] = audit_batch
    updated["updated_at"] = now
    updated["expires_at"] = preserved_expires_at

    saved = save_temp_table_record(updated)
    logger.info(
        "Cleide audit batch upload: temp_table_id=%s user_id=%s rows=%s max_rows=%s",
        saved.get("temp_table_id"),
        user_scope,
        len(normalized_rows),
        max_rows,
    )
    public = _public_temp_table(saved)
    if public is None:
        raise CleideAuditBatchError(
            ERROR_AUDIT_NO_TEMP_TABLE,
            "Não foi possível retornar a tabela temporária atualizada.",
        )
    return public


def _public_temp_table(record: dict | None) -> dict | None:
    if not record:
        return None
    ui = record.get("ui_visibility") if isinstance(record.get("ui_visibility"), dict) else {}
    public = {
        "temp_table_id": record.get("temp_table_id"),
        "status": record.get("status"),
        "source_documents": list(record.get("source_documents") or []),
        "detected_carrier": record.get("detected_carrier"),
        "origins": list(record.get("origins") or []),
        "destinations": list(record.get("destinations") or []),
        "routes": list(record.get("routes") or []),
        "freight_tables": list(record.get("freight_tables") or []),
        "freight_routes": list(record.get("freight_routes") or []),
        "weight_ranges": list(record.get("weight_ranges") or []),
        "freight_values": list(record.get("freight_values") or []),
        "accessorial_fees": list(record.get("accessorial_fees") or []),
        "charge_type_detected": record.get("charge_type_detected"),
        "extracted_items": list(record.get("extracted_items") or []),
        "uncertain_fields": list(record.get("uncertain_fields") or []),
        "reading_alerts": list(record.get("reading_alerts") or []),
        "evidence_refs": list(record.get("evidence_refs") or []),
        "created_at": record.get("created_at"),
        "updated_at": record.get("updated_at"),
        "expires_at": record.get("expires_at"),
        "session_scope": record.get("session_scope"),
        "franquia_scope": record.get("franquia_scope"),
        "user_scope": record.get("user_scope"),
        "operational_owner": record.get("operational_owner"),
        "ui_visibility": {
            "display_name": ui.get("display_name") or TEMP_TABLE_UI_DISPLAY_NAME,
            "readonly": True,
        },
        "version_marker": record.get("version_marker"),
        "human_review_status": record.get("human_review_status"),
        "human_edited_at": record.get("human_edited_at"),
        "human_edited_by_user_id": record.get("human_edited_by_user_id"),
        "edit_version": record.get("edit_version"),
    }
    coverage = _public_coverage_table(record.get("coverage_table"))
    if coverage is not None:
        public["coverage_table"] = coverage
    audit_batch = _public_audit_batch(record.get("audit_batch"))
    if audit_batch is not None:
        public["audit_batch"] = audit_batch
    return public


def _sanitize_cell_string(value) -> str | None:
    if value is None:
        return None
    if isinstance(value, bool):
        return str(value)
    if not isinstance(value, str):
        value = str(value)
    cleaned = "".join(
        ch for ch in value if ch in {"\n", "\t"} or (ord(ch) >= 32 and ord(ch) != 127)
    )
    cleaned = _HTML_TAG_RE.sub("", cleaned)
    stripped = cleaned.strip()
    return stripped if stripped else ""


def _sanitize_freight_table_context(raw_context) -> dict:
    normalized = _normalize_freight_table_context(raw_context)
    return {
        key: _sanitize_cell_string(val) if val is not None else None
        for key, val in normalized.items()
    }


def _validate_freight_table_item_for_save(item) -> dict:
    if not isinstance(item, dict):
        raise CleideAuditTempTableError(
            ERROR_TEMP_TABLE_INVALID_PAYLOAD,
            "Cada item de freight_tables deve ser um objeto.",
        )
    raw_columns = item.get("columns")
    columns: list[str] = []
    if raw_columns is not None:
        if not isinstance(raw_columns, list):
            raise CleideAuditTempTableError(
                ERROR_TEMP_TABLE_INVALID_PAYLOAD,
                "freight_tables.columns deve ser uma lista.",
            )
        for col in raw_columns:
            if not isinstance(col, str):
                raise CleideAuditTempTableError(
                    ERROR_TEMP_TABLE_INVALID_PAYLOAD,
                    "Nome de coluna inválido.",
                )
            candidate = _sanitize_cell_string(col)
            if not candidate:
                raise CleideAuditTempTableError(
                    ERROR_TEMP_TABLE_INVALID_PAYLOAD,
                    "Coluna sem nome não é permitida.",
                )
            columns.append(candidate)
    raw_rows = item.get("rows")
    rows: list[dict] = []
    if raw_rows is not None:
        if not isinstance(raw_rows, list):
            raise CleideAuditTempTableError(
                ERROR_TEMP_TABLE_INVALID_PAYLOAD,
                "freight_tables.rows deve ser uma lista.",
            )
        for row in raw_rows:
            if not isinstance(row, dict):
                raise CleideAuditTempTableError(
                    ERROR_TEMP_TABLE_INVALID_PAYLOAD,
                    "Cada linha de freight_tables deve ser um objeto.",
                )
            normalized_row: dict = {}
            if columns:
                for col in columns:
                    val = row.get(col)
                    if val is None:
                        normalized_row[col] = None
                    else:
                        normalized_row[col] = _sanitize_cell_string(val)
            else:
                for key, val in row.items():
                    if not isinstance(key, str) or not key.strip():
                        continue
                    safe_key = _sanitize_cell_string(key)
                    if not safe_key:
                        continue
                    normalized_row[safe_key] = _sanitize_cell_string(val) if val is not None else None
            if normalized_row:
                rows.append(normalized_row)
    if not columns and not rows:
        raise CleideAuditTempTableError(
            ERROR_TEMP_TABLE_INVALID_PAYLOAD,
            "Tabela principal não pode ficar completamente vazia.",
        )
    if columns and not rows:
        raise CleideAuditTempTableError(
            ERROR_TEMP_TABLE_INVALID_PAYLOAD,
            "Tabela principal não pode ficar sem linhas.",
        )
    if rows and not columns:
        raise CleideAuditTempTableError(
            ERROR_TEMP_TABLE_INVALID_PAYLOAD,
            "Tabela principal não pode ficar sem colunas.",
        )
    return {
        "table_title": _sanitize_cell_string(item.get("table_title")),
        "table_type": _sanitize_cell_string(item.get("table_type")),
        "context": _sanitize_freight_table_context(item.get("context")),
        "columns": columns,
        "rows": rows,
        "notes": _sanitize_cell_string(item.get("notes")) or "",
        "evidence_ref": _sanitize_cell_string(item.get("evidence_ref")),
        "confidence": _sanitize_cell_string(item.get("confidence")),
    }


def _validate_freight_tables_for_save(raw_tables) -> list[dict]:
    if raw_tables is None:
        return []
    if not isinstance(raw_tables, list):
        raise CleideAuditTempTableError(
            ERROR_TEMP_TABLE_INVALID_PAYLOAD,
            "freight_tables deve ser uma lista.",
        )
    if not raw_tables:
        return []
    return [_validate_freight_table_item_for_save(item) for item in raw_tables]


def _validate_freight_routes_for_save(raw_routes) -> list[dict]:
    if raw_routes is None:
        return []
    if not isinstance(raw_routes, list):
        raise CleideAuditTempTableError(
            ERROR_TEMP_TABLE_INVALID_PAYLOAD,
            "freight_routes deve ser uma lista.",
        )
    if not raw_routes:
        return []
    normalized = _normalize_freight_routes(raw_routes)
    if not normalized:
        raise CleideAuditTempTableError(
            ERROR_TEMP_TABLE_INVALID_PAYLOAD,
            "freight_routes inválido.",
        )
    sanitized: list[dict] = []
    for route in normalized:
        sanitized.append(
            {
                key: _sanitize_cell_string(val) if val is not None else None
                for key, val in route.items()
            }
        )
    return sanitized


def _normalize_accessorial_fee_item(item) -> dict | None:
    if isinstance(item, str):
        text = _optional_normalized_str(item)
        if text is None:
            return None
        return {
            "name": text,
            "value": None,
            "unit": None,
            "calculation_basis": None,
            "notes": "",
            "scope": None,
        }
    if not isinstance(item, dict):
        return None
    return {
        "name": _optional_normalized_str(item.get("name")),
        "value": _optional_normalized_str(item.get("value")),
        "unit": _optional_normalized_str(item.get("unit")),
        "calculation_basis": _optional_normalized_str(item.get("calculation_basis")),
        "notes": _optional_normalized_str(item.get("notes")) or "",
        "scope": _optional_normalized_str(item.get("scope")),
    }


def _normalize_accessorial_fees(raw_fees) -> list[dict]:
    if not isinstance(raw_fees, list):
        return []
    normalized: list[dict] = []
    for item in raw_fees:
        fee = _normalize_accessorial_fee_item(item)
        if fee is not None:
            normalized.append(fee)
    return normalized


def _validate_accessorial_fees_for_save(raw_fees) -> list[dict]:
    if raw_fees is None:
        return []
    if not isinstance(raw_fees, list):
        raise CleideAuditTempTableError(
            ERROR_TEMP_TABLE_INVALID_PAYLOAD,
            "accessorial_fees deve ser uma lista.",
        )
    if not raw_fees:
        return []
    normalized = _normalize_accessorial_fees(raw_fees)
    if len(normalized) != len(raw_fees):
        raise CleideAuditTempTableError(
            ERROR_TEMP_TABLE_INVALID_PAYLOAD,
            "accessorial_fees inválido.",
        )
    return [
        {
            "name": _sanitize_cell_string(item.get("name")),
            "value": _sanitize_cell_string(item.get("value")),
            "unit": _sanitize_cell_string(item.get("unit")),
            "calculation_basis": _sanitize_cell_string(item.get("calculation_basis")),
            "notes": _sanitize_cell_string(item.get("notes")) or "",
            "scope": _sanitize_cell_string(item.get("scope")),
        }
        for item in normalized
    ]


def _assert_temp_table_scope(record: dict, *, user_scope=None, franquia_scope=None) -> None:
    record_user = record.get("user_scope")
    if record_user is not None and user_scope is not None and record_user != user_scope:
        raise CleideAuditTempTableError(
            ERROR_TEMP_TABLE_SCOPE_MISMATCH,
            "Escopo de usuário não autorizado para esta tabela temporária.",
        )
    record_franquia = record.get("franquia_scope")
    if (
        record_franquia is not None
        and franquia_scope is not None
        and record_franquia != franquia_scope
    ):
        raise CleideAuditTempTableError(
            ERROR_TEMP_TABLE_SCOPE_MISMATCH,
            "Escopo de franquia não autorizado para esta tabela temporária.",
        )


def _validate_temp_table_save_payload(payload, *, content_length: int | None = None) -> dict:
    if content_length is not None and content_length > TEMP_TABLE_SAVE_MAX_PAYLOAD_BYTES:
        raise CleideAuditTempTableError(
            ERROR_TEMP_TABLE_PAYLOAD_TOO_LARGE,
            "Payload de edição excede o limite permitido.",
        )
    if not isinstance(payload, dict):
        raise CleideAuditTempTableError(
            ERROR_TEMP_TABLE_INVALID_PAYLOAD,
            "Payload deve ser um objeto JSON.",
        )
    try:
        serialized = json.dumps(payload, ensure_ascii=False)
    except (TypeError, ValueError):
        raise CleideAuditTempTableError(
            ERROR_TEMP_TABLE_INVALID_PAYLOAD,
            "Payload JSON inválido.",
        ) from None
    if len(serialized.encode("utf-8")) > TEMP_TABLE_SAVE_MAX_PAYLOAD_BYTES:
        raise CleideAuditTempTableError(
            ERROR_TEMP_TABLE_PAYLOAD_TOO_LARGE,
            "Payload de edição excede o limite permitido.",
        )
    temp_table_id = payload.get("temp_table_id")
    if not isinstance(temp_table_id, str) or not temp_table_id.strip():
        raise CleideAuditTempTableError(
            ERROR_TEMP_TABLE_INVALID_PAYLOAD,
            "temp_table_id é obrigatório.",
        )
    edit_target = payload.get("edit_target")
    if not isinstance(edit_target, dict):
        raise CleideAuditTempTableError(
            ERROR_TEMP_TABLE_INVALID_PAYLOAD,
            "edit_target é obrigatório.",
        )
    review_action = payload.get("review_action")
    if review_action is not None and review_action != TEMP_TABLE_REVIEW_ACTION_SAVE_AND_ADVANCE:
        raise CleideAuditTempTableError(
            ERROR_TEMP_TABLE_INVALID_PAYLOAD,
            "review_action inválida.",
        )

    has_freight_tables_key = "freight_tables" in edit_target
    has_freight_routes_key = "freight_routes" in edit_target
    has_accessorial_fees_key = "accessorial_fees" in edit_target
    has_coverage_table_key = "coverage_table" in edit_target

    freight_tables = (
        _validate_freight_tables_for_save(edit_target.get("freight_tables"))
        if has_freight_tables_key
        else None
    )
    freight_routes = (
        _validate_freight_routes_for_save(edit_target.get("freight_routes"))
        if has_freight_routes_key
        else None
    )
    accessorial_fees = (
        _validate_accessorial_fees_for_save(edit_target.get("accessorial_fees"))
        if has_accessorial_fees_key
        else None
    )
    coverage_table = (
        _validate_coverage_table_for_save(edit_target.get("coverage_table"))
        if has_coverage_table_key
        else None
    )

    has_freight_structural_edit = bool(
        (freight_tables if has_freight_tables_key else [])
        or (freight_routes if has_freight_routes_key else [])
        or (accessorial_fees if has_accessorial_fees_key else [])
    )
    has_coverage_structural_edit = bool(
        coverage_table and coverage_table.get("rows") is not None
    )
    if not (
        has_freight_tables_key
        or has_freight_routes_key
        or has_accessorial_fees_key
        or has_coverage_table_key
    ):
        raise CleideAuditTempTableError(
            ERROR_TEMP_TABLE_INVALID_PAYLOAD,
            "edit_target deve conter ao menos uma seção editável.",
        )

    return {
        "temp_table_id": temp_table_id.strip(),
        "freight_tables": freight_tables,
        "freight_routes": freight_routes,
        "accessorial_fees": accessorial_fees,
        "coverage_table": coverage_table,
        "has_freight_tables_key": has_freight_tables_key,
        "has_freight_routes_key": has_freight_routes_key,
        "has_accessorial_fees_key": has_accessorial_fees_key,
        "has_coverage_table_key": has_coverage_table_key,
        "has_freight_structural_edit": has_freight_structural_edit,
        "has_coverage_structural_edit": has_coverage_structural_edit,
        "has_structural_edit": bool(has_freight_structural_edit or has_coverage_structural_edit),
        "review_action": review_action or TEMP_TABLE_REVIEW_ACTION_SAVE_AND_ADVANCE,
    }


def save_temp_table_edit(
    payload: dict,
    *,
    user_scope=None,
    franquia_scope=None,
    content_length: int | None = None,
) -> dict:
    """
    Persiste revisão/edição humana no artefato temporário tt_*.json da sessão.

    Não cria novo artefato, não chama Gemini e não grava em banco relacional.
    """
    _require_session()
    validated = _validate_temp_table_save_payload(payload, content_length=content_length)
    sync_temp_table_with_session_documents()
    active_id = get_temp_table_id(session)
    if not active_id:
        raise CleideAuditTempTableError(
            ERROR_TEMP_TABLE_NOT_FOUND,
            "Nenhuma tabela temporária ativa nesta sessão.",
        )
    if validated["temp_table_id"] != active_id:
        raise CleideAuditTempTableError(
            ERROR_TEMP_TABLE_ID_MISMATCH,
            "temp_table_id não corresponde à tabela temporária ativa da sessão.",
        )
    cfg = get_cleiton_doc_config()
    record = load_temp_table_record(active_id, ttl_hours=cfg.upload_ttl_hours)
    if record is None:
        clear_temp_table_session_refs(session)
        _mark_session_modified()
        raise CleideAuditTempTableError(
            ERROR_TEMP_TABLE_NOT_FOUND,
            "Tabela temporária ativa não encontrada.",
        )
    status = (record.get("status") or "").strip().lower()
    if status == TEMP_TABLE_STATUS_EXPIRED:
        raise CleideAuditTempTableError(
            ERROR_TEMP_TABLE_EXPIRED,
            "A tabela temporária desta sessão expirou.",
        )
    if status in {TEMP_TABLE_STATUS_DISCARDED, TEMP_TABLE_STATUS_PROCESSING}:
        raise CleideAuditTempTableError(
            ERROR_TEMP_TABLE_NOT_FOUND,
            "Tabela temporária indisponível para revisão.",
        )
    _assert_temp_table_scope(record, user_scope=user_scope, franquia_scope=franquia_scope)

    now = _utcnow().isoformat()
    preserved_expires_at = record.get("expires_at")
    updated = dict(record)
    if validated["freight_tables"]:
        updated["freight_tables"] = validated["freight_tables"]
    if validated["freight_routes"]:
        updated["freight_routes"] = validated["freight_routes"]
    if validated["has_accessorial_fees_key"] and validated["accessorial_fees"] is not None:
        updated["accessorial_fees"] = validated["accessorial_fees"]

    has_freight_edit = (
        validated["has_freight_tables_key"]
        or validated["has_freight_routes_key"]
        or validated["has_accessorial_fees_key"]
    )
    if has_freight_edit:
        updated["human_review_status"] = (
            HUMAN_REVIEW_STATUS_EDITED
            if validated["has_freight_structural_edit"]
            else HUMAN_REVIEW_STATUS_REVIEWED
        )
        updated["human_edited_at"] = now
        if user_scope is not None:
            updated["human_edited_by_user_id"] = user_scope
        current_edit_version = updated.get("edit_version")
        if isinstance(current_edit_version, int) and current_edit_version >= 0:
            updated["edit_version"] = current_edit_version + 1
        else:
            updated["edit_version"] = 1

    if validated["has_coverage_table_key"] and validated["coverage_table"] is not None:
        existing_coverage = updated.get("coverage_table")
        if not isinstance(existing_coverage, dict):
            existing_coverage = _empty_coverage_table_shell()
        coverage = dict(existing_coverage)
        coverage["status"] = COVERAGE_TABLE_STATUS_NEEDS_REVIEW
        coverage["columns"] = list(COVERAGE_TABLE_COLUMNS)
        coverage["rows"] = validated["coverage_table"]["rows"]
        coverage["human_review_status"] = (
            HUMAN_REVIEW_STATUS_EDITED
            if validated["has_coverage_structural_edit"]
            else HUMAN_REVIEW_STATUS_REVIEWED
        )
        coverage["human_edited_at"] = now
        if user_scope is not None:
            coverage["human_edited_by_user_id"] = user_scope
        current_coverage_version = coverage.get("edit_version")
        if isinstance(current_coverage_version, int) and current_coverage_version >= 0:
            coverage["edit_version"] = current_coverage_version + 1
        else:
            coverage["edit_version"] = 1
        if "validation_warnings" not in coverage or not isinstance(coverage.get("validation_warnings"), list):
            coverage["validation_warnings"] = []
        if "uploaded_at" not in coverage:
            coverage["uploaded_at"] = now
        updated["coverage_table"] = coverage

    updated["updated_at"] = now
    updated["expires_at"] = preserved_expires_at

    saved = save_temp_table_record(updated)
    logger.info(
        "Cleide temp_table save: temp_table_id=%s user_id=%s status=%s tables=%s routes=%s fees=%s coverage_rows=%s",
        saved.get("temp_table_id"),
        user_scope,
        updated.get("human_review_status"),
        len(saved.get("freight_tables") or []),
        len(saved.get("freight_routes") or []),
        len(saved.get("accessorial_fees") or []),
        len((saved.get("coverage_table") or {}).get("rows") or []),
    )
    public = _public_temp_table(saved)
    if public is None:
        raise CleideAuditTempTableError(
            ERROR_TEMP_TABLE_NOT_FOUND,
            "Não foi possível retornar a tabela temporária atualizada.",
        )
    return public


def load_temp_table_record(temp_table_id: str, *, ttl_hours: int) -> dict | None:
    ref = (temp_table_id or "").strip()
    if not ref:
        return None
    try:
        path = _temp_table_path(ref)
    except ValueError:
        return None
    if not path.is_file():
        return None
    try:
        with open(path, "r", encoding="utf-8") as handle:
            payload = json.load(handle)
        if not isinstance(payload, dict):
            remove_temp_table_record(ref)
            return None
    except Exception:
        remove_temp_table_record(ref)
        return None

    expires_at = _parse_iso(payload.get(FIELD_EXPIRES_AT))
    if expires_at is None:
        created_at = _parse_iso(payload.get(FIELD_CREATED_AT))
        if created_at is None:
            remove_temp_table_record(ref)
            return None
        expires_at = created_at + timedelta(hours=max(1, int(ttl_hours)))
    if _utcnow() >= expires_at:
        payload["status"] = TEMP_TABLE_STATUS_EXPIRED
        payload["updated_at"] = _utcnow().isoformat()
        try:
            _write_temp_table_atomic(path, payload)
        except Exception:
            remove_temp_table_record(ref)
            return None
        return payload
    return payload


def save_temp_table_record(record: dict) -> dict:
    _require_session()
    temp_table_id = (record.get("temp_table_id") or uuid4().hex).strip()
    record = dict(record)
    record["temp_table_id"] = temp_table_id
    path = _temp_table_path(temp_table_id)
    _write_temp_table_atomic(path, record)
    set_temp_table_id(session, temp_table_id)
    set_temp_table_source_doc_ids(session, list(record.get("source_documents") or []))
    _mark_session_modified()
    return record


def remove_temp_table_record(temp_table_id: str) -> bool:
    ref = (temp_table_id or "").strip()
    if not ref:
        return False
    try:
        path = _temp_table_path(ref)
    except ValueError:
        return False
    if path.is_file():
        try:
            path.unlink()
        except Exception:
            return False
    return True


def invalidate_temp_table_for_session(*, reason: str = TEMP_TABLE_STATUS_DISCARDED) -> None:
    _require_session()
    temp_table_id = get_temp_table_id(session)
    clear_temp_table_session_refs(session)
    if temp_table_id:
        remove_temp_table_record(temp_table_id)
    _mark_session_modified()


def invalidate_temp_table_if_source_changed(
    *,
    reason: str = TEMP_TABLE_STATUS_DISCARDED,
    removed_doc_id: str | None = None,
) -> None:
    _require_session()
    temp_table_id = get_temp_table_id(session)
    if not temp_table_id:
        return
    cfg = get_cleiton_doc_config()
    record = load_temp_table_record(temp_table_id, ttl_hours=cfg.upload_ttl_hours)
    if record is None:
        clear_temp_table_session_refs(session)
        _mark_session_modified()
        return
    source_docs = list(record.get("source_documents") or [])
    active_ids = set(get_cleide_audit_doc_ids(session))
    if removed_doc_id and removed_doc_id in source_docs:
        invalidate_temp_table_for_session(reason=reason)
        return
    if source_docs and not all(doc_id in active_ids for doc_id in source_docs):
        invalidate_temp_table_for_session(reason=reason)


def _parse_iso(raw: str | None) -> datetime | None:
    if not raw:
        return None
    try:
        dt = datetime.fromisoformat(str(raw).replace("Z", "+00:00"))
        if dt.tzinfo:
            off = dt.utcoffset()
            dt = (dt.replace(tzinfo=None) - off) if off else dt.replace(tzinfo=None)
        return dt
    except Exception:
        return None


def _temp_table_expires_at(source_doc_ids: list[str]) -> str:
    cfg = get_cleiton_doc_config()
    latest: datetime | None = None
    for doc_id in source_doc_ids:
        record = load_document_record(doc_id, ttl_hours=cfg.upload_ttl_hours)
        if record is None:
            continue
        candidate = _parse_iso(record.get(FIELD_EXPIRES_AT))
        if candidate and (latest is None or candidate > latest):
            latest = candidate
    if latest is None:
        latest = _utcnow() + timedelta(hours=max(1, int(cfg.upload_ttl_hours)))
    return latest.isoformat()


def get_active_temp_table_for_session() -> dict | None:
    _require_session()
    sync_temp_table_with_session_documents()
    temp_table_id = get_temp_table_id(session)
    if not temp_table_id:
        return None
    cfg = get_cleiton_doc_config()
    record = load_temp_table_record(temp_table_id, ttl_hours=cfg.upload_ttl_hours)
    if record is None:
        clear_temp_table_session_refs(session)
        _mark_session_modified()
        return None
    status = (record.get("status") or "").strip()
    if status in {TEMP_TABLE_STATUS_DISCARDED, TEMP_TABLE_STATUS_EXPIRED}:
        return _public_temp_table(record)
    return _public_temp_table(record)


def sync_temp_table_with_session_documents() -> None:
    _require_session()
    temp_table_id = get_temp_table_id(session)
    if not temp_table_id:
        return
    cfg = get_cleiton_doc_config()
    record = load_temp_table_record(temp_table_id, ttl_hours=cfg.upload_ttl_hours)
    if record is None:
        clear_temp_table_session_refs(session)
        _mark_session_modified()
        return
    active_ids = set(get_cleide_audit_doc_ids(session))
    source_docs = list(record.get("source_documents") or [])
    if source_docs and not all(doc_id in active_ids for doc_id in source_docs):
        invalidate_temp_table_for_session(reason=TEMP_TABLE_STATUS_DISCARDED)


def should_attempt_temp_table_extraction(session_obj, source_doc_ids: list[str]) -> bool:
    normalized = _normalize_source_doc_ids(source_doc_ids)
    if not normalized:
        return False
    sync_temp_table_with_session_documents()
    temp_table_id = get_temp_table_id(session_obj)
    if not temp_table_id:
        return True
    cfg = get_cleiton_doc_config()
    record = load_temp_table_record(temp_table_id, ttl_hours=cfg.upload_ttl_hours)
    if record is None:
        return True
    bound_sources = _normalize_source_doc_ids(list(record.get("source_documents") or []))
    if bound_sources != normalized:
        return True
    raw_status = record.get("status")
    status = raw_status.strip().lower() if isinstance(raw_status, str) else ""
    if status == TEMP_TABLE_STATUS_PROCESSING:
        return False
    return False


def mark_temp_table_processing(source_doc_ids: list[str], *, user_scope=None, franquia_scope=None) -> dict:
    _require_session()
    normalized = _normalize_source_doc_ids(source_doc_ids)
    now = _utcnow()
    temp_table_id = get_temp_table_id(session) or uuid4().hex
    record = {
        "temp_table_id": temp_table_id,
        "status": TEMP_TABLE_STATUS_PROCESSING,
        "source_documents": normalized,
        "detected_carrier": None,
        "origins": [],
        "destinations": [],
        "routes": [],
        "freight_tables": [],
        "freight_routes": [],
        "weight_ranges": [],
        "freight_values": [],
        "accessorial_fees": [],
        "charge_type_detected": None,
        "extracted_items": [],
        "uncertain_fields": [],
        "reading_alerts": [],
        "evidence_refs": [],
        "created_at": now.isoformat(),
        "updated_at": now.isoformat(),
        "expires_at": _temp_table_expires_at(normalized),
        "session_scope": CLEIDE_AUDIT_DOC_IDS_SESSION_KEY,
        "franquia_scope": franquia_scope,
        "user_scope": user_scope,
        "operational_owner": TEMP_TABLE_OPERATIONAL_OWNER,
        "ui_visibility": {
            "display_name": TEMP_TABLE_UI_DISPLAY_NAME,
            "readonly": True,
        },
        "version_marker": TEMP_TABLE_VERSION_MARKER,
    }
    return save_temp_table_record(record)


def temp_table_status_message(status: str) -> str:
    mapping = {
        TEMP_TABLE_STATUS_PROCESSING: (
            "Recebi os anexos e iniciei a estruturação da tabela temporária de frete."
        ),
        TEMP_TABLE_STATUS_AWAITING_VALIDATION: (
            "A tabela temporária foi estruturada e está aguardando sua validação."
        ),
        TEMP_TABLE_STATUS_NEEDS_REVIEW: (
            "A tabela temporária foi gerada. Revise os dados antes de continuar."
        ),
        TEMP_TABLE_STATUS_FAILED: (
            "Não foi possível estruturar a tabela temporária a partir dos anexos enviados."
        ),
        TEMP_TABLE_STATUS_EXPIRED: (
            "A tabela temporária desta sessão expirou."
        ),
        TEMP_TABLE_STATUS_DISCARDED: (
            "Os documentos de origem foram alterados ou removidos, "
            "então a tabela temporária anterior foi invalidada."
        ),
    }
    return mapping.get((status or "").strip(), "")


def _normalize_temp_table_status(raw_status) -> str:
    allowed = {
        TEMP_TABLE_STATUS_AWAITING_VALIDATION,
        TEMP_TABLE_STATUS_NEEDS_REVIEW,
        TEMP_TABLE_STATUS_FAILED,
        TEMP_TABLE_STATUS_VALIDATED,
    }
    if not isinstance(raw_status, str):
        return TEMP_TABLE_STATUS_FAILED
    candidate = raw_status.strip().lower()
    if not candidate or candidate not in allowed:
        return TEMP_TABLE_STATUS_FAILED
    return candidate


def _list_field_from_raw(raw: dict, name: str) -> list:
    value = raw.get(name)
    return list(value) if isinstance(value, list) else []


def _is_useful_freight_value(item) -> bool:
    if isinstance(item, dict):
        label = item.get("label")
        if isinstance(label, str) and label.strip():
            return True
        return item.get("value") is not None
    if isinstance(item, str):
        return bool(item.strip())
    return False


def _is_useful_accessorial_fee(item) -> bool:
    if isinstance(item, dict):
        name = item.get("name")
        if isinstance(name, str) and name.strip():
            return True
        return item.get("value") is not None
    if isinstance(item, str):
        return bool(item.strip())
    return False


def _is_useful_weight_range(item) -> bool:
    if isinstance(item, dict):
        label = item.get("label")
        if isinstance(label, str) and label.strip():
            return True
        if item.get("min_weight") is not None or item.get("max_weight") is not None:
            return True
    if isinstance(item, str):
        return bool(item.strip())
    return False


def _optional_normalized_str(value) -> str | None:
    if value is None:
        return None
    if isinstance(value, str):
        candidate = value.strip()
        return candidate or None
    candidate = str(value).strip()
    return candidate or None


def _freight_route_field(item: dict, primary: str, *aliases: str) -> str | None:
    keys = (primary, *aliases)
    for key in keys:
        if key not in item:
            continue
        normalized = _optional_normalized_str(item.get(key))
        if normalized is not None:
            return normalized
    return None


def _normalize_freight_route_item(item) -> dict | None:
    if not isinstance(item, dict):
        return None
    return {
        "origin": _freight_route_field(item, "origin"),
        "destination": _freight_route_field(item, "destination"),
        "freight_type": _freight_route_field(item, "freight_type", "type"),
        "weight_30": _freight_route_field(item, "weight_30", "weight_30kg"),
        "weight_50": _freight_route_field(item, "weight_50", "weight_50kg"),
        "weight_70": _freight_route_field(item, "weight_70", "weight_70kg"),
        "weight_100": _freight_route_field(item, "weight_100", "weight_100kg"),
        "boarding_fee": _freight_route_field(item, "boarding_fee", "taxa_embarque_kg"),
        "freight_value_pct": _freight_route_field(item, "freight_value_pct", "frete_valor_pct"),
        "freight_weight_kg": _freight_route_field(item, "freight_weight_kg", "frete_peso_kg"),
        "notes": _freight_route_field(item, "notes", "observations", "observacoes") or "",
        "evidence_ref": _freight_route_field(item, "evidence_ref"),
        "confidence": _freight_route_field(item, "confidence"),
    }


def _is_useful_freight_route(item) -> bool:
    normalized = _normalize_freight_route_item(item) if isinstance(item, dict) else None
    if not normalized:
        return False
    useful_fields = (
        "origin",
        "destination",
        "freight_type",
        "weight_30",
        "weight_50",
        "weight_70",
        "weight_100",
        "boarding_fee",
        "freight_value_pct",
        "freight_weight_kg",
    )
    return any(normalized.get(field) is not None for field in useful_fields)


def _normalize_freight_routes(raw_routes) -> list[dict]:
    if not isinstance(raw_routes, list):
        return []
    normalized: list[dict] = []
    for item in raw_routes:
        route = _normalize_freight_route_item(item)
        if route is not None:
            normalized.append(route)
    return normalized


def _normalize_freight_table_context(raw_context) -> dict:
    if not isinstance(raw_context, dict):
        return {
            "route_label": None,
            "origin": None,
            "destination": None,
            "customer": None,
            "supplier": None,
            "valid_from": None,
            "valid_to": None,
            "delivery_deadline": None,
        }
    return {
        "route_label": _optional_normalized_str(raw_context.get("route_label")),
        "origin": _optional_normalized_str(raw_context.get("origin")),
        "destination": _optional_normalized_str(raw_context.get("destination")),
        "customer": _optional_normalized_str(raw_context.get("customer")),
        "supplier": _optional_normalized_str(raw_context.get("supplier")),
        "valid_from": _optional_normalized_str(raw_context.get("valid_from")),
        "valid_to": _optional_normalized_str(raw_context.get("valid_to")),
        "delivery_deadline": _optional_normalized_str(raw_context.get("delivery_deadline")),
    }


def _normalize_freight_table_row(item, columns: list[str]) -> dict:
    if not isinstance(item, dict):
        return {}
    normalized: dict = {}
    for col in columns:
        if col in item:
            val = item.get(col)
            if val is None:
                normalized[col] = None
            elif isinstance(val, str):
                normalized[col] = val
            else:
                normalized[col] = str(val)
    for key, val in item.items():
        if key not in normalized and isinstance(key, str) and key.strip():
            if val is None:
                normalized[key] = None
            elif isinstance(val, str):
                normalized[key] = val
            else:
                normalized[key] = str(val)
    return normalized


def _normalize_freight_table_item(item) -> dict | None:
    if not isinstance(item, dict):
        return None
    raw_columns = item.get("columns")
    columns: list[str] = []
    if isinstance(raw_columns, list):
        for col in raw_columns:
            if isinstance(col, str):
                candidate = col.strip()
                if candidate:
                    columns.append(candidate)
    raw_rows = item.get("rows")
    rows: list[dict] = []
    if isinstance(raw_rows, list):
        for row in raw_rows:
            normalized_row = _normalize_freight_table_row(row, columns)
            if normalized_row:
                rows.append(normalized_row)
    return {
        "table_title": _optional_normalized_str(item.get("table_title")),
        "table_type": _optional_normalized_str(item.get("table_type")),
        "context": _normalize_freight_table_context(item.get("context")),
        "columns": columns,
        "rows": rows,
        "notes": _optional_normalized_str(item.get("notes")) or "",
        "evidence_ref": _optional_normalized_str(item.get("evidence_ref")),
        "confidence": _optional_normalized_str(item.get("confidence")),
    }


def _row_has_any_value(row: dict) -> bool:
    if not isinstance(row, dict):
        return False
    for val in row.values():
        if val is None:
            continue
        if isinstance(val, str) and not val.strip():
            continue
        return True
    return False


def _is_useful_freight_table(item) -> bool:
    normalized = _normalize_freight_table_item(item) if isinstance(item, dict) else None
    if not normalized:
        return False
    if normalized.get("table_title"):
        return True
    if normalized.get("columns"):
        return True
    for row in normalized.get("rows") or []:
        if _row_has_any_value(row):
            return True
    return False


def _normalize_freight_tables(raw_tables) -> list[dict]:
    if not isinstance(raw_tables, list):
        return []
    normalized: list[dict] = []
    for item in raw_tables:
        table = _normalize_freight_table_item(item)
        if table is not None:
            normalized.append(table)
    return normalized


def _has_useful_partial_extraction_data(raw: dict) -> bool:
    for item in _list_field_from_raw(raw, "freight_tables"):
        if _is_useful_freight_table(item):
            return True
    for item in _list_field_from_raw(raw, "freight_routes"):
        if _is_useful_freight_route(item):
            return True
    for item in _list_field_from_raw(raw, "freight_values"):
        if _is_useful_freight_value(item):
            return True
    for item in _list_field_from_raw(raw, "accessorial_fees"):
        if _is_useful_accessorial_fee(item):
            return True
    for item in _list_field_from_raw(raw, "weight_ranges"):
        if _is_useful_weight_range(item):
            return True
    return False


def _has_legacy_useful_extraction_data(raw: dict) -> bool:
    for name in ("origins", "destinations", "routes", "extracted_items", "uncertain_fields"):
        items = _list_field_from_raw(raw, name)
        if items:
            return True
    carrier = raw.get("detected_carrier")
    if isinstance(carrier, str) and carrier.strip():
        return True
    charge = raw.get("charge_type_detected")
    if isinstance(charge, str) and charge.strip():
        return True
    return False


def _resolve_extraction_status(raw: dict, expanded: dict) -> str:
    """Resolve status final após normalização partial-first."""
    has_partial = _has_useful_partial_extraction_data(expanded)
    has_legacy = _has_legacy_useful_extraction_data(expanded)
    candidate = _normalize_temp_table_status(raw.get("status"))

    if has_partial:
        return TEMP_TABLE_STATUS_NEEDS_REVIEW

    if candidate == TEMP_TABLE_STATUS_AWAITING_VALIDATION:
        return TEMP_TABLE_STATUS_AWAITING_VALIDATION

    if candidate == TEMP_TABLE_STATUS_VALIDATED:
        return TEMP_TABLE_STATUS_VALIDATED

    if has_legacy and candidate == TEMP_TABLE_STATUS_FAILED:
        return TEMP_TABLE_STATUS_NEEDS_REVIEW

    if has_legacy:
        return candidate

    if candidate == TEMP_TABLE_STATUS_NEEDS_REVIEW:
        return TEMP_TABLE_STATUS_FAILED

    return TEMP_TABLE_STATUS_FAILED


def normalize_partial_first_extraction_to_temp_table(raw: dict) -> dict:
    """
    Normaliza a resposta partial-first (Etapa A) para o contrato interno da temp_table.

    A Etapa A retorna apenas custos brutos detectados; o backend completa campos
    opcionais do contrato interno sem exigir fechamento de rotas ou transportadora.
    """
    alerts = [
        str(item).strip()
        for item in _list_field_from_raw(raw, "reading_alerts")
        if isinstance(item, str) and str(item).strip()
    ]
    expanded = {
        "status": raw.get("status"),
        "detected_carrier": raw.get("detected_carrier"),
        "origins": _list_field_from_raw(raw, "origins"),
        "destinations": _list_field_from_raw(raw, "destinations"),
        "routes": _list_field_from_raw(raw, "routes"),
        "freight_tables": _normalize_freight_tables(_list_field_from_raw(raw, "freight_tables")),
        "freight_routes": _normalize_freight_routes(_list_field_from_raw(raw, "freight_routes")),
        "weight_ranges": _list_field_from_raw(raw, "weight_ranges"),
        "freight_values": _list_field_from_raw(raw, "freight_values"),
        "accessorial_fees": _list_field_from_raw(raw, "accessorial_fees"),
        "charge_type_detected": raw.get("charge_type_detected"),
        "extracted_items": _list_field_from_raw(raw, "extracted_items"),
        "uncertain_fields": _list_field_from_raw(raw, "uncertain_fields"),
        "reading_alerts": alerts,
        "evidence_refs": _list_field_from_raw(raw, "evidence_refs"),
        "franquia_scope": raw.get("franquia_scope"),
        "user_scope": raw.get("user_scope"),
    }
    expanded["status"] = _resolve_extraction_status(raw, expanded)
    return expanded


def _coerce_temp_table_payload(raw: dict, *, source_doc_ids: list[str]) -> dict:
    now = _utcnow().isoformat()
    normalized = normalize_partial_first_extraction_to_temp_table(raw)
    status = _resolve_extraction_status(raw, normalized)
    uncertain = (
        normalized.get("uncertain_fields")
        if isinstance(normalized.get("uncertain_fields"), list)
        else []
    )
    alerts = [
        str(item).strip()
        for item in (
            normalized.get("reading_alerts")
            if isinstance(normalized.get("reading_alerts"), list)
            else []
        )
        if isinstance(item, str) and str(item).strip()
    ]

    if _has_useful_partial_extraction_data(normalized):
        status = TEMP_TABLE_STATUS_NEEDS_REVIEW
    elif status == TEMP_TABLE_STATUS_NEEDS_REVIEW and not _has_legacy_useful_extraction_data(
        normalized
    ):
        status = TEMP_TABLE_STATUS_FAILED

    if status == TEMP_TABLE_STATUS_AWAITING_VALIDATION and (uncertain or alerts):
        status = TEMP_TABLE_STATUS_NEEDS_REVIEW
    if status == TEMP_TABLE_STATUS_NEEDS_REVIEW and not alerts and not uncertain:
        alerts = [
            "A extração encontrou dados parciais e precisa de validação humana."
        ]
    if status == TEMP_TABLE_STATUS_FAILED and not alerts:
        alerts = [
            "Não foi possível estruturar a tabela temporária a partir dos anexos enviados."
        ]

    temp_table_id = get_temp_table_id(session) or uuid4().hex
    existing = load_temp_table_record(temp_table_id, ttl_hours=get_cleiton_doc_config().upload_ttl_hours)
    created_at = (existing or {}).get("created_at") or now
    preserved_coverage = None
    if existing and isinstance(existing.get("coverage_table"), dict):
        preserved_coverage = existing.get("coverage_table")
    preserved_audit_batch = None
    if existing and isinstance(existing.get("audit_batch"), dict):
        preserved_audit_batch = existing.get("audit_batch")
    record = {
        "temp_table_id": temp_table_id,
        "status": status,
        "source_documents": _normalize_source_doc_ids(source_doc_ids),
        "detected_carrier": normalized.get("detected_carrier"),
        "origins": _list_field_from_raw(normalized, "origins"),
        "destinations": _list_field_from_raw(normalized, "destinations"),
        "routes": _list_field_from_raw(normalized, "routes"),
        "freight_tables": _normalize_freight_tables(_list_field_from_raw(normalized, "freight_tables")),
        "freight_routes": _normalize_freight_routes(_list_field_from_raw(normalized, "freight_routes")),
        "weight_ranges": _list_field_from_raw(normalized, "weight_ranges"),
        "freight_values": _list_field_from_raw(normalized, "freight_values"),
        "accessorial_fees": _list_field_from_raw(normalized, "accessorial_fees"),
        "charge_type_detected": normalized.get("charge_type_detected"),
        "extracted_items": _list_field_from_raw(normalized, "extracted_items"),
        "uncertain_fields": uncertain,
        "reading_alerts": alerts,
        "evidence_refs": _list_field_from_raw(normalized, "evidence_refs"),
        "created_at": created_at,
        "updated_at": now,
        "expires_at": _temp_table_expires_at(_normalize_source_doc_ids(source_doc_ids)),
        "session_scope": CLEIDE_AUDIT_DOC_IDS_SESSION_KEY,
        "franquia_scope": normalized.get("franquia_scope"),
        "user_scope": normalized.get("user_scope"),
        "operational_owner": TEMP_TABLE_OPERATIONAL_OWNER,
        "ui_visibility": {
            "display_name": TEMP_TABLE_UI_DISPLAY_NAME,
            "readonly": True,
        },
        "version_marker": TEMP_TABLE_VERSION_MARKER,
    }
    if preserved_coverage is not None:
        record["coverage_table"] = preserved_coverage
    if preserved_audit_batch is not None:
        record["audit_batch"] = preserved_audit_batch
    return record


def _has_human_review_metadata(record: dict | None) -> bool:
    if not record or not isinstance(record, dict):
        return False
    if record.get("human_review_status"):
        return True
    if record.get("human_edited_at"):
        return True
    if record.get("human_edited_by_user_id") is not None:
        return True
    edit_version = record.get("edit_version")
    if isinstance(edit_version, int) and edit_version > 0:
        return True
    if isinstance(edit_version, str) and edit_version.strip().isdigit() and int(edit_version) > 0:
        return True
    return False


def _source_documents_match(existing_sources, incoming_sources: list[str]) -> bool:
    return _normalize_source_doc_ids(list(existing_sources or [])) == _normalize_source_doc_ids(
        incoming_sources
    )


def _should_skip_extraction_overwrite(
    existing: dict | None,
    *,
    source_doc_ids: list[str],
    force_overwrite: bool = False,
) -> bool:
    if force_overwrite or not existing:
        return False
    if not _has_human_review_metadata(existing):
        return False
    return _source_documents_match(existing.get("source_documents"), source_doc_ids)


def apply_temp_table_extraction_from_model_payload(
    payload: dict | None,
    *,
    source_doc_ids: list[str],
    force_overwrite: bool = False,
) -> dict | None:
    _require_session()
    normalized = _normalize_source_doc_ids(source_doc_ids)
    if not normalized:
        return None

    cfg = get_cleiton_doc_config()
    temp_table_id = get_temp_table_id(session)
    existing = None
    if temp_table_id:
        existing = load_temp_table_record(temp_table_id, ttl_hours=cfg.upload_ttl_hours)

    if _should_skip_extraction_overwrite(
        existing,
        source_doc_ids=normalized,
        force_overwrite=force_overwrite,
    ):
        logger.info(
            "Cleide temp_table extraction skipped because human-reviewed artifact already exists "
            "(temp_table_id=%s edit_version=%s human_review_status=%s).",
            existing.get("temp_table_id") if existing else None,
            (existing or {}).get("edit_version"),
            (existing or {}).get("human_review_status"),
        )
        return existing

    if not isinstance(payload, dict):
        record = _coerce_temp_table_payload({"status": TEMP_TABLE_STATUS_FAILED}, source_doc_ids=normalized)
        return save_temp_table_record(record)
    record = _coerce_temp_table_payload(payload, source_doc_ids=normalized)
    return save_temp_table_record(record)


def split_temp_table_block_from_answer(answer_text: str) -> tuple[str, dict | None]:
    text = answer_text or ""
    begin = text.find(TEMP_TABLE_JSON_BEGIN)
    if begin < 0:
        return text.strip(), None
    end = text.find(TEMP_TABLE_JSON_END, begin)
    if end < 0:
        return text.strip(), None
    json_chunk = text[begin + len(TEMP_TABLE_JSON_BEGIN) : end].strip()
    visible = (text[:begin] + text[end + len(TEMP_TABLE_JSON_END) :]).strip()
    if not json_chunk:
        return visible, None
    try:
        parsed = json.loads(json_chunk)
    except (TypeError, ValueError, json.JSONDecodeError):
        return visible, None
    return visible, parsed if isinstance(parsed, dict) else None


def build_document_status_metadata() -> dict:
    """Metadados básicos para futuro endpoint de status da Cleide Auditoria."""
    _require_session()
    maybe_cleanup_expired_cleiton_docs()
    sync_temp_table_with_session_documents()
    totals = get_document_session_totals()
    temp_table = get_active_temp_table_for_session()
    return {
        "domain": CLEIDE_AUDIT_DOMAIN,
        "flow_types": {
            "upload": CLEIDE_AUDIT_DOCUMENT_UPLOAD_FLOW_TYPE,
            "prepare": CLEIDE_AUDIT_DOCUMENT_PREPARE_FLOW_TYPE,
            "chat": CLEIDE_AUDIT_CHAT_FLOW_TYPE,
            "temp_table_extraction": CLEIDE_AUDIT_TEMP_TABLE_EXTRACTION_FLOW_TYPE,
        },
        "documents": get_active_documents_for_session(),
        "temp_table": temp_table,
        "allowed_formats": get_allowed_document_formats(),
        "session": {
            "count": totals["active_count"],
            "max_files": totals["max_files_per_session"],
            "total_bytes": totals["total_bytes"],
            "session_max_bytes": totals["session_max_bytes"],
        },
    }
