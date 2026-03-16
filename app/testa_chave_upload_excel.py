import openpyxl

# Caminho do arquivo Excel de upload (ajuste conforme necessário)
CAMINHO_EXCEL = r'C:\Users\User\Desktop\logcompleta_docs\base id\inteligencia de frete\ARQUIVO.xlsx'  # <-- Substitua pelo caminho real do arquivo de upload

print("\n--- TESTE DE EXTRAÇÃO DE CHAVES DO EXCEL DE UPLOAD ---\n")

try:
    wb = openpyxl.load_workbook(CAMINHO_EXCEL, read_only=True, data_only=True)
    ws = wb.active
    colunas = [str(c).strip().lower() for c in next(ws.iter_rows(min_row=1, max_row=1, values_only=True), [])]
    indices = {c: colunas.index(c) for c in colunas}
    for num_linha, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not any(v is not None and str(v).strip() for v in row):
            continue
        cidade_origem = str(row[indices["cidade_origem"]]).strip().lower()
        uf_origem = str(row[indices["uf_origem"]]).strip().lower()
        chave_origem = f"{cidade_origem}-{uf_origem}"
        print(f"Linha {num_linha}: cidade_origem='{cidade_origem}', uf_origem='{uf_origem}', chave_origem='{chave_origem}'")
except Exception as e:
    print("Erro ao processar o Excel:", e)

print("\n--- FIM DO TESTE ---\n")
