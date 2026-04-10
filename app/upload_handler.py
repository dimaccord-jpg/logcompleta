"""
Upload e gerenciamento temporário de dados de fretes para o Roberto Intelligence.
Valida Excel, normaliza texto, resolve localidades em base_localidades (id_cidade, id_uf, UF textual)
e salva os dados em persistência temporária dedicada (referência leve em sessão).
"""
import logging
import os
import time
from datetime import datetime
from typing import Any
from uuid import uuid4

from flask import request, session, jsonify
from werkzeug.datastructures import FileStorage

from app.infra import carregar_localidades_por_chaves
from app.roberto_upload_store import (
    clear_upload_data as clear_upload_data_store,
    maybe_cleanup_expired_uploads,
    read_upload_data,
    save_upload_data,
)
from app.services.roberto_config_service import get_roberto_config

logger = logging.getLogger(__name__)

# Diretório dedicado para uploads temporários do Roberto (isolado do flask_session)
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
UPLOAD_DIR = os.path.join(BASE_DIR, "roberto_uploads")
os.makedirs(UPLOAD_DIR, exist_ok=True)

# Colunas obrigatórias no Excel (nomes exatos após normalização do cabeçalho)
COLUNAS_OBRIGATORIAS = {
    "data_emissao",
    "cidade_origem",
    "uf_origem",
    "cidade_destino",
    "uf_destino",
    "peso_real",
    "valor_nf",
    "valor_frete_total",
    "modal",
}
COLUNA_OPCIONAL_IMPOSTO = "valor_imposto"

# Chave de sessão com referência ao payload temporário dedicado.
SESSION_KEY_UPLOAD_REF = "roberto_upload_ref"


def _normalizar_texto(val: Any) -> str:
    """Converte valor para string e coloca em minúsculas, sem remover acentos."""
    if val is None:
        return ""
    return str(val).strip().lower()


def _uf_para_payload(uf_nome: str | None) -> str:
    """Alinha UF textual ao contrato do BI (2 letras maiúsculas)."""
    s = (uf_nome or "").strip().upper()
    return s[:2] if s else ""


def _ler_cabecalho_normalizado(ws) -> list[str]:
    """Lê primeira linha da planilha e retorna colunas em minúsculas."""
    row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
    if not row:
        return []
    return [_normalizar_texto(c) for c in row]


def _validar_arquivo_excel(arquivo: FileStorage) -> tuple[bool, str]:
    """Valida extensão e tipo MIME básico."""
    if not arquivo or not arquivo.filename:
        return False, "Nenhum arquivo enviado."
    fn = (arquivo.filename or "").strip().lower()
    if not fn.endswith(".xlsx"):
        return False, "Apenas arquivos Excel (.xlsx) são aceitos."
    return True, ""


def _validar_colunas(colunas: list[str]) -> tuple[bool, str]:
    """Verifica se todas as colunas obrigatórias estão presentes."""
    conjunto = set(colunas)
    faltando = COLUNAS_OBRIGATORIAS - conjunto
    if faltando:
        return False, f"Colunas obrigatórias ausentes: {', '.join(sorted(faltando))}."
    return True, ""


def _resolve_execution_id() -> str:
    """
    Resolve identidade da execução:
    - Header: X-Execution-ID
    - Form: execution_id
    - Fallback: UUID da própria request (compatível com clientes antigos)
    """
    execution_id = (request.headers.get("X-Execution-ID") or "").strip()
    if not execution_id:
        execution_id = (request.form.get("execution_id") or "").strip()
    if not execution_id:
        execution_id = str(uuid4())
    return execution_id[:120]


def _sort_key_recente(reg: dict[str, Any]) -> str:
    val = reg.get("data_emissao")
    if val is None:
        return ""
    if hasattr(val, "isoformat"):
        return val.isoformat()[:10]
    return str(val).strip()[:10]


def _chave_mes_data(reg: dict[str, Any]) -> str:
    val = reg.get("data_emissao")
    if val is None:
        return "sem_data"
    if hasattr(val, "strftime"):
        return val.strftime("%Y-%m")
    s = str(val).strip()[:10]
    return s[:7] if len(s) >= 7 else "sem_data"


def _selecionar_meses_representativos(meses_ordenados: list[str], limite_meses: int) -> list[str]:
    if limite_meses >= len(meses_ordenados):
        return list(meses_ordenados)
    if limite_meses <= 1:
        # Melhor esforço para preservar representatividade temporal:
        # quando só cabe 1 mês, prioriza o mais recente.
        return [meses_ordenados[-1]]

    total = len(meses_ordenados)
    escolhidos_idx: set[int] = set()
    for i in range(limite_meses):
        idx = round(i * (total - 1) / (limite_meses - 1))
        escolhidos_idx.add(idx)
    return [meses_ordenados[i] for i in sorted(escolhidos_idx)]


def _aplicar_limite_upload_total(
    linhas: list[dict],
    upload_total_max: int,
) -> tuple[list[dict], dict[str, int]]:
    """
    Limita o volume total de linhas do upload para manter segurança operacional.
    Estratégia: preserva representatividade temporal por mês e respeita teto global.
    """
    limite = max(1, int(upload_total_max))
    if len(linhas) <= limite:
        return linhas, {
            "registros_recebidos": len(linhas),
            "registros_utilizados": len(linhas),
            "registros_descartados": 0,
        }

    por_mes: dict[str, list[dict]] = {}
    for row in linhas:
        chave_mes = _chave_mes_data(row)
        por_mes.setdefault(chave_mes, []).append(row)
    meses = sorted(por_mes.keys())
    meses_ativos = _selecionar_meses_representativos(meses, min(limite, len(meses)))

    # Dentro de cada mês, mantém ordem por recência para estabilidade operacional.
    for mes in meses_ativos:
        por_mes[mes] = sorted(por_mes[mes], key=_sort_key_recente, reverse=True)

    usadas: list[dict] = []
    idx_por_mes: dict[str, int] = {mes: 0 for mes in meses_ativos}
    while len(usadas) < limite:
        adicionou = False
        for mes in meses_ativos:
            idx = idx_por_mes[mes]
            rows_mes = por_mes[mes]
            if idx < len(rows_mes):
                usadas.append(rows_mes[idx])
                idx_por_mes[mes] = idx + 1
                adicionou = True
                if len(usadas) >= limite:
                    break
        if not adicionou:
            break

    return usadas, {
        "registros_recebidos": len(linhas),
        "registros_utilizados": len(usadas),
        "registros_descartados": len(linhas) - len(usadas),
    }


def processar_upload_frete_excel() -> tuple[dict, int]:
    """
    Processa upload de arquivo Excel (.xlsx): valida colunas, normaliza dados,
    resolve IDs de localidade e armazena em sessão.

    :return: (resposta JSON, código HTTP)
    """
    arquivo = request.files.get("file") or request.files.get("arquivo")
    execution_id = _resolve_execution_id()
    cfg = get_roberto_config()
    ok, msg = _validar_arquivo_excel(arquivo)
    if not ok:
        return jsonify({"success": False, "error": msg}), 400

    t0 = time.perf_counter()
    emitted = False

    def _emit_upload_proc(status: str, rows: int, err: str | None = None) -> None:
        nonlocal emitted
        if emitted:
            return
        emitted = True
        ms = int((time.perf_counter() - t0) * 1000)
        from app.run_cleiton_processing_governance import cleiton_register_processing_event

        cleiton_register_processing_event(
            agent="roberto",
            flow_type="upload_bi",
            processing_type="non_llm",
            rows_processed=rows,
            processing_time_ms=ms,
            status=status,
            error_summary=err,
            execution_id=execution_id,
        )

    try:
        import openpyxl
    except ImportError:
        logger.exception("openpyxl não instalado")
        _emit_upload_proc("failure", 0, "openpyxl não instalado")
        return jsonify({"success": False, "error": "Serviço de planilhas indisponível."}), 500

    # Salva o arquivo em diretório próprio de uploads do Roberto, separado do diretório
    # de sessão do Flask, para evitar conflitos com a pasta usada pelo flask_session.
    nome_base = (arquivo.filename or "upload.xlsx").rsplit(".", 1)[0]
    nome_seguro = "".join(ch for ch in nome_base if ch.isalnum() or ch in ("-", "_")) or "upload"
    nome_arquivo = f"{nome_seguro}_{datetime.utcnow().strftime('%Y%m%d%H%M%S')}_{uuid4().hex}.xlsx"
    caminho_arquivo = os.path.join(UPLOAD_DIR, nome_arquivo)

    try:
        arquivo.save(caminho_arquivo)
    except Exception as e:
        logger.exception("Falha ao salvar arquivo de upload em %s: %s", caminho_arquivo, e)
        _emit_upload_proc("failure", 0, "Falha ao salvar arquivo de upload.")
        return jsonify({"success": False, "error": "Falha ao salvar arquivo de upload."}), 500

    try:
        wb = openpyxl.load_workbook(caminho_arquivo, read_only=True, data_only=True)
        ws = wb.active
        if ws is None:
            _emit_upload_proc("failure", 0, "Planilha vazia.")
            return jsonify({"success": False, "error": "Planilha vazia."}), 400
    except Exception as e:
        logger.debug("Erro ao abrir Excel em %s: %s", caminho_arquivo, e)
        _emit_upload_proc("failure", 0, "Arquivo Excel inválido ou corrompido.")
        return jsonify({"success": False, "error": "Arquivo Excel inválido ou corrompido."}), 400

    colunas = _ler_cabecalho_normalizado(ws)
    ok, msg = _validar_colunas(colunas)
    if not ok:
        _emit_upload_proc("failure", 0, msg)
        return jsonify({"success": False, "error": msg}), 400

    idx_imposto = colunas.index(COLUNA_OPCIONAL_IMPOSTO) if COLUNA_OPCIONAL_IMPOSTO in colunas else None
    indices = {c: colunas.index(c) for c in colunas if c in COLUNAS_OBRIGATORIAS or c == COLUNA_OPCIONAL_IMPOSTO}

    # read_only: o worksheet só pode ser iterado uma vez; materializamos as linhas para duas passagens.
    linhas_planilha = list(enumerate(ws.iter_rows(min_row=2, values_only=True), start=2))

    chaves_unicas: set[str] = set()
    for _, row in linhas_planilha:
        if not any(v is not None and str(v).strip() for v in row):
            continue
        try:
            cidade_origem = _normalizar_texto(row[indices["cidade_origem"]])
            uf_origem = _normalizar_texto(row[indices["uf_origem"]])
            cidade_destino = _normalizar_texto(row[indices["cidade_destino"]])
            uf_destino = _normalizar_texto(row[indices["uf_destino"]])
            chaves_unicas.add(f"{cidade_origem}-{uf_origem}")
            chaves_unicas.add(f"{cidade_destino}-{uf_destino}")
        except (IndexError, KeyError, TypeError):
            continue

    loc_map = carregar_localidades_por_chaves(chaves_unicas)

    linhas_processadas = []
    erros_linha = []

    try:
        for num_linha, row in linhas_planilha:
            if not any(v is not None and str(v).strip() for v in row):
                continue
            try:
                cidade_origem = _normalizar_texto(row[indices["cidade_origem"]])
                uf_origem = _normalizar_texto(row[indices["uf_origem"]])
                cidade_destino = _normalizar_texto(row[indices["cidade_destino"]])
                uf_destino = _normalizar_texto(row[indices["uf_destino"]])

                # Chave de localidade padronizada: municipio-uf em minúsculo,
                # mantendo acentos e caracteres especiais. Ex.: "cariacica-es".
                chave_origem = f"{cidade_origem}-{uf_origem}"
                chave_destino = f"{cidade_destino}-{uf_destino}"

                loc_origem = loc_map.get(chave_origem)
                loc_destino = loc_map.get(chave_destino)

                if loc_origem is None or loc_origem.get("id_cidade") is None:
                    erros_linha.append(f"Linha {num_linha}: localidade de origem '{chave_origem}' não encontrada.")
                    continue
                if loc_destino is None or loc_destino.get("id_cidade") is None:
                    erros_linha.append(f"Linha {num_linha}: localidade de destino '{chave_destino}' não encontrada.")
                    continue

                data_val = row[indices["data_emissao"]]
                if hasattr(data_val, "strftime"):
                    data_emissao = data_val
                else:
                    data_str = _normalizar_texto(data_val)
                    data_emissao = None
                    if data_str:
                        for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y"):
                            try:
                                data_emissao = datetime.strptime(data_str, fmt).date()
                                break
                            except ValueError:
                                continue
                        if data_emissao is None:
                            erros_linha.append(f"Linha {num_linha}: data inválida '{data_str}'.")
                            continue

                peso = float(row[indices["peso_real"]] or 0)
                valor_nf = float(row[indices["valor_nf"]] or 0)
                valor_frete = float(row[indices["valor_frete_total"]] or 0)
                modal = _normalizar_texto(row[indices["modal"]])
                valor_imposto = None
                if idx_imposto is not None and row[idx_imposto] is not None:
                    try:
                        valor_imposto = float(row[idx_imposto])
                    except (TypeError, ValueError):
                        pass

                linhas_processadas.append({
                    "data_emissao": data_emissao.isoformat() if hasattr(data_emissao, "isoformat") else str(data_emissao),
                    "id_cidade_origem": loc_origem["id_cidade"],
                    "id_uf_origem": loc_origem.get("id_uf"),
                    "uf_origem": _uf_para_payload(loc_origem.get("uf_nome")),
                    "id_cidade_destino": loc_destino["id_cidade"],
                    "id_uf_destino": loc_destino.get("id_uf"),
                    "uf_destino": _uf_para_payload(loc_destino.get("uf_nome")),
                    "peso_real": peso,
                    "valor_nf": valor_nf,
                    "valor_frete_total": valor_frete,
                    "valor_imposto": valor_imposto,
                    "modal": modal or "",
                })
            except (IndexError, KeyError, TypeError, ValueError) as e:
                erros_linha.append(f"Linha {num_linha}: erro de dados - {e}.")
    finally:
        # Remove o arquivo físico após o processamento para manter o diretório limpo.
        try:
            if os.path.exists(caminho_arquivo):
                os.remove(caminho_arquivo)
        except Exception:
            logger.debug("Falha ao remover arquivo temporário %s", caminho_arquivo)

    if not linhas_processadas:
        _emit_upload_proc(
            "failure",
            0,
            "Nenhuma linha válida após processamento.",
        )
        return jsonify({
            "success": False,
            "error": "Nenhuma linha válida após processamento.",
            "detalhes": erros_linha[:20],
        }), 400

    linhas_utilizadas, stats_upload = _aplicar_limite_upload_total(
        linhas_processadas,
        cfg.upload_total_max,
    )
    maybe_cleanup_expired_uploads(cfg.upload_ttl_minutes)
    upload_id_anterior = session.get(SESSION_KEY_UPLOAD_REF)
    if isinstance(upload_id_anterior, str) and upload_id_anterior.strip():
        clear_upload_data_store(upload_id_anterior.strip())
    upload_ref = save_upload_data(linhas_utilizadas)
    session[SESSION_KEY_UPLOAD_REF] = upload_ref
    session.modified = True

    try:
        from app.services.cleiton_upload_billing_service import apropriar_billing_upload_roberto

        rows_processed = len(linhas_processadas)
        processing_time_ms = int((time.perf_counter() - t0) * 1000)
        idempotency_key = f"roberto-upload:{execution_id}"

        apropriar_billing_upload_roberto(
            idempotency_key=idempotency_key,
            rows_processed=rows_processed,
            processing_time_ms=processing_time_ms,
            status="success",
            execution_id=execution_id,
        )
    except Exception:
        logger.exception("Falha ao apropriar billing do upload Roberto.")
        _emit_upload_proc("success", len(linhas_processadas))

    return jsonify({
        "success": True,
        "registros": len(linhas_utilizadas),
        "registros_recebidos": stats_upload["registros_recebidos"],
        "registros_utilizados": stats_upload["registros_utilizados"],
        "registros_descartados": stats_upload["registros_descartados"],
        "upload_total_max": cfg.upload_total_max,
        "avisos": erros_linha[:15] if erros_linha else None,
    }), 200


def get_dados_upload_cliente() -> list[dict] | None:
    """
    Retorna os dados temporários do upload do cliente (lista de dicts com ids de cidade/UF
    e UF textual vindos de base_localidades). Retorna None se não houver dados ou se expirados (TTL).
    """
    upload_ref = session.get(SESSION_KEY_UPLOAD_REF)
    if not isinstance(upload_ref, str) or not upload_ref.strip():
        return None
    cfg = get_roberto_config()
    dados = read_upload_data(upload_ref.strip(), cfg.upload_ttl_minutes)
    if dados is None:
        clear_upload_data()
    return dados


def clear_upload_data() -> None:
    """Remove referência de sessão e payload temporário dedicado do upload."""
    upload_ref = session.get(SESSION_KEY_UPLOAD_REF)
    if isinstance(upload_ref, str) and upload_ref.strip():
        clear_upload_data_store(upload_ref.strip())
    session.pop(SESSION_KEY_UPLOAD_REF, None)
    session.modified = True


def roberto_clear_upload_endpoint():
    """Endpoint para o frontend limpar dados ao sair da tela /fretes (evento de navegação)."""
    clear_upload_data()
    return jsonify({"success": True}), 200
