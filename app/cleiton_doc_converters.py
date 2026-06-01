"""
Conversores técnicos de formatos documentais do Cleiton (Fase 3).

Prepara texto ou placeholder sem interpretação de negócio, layout fixo
ou extração semântica por palavra-chave.
"""
from __future__ import annotations

import csv
import io
import zipfile
from dataclasses import dataclass, field

from defusedxml import ElementTree as DefusedET
from defusedxml.common import EntitiesForbidden

from app.cleiton_doc_contracts import (
    CONTEXT_KIND_GEMINI_FILE,
    CONTEXT_KIND_TEXT,
    DOC_TYPE_CSV,
    DOC_TYPE_DOCX,
    DOC_TYPE_PDF,
    DOC_TYPE_TXT,
    DOC_TYPE_XLSX,
    DOC_TYPE_XML,
    ERROR_CONVERSION_FAILED,
    ERROR_CORRUPTED_FILE,
    ERROR_TOO_DEEP_XML,
    ERROR_TOO_MANY_COLUMNS,
    ERROR_TOO_MANY_NODES,
    ERROR_TOO_MANY_PAGES,
    ERROR_TOO_MANY_PARAGRAPHS,
    ERROR_TOO_MANY_ROWS,
    ERROR_UNSUPPORTED_TYPE,
)
from app.cleiton_doc_gemini_files import build_pdf_gemini_placeholder, estimate_pdf_page_count
from app.cleiton_doc_security import CleitonDocSecurityError
from app.services.cleiton_doc_config_service import CleitonDocConfig

TXT_SAFE_ENCODINGS = ("utf-8-sig", "utf-8", "cp1252")


@dataclass
class ConversionResult:
    prepared_context: str
    context_kind: str = CONTEXT_KIND_TEXT
    truncated: bool = False
    char_count: int = 0
    row_count: int | None = None
    column_count: int | None = None
    page_count: int | None = None
    node_count: int | None = None
    max_depth: int | None = None
    warnings: list[str] = field(default_factory=list)


def _truncate_text(text: str, max_chars: int) -> tuple[str, bool]:
    if max_chars <= 0:
        return "", True
    if len(text) <= max_chars:
        return text, False
    return text[:max_chars], True


def _raise_conversion(message: str) -> None:
    raise CleitonDocSecurityError(ERROR_CONVERSION_FAILED, message)


def _decode_txt_bytes(file_bytes: bytes) -> str:
    for encoding in TXT_SAFE_ENCODINGS:
        try:
            return file_bytes.decode(encoding)
        except UnicodeDecodeError:
            continue
    _raise_conversion("Encoding de TXT não suportado pelos encodings seguros configurados.")


def convert_txt(file_bytes: bytes, cfg: CleitonDocConfig) -> ConversionResult:
    text = _decode_txt_bytes(file_bytes)
    text, truncated = _truncate_text(text, int(cfg.txt_max_chars))
    return ConversionResult(
        prepared_context=text,
        truncated=truncated,
        char_count=len(text),
    )


def _xml_node_stats(element, depth: int = 1) -> tuple[int, int]:
    count = 1
    max_depth = depth
    for child in element:
        child_count, child_depth = _xml_node_stats(child, depth + 1)
        count += child_count
        max_depth = max(max_depth, child_depth)
    return count, max_depth


def convert_xml(file_bytes: bytes, cfg: CleitonDocConfig) -> ConversionResult:
    try:
        root = DefusedET.fromstring(file_bytes)
    except EntitiesForbidden as exc:
        raise CleitonDocSecurityError(
            ERROR_CORRUPTED_FILE,
            "XML malicioso ou entidade externa bloqueada.",
        ) from exc
    except Exception as exc:
        raise CleitonDocSecurityError(
            ERROR_CORRUPTED_FILE,
            "XML inválido ou corrompido.",
        ) from exc

    node_count, max_depth = _xml_node_stats(root)
    if node_count > int(cfg.xml_max_nodes):
        raise CleitonDocSecurityError(
            ERROR_TOO_MANY_NODES,
            "XML excede o número máximo de nós configurado.",
        )
    if max_depth > int(cfg.xml_max_depth):
        raise CleitonDocSecurityError(
            ERROR_TOO_DEEP_XML,
            "XML excede a profundidade máxima configurada.",
        )

    try:
        serialized = DefusedET.tostring(root, encoding="unicode")
    except Exception as exc:
        raise CleitonDocSecurityError(
            ERROR_CONVERSION_FAILED,
            "Falha ao serializar XML para contexto textual.",
        ) from exc

    text, truncated = _truncate_text(serialized, int(cfg.xml_max_chars))
    return ConversionResult(
        prepared_context=text,
        truncated=truncated,
        char_count=len(text),
        node_count=node_count,
        max_depth=max_depth,
    )


def _csv_rows_from_bytes(file_bytes: bytes) -> list[list[str]]:
    text = _decode_txt_bytes(file_bytes)
    reader = csv.reader(io.StringIO(text))
    return [row for row in reader]


def convert_csv(file_bytes: bytes, cfg: CleitonDocConfig) -> ConversionResult:
    try:
        rows = _csv_rows_from_bytes(file_bytes)
    except CleitonDocSecurityError:
        raise
    except Exception as exc:
        raise CleitonDocSecurityError(
            ERROR_CORRUPTED_FILE,
            "CSV inválido ou corrompido.",
        ) from exc

    row_count = len(rows)
    if row_count > int(cfg.csv_max_rows):
        raise CleitonDocSecurityError(
            ERROR_TOO_MANY_ROWS,
            "CSV excede o número máximo de linhas configurado.",
        )

    max_columns = max((len(row) for row in rows), default=0)
    if max_columns > int(cfg.csv_max_columns):
        raise CleitonDocSecurityError(
            ERROR_TOO_MANY_COLUMNS,
            "CSV excede o número máximo de colunas configurado.",
        )

    buffer = io.StringIO()
    writer = csv.writer(buffer, lineterminator="\n")
    for row in rows:
        writer.writerow(row)
    normalized = buffer.getvalue()
    text, truncated = _truncate_text(normalized, int(cfg.csv_max_chars))
    return ConversionResult(
        prepared_context=text,
        truncated=truncated,
        char_count=len(text),
        row_count=row_count,
        column_count=max_columns,
    )


def _xlsx_is_zip_bomb(file_bytes: bytes, max_uncompressed: int) -> bool:
    try:
        with zipfile.ZipFile(io.BytesIO(file_bytes)) as zf:
            total = sum(info.file_size for info in zf.infolist())
            return total > max_uncompressed
    except zipfile.BadZipFile:
        return True


def convert_xlsx(file_bytes: bytes, cfg: CleitonDocConfig) -> ConversionResult:
    if _xlsx_is_zip_bomb(file_bytes, int(cfg.excel_max_bytes) * 20):
        raise CleitonDocSecurityError(
            ERROR_CORRUPTED_FILE,
            "Arquivo XLSX inválido ou com compressão suspeita.",
        )

    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        _raise_conversion("Dependência openpyxl indisponível para XLSX.")
        raise

    try:
        workbook = load_workbook(
            filename=io.BytesIO(file_bytes),
            read_only=True,
            data_only=True,
        )
    except zipfile.BadZipFile as exc:
        raise CleitonDocSecurityError(
            ERROR_CORRUPTED_FILE,
            "Arquivo XLSX corrompido.",
        ) from exc
    except Exception as exc:
        raise CleitonDocSecurityError(
            ERROR_CORRUPTED_FILE,
            "Arquivo XLSX inválido ou corrompido.",
        ) from exc

    total_rows = 0
    max_columns = 0
    chunks: list[str] = []
    warnings: list[str] = []

    try:
        for sheet in workbook.worksheets:
            sheet_rows: list[list[str]] = []
            for row in sheet.iter_rows(values_only=True):
                values = ["" if cell is None else str(cell) for cell in row]
                while values and values[-1] == "":
                    values.pop()
                if not values:
                    continue
                sheet_rows.append(values)
                total_rows += 1
                max_columns = max(max_columns, len(values))
                if total_rows > int(cfg.excel_max_rows):
                    raise CleitonDocSecurityError(
                        ERROR_TOO_MANY_ROWS,
                        "XLSX excede o número máximo de linhas configurado.",
                    )
                if max_columns > int(cfg.excel_max_columns):
                    raise CleitonDocSecurityError(
                        ERROR_TOO_MANY_COLUMNS,
                        "XLSX excede o número máximo de colunas configurado.",
                    )

            if sheet_rows:
                buffer = io.StringIO()
                writer = csv.writer(buffer, lineterminator="\n")
                for values in sheet_rows:
                    writer.writerow(values)
                chunks.append(buffer.getvalue())
    finally:
        workbook.close()

    normalized = "\n".join(chunks)
    text, truncated = _truncate_text(normalized, int(cfg.excel_max_chars))
    return ConversionResult(
        prepared_context=text,
        truncated=truncated,
        char_count=len(text),
        row_count=total_rows,
        column_count=max_columns,
        warnings=warnings,
    )


def _docx_is_zip_bomb(file_bytes: bytes, max_uncompressed: int) -> bool:
    return _xlsx_is_zip_bomb(file_bytes, max_uncompressed)


def convert_docx(file_bytes: bytes, cfg: CleitonDocConfig) -> ConversionResult:
    if _docx_is_zip_bomb(file_bytes, int(cfg.docx_max_bytes) * 20):
        raise CleitonDocSecurityError(
            ERROR_CORRUPTED_FILE,
            "Arquivo DOCX inválido ou com compressão suspeita.",
        )

    try:
        from docx import Document
    except ImportError as exc:
        _raise_conversion("Dependência python-docx indisponível para DOCX.")
        raise

    try:
        document = Document(io.BytesIO(file_bytes))
    except zipfile.BadZipFile as exc:
        raise CleitonDocSecurityError(
            ERROR_CORRUPTED_FILE,
            "Arquivo DOCX corrompido.",
        ) from exc
    except Exception as exc:
        raise CleitonDocSecurityError(
            ERROR_CORRUPTED_FILE,
            "Arquivo DOCX inválido ou corrompido.",
        ) from exc

    paragraphs: list[str] = []
    for paragraph in document.paragraphs:
        paragraphs.append(paragraph.text or "")
        if len(paragraphs) > int(cfg.docx_max_paragraphs):
            raise CleitonDocSecurityError(
                ERROR_TOO_MANY_PARAGRAPHS,
                "DOCX excede o número máximo de parágrafos configurado.",
            )

    text_raw = "\n".join(paragraphs)
    text, truncated = _truncate_text(text_raw, int(cfg.docx_max_chars))
    return ConversionResult(
        prepared_context=text,
        truncated=truncated,
        char_count=len(text),
    )


def convert_pdf(file_bytes: bytes, cfg: CleitonDocConfig) -> ConversionResult:
    page_count = estimate_pdf_page_count(file_bytes)
    if page_count is not None and page_count > int(cfg.pdf_max_pages):
        raise CleitonDocSecurityError(
            ERROR_TOO_MANY_PAGES,
            "PDF excede o número máximo de páginas configurado.",
        )

    placeholder = build_pdf_gemini_placeholder(
        size_bytes=len(file_bytes),
        mime_type="application/pdf",
        page_count=page_count,
        max_pages=int(cfg.pdf_max_pages),
    )
    return ConversionResult(
        prepared_context=placeholder.prepared_context,
        context_kind=CONTEXT_KIND_GEMINI_FILE,
        page_count=page_count,
        warnings=list(placeholder.warnings),
        char_count=len(placeholder.prepared_context),
    )


def convert_document(
    doc_type: str,
    file_bytes: bytes,
    cfg: CleitonDocConfig,
) -> ConversionResult:
    converters = {
        DOC_TYPE_TXT: convert_txt,
        DOC_TYPE_XML: convert_xml,
        DOC_TYPE_CSV: convert_csv,
        DOC_TYPE_XLSX: convert_xlsx,
        DOC_TYPE_DOCX: convert_docx,
        DOC_TYPE_PDF: convert_pdf,
    }
    converter = converters.get(doc_type)
    if converter is None:
        raise CleitonDocSecurityError(
            ERROR_UNSUPPORTED_TYPE,
            "Tipo documental não suportado para conversão.",
        )
    return converter(file_bytes, cfg)
