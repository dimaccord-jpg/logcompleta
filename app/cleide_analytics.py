from __future__ import annotations

import csv
import io
import logging
import math
import re
import zipfile
from dataclasses import dataclass
from datetime import date, datetime
from decimal import Decimal, InvalidOperation, ROUND_HALF_UP
from typing import Any

from app.cleide_structural_validation import ALIASES_BY_REQUIRED, canonicalize_header

_CSV_DELIMITERS = (",", ";", "\t", "|")
_MAX_SAMPLE = 4096
_MAX_NUMERIC_ISSUE_SAMPLES = 10
_MAX_VALUE_PREVIEW_LEN = 32
_Q2 = Decimal("0.01")
_Q4 = Decimal("0.0001")
logger = logging.getLogger(__name__)


@dataclass(frozen=True)
class AnalyticsProcessingError(Exception):
    code: str
    message: str

    def __str__(self) -> str:
        return self.message


def build_analytics_context(
    *,
    raw_bytes: bytes,
    extension: str,
    structural_context: dict[str, Any],
    delimiter_default: str,
    max_rows: int,
    max_group_items: int = 30,
    filters: dict[str, Any] | None = None,
) -> dict[str, Any]:
    if not bool(structural_context.get("dataset_validado")):
        return _empty_analytics()

    ext = (extension or "").strip().lower()
    header = list(structural_context.get("raw_headers") or [])
    if not header:
        raise AnalyticsProcessingError(code="invalid_analytics", message="Cabecalho indisponivel para analytics.")

    col_map = _resolve_required_indexes(
        header=header,
        aliases_resolvidos=structural_context.get("aliases_resolvidos") or {},
    )
    if len(col_map) < 6:
        return _empty_analytics()

    active_filters = _prepare_filters(filters)
    state = _new_state()
    safe_max_rows = max(100, int(max_rows))
    if ext == ".csv":
        rows = _iter_csv_rows(
            raw_bytes=raw_bytes,
            delimiter_default=delimiter_default,
            delimiter_hint=structural_context.get("delimiter_detectado"),
        )
    elif ext == ".xlsx":
        rows = _iter_xlsx_rows(raw_bytes=raw_bytes, sheet_hint=structural_context.get("sheet_detectada"))
    else:
        raise AnalyticsProcessingError(code="invalid_analytics", message="Formato invalido para analytics.")

    consumed = 0
    try:
        for row in rows:
            if consumed >= safe_max_rows:
                break
            consumed += 1
            _consume_row(state=state, row=row, col_map=col_map, active_filters=active_filters)
    except (ValueError, TypeError, KeyError, IndexError, InvalidOperation, ArithmeticError) as exc:
        logger.warning("Erro tipado na agregacao Cleide: %s", exc.__class__.__name__)
        raise AnalyticsProcessingError(code="invalid_analytics", message="Falha ao processar agregacoes.") from exc

    return _finalize_state(state=state, max_group_items=max_group_items)


def build_filtered_analytics_context(
    *,
    raw_bytes: bytes,
    extension: str,
    structural_context: dict[str, Any],
    delimiter_default: str,
    max_rows: int,
    max_group_items: int = 30,
    filters: dict[str, Any] | None = None,
) -> dict[str, Any]:
    return build_analytics_context(
        raw_bytes=raw_bytes,
        extension=extension,
        structural_context=structural_context,
        delimiter_default=delimiter_default,
        max_rows=max_rows,
        max_group_items=max_group_items,
        filters=filters,
    )


def _iter_csv_rows(*, raw_bytes: bytes, delimiter_default: str, delimiter_hint: str | None = None):
    text = _decode_csv(raw_bytes)
    delimiter = delimiter_hint if delimiter_hint in _CSV_DELIMITERS else _detect_delimiter(text, delimiter_default)
    try:
        reader = csv.reader(io.StringIO(text), delimiter=delimiter)
        header_consumed = False
        for row in reader:
            clean = [str(cell).strip() if cell is not None else "" for cell in row]
            if not any(clean):
                continue
            if not header_consumed:
                header_consumed = True
                continue
            yield clean
    except csv.Error as exc:
        raise AnalyticsProcessingError(code="invalid_analytics", message=f"CSV invalido para analytics: {exc}") from exc


def _iter_xlsx_rows(*, raw_bytes: bytes, sheet_hint: str | None = None):
    try:
        import openpyxl
    except ImportError as exc:
        raise AnalyticsProcessingError(code="invalid_analytics", message="Servico XLSX indisponivel.") from exc
    try:
        workbook = openpyxl.load_workbook(io.BytesIO(raw_bytes), read_only=True, data_only=True)
    except (OSError, ValueError, KeyError, RuntimeError, zipfile.BadZipFile) as exc:
        raise AnalyticsProcessingError(
            code="invalid_analytics",
            message="Arquivo XLSX invalido para analytics.",
        ) from exc

    try:
        sheet_names = [sheet_hint] if sheet_hint in workbook.sheetnames else list(workbook.sheetnames)
        for name in sheet_names:
            sheet = workbook[name]
            header_consumed = False
            has_data = False
            for raw_row in sheet.iter_rows(values_only=True):
                row = [str(cell).strip() if cell is not None else "" for cell in raw_row]
                if not any(row):
                    continue
                if not header_consumed:
                    header_consumed = True
                    continue
                has_data = True
                yield row
            if has_data:
                break
    finally:
        workbook.close()


def _resolve_required_indexes(header: list[str], aliases_resolvidos: dict[str, str]) -> dict[str, int]:
    canonical = [canonicalize_header(col) for col in header]
    result: dict[str, int] = {}
    if aliases_resolvidos:
        by_alias = {str(v).strip(): k for k, v in aliases_resolvidos.items() if str(v).strip()}
        for idx, raw in enumerate(header):
            required = by_alias.get(str(raw).strip())
            if required and required in ALIASES_BY_REQUIRED:
                result[required] = idx
    if len(result) >= 6:
        return result
    for required, aliases in ALIASES_BY_REQUIRED.items():
        if required in result:
            continue
        options = {canonicalize_header(a) for a in aliases}
        for idx, name in enumerate(canonical):
            if name in options:
                result[required] = idx
                break
    return result


def _new_state() -> dict[str, Any]:
    return {
        "total_documentos": 0,
        "valor_total": Decimal("0"),
        "peso_total": Decimal("0"),
        "fretes_zerados": 0,
        "pesos_zerados": 0,
        "docs_valor_validos": 0,
        "transportadoras": set(),
        "ufs_origem": set(),
        "ufs_destino": set(),
        "transportadora_stats": {},
        "uf_origem_stats": {},
        "uf_destino_stats": {},
        "temporal_stats": {},
        "pareto_fretes_zerados_uf_destino": {},
        "pareto_fretes_zerados_transportadora": {},
        "invalid_numeric_rows": 0,
        "invalid_date_rows": 0,
        "negative_value_rows": 0,
        "numeric_issue_rows_total": 0,
        "numeric_issue_by_column": {"valor_frete": 0, "peso": 0, "both": 0},
        "numeric_issue_by_reason": {"empty": 0, "invalid_format": 0, "negative": 0},
        "numeric_issue_samples": [],
        "periodo_min": None,
        "periodo_max": None,
    }


def _consume_row(
    *,
    state: dict[str, Any],
    row: list[str],
    col_map: dict[str, int],
    active_filters: dict[str, Any],
) -> None:
    carrier = _safe_cell(row, col_map["transportadora"]).strip()
    uf_origem = _normalize_uf(_safe_cell(row, col_map["uf_origem"]))
    uf_destino = _normalize_uf(_safe_cell(row, col_map["uf_destino"]))
    dt = _parse_date(_safe_cell(row, col_map["data_emissao"]))
    if not _row_matches_filters(
        carrier=carrier,
        uf_origem=uf_origem,
        uf_destino=uf_destino,
        data_emissao=dt,
        active_filters=active_filters,
    ):
        return
    line_number = int(state["total_documentos"]) + 2
    state["total_documentos"] += 1
    valor_raw = _safe_cell(row, col_map["valor_frete"])
    peso_raw = _safe_cell(row, col_map["peso"])
    valor, valor_invalid, valor_negative = _parse_non_negative_decimal(
        valor_raw,
        empty_as_zero=True,
    )
    peso, peso_invalid, peso_negative = _parse_non_negative_decimal(peso_raw)

    if valor_invalid or peso_invalid:
        state["invalid_numeric_rows"] += 1
    if valor_negative or peso_negative:
        state["negative_value_rows"] += 1
    _track_numeric_issues(
        state=state,
        line_number=line_number,
        valor_raw=valor_raw,
        peso_raw=peso_raw,
        valor_invalid=valor_invalid,
        peso_invalid=peso_invalid,
        valor_negative=valor_negative,
        peso_negative=peso_negative,
    )
    if dt is None:
        state["invalid_date_rows"] += 1

    if carrier:
        state["transportadoras"].add(carrier)
        _bump_group(state["transportadora_stats"], carrier, valor=valor, peso=peso)
    if uf_origem:
        state["ufs_origem"].add(uf_origem)
        _bump_group(state["uf_origem_stats"], uf_origem, valor=valor)
    if uf_destino:
        state["ufs_destino"].add(uf_destino)
        _bump_group(state["uf_destino_stats"], uf_destino, valor=valor)

    if valor is not None:
        state["valor_total"] += valor
        state["docs_valor_validos"] += 1
        if valor == 0:
            state["fretes_zerados"] += 1
            if carrier:
                state["pareto_fretes_zerados_transportadora"][carrier] = (
                    int(state["pareto_fretes_zerados_transportadora"].get(carrier, 0)) + 1
                )
            if uf_destino:
                state["pareto_fretes_zerados_uf_destino"][uf_destino] = (
                    int(state["pareto_fretes_zerados_uf_destino"].get(uf_destino, 0)) + 1
                )
    if peso is not None:
        state["peso_total"] += peso
        if peso == 0:
            state["pesos_zerados"] += 1

    if dt is not None:
        period_min = state["periodo_min"]
        period_max = state["periodo_max"]
        if period_min is None or dt < period_min:
            state["periodo_min"] = dt
        if period_max is None or dt > period_max:
            state["periodo_max"] = dt
        key = dt.isoformat()
        bucket = state["temporal_stats"].setdefault(key, {"quantidade": 0, "valor_total": Decimal("0")})
        bucket["quantidade"] += 1
        if valor is not None:
            bucket["valor_total"] += valor


def _bump_group(target: dict[str, Any], key: str, *, valor: Decimal | None, peso: Decimal | None = None) -> None:
    bucket = target.setdefault(key, {"quantidade": 0, "valor_total": Decimal("0"), "peso_total": Decimal("0")})
    bucket["quantidade"] += 1
    if valor is not None:
        bucket["valor_total"] += valor
    if peso is not None:
        bucket["peso_total"] += peso


def _finalize_state(*, state: dict[str, Any], max_group_items: int) -> dict[str, Any]:
    total_docs = int(state["total_documentos"])
    valor_total = state["valor_total"]
    peso_total = state["peso_total"]
    docs_valor = max(0, int(state["docs_valor_validos"]))
    ticket_medio = _safe_div(valor_total, Decimal(docs_valor)) if docs_valor > 0 else Decimal("0")

    frete_zero_pct = _safe_div(Decimal(int(state["fretes_zerados"])) * Decimal("100"), Decimal(total_docs))
    peso_zero_pct = _safe_div(Decimal(int(state["pesos_zerados"])) * Decimal("100"), Decimal(total_docs))

    kpis = {
        "total_documentos": total_docs,
        "valor_total_frete": _to_float(valor_total, 2),
        "peso_total": _to_float(peso_total, 2),
        "ticket_medio_frete": _to_float(ticket_medio, 2),
        "transportadoras_unicas": len(state["transportadoras"]),
        "ufs_origem_unicas": len(state["ufs_origem"]),
        "ufs_destino_unicas": len(state["ufs_destino"]),
        "percentual_fretes_zerados": _to_float(frete_zero_pct, 2),
        "percentual_peso_zerado": _to_float(peso_zero_pct, 2),
        "periodo_dataset": {
            "inicio": state["periodo_min"].isoformat() if state["periodo_min"] else None,
            "fim": state["periodo_max"].isoformat() if state["periodo_max"] else None,
        },
    }

    transportadora_stats = _finalize_group_table(
        state["transportadora_stats"],
        max_group_items=max_group_items,
        include_peso=True,
    )
    uf_origem_stats = _finalize_group_table(
        state["uf_origem_stats"],
        max_group_items=max_group_items,
        include_peso=False,
    )
    uf_destino_stats = _finalize_group_table(
        state["uf_destino_stats"],
        max_group_items=max_group_items,
        include_peso=False,
    )

    temporal_stats = []
    for key in sorted(state["temporal_stats"].keys()):
        item = state["temporal_stats"][key]
        temporal_stats.append(
            {
                "data": key,
                "quantidade": int(item["quantidade"]),
                "valor_total": _to_float(item["valor_total"], 2),
            }
        )
    temporal_stats = temporal_stats[: max(1, int(max_group_items))]
    pareto_uf_destino = _build_pareto_table(
        state["pareto_fretes_zerados_uf_destino"],
        max_group_items=max_group_items,
    )
    pareto_transportadora = _build_pareto_table(
        state["pareto_fretes_zerados_transportadora"],
        max_group_items=max_group_items,
    )

    dataset_summary = {
        "linhas_processadas": total_docs,
        "invalid_numeric_rows": int(state["invalid_numeric_rows"]),
        "invalid_date_rows": int(state["invalid_date_rows"]),
        "negative_value_rows": int(state["negative_value_rows"]),
        "numeric_issue_details": {
            "invalid_rows_total": int(state["numeric_issue_rows_total"]),
            "by_column": dict(state["numeric_issue_by_column"]),
            "by_reason": dict(state["numeric_issue_by_reason"]),
            "samples": list(state["numeric_issue_samples"]),
        },
    }

    aggregate_counts = {
        "transportadora_stats": len(transportadora_stats),
        "uf_origem_stats": len(uf_origem_stats),
        "uf_destino_stats": len(uf_destino_stats),
        "temporal_stats": len(temporal_stats),
        "pareto_fretes_zerados_uf_destino": len(pareto_uf_destino),
        "pareto_fretes_zerados_transportadora": len(pareto_transportadora),
    }

    return {
        "analytics_ready": total_docs > 0,
        "kpis": kpis,
        "dataset_summary": dataset_summary,
        "aggregate_counts": aggregate_counts,
        "transportadora_stats": transportadora_stats,
        "uf_origem_stats": uf_origem_stats,
        "uf_destino_stats": uf_destino_stats,
        "temporal_stats": temporal_stats,
        "pareto_fretes_zerados_uf_destino": pareto_uf_destino,
        "pareto_fretes_zerados_transportadora": pareto_transportadora,
        "normalized_metrics": {
            "valor_total_frete_dec": str(valor_total),
            "peso_total_dec": str(peso_total),
            "ticket_medio_frete_dec": str(ticket_medio),
        },
        "aggregate_tables": {
            "transportadora": transportadora_stats,
            "uf_origem": uf_origem_stats,
            "uf_destino": uf_destino_stats,
            "temporal": temporal_stats,
            "pareto_fretes_zerados_uf_destino": pareto_uf_destino,
            "pareto_fretes_zerados_transportadora": pareto_transportadora,
        },
        "temporal_aggregates": temporal_stats,
    }


def _build_pareto_table(table: dict[str, int], *, max_group_items: int) -> list[dict[str, Any]]:
    safe_items = [(str(key), max(0, int(qty))) for key, qty in table.items() if str(key).strip() and int(qty or 0) > 0]
    if not safe_items:
        return []
    ordered = sorted(safe_items, key=lambda item: (-item[1], item[0]))
    capped = ordered[: max(1, int(max_group_items))]
    total = sum(item[1] for item in capped)
    if total <= 0:
        return []
    result: list[dict[str, Any]] = []
    acumulado = Decimal("0")
    total_dec = Decimal(total)
    for key, quantidade in capped:
        percentual = _safe_div(Decimal(quantidade) * Decimal("100"), total_dec)
        acumulado += percentual
        result.append(
            {
                "chave": key,
                "quantidade": quantidade,
                "percentual": _to_float(percentual, 2),
                "percentual_acumulado": _to_float(acumulado, 2),
            }
        )
    if result:
        result[-1]["percentual_acumulado"] = 100.0
    return result


def _finalize_group_table(
    table: dict[str, Any],
    *,
    max_group_items: int,
    include_peso: bool,
) -> list[dict[str, Any]]:
    ordered = sorted(
        table.items(),
        key=lambda kv: (kv[1]["valor_total"], kv[1]["quantidade"], kv[0]),
        reverse=True,
    )
    result: list[dict[str, Any]] = []
    for key, item in ordered[: max(1, int(max_group_items))]:
        row = {
            "chave": key,
            "quantidade": int(item["quantidade"]),
            "valor_total": _to_float(item["valor_total"], 2),
        }
        if include_peso:
            row["peso_total"] = _to_float(item["peso_total"], 2)
        result.append(row)
    return result


def _decode_csv(raw_bytes: bytes) -> str:
    last_exc: UnicodeDecodeError | None = None
    for encoding in ("utf-8-sig", "utf-8", "latin1"):
        try:
            return raw_bytes.decode(encoding)
        except UnicodeDecodeError as exc:
            last_exc = exc
            continue
    if last_exc is not None:
        raise AnalyticsProcessingError(
            code="invalid_analytics",
            message="Nao foi possivel decodificar CSV para analytics.",
        ) from last_exc
    raise AnalyticsProcessingError(code="invalid_analytics", message="CSV invalido para analytics.")


def _detect_delimiter(text: str, fallback: str) -> str:
    sample = text[:_MAX_SAMPLE]
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=_CSV_DELIMITERS)
        if dialect.delimiter in _CSV_DELIMITERS:
            return dialect.delimiter
    except csv.Error:
        pass
    return fallback if fallback in _CSV_DELIMITERS else ","


def _safe_cell(row: list[str], index: int) -> str:
    if index < 0 or index >= len(row):
        return ""
    return str(row[index] or "").strip()


def _track_numeric_issues(
    *,
    state: dict[str, Any],
    line_number: int,
    valor_raw: str,
    peso_raw: str,
    valor_invalid: bool,
    peso_invalid: bool,
    valor_negative: bool,
    peso_negative: bool,
) -> None:
    valor_reason = _numeric_issue_reason(raw_value=valor_raw, invalid=valor_invalid, negative=valor_negative)
    peso_reason = _numeric_issue_reason(raw_value=peso_raw, invalid=peso_invalid, negative=peso_negative)
    has_valor_issue = bool(valor_reason)
    has_peso_issue = bool(peso_reason)
    if not has_valor_issue and not has_peso_issue:
        return

    state["numeric_issue_rows_total"] += 1
    if has_valor_issue and has_peso_issue:
        state["numeric_issue_by_column"]["both"] += 1
    elif has_valor_issue:
        state["numeric_issue_by_column"]["valor_frete"] += 1
    else:
        state["numeric_issue_by_column"]["peso"] += 1

    if has_valor_issue:
        state["numeric_issue_by_reason"][valor_reason] += 1
    if has_peso_issue:
        state["numeric_issue_by_reason"][peso_reason] += 1

    samples: list[dict[str, Any]] = state["numeric_issue_samples"]
    if len(samples) >= _MAX_NUMERIC_ISSUE_SAMPLES:
        return
    if has_valor_issue:
        samples.append(
            {
                "line": int(line_number),
                "column": "valor_frete",
                "reason": valor_reason,
                "value_preview": _preview_value(valor_raw),
            }
        )
    if len(samples) >= _MAX_NUMERIC_ISSUE_SAMPLES:
        return
    if has_peso_issue:
        samples.append(
            {
                "line": int(line_number),
                "column": "peso",
                "reason": peso_reason,
                "value_preview": _preview_value(peso_raw),
            }
        )


def _numeric_issue_reason(*, raw_value: str, invalid: bool, negative: bool) -> str | None:
    if negative:
        return "negative"
    if not invalid:
        return None
    if not str(raw_value or "").strip():
        return "empty"
    return "invalid_format"


def _preview_value(value: str) -> str:
    text = str(value or "").strip()
    if not text:
        return ""
    if len(text) <= _MAX_VALUE_PREVIEW_LEN:
        return text
    return text[: _MAX_VALUE_PREVIEW_LEN - 3] + "..."


def _normalize_uf(value: str) -> str:
    text = (value or "").strip().upper()
    if len(text) > 2:
        text = text[:2]
    if not re.fullmatch(r"[A-Z]{2}", text):
        return ""
    return text


def _normalize_text_filter(value: Any) -> str:
    return str(value or "").strip().casefold()


def _prepare_filters(filters: dict[str, Any] | None) -> dict[str, Any]:
    data = filters if isinstance(filters, dict) else {}
    prepared = {
        "transportadora": _normalize_text_filter(data.get("transportadora")),
        "uf_origem": _normalize_uf(str(data.get("uf_origem") or "")),
        "uf_destino": _normalize_uf(str(data.get("uf_destino") or "")),
        "data_inicio": _parse_date(str(data.get("data_inicio") or "")),
        "data_fim": _parse_date(str(data.get("data_fim") or "")),
    }
    if prepared["data_inicio"] and prepared["data_fim"] and prepared["data_inicio"] > prepared["data_fim"]:
        prepared["data_inicio"], prepared["data_fim"] = prepared["data_fim"], prepared["data_inicio"]
    return prepared


def _row_matches_filters(
    *,
    carrier: str,
    uf_origem: str,
    uf_destino: str,
    data_emissao: date | None,
    active_filters: dict[str, Any],
) -> bool:
    carrier_filter = str(active_filters.get("transportadora") or "")
    if carrier_filter and _normalize_text_filter(carrier) != carrier_filter:
        return False
    origin_filter = str(active_filters.get("uf_origem") or "")
    if origin_filter and uf_origem != origin_filter:
        return False
    dest_filter = str(active_filters.get("uf_destino") or "")
    if dest_filter and uf_destino != dest_filter:
        return False
    date_start = active_filters.get("data_inicio")
    date_end = active_filters.get("data_fim")
    if date_start or date_end:
        if data_emissao is None:
            return False
        if date_start and data_emissao < date_start:
            return False
        if date_end and data_emissao > date_end:
            return False
    return True


def _parse_non_negative_decimal(value: str, *, empty_as_zero: bool = False) -> tuple[Decimal | None, bool, bool]:
    text = (value or "").strip()
    if not text:
        if empty_as_zero:
            return Decimal("0"), False, False
        return None, True, False
    normalized = text.replace("R$", "").replace(" ", "")
    if "," in normalized and "." in normalized:
        if normalized.rfind(",") > normalized.rfind("."):
            normalized = normalized.replace(".", "").replace(",", ".")
        else:
            normalized = normalized.replace(",", "")
    elif "," in normalized:
        normalized = normalized.replace(",", ".")
    try:
        dec = Decimal(normalized)
    except InvalidOperation:
        return None, True, False
    if not dec.is_finite():
        return None, True, False
    if dec < 0:
        return None, False, True
    return dec, False, False


def _parse_date(value: str) -> date | None:
    text = (value or "").strip()
    if not text:
        return None
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%Y/%m/%d"):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            continue
    return None


def _safe_div(num: Decimal, den: Decimal) -> Decimal:
    if den == 0:
        return Decimal("0")
    out = num / den
    if not out.is_finite():
        return Decimal("0")
    return out


def _to_float(value: Decimal, scale: int) -> float:
    quantum = _Q2 if scale <= 2 else _Q4
    val = value.quantize(quantum, rounding=ROUND_HALF_UP)
    if not math.isfinite(float(val)):
        return 0.0
    return float(val)


def _empty_analytics() -> dict[str, Any]:
    return {
        "analytics_ready": False,
        "kpis": {
            "total_documentos": 0,
            "valor_total_frete": 0.0,
            "peso_total": 0.0,
            "ticket_medio_frete": 0.0,
            "transportadoras_unicas": 0,
            "ufs_origem_unicas": 0,
            "ufs_destino_unicas": 0,
            "percentual_fretes_zerados": 0.0,
            "percentual_peso_zerado": 0.0,
            "periodo_dataset": {"inicio": None, "fim": None},
        },
        "dataset_summary": {
            "linhas_processadas": 0,
            "invalid_numeric_rows": 0,
            "invalid_date_rows": 0,
            "negative_value_rows": 0,
            "numeric_issue_details": {
                "invalid_rows_total": 0,
                "by_column": {"valor_frete": 0, "peso": 0, "both": 0},
                "by_reason": {"empty": 0, "invalid_format": 0, "negative": 0},
                "samples": [],
            },
        },
        "aggregate_counts": {
            "transportadora_stats": 0,
            "uf_origem_stats": 0,
            "uf_destino_stats": 0,
            "temporal_stats": 0,
            "pareto_fretes_zerados_uf_destino": 0,
            "pareto_fretes_zerados_transportadora": 0,
        },
        "transportadora_stats": [],
        "uf_origem_stats": [],
        "uf_destino_stats": [],
        "temporal_stats": [],
        "pareto_fretes_zerados_uf_destino": [],
        "pareto_fretes_zerados_transportadora": [],
        "normalized_metrics": {},
        "aggregate_tables": {
            "transportadora": [],
            "uf_origem": [],
            "uf_destino": [],
            "temporal": [],
            "pareto_fretes_zerados_uf_destino": [],
            "pareto_fretes_zerados_transportadora": [],
        },
        "temporal_aggregates": [],
    }
